# myapp/views.py

import os
import uuid
import json
import logging
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import random
from decimal import Decimal, InvalidOperation
from django.core.exceptions import ObjectDoesNotExist
import pandas as pd
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date
from django.views.decorators.cache import never_cache
from Church.models import Church, ChurchApplication, DenominationLiquidationAccessRequest
from .models import TemporaryOtherIncomeFile, TemporaryDonationFile
from Church.models import DenominationLiquidationAccessRequest
from .forms import ForgotPasswordForm, ResetPasswordConfirmForm, ExpenseFraudDetectionSettingsForm
from django import forms
from django.db import connection
from openai import OpenAI
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from django.core.files.base import File
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from calendar import monthrange
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from django.db.models.functions import Coalesce
from Register.models import Ministry, MinistryBudget, ReleasedBudget  # adjust imports if your app label differs
from ai_analytics.forecasting import analyze_ministry_performance
from Register.utils import analyze_ministry_performance
from django.contrib.auth import get_user_model
from datetime import date, datetime
from django.core.exceptions import ValidationError
from django.db.models.functions import TruncWeek, TruncMonth, TruncYear
from ai_analytics.optimizer import optimize_budget_distribution_ai
from .ocr import run_ocr
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.forms import formset_factory
from django.utils.dateparse import parse_date
from .models import  Fund, AuditLog
from .forms import DonationsForm, SelectDonationsNumberForm, AdminLoginForm
from collections import defaultdict
import openpyxl
from .forms import RoleAwareAuthenticationForm
from django.apps import apps
from django.contrib.auth.decorators import login_required
from openpyxl.styles import Font
from django.http import HttpResponse
from collections import defaultdict
from .models import BankAccount
from .models import Expense, ExpenseCategory, TemporaryExpenseFile
from .forms import BankInfoForm, BankBalanceForm, BankDepositForm
import calendar
from django.db.models import Max
from django.apps import apps
from datetime import datetime
from django.views.generic.edit import UpdateView
from django.views.generic.detail import DetailView
from .forms import CustomUserForm
from .forms import TreasurerLoginForm
from django.contrib.auth.views import LogoutView
from .models import Tithe, Offering, Expense, Donations
from .utils import analyze_receipt_with_openai
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
import traceback
import copy
from .forms import BankMovementUnifiedForm
from .models import CashBankMovement, TemporaryBankMovementFile
from .models import CashBankMovement, Tithe, Offering, Donations, OtherIncome, Expense
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.mail import send_mail
from django.conf import settings
from django.views import View
from .forms import TitheForm, OfferingForm, DonationsForm
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib import messages
from .forms import SelectExpensesNumberForm
from django.views.generic import TemplateView
from .models import HomePageContent
from .forms import HomeContentForm
import json
import base64
from django.views.decorators.http import require_POST
from .models import PendingReference, ApprovedReference  # add
from .forms import PendingReferenceForm
from decimal import Decimal
from django.views.generic import ListView, UpdateView, DeleteView
from django.shortcuts import get_object_or_404
from datetime import timedelta
from django.http import JsonResponse
from .models import CustomUser
from .forms import CustomUserRegistrationForm
from .models import Tithe, Offering, Expense, Donations, AccountingSettings
from .forms import ExpenseForm
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from .forms import MemberLoginForm
from django.urls import reverse_lazy
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.views.generic import FormView
from django.views.generic import CreateView
from django.template.loader import render_to_string
from django.template import TemplateDoesNotExist
from django.db.models import Q
from django.views.generic import ListView
from django.utils.html import strip_tags
from .forms import MinistryForm, BudgetForm, BudgetExpenseForm
import re
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
import json
import fitz
from .models import OtherIncome, OtherIncomeCategory
from .forms import OtherIncomeForm, OtherIncomeCategoryForm
from .forms import BudgetReleaseRequestForm, BankWithdrawForm
from .models import ApprovedReference, ThankYouLetter
from .forms import DirectBankReceiptForm
from .forms import OfferingForm, LooseOfferingForm, EnvelopeFormSet
from .models import CustomUser
from .models import Expense, ExpenseCategory  # <--- Make sure ExpenseCategory is added here
from .forms import ExpenseForm, ExpenseCategoryForm
from .models import DonationCategory
from .forms import DonationCategoryForm
from ai_analytics.models import OptimizerRunLog
from django.db.models import Sum, Case, When, Value, CharField
from django.db.models.functions import ExtractYear, TruncMonth
from ai_analytics.priority_predictor import predict_priority

from ai_analytics.services.monthly_yearly_analysis import get_monthly_yearly_dashboard_summary
from ai_analytics.services.openai_budget_insights import build_budget_insights
from .models import (
    BudgetBalance,  # <--- MAKE SURE THIS IS HERE
    DeclinedReleaseRequest
)
from ai_analytics.services.monthly_yearly_analysis import (
    get_monthly_yearly_dashboard_summary,
)
from .models import (
    MinistryBudget,
    BudgetReleaseRequest, ApprovedReleaseRequest, DeclinedReleaseRequest,
    ReleasedBudget, BudgetExpense, CashTransaction
)
from .models import Tithe, Offering, Donations, OtherIncome, BudgetExpense
from ai_analytics.services.openai_budget_insights import build_budget_insights
from .forms import (
    # ... other forms ...
    BudgetExpenseForm,
    BudgetExpenseFormSet,  # <--- Add this line
)
from .forms import ThankYouLetterForm
from ai_analytics.forecasting import (
    calculate_financial_health,
    get_projected_expenses,
    analyze_ministry_performance
)
# --- IMPORT OPTIMIZER HERE (CRITICAL) ---
from ai_analytics.optimizer import optimize_budget_distribution
from django.db import transaction, IntegrityError
from ai_analytics.priority_model_runtime import predict_priority
from .models import Ministry
from .utils import verify_receipt_with_openai
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import render, redirect
from django.views import View
from django.utils.crypto import get_random_string
from django.contrib import messages
from .models import AccountingSettings
from datetime import date as date_cls, datetime as datetime_cls, timedelta
from .forms import AccountingLockForm
from django.http import HttpResponseForbidden
logger = logging.getLogger(__name__)

class IsDenominationAdmin(UserPassesTestMixin):
    def test_func(self):
        u = self.request.user
        return u.is_authenticated and u.user_type in ['DenominationAdmin', 'Admin'] and u.denomination_id

class FinanceRoleRequiredMixin(LoginRequiredMixin):
    """
    Allows access to: Admin, ChurchAdmin, and Treasurer.
    Redirects unauthorized users to the home dashboard.
    """

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        # Define authorized roles
        allowed_roles = ['Admin', 'ChurchAdmin', 'Treasurer', 'DenominationAdmin','Pastor']

        if not user.is_authenticated or getattr(user, "user_type", None) not in allowed_roles:
            messages.error(request, "Access Denied: You must be a Treasurer or Admin.")
            return redirect("home")

        return super().dispatch(request, *args, **kwargs)

class IsChurchAdmin(UserPassesTestMixin):
    def test_func(self):
        u = self.request.user
        return u.is_authenticated and u.user_type in ['ChurchAdmin', 'Admin'] and u.church_id

class RegisterView(LoginRequiredMixin, FormView):
    template_name = 'register.html'
    form_class = CustomUserRegistrationForm
    success_url = reverse_lazy('register')

    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type not in ['Admin', 'ChurchAdmin']:
            messages.error(request, "You do not have permission to access the registration page.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request_user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        user = form.save(commit=False)

        if self.request.user.is_authenticated:
            user.created_by = self.request.user

            # Auto-assign church from creator
            if getattr(self.request.user, 'church_id', None):
                user.church = self.request.user.church

            # Optional: auto-assign denomination too
            if getattr(self.request.user, 'denomination_id', None):
                user.denomination = self.request.user.denomination

        try:
            user.save()
            messages.success(self.request, f"Account for {user.username} created successfully!")
            return super().form_valid(form)

        except IntegrityError:
            form.add_error('email', "Database Error: This email is already registered.")
            return self.form_invalid(form)

        except ValidationError as e:
            if hasattr(e, 'message_dict'):
                for field, errors in e.message_dict.items():
                    if field in form.fields:
                        for err in errors:
                            form.add_error(field, err)
                    else:
                        for err in errors:
                            form.add_error(None, err)
            else:
                form.add_error(None, e)
            return self.form_invalid(form)


class CustomLoginView(LoginView):
    template_name = 'landing_Page.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        user = self.request.user

        if user.user_type == user.UserType.MEMBER:
            return reverse('weekly_financial_report')

        return self.get_redirect_url() or reverse('home')



class HomeView(View):
    template_name = 'success.html'

    def can_edit_homepage(self, user):
        return (
            user.is_authenticated
            and user.user_type in ['Admin', 'ChurchAdmin', 'DenominationAdmin']
            and hasattr(user, 'church')
            and user.church
        )

    def get_content(self, user):
        """Helper to get the correct content based on the user."""
        if user.is_authenticated and hasattr(user, 'church') and user.church:
            # Get content specific to the logged-in user's church
            content, _ = HomePageContent.objects.get_or_create(church=user.church)
            return content
        else:
            # Public / Fallback content
            class PublicContent:
                title = "Welcome to Church Financial Management System"
                subtitle = "Disciple making Disciples"
                hero_image = None

            return PublicContent()

    def get(self, request):
        content = self.get_content(request.user)
        form = None

        # Allow Admin, ChurchAdmin, and DenominationAdmin to edit
        if self.can_edit_homepage(request.user):
            if hasattr(content, 'pk'):
                form = HomeContentForm(instance=content)

        # Retrieve session flags for welcome popup
        show_guide = request.session.pop('show_welcome_guide', False)
        org_name = request.session.pop('registered_org_name', 'Your Organization')

        context = {
            'content': content,
            'form': form,
            'show_welcome_guide': show_guide,
            'registered_org_name': org_name
        }

        return render(request, self.template_name, context)

    def post(self, request):
        # Allow Admin, ChurchAdmin, and DenominationAdmin to edit
        if not self.can_edit_homepage(request.user):
            messages.error(request, "You are not authorized to edit this page.")
            return redirect('home')

        content = self.get_content(request.user)

        # Bind the form to the existing content
        form = HomeContentForm(request.POST, request.FILES, instance=content)

        if form.is_valid():
            form.save()
            messages.success(request, "Home page updated successfully!")
            return redirect('home')

        # If form is invalid, re-render the page with errors
        return render(request, self.template_name, {
            'content': content,
            'form': form,
            'show_welcome_guide': False,
            'registered_org_name': 'Your Organization',
        })

def _is_tithe_fraud_detection_enabled(user) -> bool:
    church = getattr(user, "church", None)
    if not church:
        return False

    try:
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=church)
        return bool(settings_obj.enable_tithe_fraud_detection)
    except Exception as e:
        print("TITHE FRAUD TOGGLE READ ERROR:", repr(e))
        return False


def _get_uploaded_tithe_proof(request, cleaned_data=None):
    """
    Tries common file field names.
    Adjust if your TitheForm uses a different file field name.
    """
    cleaned_data = cleaned_data or {}

    return (
        cleaned_data.get("file")
        or cleaned_data.get("proof_file")
        or cleaned_data.get("receipt")
        or cleaned_data.get("receipt_image")
        or request.FILES.get("file")
        or request.FILES.get("proof_file")
        or request.FILES.get("receipt")
        or request.FILES.get("receipt_image")
    )


def _verify_tithe_receipt(uploaded_file, *, amount, tithe_date, description, church, cleaned_data=None):
    """
    OCR can still handle PDF separately.
    Fraud detection checks image proofs only.
    """
    cleaned_data = cleaned_data or {}

    if not uploaded_file:
        return True, "No proof uploaded; fraud detection skipped."

    ext = (os.path.splitext(getattr(uploaded_file, "name", ""))[1] or "").lower()

    if ext == ".pdf":
        return True, "PDF proof detected. Fraud detection skipped; OCR remains available."

    if ext not in (".jpg", ".jpeg", ".png"):
        return True, f"Unsupported proof type '{ext}'. Fraud detection skipped."

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        is_valid, reason = verify_receipt_with_openai(
            image_file=uploaded_file,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": tithe_date.isoformat() if tithe_date else "",
                "context": {
                    "transaction_type": "tithe",
                    "description": description or "",
                    "church_id": getattr(church, "id", None),
                    "member_name": str(
                        cleaned_data.get("member")
                        or cleaned_data.get("user")
                        or cleaned_data.get("giver_name")
                        or ""
                    ),
                }
            }
        )
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return is_valid, reason


@method_decorator(never_cache, name='dispatch')
class TitheCreateView(FinanceRoleRequiredMixin, View):
    template_name = 'EnterTithes.html'
    unified_template_name = 'unified_income_entry.html'

    REVIEW_ID_SESSION_KEY = "review_tithe_id"
    REVIEW_SOURCE_SESSION_KEY = "review_tithe_source"
    REVIEW_PANE_SESSION_KEY = "review_tithe_pane"

    UNIFIED_PANE = "pane-tithe"

    def _clear_review_session(self, request):
        request.session.pop(self.REVIEW_ID_SESSION_KEY, None)
        request.session.pop(self.REVIEW_SOURCE_SESSION_KEY, None)
        request.session.pop(self.REVIEW_PANE_SESSION_KEY, None)
        request.session.modified = True

    def _render_form(self, request, tithe_form):
        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return render(
                request,
                self.unified_template_name,
                build_unified_income_context(
                    request,
                    tithe_form=tithe_form,
                    active_pane=self.UNIFIED_PANE,
                ),
            )
        return render(request, self.template_name, {'tithe_form': tithe_form})

    def get(self, request):
        # Delete abandoned draft when user returns from review page
        staged_id = request.session.get(self.REVIEW_ID_SESSION_KEY)

        if staged_id:
            try:
                Tithe.objects.filter(id=staged_id).delete()
            except Exception:
                pass

            self._clear_review_session(request)

        tithe_form = TitheForm()
        return render(request, self.template_name, {'tithe_form': tithe_form})

    def post(self, request):
        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)

        # OCR branch
        if request.POST.get("action") == "OCR":
            uploaded = (
                request.FILES.get("receipt")
                or request.FILES.get("file")
                or request.FILES.get("receipt_image")
                or request.FILES.get("proof_file")
            )

            if not uploaded:
                return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

            try:
                result = analyze_receipt_with_openai(uploaded)

                if isinstance(result, dict) and "error" in result:
                    return JsonResponse({"success": False, "error": result["error"]}, status=500)

                return JsonResponse({"success": True, "data": result})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)

        tithe_form = TitheForm(request.POST, request.FILES)

        if tithe_form.is_valid():
            try:
                tithe = tithe_form.save(commit=False)
                tithe.created_by = request.user
                tithe.user = tithe_form.cleaned_data.get('user', None)

                church = getattr(request.user, "church", None)
                if church:
                    tithe.church = church
                else:
                    messages.error(request, "Access Denied: You are not linked to a church.")
                    return redirect("home")

                fraud_detection_enabled = _is_tithe_fraud_detection_enabled(request.user)
                uploaded_file = _get_uploaded_tithe_proof(request, tithe_form.cleaned_data)

                # Require proof file when fraud detection is ON
                if fraud_detection_enabled and not uploaded_file:
                    proof_field = None
                    for candidate in ["file", "proof_file", "receipt", "receipt_image"]:
                        if candidate in tithe_form.fields:
                            proof_field = candidate
                            break

                    if proof_field:
                        tithe_form.add_error(
                            proof_field,
                            "Receipt file is required while tithe fraud detection is enabled."
                        )

                    messages.error(
                        request,
                        "Please upload a proof file for tithe while fraud detection is enabled."
                    )
                    return self._render_form(request, tithe_form)

                # Run fraud detection only when proof exists
                if fraud_detection_enabled and uploaded_file:
                    is_valid, reason = _verify_tithe_receipt(
                        uploaded_file=uploaded_file,
                        amount=tithe_form.cleaned_data.get("amount"),
                        tithe_date=tithe_form.cleaned_data.get("date"),
                        description=tithe_form.cleaned_data.get("description", ""),
                        church=church,
                        cleaned_data=tithe_form.cleaned_data,
                    )

                    if not is_valid:
                        proof_field = None
                        for candidate in ["file", "proof_file", "receipt", "receipt_image"]:
                            if candidate in tithe_form.fields:
                                proof_field = candidate
                                break

                        if proof_field:
                            tithe_form.add_error(
                                proof_field,
                                f"Fraud detection failed: {reason}"
                            )

                        messages.error(request, f"Potential Fraud Detected: {reason}")
                        return self._render_form(request, tithe_form)

                tithe.full_clean()
                tithe.save()

                request.session[self.REVIEW_ID_SESSION_KEY] = tithe.id
                request.session[self.REVIEW_SOURCE_SESSION_KEY] = request.POST.get("entry_source", "")
                request.session[self.REVIEW_PANE_SESSION_KEY] = request.POST.get("active_pane", self.UNIFIED_PANE)
                request.session.modified = True

                messages.success(request, "Draft saved. Please review and confirm.")
                return redirect('review_tithe')

            except ValidationError as e:
                if hasattr(e, 'message_dict') and '__all__' in e.message_dict:
                    error_msg = e.message_dict['__all__'][0]
                elif hasattr(e, 'messages') and e.messages:
                    error_msg = e.messages[0]
                else:
                    error_msg = str(e)

                messages.error(request, error_msg)
                return self._render_form(request, tithe_form)

            except Exception as e:
                messages.error(request, f"System Error: {e}")
                return self._render_form(request, tithe_form)

        messages.error(request, "Please correct the errors below.")
        return self._render_form(request, tithe_form)


@method_decorator(never_cache, name='dispatch')
class ReviewTitheView(FinanceRoleRequiredMixin, View):
    template_name = 'review_tithes.html'

    REVIEW_ID_SESSION_KEY = "review_tithe_id"
    REVIEW_SOURCE_SESSION_KEY = "review_tithe_source"
    REVIEW_PANE_SESSION_KEY = "review_tithe_pane"

    DEFAULT_PANE = "pane-tithe"

    def _clear_review_session(self, request):
        request.session.pop(self.REVIEW_ID_SESSION_KEY, None)
        request.session.pop(self.REVIEW_SOURCE_SESSION_KEY, None)
        request.session.pop(self.REVIEW_PANE_SESSION_KEY, None)
        request.session.modified = True

    def _entry_redirect(self, request, clear=False):
        source = request.session.get(self.REVIEW_SOURCE_SESSION_KEY, "")
        pane = request.session.get(self.REVIEW_PANE_SESSION_KEY) or self.DEFAULT_PANE

        if clear:
            self._clear_review_session(request)

        if source == "unified_income_entry":
            set_unified_income_active_pane(request, pane)
            return redirect_unified_income_entry(pane)

        return redirect("add_tithe")

    def _get_tithe(self, request):
        tithe_id = request.session.get(self.REVIEW_ID_SESSION_KEY)
        if not tithe_id:
            return None

        return Tithe.objects.filter(
            id=tithe_id,
            church=getattr(request.user, "church", None)
        ).first()

    def get(self, request):
        tithe = self._get_tithe(request)

        if not tithe:
            self._clear_review_session(request)
            messages.warning(request, "No draft found. Please enter tithe details first.")
            return self._entry_redirect(request, clear=False)

        return render(request, self.template_name, {"tithe": tithe})

    def post(self, request):
        tithe = self._get_tithe(request)

        if not tithe:
            self._clear_review_session(request)
            messages.warning(request, "No draft found. Please enter tithe details first.")
            return self._entry_redirect(request, clear=False)

        if "confirm" in request.POST:
            messages.success(
                request,
                f"Tithe confirmed and saved successfully. Amount saved: ₱{Decimal(str(tithe.amount or '0.00')):,.2f}."
            )
            return self._entry_redirect(request, clear=True)

        if "cancel" in request.POST:
            try:
                tithe.delete()
                messages.info(request, "Entry cancelled and deleted.")
                return self._entry_redirect(request, clear=True)

            except Exception as e:
                messages.error(request, f"System Error: {e}")
                return redirect("review_tithe")

        return self._entry_redirect(request, clear=False)

def _is_offering_fraud_detection_enabled(user) -> bool:
    """
    Returns True if the current user's church has offering fraud detection enabled.
    Safely creates AccountingSettings if missing.
    """
    church = getattr(user, "church", None)
    if not church:
        return False

    try:
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=church)
        return bool(settings_obj.enable_offering_fraud_detection)
    except Exception:
        logger.exception("OFFERING FRAUD TOGGLE READ ERROR")
        return False


def _verify_offering_proof(
    uploaded_file,
    *,
    amount,
    offering_date,
    description,
    church,
    is_batch=False
):
    """
    Verifies an offering proof image using your OpenAI receipt checker.

    Rules:
    - No file     -> skip
    - PDF         -> skip fraud detection, allow OCR separately
    - JPG/JPEG/PNG -> run fraud detection
    - other file  -> skip

    Returns:
        (is_valid: bool, reason: str)
    """
    if not uploaded_file:
        return True, "No proof uploaded; fraud detection skipped."

    ext = (os.path.splitext(getattr(uploaded_file, "name", ""))[1] or "").lower()

    if ext == ".pdf":
        return True, "PDF proof detected. Fraud detection skipped; OCR remains available."

    if ext not in (".jpg", ".jpeg", ".png"):
        return True, f"Unsupported proof type '{ext}'. Fraud detection skipped."

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        is_valid, reason = verify_receipt_with_openai(
            image_file=uploaded_file,
            check_type="TRANSACTION",
            expected_data={
                "amount": str(amount or ""),
                "date": offering_date.isoformat() if offering_date else "",
                "context": {
                    "transaction_type": "offering_batch" if is_batch else "offering",
                    "description": description or "",
                    "church_id": getattr(church, "id", None),
                }
            }
        )
    except Exception as e:
        logger.exception("OFFERING FRAUD DETECTION ERROR")
        return False, f"System Error: {e}"
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return bool(is_valid), str(reason or "")


def _extract_offering_proof_for_ocr(request):
    """
    Helper for OCR AJAX requests.
    Tries common field names from both single and batch offering forms.
    """
    return (
        request.FILES.get("single-proof_document")
        or request.FILES.get("loose-proof_document")
        or request.FILES.get("proof_document")
        or request.FILES.get("receipt")
        or request.FILES.get("file")
    )


def _run_offering_ocr(uploaded_file):
    """
    Runs OCR on the uploaded offering proof.
    Uses your existing analyze_receipt_with_openai() helper.

    Returns:
        {
            "success": bool,
            "data": {...} OR None,
            "error": str OR None
        }
    """
    if not uploaded_file:
        return {"success": False, "data": None, "error": "No file uploaded"}

    try:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

        result = analyze_receipt_with_openai(uploaded_file)

        if isinstance(result, dict) and "error" in result:
            return {"success": False, "data": None, "error": result["error"]}

        return {"success": True, "data": result, "error": None}

    except Exception as e:
        logger.exception("OFFERING OCR ERROR")
        return {"success": False, "data": None, "error": str(e)}
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass


@method_decorator(never_cache, name="dispatch")
class UnifiedOfferingView(LoginRequiredMixin, View):
    template_name = "add_offering.html"
    unified_template_name = "unified_income_entry.html"

    STAGED_IDS_SESSION_KEY = "staged_offering_ids"
    STAGED_SOURCE_SESSION_KEY = "staged_offering_source"
    STAGED_PANE_SESSION_KEY = "staged_offering_pane"

    UNIFIED_PANE_SINGLE = "pane-offering-single"
    UNIFIED_PANE_BATCH = "pane-offering-batch"

    def get_context(self, request):
        single_form = OfferingForm(prefix="single", user=request.user)
        loose_form = LooseOfferingForm(
            initial={"date": timezone.now().date()},
            prefix="loose"
        )
        batch_formset = EnvelopeFormSet(
            prefix="envelope",
            form_kwargs={"user": request.user}
        )

        return {
            "single_form": single_form,
            "loose_form": loose_form,
            "batch_formset": batch_formset,
        }

    def _get_pane_for_form_type(self, form_type):
        return self.UNIFIED_PANE_BATCH if form_type == "batch" else self.UNIFIED_PANE_SINGLE

    def _cleanup_staged_offerings(self, staged_ids):
        if staged_ids:
            try:
                Offering.objects.filter(id__in=staged_ids).delete()
            except Exception:
                pass

    def _clear_staging_session(self, request):
        request.session.pop(self.STAGED_IDS_SESSION_KEY, None)
        request.session.pop(self.STAGED_SOURCE_SESSION_KEY, None)
        request.session.pop(self.STAGED_PANE_SESSION_KEY, None)
        request.session.modified = True

    def _render_form(self, request, form_type=None, single_form=None, loose_form=None, batch_formset=None):
        if single_form is None:
            single_form = OfferingForm(prefix="single", user=request.user)

        if loose_form is None:
            loose_form = LooseOfferingForm(
                initial={"date": timezone.now().date()},
                prefix="loose"
            )

        if batch_formset is None:
            batch_formset = EnvelopeFormSet(
                prefix="envelope",
                form_kwargs={"user": request.user}
            )

        active_pane = self._get_pane_for_form_type(form_type)

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, active_pane)
            return render(
                request,
                self.unified_template_name,
                build_unified_income_context(
                    request,
                    single_form=single_form,
                    loose_form=loose_form,
                    batch_formset=batch_formset,
                    active_pane=active_pane,
                ),
            )

        return render(request, self.template_name, {
            "single_form": single_form,
            "loose_form": loose_form,
            "batch_formset": batch_formset,
        })

    def get(self, request):
        staged_ids = request.session.get(self.STAGED_IDS_SESSION_KEY, [])
        if staged_ids:
            self._cleanup_staged_offerings(staged_ids)
            self._clear_staging_session(request)

        return render(request, self.template_name, self.get_context(request))

    def post(self, request):
        current_church = getattr(request.user, "church", None)

        if not current_church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if request.POST.get("action") == "OCR":
            uploaded = _extract_offering_proof_for_ocr(request)
            ocr_result = _run_offering_ocr(uploaded)

            if not ocr_result["success"]:
                return JsonResponse(
                    {"success": False, "error": ocr_result["error"]},
                    status=400
                )

            return JsonResponse({"success": True, "data": ocr_result["data"]})

        form_type = request.POST.get("form_type")
        active_pane = self._get_pane_for_form_type(form_type)
        fraud_detection_enabled = _is_offering_fraud_detection_enabled(request.user)
        saved_ids = []

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, active_pane)

        try:
            if form_type == "single":
                form = OfferingForm(
                    request.POST,
                    request.FILES,
                    prefix="single",
                    user=request.user
                )

                if not form.is_valid():
                    messages.error(request, "Please check the form for errors.")
                    return self._render_form(
                        request,
                        form_type="single",
                        single_form=form,
                    )

                offering = form.save(commit=False)
                uploaded_file = form.cleaned_data.get("proof_document")

                if fraud_detection_enabled and not uploaded_file:
                    form.add_error(
                        "proof_document",
                        "Proof file is required while offering fraud detection is enabled."
                    )
                    messages.error(
                        request,
                        "Please upload a proof file for this offering while fraud detection is enabled."
                    )
                    return self._render_form(
                        request,
                        form_type="single",
                        single_form=form,
                    )

                if fraud_detection_enabled and uploaded_file:
                    is_valid, reason = _verify_offering_proof(
                        uploaded_file=uploaded_file,
                        amount=offering.amount,
                        offering_date=offering.date,
                        description=offering.description,
                        church=current_church,
                        is_batch=False,
                    )

                    if not is_valid:
                        form.add_error(
                            "proof_document",
                            f"Fraud detection failed: {reason}"
                        )
                        messages.error(
                            request,
                            f"Potential Fraud Detected in offering proof: {reason}"
                        )
                        return self._render_form(
                            request,
                            form_type="single",
                            single_form=form,
                        )

                offering.created_by = request.user
                offering.church = current_church
                offering.status = "Pending"
                offering.full_clean()
                offering.save()
                saved_ids.append(offering.id)

            elif form_type == "batch":
                loose_form = LooseOfferingForm(
                    request.POST,
                    request.FILES,
                    prefix="loose"
                )
                batch_formset = EnvelopeFormSet(
                    request.POST,
                    request.FILES,
                    prefix="envelope",
                    form_kwargs={"user": request.user}
                )

                if not (loose_form.is_valid() and batch_formset.is_valid()):
                    messages.error(request, "Please check the batch offering form for errors.")
                    return self._render_form(
                        request,
                        form_type="batch",
                        loose_form=loose_form,
                        batch_formset=batch_formset,
                    )

                main_date = loose_form.cleaned_data["date"]
                loose_amt = loose_form.cleaned_data.get("loose_total") or Decimal("0.00")
                batch_proof = loose_form.cleaned_data.get("proof_document")

                valid_rows = []
                envelope_total = Decimal("0.00")

                for f in batch_formset:
                    if not getattr(f, "cleaned_data", None):
                        continue
                    if f.cleaned_data.get("DELETE", False):
                        continue

                    amount = f.cleaned_data.get("amount")
                    if amount:
                        valid_rows.append(f)
                        envelope_total += amount

                batch_total = loose_amt + envelope_total

                if batch_total <= 0:
                    messages.error(request, "Please enter at least one offering amount.")
                    return self._render_form(
                        request,
                        form_type="batch",
                        loose_form=loose_form,
                        batch_formset=batch_formset,
                    )

                if fraud_detection_enabled and not batch_proof:
                    loose_form.add_error(
                        "proof_document",
                        "A tally sheet or deposit slip is required while offering fraud detection is enabled."
                    )
                    messages.error(
                        request,
                        "Please upload one proof file for this batch while fraud detection is enabled."
                    )
                    return self._render_form(
                        request,
                        form_type="batch",
                        loose_form=loose_form,
                        batch_formset=batch_formset,
                    )

                if fraud_detection_enabled and batch_proof:
                    is_valid, reason = _verify_offering_proof(
                        uploaded_file=batch_proof,
                        amount=batch_total,
                        offering_date=main_date,
                        description=f"Batch offering: loose={loose_amt}, envelopes={envelope_total}",
                        church=current_church,
                        is_batch=True,
                    )

                    if not is_valid:
                        loose_form.add_error(
                            "proof_document",
                            f"Fraud detection failed: {reason}"
                        )
                        messages.error(
                            request,
                            f"Potential Fraud Detected in batch offering proof: {reason}"
                        )
                        return self._render_form(
                            request,
                            form_type="batch",
                            loose_form=loose_form,
                            batch_formset=batch_formset,
                        )

                with transaction.atomic():
                    proof_attached = False

                    if loose_amt > 0:
                        loose = Offering(
                            amount=loose_amt,
                            date=main_date,
                            description="Sunday Loose Offering",
                            church=current_church,
                            created_by=request.user,
                            status="Pending",
                        )

                        if batch_proof and not proof_attached:
                            try:
                                batch_proof.seek(0)
                            except Exception:
                                pass
                            loose.proof_document = batch_proof
                            proof_attached = True

                        loose.full_clean()
                        loose.save()
                        saved_ids.append(loose.id)

                    for f in valid_rows:
                        member_obj = f.cleaned_data.get("member")
                        row_type = f.cleaned_data.get("type") or "Cash"
                        check_number = (f.cleaned_data.get("check_number") or "").strip()

                        desc = f"Envelope: {row_type}"
                        if row_type == "Check" and check_number:
                            desc = f"Envelope: {row_type} #{check_number}"

                        env = Offering(
                            amount=f.cleaned_data.get("amount"),
                            date=f.cleaned_data.get("date") or main_date,
                            user=member_obj,
                            description=desc,
                            church=current_church,
                            created_by=request.user,
                            status="Pending",
                        )

                        if batch_proof and not proof_attached:
                            try:
                                batch_proof.seek(0)
                            except Exception:
                                pass
                            env.proof_document = batch_proof
                            proof_attached = True

                        env.full_clean()
                        env.save()
                        saved_ids.append(env.id)

            else:
                messages.error(request, "Invalid offering form submission.")
                return self._render_form(request, form_type="single")

            if saved_ids:
                request.session[self.STAGED_IDS_SESSION_KEY] = saved_ids
                request.session[self.STAGED_SOURCE_SESSION_KEY] = request.POST.get("entry_source", "")
                request.session[self.STAGED_PANE_SESSION_KEY] = active_pane
                request.session.modified = True
                return redirect("review_offerings")

            messages.warning(request, "No offering data was saved.")
            return self._render_form(request, form_type=form_type or "single")

        except ValidationError as e:
            msg = str(e)

            if hasattr(e, "message_dict") and "__all__" in e.message_dict:
                msg = e.message_dict["__all__"][0]
            elif hasattr(e, "messages") and e.messages:
                msg = e.messages[0]

            messages.error(request, msg)

        except Exception as e:
            messages.error(request, f"Error: {e}")

        if form_type == "single":
            return self._render_form(
                request,
                form_type="single",
                single_form=OfferingForm(
                    request.POST,
                    request.FILES,
                    prefix="single",
                    user=request.user
                ),
            )

        if form_type == "batch":
            return self._render_form(
                request,
                form_type="batch",
                loose_form=LooseOfferingForm(
                    request.POST,
                    request.FILES,
                    prefix="loose"
                ),
                batch_formset=EnvelopeFormSet(
                    request.POST,
                    request.FILES,
                    prefix="envelope",
                    form_kwargs={"user": request.user}
                ),
            )

        return self._render_form(request, form_type="single")


@method_decorator(never_cache, name="dispatch")
class UnifiedOfferingReviewView(LoginRequiredMixin, View):
    template_name = "review_offerings.html"
    STAGED_IDS_SESSION_KEY = "staged_offering_ids"
    STAGED_SOURCE_SESSION_KEY = "staged_offering_source"
    STAGED_PANE_SESSION_KEY = "staged_offering_pane"

    DEFAULT_PANE = "pane-offering-single"

    def _entry_redirect(self, request, clear=False):
        source = request.session.get(self.STAGED_SOURCE_SESSION_KEY, "")
        pane = request.session.get(self.STAGED_PANE_SESSION_KEY) or self.DEFAULT_PANE

        if clear:
            request.session.pop(self.STAGED_IDS_SESSION_KEY, None)
            request.session.pop(self.STAGED_SOURCE_SESSION_KEY, None)
            request.session.pop(self.STAGED_PANE_SESSION_KEY, None)
            request.session.modified = True

        if source == "unified_income_entry":
            set_unified_income_active_pane(request, pane)
            return redirect_unified_income_entry(pane)

        return redirect("add_offerings")

    def _get_staged_queryset(self, request):
        staged_ids = request.session.get(self.STAGED_IDS_SESSION_KEY, [])
        return Offering.objects.filter(
            id__in=staged_ids,
            church=getattr(request.user, "church", None),
        ).order_by("id")

    def _build_success_message(self, offerings):
        saved_count = offerings.count()
        total_amount = offerings.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        if saved_count <= 0:
            return "No offerings were saved."

        if saved_count == 1:
            return f"Offering saved successfully. Amount saved: ₱{total_amount:,.2f}."

        return (
            f"{saved_count} offerings saved successfully. "
            f"Total amount saved: ₱{total_amount:,.2f}."
        )

    def get(self, request):
        staged_ids = request.session.get(self.STAGED_IDS_SESSION_KEY, [])

        if not staged_ids:
            messages.warning(request, "No pending offerings to review.")
            return self._entry_redirect(request, clear=True)

        offerings = self._get_staged_queryset(request)

        if not offerings.exists():
            messages.warning(request, "No pending offerings were found.")
            return self._entry_redirect(request, clear=True)

        total_amount = offerings.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        return render(request, self.template_name, {
            "offerings": offerings,
            "total_amount": total_amount,
        })

    def post(self, request):
        staged_ids = request.session.get(self.STAGED_IDS_SESSION_KEY, [])
        offerings = self._get_staged_queryset(request)

        if "confirm" in request.POST:
            if offerings.exists():
                messages.success(request, self._build_success_message(offerings))
            else:
                messages.warning(request, "Session expired or no offering data was found.")

            return self._entry_redirect(request, clear=True)

        if "cancel" in request.POST:
            if offerings.exists():
                deleted_count, _ = offerings.delete()
                messages.info(request, f"Entry cancelled. {deleted_count} records removed.")
            else:
                messages.info(request, "Entry cancelled. No staged offerings were found.")

            return self._entry_redirect(request, clear=True)

        return self._entry_redirect(request, clear=False)

class SelectExpensesNumberView(LoginRequiredMixin, View):
    template_name = "select_expenses_number.html"

    def get(self, request):
        form = SelectExpensesNumberForm()
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = SelectExpensesNumberForm(request.POST)
        if form.is_valid():
            num_expenses = form.cleaned_data["number_of_expenses"]
            return redirect("add_expenses", number_of_expenses=num_expenses)
        return render(request, self.template_name, {"form": form})



def _is_expense_fraud_detection_enabled(user) -> bool:
    church = getattr(user, "church", None)
    if not church:
        return False

    try:
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=church)
        return bool(settings_obj.enable_expense_fraud_detection)
    except Exception as e:
        print("FRAUD TOGGLE READ ERROR:", repr(e))
        return False


def _verify_expense_receipt(uploaded_file, *, amount, expense_date, category_obj, description, church):
    """
    OCR still handles PDF separately.
    Fraud detection checks image receipts only.
    """
    if not uploaded_file:
        return True, "No proof uploaded; fraud detection skipped."

    ext = (os.path.splitext(getattr(uploaded_file, "name", ""))[1] or "").lower()

    if ext == ".pdf":
        return True, "PDF proof detected. Fraud detection skipped; OCR remains available."

    if ext not in (".jpg", ".jpeg", ".png"):
        return True, f"Unsupported proof type '{ext}'. Fraud detection skipped."

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        is_valid, reason = verify_receipt_with_openai(
            image_file=uploaded_file,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": expense_date.isoformat() if expense_date else "",
                "context": {
                    "transaction_type": "expense",
                    "category_name": getattr(category_obj, "name", "") if category_obj else "",
                    "description": description or "",
                    "church_id": getattr(church, "id", None),
                }
            }
        )
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return is_valid, reason

def sp_get_cash_on_hand(church_id: int) -> Decimal:
    with connection.cursor() as cursor:
        cursor.execute("CALL Calculate_CashOnHand(%s)", [church_id])
        row = cursor.fetchone()
        _drain_result_sets(cursor)

    if not row:
        return Decimal("0.00")

    return Decimal(str(row[0] or "0.00")).quantize(Decimal("0.01"))

@method_decorator(never_cache, name="dispatch")
class AddExpenseView(LoginRequiredMixin, View):
    template_name = "add_expenses.html"
    SESSION_KEY = "expenses_to_review"
    ExpenseFormSet = formset_factory(ExpenseForm, extra=1, can_delete=True)

    def _build_formset(self, *args, **kwargs):
        return self.ExpenseFormSet(
            *args,
            **kwargs,
            prefix="expense",
            form_kwargs={"user": self.request.user},
        )

    def _cleanup_temp_files(self, staged_items):
        for item in staged_items:
            temp_id = item.get("temp_file_id")
            if temp_id:
                TemporaryExpenseFile.objects.filter(id=temp_id).delete()

    def _to_decimal(self, value, default="0.00"):
        try:
            return Decimal(str(value if value not in (None, "") else default)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal(default).quantize(Decimal("0.01"))

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if request.GET.get("new_batch") == "1":
            staged = request.session.get(self.SESSION_KEY, [])
            self._cleanup_temp_files(staged)

            request.session.pop(self.SESSION_KEY, None)
            request.session.modified = True
            messages.info(request, "Started a new batch.")

        formset = self._build_formset()
        return render(request, self.template_name, {"formset": formset})

    def post(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        # Keep OCR flow unchanged
        if request.POST.get("action") == "OCR":
            uploaded = (
                request.FILES.get("receipt")
                or request.FILES.get("file")
                or request.FILES.get("receipt_image")
            )

            if not uploaded:
                return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

            try:
                result = analyze_receipt_with_openai(uploaded)

                if isinstance(result, dict) and "error" in result:
                    return JsonResponse({"success": False, "error": result["error"]}, status=500)

                return JsonResponse({"success": True, "data": result})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)

        formset = self._build_formset(request.POST, request.FILES)

        if not formset.is_valid():
            messages.error(request, "Please check the form for errors.")
            return render(request, self.template_name, {"formset": formset})

        existing_staged = request.session.get(self.SESSION_KEY, [])
        fraud_detection_enabled = _is_expense_fraud_detection_enabled(request.user)

        # ---------------------------------------------------------
        # PASS 1: collect valid rows first, validate required file,
        # and compute projected cash impact BEFORE saving temp files
        # ---------------------------------------------------------
        prepared_rows = []
        new_batch_total = Decimal("0.00")

        for form in formset:
            if not form.cleaned_data:
                continue

            if form.cleaned_data.get("DELETE", False):
                continue

            data = form.cleaned_data
            category_obj = data.get("category")
            uploaded_file = data.get("file")

            # Skip fully empty rows
            if not any(
                data.get(k)
                for k in ["amount", "expense_date", "description", "category", "file", "ocr_text"]
            ):
                continue

            amount = self._to_decimal(data.get("amount") or "0.00")
            if amount <= Decimal("0.00"):
                continue

            if fraud_detection_enabled and not uploaded_file:
                form.add_error("file", "Receipt file is required while expense fraud detection is enabled.")
                messages.error(
                    request,
                    "Please upload a receipt file for every expense while fraud detection is enabled."
                )
                return render(request, self.template_name, {"formset": formset})

            prepared_rows.append({
                "form": form,
                "data": data,
                "category_obj": category_obj,
                "uploaded_file": uploaded_file,
                "amount": amount,
            })
            new_batch_total += amount

        if not prepared_rows:
            messages.warning(request, "No valid expense data found. Please fill at least one row.")
            return render(request, self.template_name, {"formset": formset})

        existing_staged_total = sum(
            (self._to_decimal(item.get("amount") or "0.00") for item in existing_staged),
            Decimal("0.00")
        ).quantize(Decimal("0.01"))

        current_cash_on_hand = sp_get_cash_on_hand(request.user.church.id)
        total_pending_expenses = (existing_staged_total + new_batch_total).quantize(Decimal("0.01"))
        projected_cash_on_hand = (current_cash_on_hand - total_pending_expenses).quantize(Decimal("0.01"))

        if projected_cash_on_hand < Decimal("0.00"):
            messages.error(
                request,
                (
                    f"Cannot continue to review. Current physical cash on hand is "
                    f"₱{current_cash_on_hand:,.2f}. Already staged expenses total "
                    f"₱{existing_staged_total:,.2f}, and this new batch totals "
                    f"₱{new_batch_total:,.2f}. Projected cash on hand would become "
                    f"₱{projected_cash_on_hand:,.2f}."
                )
            )
            return render(request, self.template_name, {"formset": formset})

        # ---------------------------------------------------------
        # PASS 2: fraud check + temp file save + stage items
        # ---------------------------------------------------------
        expenses_data = list(existing_staged)

        for row in prepared_rows:
            form = row["form"]
            data = row["data"]
            category_obj = row["category_obj"]
            uploaded_file = row["uploaded_file"]

            temp_file_id = None
            file_url = ""

            fraud_checked = False
            if not uploaded_file:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            else:
                fraud_reason = ""

            if fraud_detection_enabled and uploaded_file:
                is_valid, reason = _verify_expense_receipt(
                    uploaded_file=uploaded_file,
                    amount=data.get("amount"),
                    expense_date=data.get("expense_date"),
                    category_obj=category_obj,
                    description=data.get("description"),
                    church=request.user.church,
                )

                fraud_checked = True
                fraud_reason = reason or ""

                if not is_valid:
                    form.add_error("file", f"Fraud detection failed: {reason}")
                    messages.error(request, f"Potential Fraud Detected in one expense row: {reason}")
                    return render(request, self.template_name, {"formset": formset})

            if uploaded_file:
                try:
                    try:
                        uploaded_file.seek(0)
                    except Exception:
                        pass

                    temp_file = TemporaryExpenseFile.objects.create(file=uploaded_file)
                    temp_file_id = int(temp_file.id)
                    file_url = str(temp_file.file.url)
                except Exception as e:
                    messages.error(request, f"File upload failed: {e}")
                    return render(request, self.template_name, {"formset": formset})

            category_is_restricted = bool(getattr(category_obj, "is_restricted", False)) if category_obj else False
            restricted_source = getattr(category_obj, "restricted_source", None) if category_obj else None
            restricted_category_id = getattr(category_obj, "restricted_category_id", None) if category_obj else None

            restricted_fund_key = None
            if category_is_restricted and restricted_source and restricted_category_id:
                prefix_key = "D" if restricted_source == "DONATION" else "I"
                restricted_fund_key = f"{prefix_key}:{int(restricted_category_id)}"

            expense_item = {
                "amount": str(data.get("amount") or ""),
                "expense_date": str(data.get("expense_date") or ""),
                "description": str(data.get("description") or ""),
                "vendor": str(data.get("vendor") or ""),
                "category_id": int(category_obj.id) if category_obj else None,
                "category_name": str(category_obj.name) if category_obj else "Uncategorized",

                "has_file": bool(temp_file_id),
                "temp_file_id": temp_file_id,
                "file_url": file_url,

                "category_is_restricted": category_is_restricted,
                "restricted_source": str(restricted_source) if restricted_source else None,
                "restricted_category_id": int(restricted_category_id) if restricted_category_id else None,
                "restricted_fund_key": restricted_fund_key,

                "ocr_text": str(data.get("ocr_text", "") or ""),

                "fraud_checked": fraud_checked,
                "fraud_reason": fraud_reason,
                "fraud_detection_enabled": fraud_detection_enabled,
            }

            expenses_data.append(expense_item)

        if not expenses_data:
            messages.warning(request, "No valid expense data found. Please fill at least one row.")
            return render(request, self.template_name, {"formset": formset})

        try:
            json.dumps(expenses_data, cls=DjangoJSONEncoder)
        except TypeError as e:
            messages.error(request, f"Data Error: {e}")
            return redirect("add_expenses")

        request.session[self.SESSION_KEY] = expenses_data
        request.session.modified = True
        return redirect("review_expenses")



def session_test(request):
    # 1. Try to write to the session
    if 'test_counter' not in request.session:
        request.session['test_counter'] = 0
        request.session['test_timestamp'] = str(timezone.now())
        msg = "New Session Created. value=0. Please REFRESH this page."
    else:
        # 2. Try to read and update the session
        request.session['test_counter'] += 1
        msg = f"Session is WORKING! Counter: {request.session['test_counter']} (Started: {request.session['test_timestamp']})"

    # Force save
    request.session.modified = True
    request.session.save()

    return HttpResponse(f"""
    <html>
        <body>
            <h1>Session Diagnosis</h1>
            <p style='font-size: 20px; font-weight: bold;'>{msg}</p>
            <p>If the counter stays at 0 every time you refresh, your settings.py is blocking cookies.</p>
        </body>
    </html>
    """)


class ManageExpenseCategoriesView(IsChurchAdmin, View):
    template_name = 'manage_expense_categories.html'

    def get(self, request):
        # ✅ pass user so restricted fund choices are filtered by church
        form = ExpenseCategoryForm(user=request.user)

        categories = ExpenseCategory.objects.filter(
            church=request.user.church
        ).order_by('name')

        return render(request, self.template_name, {
            'form': form,
            'categories': categories
        })

    def post(self, request):
        # ✅ pass user in POST too
        form = ExpenseCategoryForm(request.POST, user=request.user)

        if form.is_valid():
            name = form.cleaned_data['name']

            if ExpenseCategory.objects.filter(
                church=request.user.church,
                name__iexact=name
            ).exists():
                messages.error(request, f"The category '{name}' already exists.")
            else:
                category = form.save(commit=False)
                category.church = request.user.church
                category.save()
                messages.success(request, f"Category '{name}' added successfully.")
                return redirect('manage_expense_categories')

        categories = ExpenseCategory.objects.filter(
            church=request.user.church
        ).order_by('name')

        return render(request, self.template_name, {
            'form': form,
            'categories': categories
        })

# ==========================================
# 3. OCR API (Handles PDF & Images)
# ==========================================


# ---------- 2) Add expenses with a FormSet + dynamic "Add row" ----------
def _drain_result_sets(cursor):
    # MySQL stored procedures can leave extra result sets.
    try:
        while cursor.nextset():
            pass
    except Exception:
        pass

def sp_get_restricted_balances(church_id: int) -> dict:
    """
    Uses your Finance_RestrictedNetBalance stored proc.
    Returns dict {CategoryName: CurrentBalance}
    """
    balances = {}
    with connection.cursor() as cursor:
        cursor.execute("CALL Finance_RestrictedNetBalance(%s)", [church_id])
        rows = cursor.fetchall()
        _drain_result_sets(cursor)

    # rows: CategoryName, TotalCollected, TotalSpent, CurrentBalance
    for r in rows or []:
        name = str(r[0])
        bal = Decimal(str(r[5] or "0.00"))
        balances[name] = bal

    return balances

# ---------------------------
# Helpers (allocation engine)
# ---------------------------
def allocate_amount(
    amount: Decimal,
    preferred: dict,
    unres_balance: Decimal,
    res_balances: dict
):
    """
    preferred:
      {"type": "UNRESTRICTED"} OR {"type": "RESTRICTED", "name": "<fund_name>"}
    Returns:
      (ok, allocations, shortage, flags, new_unres, new_res_dict)
    allocations: list of {"fund_type": "...", "fund_name": "...", "amount": Decimal}
    """
    # Only positive balances are "available"
    U = max(Decimal("0.00"), unres_balance)
    R = {k: max(Decimal("0.00"), v) for k, v in (res_balances or {}).items()}

    rem = max(Decimal("0.00"), amount)
    allocations = []
    flags = {"split": False, "used_restricted": False, "used_unrestricted": False}

    def take_unres(x):
        nonlocal rem, U
        if rem <= 0 or U <= 0:
            return
        take = min(rem, U)
        if take > 0:
            allocations.append({"fund_type": "UNRESTRICTED", "fund_name": "Unrestricted", "amount": take})
            rem -= take
            U -= take
            flags["used_unrestricted"] = True

    def take_restricted(fund_name):
        nonlocal rem, R
        if rem <= 0:
            return
        avail = R.get(fund_name, Decimal("0.00"))
        if avail <= 0:
            return
        take = min(rem, avail)
        if take > 0:
            allocations.append({"fund_type": "RESTRICTED", "fund_name": fund_name, "amount": take})
            rem -= take
            R[fund_name] = avail - take
            flags["used_restricted"] = True

    # order restricted funds by balance desc
    sorted_funds = sorted(R.items(), key=lambda kv: kv[1], reverse=True)

    if preferred.get("type") == "RESTRICTED":
        pref_name = preferred.get("name")
        if pref_name:
            take_restricted(pref_name)

        # overflow -> unrestricted
        take_unres(rem)

        # overflow -> other restricted funds (multiple if needed)
        for fname, _bal in sorted_funds:
            if rem <= 0:
                break
            if pref_name and fname == pref_name:
                continue
            take_restricted(fname)

    else:
        # preferred unrestricted
        take_unres(rem)

        # overflow -> restricted funds (multiple if needed)
        for fname, _bal in sorted_funds:
            if rem <= 0:
                break
            take_restricted(fname)

    ok = (rem <= 0)
    shortage = rem if rem > 0 else Decimal("0.00")
    if len(allocations) > 1:
        flags["split"] = True

    return ok, allocations, shortage, flags, U, R

def guess_preferred_fund(item: dict) -> dict:
    """
    If you later add UI fields, store them in session and read here.
    For now, default to UNRESTRICTED unless item provides restricted_fund_name.
    """
    # OPTIONAL: If you store this from AddExpenseView
    # item["preferred_fund_type"] = "UNRESTRICTED" or "RESTRICTED"
    # item["preferred_restricted_name"] = "Building Fund"
    ptype = (item.get("preferred_fund_type") or "UNRESTRICTED").upper()
    if ptype == "RESTRICTED":
        name = item.get("preferred_restricted_name") or item.get("category_name")
        return {"type": "RESTRICTED", "name": str(name)}
    return {"type": "UNRESTRICTED"}

def resolve_restricted_fund_by_name(church_id: int, fund_name: str):
    """
    Returns (restricted_source, restricted_category_id, fund_key)
    Based on restricted fund name in either donationcategory or otherincomecategory.
    Blocks if ambiguous (exists in both tables).
    """
    fund_name = (fund_name or "").strip()
    if not fund_name:
        raise ValueError("Restricted fund name is missing.")

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id
            FROM register_donationcategory
            WHERE church_id = %s AND COALESCE(is_restricted,0)=1 AND name = %s
            LIMIT 2
        """, [church_id, fund_name])
        d_rows = cursor.fetchall()

        cursor.execute("""
            SELECT id
            FROM register_otherincomecategory
            WHERE church_id = %s AND COALESCE(is_restricted,0)=1 AND name = %s
            LIMIT 2
        """, [church_id, fund_name])
        i_rows = cursor.fetchall()

    d_id = d_rows[0][0] if d_rows else None
    i_id = i_rows[0][0] if i_rows else None

    if d_id and i_id:
        raise ValueError(
            f"Restricted fund name '{fund_name}' is ambiguous (exists in Donation + Other Income). "
            f"Please rename one of them so names are unique."
        )

    if d_id:
        return "DONATION", int(d_id), f"D:{int(d_id)}"
    if i_id:
        return "OTHER_INCOME", int(i_id), f"I:{int(i_id)}"

    raise ValueError(
        f"Restricted fund '{fund_name}' not found for this church. "
        "Make sure it is marked restricted and belongs to the same church."
    )

# ---------------------------
# View
# ---------------------------



def get_primary_fund_for_expense_item(item):
    """
    Decide primary fund based on the selected ExpenseCategory flags.
    For now:
      - if item['is_restricted'] True -> primary = restricted category name
      - else primary = 'U'
    """
    # If you already stage these values in session, use them.
    # else rely on category_id lookup during GET.
    return item.get("primary_fund_key") or "U"


def sp_get_unrestricted_net(church_id) -> Decimal:
    """
    Calls Finance_UnrestrictedNet and returns NetGrandTotalUnrestricted.

    New SP columns:
    0 TotalTithes
    1 TotalOfferings
    2 TotalUnrestrictedDonations
    3 TotalUnrestrictedOtherIncome
    4 TotalBudgetReturnsToUnrestricted
    5 GrandTotalUnrestricted
    6 TotalUnrestrictedExpenses
    7 NetGrandTotalUnrestricted   <-- use this
    8 TotalRealExpensesAll
    9 TotalRestrictedExpenses
    """
    with connection.cursor() as cursor:
        cursor.execute("CALL Finance_UnrestrictedNet(%s)", [church_id])
        row = cursor.fetchone()

        while cursor.nextset():
            pass

    if not row:
        return Decimal("0.00")

    return Decimal(str(row[7] or "0.00"))


class UploadTempReceiptView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        receipt_file = request.FILES.get('receipt_file')
        if not receipt_file:
            return JsonResponse({'error': 'No file provided.'}, status=400)

        # Save to the temporary model
        temp_file = apps.get_model("Register", "TemporaryExpenseFile").objects.create(file=receipt_file)

        return JsonResponse({
            'message': 'success',
            'temp_file_id': temp_file.id,
            'file_name': temp_file.file.name
        })

class ReviewExpensesView(LoginRequiredMixin, View):
    template_name = "review_expenses.html"

    def _dec(self, value, default="0.00"):
        try:
            return Decimal(str(value if value not in (None, "") else default)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal(default).quantize(Decimal("0.01"))

    def _normalize_staged_item(self, item):
        """
        Ensures fraud-related keys always exist for old and new session items.
        """
        p = dict(item)

        has_file = bool(
            p.get("has_file")
            or p.get("temp_file_id")
            or p.get("file_url")
        )

        fraud_detection_enabled = bool(p.get("fraud_detection_enabled", False))
        fraud_checked = bool(p.get("fraud_checked", False))
        fraud_reason = str(p.get("fraud_reason") or "").strip()

        if not fraud_reason:
            if not has_file:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            elif fraud_checked:
                fraud_reason = "Fraud check completed."
            else:
                fraud_reason = "Fraud check not available for this staged item."

        p["has_file"] = has_file
        p["fraud_detection_enabled"] = fraud_detection_enabled
        p["fraud_checked"] = fraud_checked
        p["fraud_reason"] = fraud_reason

        return p

    def _default_lines_for_item(self, item, amt, valid_restricted_names):
        default_cat_id = item.get("category_id")
        temp_file_id = str(item.get("temp_file_id") or "")

        if item.get("category_is_restricted"):
            primary_raw = str(item.get("category_name") or "").strip()
            exact_match = valid_restricted_names.get(primary_raw.lower())

            if exact_match:
                return [{
                    "fund": f"R:{exact_match}",
                    "amount": str(amt),
                    "category_id": default_cat_id,
                    "temp_file_id": temp_file_id,
                }]

        return [{
            "fund": "U",
            "amount": str(amt),
            "category_id": default_cat_id,
            "temp_file_id": temp_file_id,
        }]

    def _build_balance_preview(self, staged_data, saved_allocs, U_start, restricted_balances, valid_restricted_names):
        summary = {}

        for idx, item in enumerate(staged_data):
            amt = self._dec(item.get("amount") or "0.00")
            lines = saved_allocs.get(str(idx), [])

            if not lines:
                lines = self._default_lines_for_item(item, amt, valid_restricted_names)

            for line in lines:
                fund = (line.get("fund") or "").strip()
                amount = self._dec(line.get("amount") or "0.00")

                if not fund or amount <= 0:
                    continue

                category_name = str(item.get("category_name") or "Uncategorized").strip()

                if fund == "U":
                    key = "U"
                    label = "Unrestricted"
                    current_balance = U_start
                elif fund.startswith("R:"):
                    fund_name = fund[2:].strip()
                    key = f"R:{fund_name}"
                    label = f"{fund_name} (Restricted)"
                    current_balance = self._dec(restricted_balances.get(fund_name, "0.00"))
                else:
                    continue

                if key not in summary:
                    summary[key] = {
                        "fund_key": key,
                        "fund_label": label,
                        "current_balance": current_balance,
                        "requested_total": Decimal("0.00"),
                        "categories": set(),
                    }

                summary[key]["requested_total"] += amount
                summary[key]["categories"].add(category_name)

        rows = []
        alerts = []

        def sort_key(item):
            return (0 if item[0] == "U" else 1, item[0])

        for key, entry in sorted(summary.items(), key=sort_key):
            projected_remaining = (entry["current_balance"] - entry["requested_total"]).quantize(Decimal("0.01"))
            insufficient = projected_remaining < 0

            row = {
                "fund_key": entry["fund_key"],
                "fund_label": entry["fund_label"],
                "current_balance": str(entry["current_balance"]),
                "requested_total": str(entry["requested_total"].quantize(Decimal("0.01"))),
                "projected_remaining": str(projected_remaining),
                "categories": ", ".join(sorted(entry["categories"])),
                "insufficient": insufficient,
            }
            rows.append(row)

            if insufficient:
                alerts.append(
                    f"{entry['fund_label']} has insufficient balance. "
                    f"Current: {entry['current_balance']}, "
                    f"Requested: {entry['requested_total'].quantize(Decimal('0.01'))}, "
                    f"Projected: {projected_remaining}."
                )

        return rows, alerts

    def get(self, request):
        staged_data = request.session.get("expenses_to_review")
        if not staged_data:
            messages.warning(request, "No expenses found to review.")
            return redirect("add_expenses")

        # normalize older session items so the template always has fraud fields
        staged_data = [self._normalize_staged_item(item) for item in staged_data]
        request.session["expenses_to_review"] = staged_data
        request.session.modified = True

        church_id = getattr(getattr(request.user, "church", None), "id", None) or staged_data[0].get("church_id")
        if not church_id:
            messages.error(request, "Church context missing.")
            return redirect("add_expenses")

        U_start = self._dec(sp_get_unrestricted_net(church_id) or "0.00")
        restricted_list = sp_get_restricted_balances(church_id) or {}

        restricted_balances = {
            str(name).strip(): self._dec(bal)
            for name, bal in restricted_list.items()
        }

        fund_options = [{"key": "U", "label": f"Unrestricted (Available: {U_start})"}]

        valid_restricted_names = {}
        for name, bal in restricted_balances.items():
            valid_restricted_names[name.strip().lower()] = name
            fund_options.append({
                "key": f"R:{name}",
                "label": f"{name} (Restricted) (Available: {bal})"
            })

        ExpenseCategory = apps.get_model("Register", "ExpenseCategory")
        all_categories = ExpenseCategory.objects.filter(
            church_id=church_id,
            is_restricted=False,
            is_transfer=False
        ).order_by("name")
        category_options = [{"id": str(c.id), "name": c.name} for c in all_categories]

        saved_allocs = request.session.get("expense_allocations_preview", {})

        preview = []
        has_missing = False

        for idx, item in enumerate(staged_data):
            amt = self._dec(item.get("amount") or "0.00")
            str_idx = str(idx)
            default_cat_id = item.get("category_id")

            if str_idx in saved_allocs:
                lines = saved_allocs[str_idx]
            else:
                lines = self._default_lines_for_item(item, amt, valid_restricted_names)

            total_alloc = Decimal("0.00")
            cleaned_lines = []
            item_has_temp = False

            for ln in lines:
                f = (ln.get("fund") or "").strip()
                a = self._dec(ln.get("amount") or "0.00")
                c_id = ln.get("category_id") or default_cat_id
                t_id = ln.get("temp_file_id") or ""

                if t_id:
                    item_has_temp = True

                if f and a > 0:
                    f_label = "Unrestricted"
                    if f.startswith("R:"):
                        f_label = f[2:] + " (Restricted)"

                    cleaned_lines.append({
                        "fund": f,
                        "amount": str(a),
                        "fund_label": f_label,
                        "category_id": str(c_id) if c_id else "",
                        "temp_file_id": str(t_id) if t_id else "",
                    })
                    total_alloc += a

            if not cleaned_lines:
                cleaned_lines.append({
                    "fund": "",
                    "amount": "0.00",
                    "fund_label": "Unallocated",
                    "category_id": "",
                    "temp_file_id": "",
                })

            remaining = (amt - total_alloc).quantize(Decimal("0.01"))
            if remaining != Decimal("0.00"):
                has_missing = True

            p = dict(item)
            p["idx"] = idx
            p["alloc_lines"] = cleaned_lines
            p["remaining"] = str(remaining)
            p["amount_str"] = str(amt)
            p["has_temp_file"] = item_has_temp
            preview.append(p)

        balance_preview_rows, balance_alerts = self._build_balance_preview(
            staged_data=staged_data,
            saved_allocs=saved_allocs,
            U_start=U_start,
            restricted_balances=restricted_balances,
            valid_restricted_names=valid_restricted_names,
        )

        if has_missing:
            messages.warning(
                request,
                "Some expenses need allocation adjustments. Please allocate the remaining balance before confirming."
            )

        for alert in balance_alerts:
            messages.warning(request, alert)

        return render(request, self.template_name, {
            "expenses": preview,
            "fund_options": fund_options,
            "category_options": category_options,
            "U_start": str(U_start),
            "balance_preview_rows": balance_preview_rows,
            "balance_alerts": balance_alerts,
        })

    def post(self, request):
        staged_data = request.session.get("expenses_to_review")
        if not staged_data:
            messages.warning(request, "Session expired. Please enter expenses again.")
            return redirect("add_expenses")

        staged_data = [self._normalize_staged_item(item) for item in staged_data]

        church_id = getattr(getattr(request.user, "church", None), "id", None) or staged_data[0].get("church_id")
        if not church_id:
            messages.error(request, "Church context missing.")
            return redirect("add_expenses")

        if "add_another" in request.POST:
            return redirect("add_expenses")

        preview_allocs = {}

        if "apply" in request.POST or "confirm" in request.POST:
            for idx in range(len(staged_data)):
                funds = request.POST.getlist(f"alloc_{idx}_fund") or request.POST.getlist(f"alloc_{idx}_fund[]")
                amts = request.POST.getlist(f"alloc_{idx}_amount") or request.POST.getlist(f"alloc_{idx}_amount[]")
                cats = request.POST.getlist(f"alloc_{idx}_category") or request.POST.getlist(f"alloc_{idx}_category[]")
                temp_files = request.POST.getlist(f"alloc_{idx}_temp_file_id") or request.POST.getlist(f"alloc_{idx}_temp_file_id[]")

                lines = []
                if funds:
                    cat_idx = 0
                    for i, (f, a) in enumerate(zip(funds, amts)):
                        f = (f or "").strip()
                        c_id = None
                        t_id = temp_files[i] if i < len(temp_files) and temp_files[i] else ""

                        if f.startswith("R:") and len(cats) < len(funds):
                            c_id = None
                        else:
                            if cats and cat_idx < len(cats):
                                c_id = cats[cat_idx]
                                cat_idx += 1

                        if c_id == "":
                            c_id = None

                        a_dec = self._dec(str(a or "0.00").replace(",", ""))

                        if f and a_dec > 0:
                            lines.append({
                                "fund": f,
                                "amount": str(a_dec),
                                "category_id": c_id,
                                "temp_file_id": t_id,
                            })

                preview_allocs[str(idx)] = lines

            request.session["expense_allocations_preview"] = preview_allocs
            request.session.modified = True

        if "apply" in request.POST:
            U_start = self._dec(sp_get_unrestricted_net(church_id) or "0.00")
            restricted_list = sp_get_restricted_balances(church_id) or {}
            restricted_balances = {
                str(name).strip(): self._dec(bal)
                for name, bal in restricted_list.items()
            }
            valid_restricted_names = {name.strip().lower(): name for name in restricted_balances.keys()}

            _, balance_alerts = self._build_balance_preview(
                staged_data=staged_data,
                saved_allocs=preview_allocs,
                U_start=U_start,
                restricted_balances=restricted_balances,
                valid_restricted_names=valid_restricted_names,
            )

            messages.success(request, "Allocation adjustments applied to preview.")
            for alert in balance_alerts:
                messages.warning(request, alert)

            return redirect("review_expenses")

        if "confirm" in request.POST:
            U_initial = self._dec(sp_get_unrestricted_net(church_id) or "0.00")
            U = U_initial

            restricted_list = sp_get_restricted_balances(church_id) or {}
            R_initial = {str(name).strip(): self._dec(bal) for name, bal in restricted_list.items()}
            R = R_initial.copy()

            for idx, item in enumerate(staged_data):
                amt = self._dec(item.get("amount") or "0.00")
                lines = preview_allocs.get(str(idx), [])

                if not lines:
                    lines = [{
                        "fund": "U",
                        "amount": str(amt),
                        "category_id": item.get("category_id"),
                        "temp_file_id": item.get("temp_file_id") or "",
                    }]

                total = sum((self._dec(l["amount"]) for l in lines), Decimal("0.00")).quantize(Decimal("0.01"))

                if total != amt:
                    messages.error(
                        request,
                        f"Expense #{idx + 1} allocation mismatch. Total is {total}, but must be exactly {amt}."
                    )
                    return redirect("review_expenses")

                for l in lines:
                    f = (l.get("fund") or "").strip()
                    a = self._dec(l.get("amount") or "0.00")

                    if f == "U":
                        U -= a
                    elif f.startswith("R:"):
                        fund_name = f[2:].strip()
                        if fund_name not in R:
                            messages.error(
                                request,
                                f"Restricted fund '{fund_name}' does not exist or has no balance."
                            )
                            return redirect("review_expenses")
                        R[fund_name] -= a

            shortage_messages = []

            if U < 0:
                shortage_messages.append(
                    f"Unrestricted has insufficient balance. Current: {U_initial}, projected: {U}."
                )

            for k, bal in R.items():
                if bal < 0:
                    shortage_messages.append(
                        f"Restricted fund '{k}' has insufficient balance. "
                        f"Current: {R_initial.get(k, Decimal('0.00'))}, projected: {bal}."
                    )

            if shortage_messages:
                for msg in shortage_messages:
                    messages.error(request, msg)
                return redirect("review_expenses")

            ExpenseAllocation = apps.get_model("Register", "ExpenseAllocation")
            ChurchModel = apps.get_model("Church", "Church")
            Expense = apps.get_model("Register", "Expense")
            TemporaryExpenseFile = apps.get_model("Register", "TemporaryExpenseFile")
            ExpenseCategoryModel = apps.get_model("Register", "ExpenseCategory")

            temp_ids_to_delete = []

            try:
                with transaction.atomic():
                    ChurchModel.objects.select_for_update().get(id=church_id)

                    for idx, item in enumerate(staged_data):
                        allocations = preview_allocs.get(str(idx), [])

                        if not allocations:
                            allocations = [{
                                "fund": "U",
                                "amount": str(self._dec(item.get("amount", "0.00"))),
                                "category_id": item.get("category_id"),
                                "temp_file_id": item.get("temp_file_id") or "",
                            }]

                        for l in allocations:
                            f = (l.get("fund") or "").strip()
                            a = self._dec(l.get("amount"))
                            new_cat_id = l.get("category_id")
                            t_id = l.get("temp_file_id")

                            exp = Expense()
                            exp.amount = a
                            exp.expense_date = parse_date(str(item.get("expense_date") or "")) or None

                            original_desc = (item.get("description", "") or "").strip()
                            exp.description = f"{original_desc} (Split)".strip() if len(allocations) > 1 else original_desc
                            # assign vendor if provided
                            exp.vendor = str(item.get("vendor") or "").strip()

                            exp.created_by = request.user
                            exp.user = request.user
                            exp.status = "Pending"
                            exp.church_id = church_id

                            if t_id:
                                try:
                                    temp_record = TemporaryExpenseFile.objects.get(id=t_id)
                                    if temp_record.file:
                                        filename = os.path.basename(temp_record.file.name)
                                        temp_record.file.open("rb")
                                        exp.file.save(filename, File(temp_record.file), save=False)
                                        temp_record.file.close()
                                        temp_ids_to_delete.append(temp_record.id)
                                except TemporaryExpenseFile.DoesNotExist:
                                    pass

                            if f == "U":
                                if new_cat_id:
                                    exp.category_id = int(new_cat_id)
                                else:
                                    exp.category_id = int(item["category_id"]) if item.get("category_id") else None

                                exp.save()

                                ExpenseAllocation.objects.create(
                                    expense=exp,
                                    church_id=church_id,
                                    fund_type="UNRESTRICTED",
                                    fund_key="U",
                                    fund_name="Unrestricted",
                                    amount=a,
                                    created_by=request.user,
                                )
                            else:
                                fund_name = f[2:].strip()
                                source, restrict_cat_id, resolved_fund_key = resolve_restricted_fund_by_name(
                                    church_id, fund_name
                                )

                                matching_cat = ExpenseCategoryModel.objects.filter(
                                    church_id=church_id,
                                    name__iexact=fund_name,
                                    is_restricted=True
                                ).first()

                                if matching_cat:
                                    exp.category_id = matching_cat.id
                                else:
                                    exp.category_id = int(item["category_id"]) if item.get("category_id") else None

                                exp.save()

                                ExpenseAllocation.objects.create(
                                    expense=exp,
                                    church_id=church_id,
                                    fund_type="RESTRICTED",
                                    fund_key=resolved_fund_key,
                                    fund_name=fund_name,
                                    amount=a,
                                    restricted_source=source,
                                    restricted_category_id=restrict_cat_id,
                                    created_by=request.user,
                                )

                if temp_ids_to_delete:
                    TemporaryExpenseFile.objects.filter(id__in=temp_ids_to_delete).delete()

                request.session.pop("expenses_to_review", None)
                request.session.pop("expense_allocations_preview", None)
                request.session.modified = True

                messages.success(request, "Expenses and adjustments saved successfully.")
                return redirect("add_expenses")

            except Exception as e:
                messages.error(request, f"Error saving: {e}")
                return redirect("review_expenses")

        if "cancel" in request.POST:
            temp_ids = set()

            for item in staged_data:
                if item.get("temp_file_id"):
                    temp_ids.add(item.get("temp_file_id"))

            preview_existing = request.session.get("expense_allocations_preview", {})
            for lines in preview_existing.values():
                for line in lines:
                    if line.get("temp_file_id"):
                        temp_ids.add(line.get("temp_file_id"))

            if temp_ids:
                TemporaryExpenseFile = apps.get_model("Register", "TemporaryExpenseFile")
                TemporaryExpenseFile.objects.filter(id__in=temp_ids).delete()

            request.session.pop("expenses_to_review", None)
            request.session.pop("expense_allocations_preview", None)
            request.session.modified = True
            messages.info(request, "Entry cancelled.")
            return redirect("add_expenses")

        return redirect("review_expenses")


class ExpenseFraudDetectionSettingsView(IsChurchAdmin, View):
    template_name = "expense_fraud_detection_settings.html"

    TOGGLE_LABELS = {
        "enable_expense_fraud_detection": "Expense fraud detection",
        "enable_tithe_fraud_detection": "Tithe fraud detection",
        "enable_offering_fraud_detection": "Offering fraud detection",
        "enable_other_income_fraud_detection": "Other income fraud detection",
        "enable_donation_fraud_detection": "Donation fraud detection",
    }

    def _get_settings(self, request):
        settings_obj, _ = AccountingSettings.objects.get_or_create(
            church=request.user.church
        )
        return settings_obj

    def get(self, request):
        settings_obj = self._get_settings(request)
        form = ExpenseFraudDetectionSettingsForm(instance=settings_obj)

        return render(request, self.template_name, {
            "form": form,
            "settings_obj": settings_obj,
        })

    def post(self, request):
        settings_obj = self._get_settings(request)

        old_values = {
            field: getattr(settings_obj, field)
            for field in self.TOGGLE_LABELS
        }

        form = ExpenseFraudDetectionSettingsForm(request.POST, instance=settings_obj)

        if form.is_valid():
            updated_settings = form.save()

            changes = []
            turned_on = []
            turned_off = []

            for field, label in self.TOGGLE_LABELS.items():
                old_val = bool(old_values.get(field))
                new_val = bool(getattr(updated_settings, field, False))

                if old_val != new_val:
                    changes.append((label, new_val))
                    if new_val:
                        turned_on.append(label)
                    else:
                        turned_off.append(label)

            if not changes:
                messages.info(request, "No fraud detection settings were changed.")
            elif len(changes) == 1:
                label, state = changes[0]
                messages.success(request, f"{label} is now {'ON' if state else 'OFF'}.")
            else:
                parts = []

                if turned_on:
                    parts.append("Turned ON: " + ", ".join(turned_on))
                if turned_off:
                    parts.append("Turned OFF: " + ", ".join(turned_off))

                messages.success(request, "Settings updated successfully. " + " | ".join(parts))

            return redirect("expense_fraud_detection_settings")

        return render(request, self.template_name, {
            "form": form,
            "settings_obj": settings_obj,
        })

class DecimalEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return str(obj)
        return super().default(obj)

def _is_donation_fraud_detection_enabled(user) -> bool:
    church = getattr(user, "church", None)
    if not church:
        return False

    try:
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=church)
        return bool(settings_obj.enable_donation_fraud_detection)
    except Exception as e:
        print("DONATION FRAUD TOGGLE READ ERROR:", repr(e))
        return False

def _verify_donation_receipt(uploaded_file, *, amount, donations_date, category_obj, donor, church):
    """
    OCR still handles PDF separately.
    Fraud detection checks image receipts only.
    """
    if not uploaded_file:
        return True, "No proof uploaded; fraud detection skipped."

    ext = (os.path.splitext(getattr(uploaded_file, "name", ""))[1] or "").lower()

    if ext == ".pdf":
        return True, "PDF proof detected. Fraud detection skipped; OCR remains available."

    if ext not in (".jpg", ".jpeg", ".png"):
        return True, f"Unsupported proof type '{ext}'. Fraud detection skipped."

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        is_valid, reason = verify_receipt_with_openai(
            image_file=uploaded_file,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": donations_date.isoformat() if donations_date else "",
                "context": {
                    "transaction_type": "donation",
                    "category_name": getattr(category_obj, "name", "") if category_obj else "",
                    "donor": donor or "",
                    "church_id": getattr(church, "id", None),
                }
            }
        )
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return is_valid, reason


@method_decorator(never_cache, name='dispatch')
class AddDonationCategoryView(IsChurchAdmin, CreateView):
    model = DonationCategory
    form_class = DonationCategoryForm
    template_name = 'add_category.html'
    success_url = reverse_lazy('add_donation_category')

    def form_valid(self, form):
        user_church = getattr(self.request.user, 'church', None)

        if not user_church:
            messages.error(self.request, "Error: Your account is not linked to a church.")
            return self.form_invalid(form)

        # --- DUPLICATE CHECK ---
        category_name = form.cleaned_data['name'].strip()

        # Check if exists for THIS church (Case-insensitive)
        name_exists = DonationCategory.objects.filter(
            church=user_church,
            name__iexact=category_name
        ).exists()

        if name_exists:
            messages.error(self.request, f"The category '{category_name}' already exists for your church.")
            return self.form_invalid(form)

        # Assign Church automatically
        form.instance.church = user_church

        # Save the form
        response = super().form_valid(form)

        # Success Message with Fund Type
        fund_type = "Restricted Fund" if form.instance.is_restricted else "General Fund"
        messages.success(self.request, f"Success! '{category_name}' added as a {fund_type}.")

        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_church = getattr(self.request.user, 'church', None)

        if user_church:
            context['existing_categories'] = DonationCategory.objects.filter(
                church=user_church
            ).order_by('name')
        else:
            context['existing_categories'] = DonationCategory.objects.none()

        return context


@login_required
def api_get_notifications(request):
    """
    API endpoint to fetch notifications based on user role
    - Pastors: See pending fund requests (BudgetReleaseRequest with status='Pending')
    - Treasurers: See funds ready for release (ReleasedBudget not yet liquidated)
    """
    from django.urls import reverse

    notifications = []
    user = request.user

    try:
        if not user.church_id:
            return JsonResponse({
                'notifications': [],
                'total_count': 0,
                'view_all_link': '#'
            })

        # Fund Requests (for Pastor) - Pending budget release requests
        if user.user_type == 'Pastor':
            try:
                pending_fund_requests = BudgetReleaseRequest.objects.filter(
                    status='Pending',
                    budget__church_id=user.church_id
                ).select_related('budget', 'budget__ministry').order_by('-created_at')[:5]

                for req in pending_fund_requests:
                    ministry_name = getattr(req.budget.ministry, 'name', 'N/A') if req.budget and req.budget.ministry else 'N/A'
                    notifications.append({
                        'id': f'fund_request_{req.id}',
                        'type': 'fund_request',
                        'category': 'Fund Request',
                        'title': 'Fund Release Request Awaiting Review',
                        'message': f'Ministry: {ministry_name} - Amount: ₱{req.amount:,.2f}',
                        'link': reverse('budgets_pending_approvals'),
                        'unread': True
                    })
            except Exception as e:
                print(f"Error fetching fund requests for Pastor: {e}")

        # Released Funds Ready for Liquidation (for Treasurer)
        if user.user_type == 'Treasurer':
            try:
                released_budgets = ReleasedBudget.objects.filter(
                    church_id=user.church_id,
                    is_liquidated=False
                ).select_related('budget', 'budget__ministry').order_by('-requested_at')[:5]

                for release in released_budgets:
                    ministry_name = getattr(release.budget.ministry, 'name', 'N/A') if release.budget and release.budget.ministry else 'N/A'
                    notifications.append({
                        'id': f'released_fund_{release.id}',
                        'type': 'released_fund',
                        'category': 'Fund Released',
                        'title': 'Fund Released - Ready for Liquidation',
                        'message': f'Ministry: {ministry_name} - Amount: ₱{release.amount:,.2f}',
                        'link': reverse('budgets_liquidation'),
                        'unread': True
                    })
            except Exception as e:
                print(f"Error fetching released funds for Treasurer: {e}")

        # Liquidation Access Requests (for Church Admin)
        if user.user_type in ['ChurchAdmin', 'Admin']:
            try:
                pending_liquidation = DenominationLiquidationAccessRequest.objects.filter(
                    status='Pending',
                    church_id=user.church_id
                ).order_by('-created_at')[:5]

                for req in pending_liquidation:
                    notifications.append({
                        'id': f'liquidation_{req.id}',
                        'type': 'liquidation',
                        'category': 'Liquidation Access',
                        'title': 'Liquidation Access Request',
                        'message': f'User: {req.user.username if req.user else "N/A"} - Status: {req.status}',
                        'link': reverse('church_admin_liquidation_requests'),
                        'unread': True
                    })
            except Exception as e:
                print(f"Error fetching liquidation requests: {e}")

        # Sort notifications by ID (most recent first)
        notifications.sort(key=lambda x: x['id'], reverse=True)

        return JsonResponse({
            'notifications': notifications[:8],  # Limit to 8 notifications
            'total_count': len(notifications),
            'view_all_link': '#'  # Can be replaced with a dedicated notifications page
        })

    except Exception as e:
        print(f"Error fetching notifications: {e}")
        return JsonResponse({
            'notifications': [],
            'total_count': 0,
            'view_all_link': '#'
        }, status=200)  # Return 200 even on error to not break the UI


# ==========================================
# 2. Add Donations (Entry Step)
# ==========================================
@method_decorator(never_cache, name="dispatch")
class AddDonationsView(LoginRequiredMixin, View):
    template_name = "add_donations.html"
    unified_template_name = "unified_income_entry.html"

    SESSION_KEY = "staged_donations"
    SOURCE_SESSION_KEY = "staged_donations_source"
    UNIFIED_PANE = "pane-donations"

    DonationFormSet = formset_factory(DonationsForm, extra=1)

    def _member_display_name(self, user_obj):
        if not user_obj:
            return ""
        if hasattr(user_obj, "get_full_name"):
            full_name = (user_obj.get_full_name() or "").strip()
            if full_name:
                return full_name
        return getattr(user_obj, "username", str(user_obj))

    def _cleanup_temp_files(self, staged_items):
        for item in staged_items or []:
            temp_ids = [
                item.get("temp_file_id"),
                item.get("temp_receipt_image_id"),
            ]
            for temp_id in temp_ids:
                if temp_id:
                    TemporaryDonationFile.objects.filter(id=temp_id).delete()

    def _render_form(self, request, formset):
        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return render(
                request,
                self.unified_template_name,
                build_unified_income_context(
                    request,
                    donation_formset=formset,
                    active_pane=self.UNIFIED_PANE,
                ),
            )
        return render(request, self.template_name, {"formset": formset})

    def _abort_with_cleanup(self, request, formset, new_items, message=None):
        self._cleanup_temp_files(new_items)
        if message:
            messages.error(request, message)
        return self._render_form(request, formset)

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if request.GET.get("new_batch") == "1":
            staged = request.session.get(self.SESSION_KEY, [])
            self._cleanup_temp_files(staged)

            request.session.pop(self.SESSION_KEY, None)
            request.session.pop(self.SOURCE_SESSION_KEY, None)
            request.session.modified = True
            messages.info(request, "Started a new batch.")

        formset = self.DonationFormSet(
            prefix="donation",
            form_kwargs={"user": request.user}
        )

        return render(request, self.template_name, {"formset": formset})

    def post(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)

        # OCR endpoint
        if request.POST.get("action") == "OCR":
            uploaded = (
                request.FILES.get("receipt")
                or request.FILES.get("file")
                or request.FILES.get("receipt_image")
            )
            if not uploaded:
                return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

            try:
                result = analyze_receipt_with_openai(uploaded)
                if isinstance(result, dict) and "error" in result:
                    return JsonResponse({"success": False, "error": result["error"]}, status=500)
                return JsonResponse({"success": True, "data": result})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)

        formset = self.DonationFormSet(
            request.POST,
            request.FILES,
            prefix="donation",
            form_kwargs={"user": request.user}
        )

        if not formset.is_valid():
            messages.error(request, "Please check the form for errors.")
            return self._render_form(request, formset)

        existing_staged_list = request.session.get(self.SESSION_KEY, [])
        new_items = []
        fraud_detection_enabled = _is_donation_fraud_detection_enabled(request.user)

        for form in formset:
            if not getattr(form, "cleaned_data", None):
                continue

            if form.cleaned_data.get("DELETE", False):
                continue

            data = form.cleaned_data

            # Do NOT include donor_type here because it can have a default value
            if not any(
                data.get(k)
                for k in [
                    "amount",
                    "donations_type",
                    "donations_date",
                    "donor",
                    "user",
                    "other_donations_type",
                    "file",
                    "receipt_image",
                    "ocr_text",
                ]
            ):
                continue

            donor_type = data.get("donor_type") or Donations.DonorType.ANONYMOUS
            category_obj = data.get("donations_type")
            user_obj = data.get("user")
            manual_donor_name = (data.get("donor") or "").strip()
            member_full_name = self._member_display_name(user_obj)

            if donor_type == Donations.DonorType.MEMBER:
                donor_display = member_full_name or "Selected Member"
            elif donor_type == Donations.DonorType.NON_MEMBER:
                donor_display = manual_donor_name
            else:
                donor_display = "Anonymous"

            donation = form.save(commit=False)
            donation.church = request.user.church
            donation.created_by = request.user
            donation.donor_type = donor_type

            if donor_type == Donations.DonorType.MEMBER:
                donation.user = user_obj
                donation.donor = None
            elif donor_type == Donations.DonorType.NON_MEMBER:
                donation.user = None
                donation.donor = manual_donor_name
            else:
                donation.user = None
                donation.donor = None

            try:
                donation.clean()
            except ValidationError as e:
                form.add_error(None, e)
                return self._abort_with_cleanup(request, formset, new_items)

            uploaded_file = data.get("file")
            uploaded_receipt_image = data.get("receipt_image")

            if fraud_detection_enabled and not uploaded_file and not uploaded_receipt_image:
                form.add_error(
                    "file",
                    "A proof file or receipt image is required while donation fraud detection is enabled."
                )
                form.add_error(
                    "receipt_image",
                    "A proof file or receipt image is required while donation fraud detection is enabled."
                )
                messages.error(
                    request,
                    "Please upload a proof file or receipt image for every donation while donation fraud detection is enabled."
                )
                return self._abort_with_cleanup(request, formset, new_items)

            fraud_checked = False
            fraud_reason = ""
            fraud_proof_used = ""

            proof_for_verification = uploaded_receipt_image or uploaded_file

            if not proof_for_verification:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            else:
                fraud_proof_used = "receipt_image" if uploaded_receipt_image else "file"

                is_valid, reason = _verify_donation_receipt(
                    uploaded_file=proof_for_verification,
                    amount=data.get("amount"),
                    donations_date=data.get("donations_date"),
                    category_obj=category_obj,
                    donor=donor_display,
                    church=request.user.church,
                )

                fraud_checked = True
                fraud_reason = reason or ""

                if not is_valid:
                    if uploaded_receipt_image:
                        form.add_error("receipt_image", f"Fraud detection failed: {reason}")
                    else:
                        form.add_error("file", f"Fraud detection failed: {reason}")

                    messages.error(request, f"Potential Fraud Detected in one donation row: {reason}")
                    return self._abort_with_cleanup(request, formset, new_items)

            temp_file_id = None
            file_url = ""
            if uploaded_file:
                try:
                    try:
                        uploaded_file.seek(0)
                    except Exception:
                        pass

                    temp_file = TemporaryDonationFile.objects.create(file=uploaded_file)
                    temp_file_id = int(temp_file.id)
                    file_url = str(temp_file.file.url)
                except Exception as e:
                    return self._abort_with_cleanup(
                        request,
                        formset,
                        new_items,
                        message=f"Receipt file upload failed: {e}"
                    )

            temp_receipt_image_id = None
            receipt_image_url = ""
            if uploaded_receipt_image:
                try:
                    try:
                        uploaded_receipt_image.seek(0)
                    except Exception:
                        pass

                    temp_img = TemporaryDonationFile.objects.create(file=uploaded_receipt_image)
                    temp_receipt_image_id = int(temp_img.id)
                    receipt_image_url = str(temp_img.file.url)
                except Exception as e:
                    return self._abort_with_cleanup(
                        request,
                        formset,
                        new_items,
                        message=f"Receipt image upload failed: {e}"
                    )

            staged_item = {
                "amount": str(donation.amount),
                "donations_type_id": int(category_obj.id) if category_obj else None,
                "donations_type_name": str(category_obj.name) if category_obj else "Uncategorized",
                "is_restricted": bool(getattr(category_obj, "is_restricted", False)) if category_obj else False,

                "donor_type": donor_type,
                "donor": str(donation.donor or ""),
                "donor_display": donor_display,
                "user_id": int(user_obj.id) if (donor_type == Donations.DonorType.MEMBER and user_obj) else None,
                "user_name": member_full_name if donor_type == Donations.DonorType.MEMBER else "",

                "donations_date": str(donation.donations_date),
                "other_donations_type": str(data.get("other_donations_type") or ""),

                "has_file": bool(temp_file_id),
                "temp_file_id": temp_file_id,
                "file_url": file_url,

                "has_receipt_image": bool(temp_receipt_image_id),
                "temp_receipt_image_id": temp_receipt_image_id,
                "receipt_image_url": receipt_image_url,

                "ocr_text": str(data.get("ocr_text", "") or ""),

                "fraud_checked": fraud_checked,
                "fraud_reason": fraud_reason,
                "fraud_detection_enabled": fraud_detection_enabled,
                "fraud_proof_used": fraud_proof_used,
            }

            new_items.append(staged_item)

        combined_staged_list = existing_staged_list + new_items

        if not combined_staged_list:
            messages.warning(request, "No valid data found. Please fill at least one row.")
            return self._render_form(request, formset)

        try:
            json.dumps(combined_staged_list, cls=DjangoJSONEncoder)
        except TypeError as e:
            self._cleanup_temp_files(new_items)
            messages.error(request, f"Data Error: {e}")
            return self._render_form(request, formset)

        request.session[self.SESSION_KEY] = combined_staged_list
        request.session[self.SOURCE_SESSION_KEY] = request.POST.get("entry_source", "")
        request.session.modified = True

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)

        return redirect("confirm_donations")


@method_decorator(never_cache, name="dispatch")
class ConfirmDonationsView(FinanceRoleRequiredMixin, View):
    template_name = "confirm_donations.html"
    SESSION_KEY = "staged_donations"
    SOURCE_SESSION_KEY = "staged_donations_source"
    UNIFIED_PANE = "pane-donations"

    def _member_display_name(self, user_obj):
        if not user_obj:
            return ""
        if hasattr(user_obj, "get_full_name"):
            full_name = (user_obj.get_full_name() or "").strip()
            if full_name:
                return full_name
        return getattr(user_obj, "username", str(user_obj))

    def _entry_redirect(self, request, clear=False):
        source = request.session.get(self.SOURCE_SESSION_KEY, "")

        if clear:
            request.session.pop(self.SESSION_KEY, None)
            request.session.pop(self.SOURCE_SESSION_KEY, None)
            request.session.modified = True

        if source == "unified_income_entry":
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return redirect_unified_income_entry(self.UNIFIED_PANE)

        return redirect("add_donations")

    def _cleanup_temp_files(self, staged_items):
        for item in staged_items or []:
            temp_ids = [
                item.get("temp_file_id"),
                item.get("temp_receipt_image_id"),
            ]
            for temp_id in temp_ids:
                if temp_id:
                    TemporaryDonationFile.objects.filter(id=temp_id).delete()

    def _normalize_staged_item(self, item):
        p = dict(item)

        has_file = bool(
            p.get("has_file")
            or p.get("temp_file_id")
            or p.get("file_url")
        )
        has_receipt_image = bool(
            p.get("has_receipt_image")
            or p.get("temp_receipt_image_id")
            or p.get("receipt_image_url")
        )

        fraud_detection_enabled = bool(p.get("fraud_detection_enabled", False))
        fraud_checked = bool(p.get("fraud_checked", False))
        fraud_reason = str(p.get("fraud_reason") or "").strip()
        fraud_proof_used = str(p.get("fraud_proof_used") or "").strip()

        if not fraud_reason:
            if not has_file and not has_receipt_image:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            elif fraud_checked:
                fraud_reason = "Fraud check completed."
            else:
                fraud_reason = "Fraud check not available for this staged item."

        donor_type = p.get("donor_type")
        donor = str(p.get("donor") or "").strip()
        donor_display = str(p.get("donor_display") or "").strip()
        user_id = p.get("user_id")
        user_name = str(p.get("user_name") or "").strip()

        if donor_type not in {
            Donations.DonorType.MEMBER,
            Donations.DonorType.NON_MEMBER,
            Donations.DonorType.ANONYMOUS,
        }:
            if user_id:
                donor_type = Donations.DonorType.MEMBER
            elif donor and donor.lower() != "anonymous":
                donor_type = Donations.DonorType.NON_MEMBER
            else:
                donor_type = Donations.DonorType.ANONYMOUS

        if donor_type == Donations.DonorType.MEMBER:
            donor = donor or user_name or donor_display or "Selected Member"
            donor_display = donor
        elif donor_type == Donations.DonorType.NON_MEMBER:
            donor = donor or donor_display or ""
            donor_display = donor or "Non-member"
            user_id = None
            user_name = ""
        else:
            donor = ""
            user_id = None
            user_name = ""
            donor_display = "Anonymous"

        p["has_file"] = has_file
        p["has_receipt_image"] = has_receipt_image
        p["fraud_detection_enabled"] = fraud_detection_enabled
        p["fraud_checked"] = fraud_checked
        p["fraud_reason"] = fraud_reason
        p["fraud_proof_used"] = fraud_proof_used

        p["donor_type"] = donor_type
        p["donor"] = donor
        p["donor_display"] = donor_display
        p["user_id"] = user_id
        p["user_name"] = user_name

        return p

    def _build_success_message(self, saved_amounts):
        saved_amounts = [Decimal(str(x or "0.00")) for x in saved_amounts]
        saved_count = len(saved_amounts)
        total_saved = sum(saved_amounts, Decimal("0.00"))

        if saved_count <= 0:
            return "No donations were saved."

        if saved_count == 1:
            return f"Donation saved successfully. Amount saved: ₱{total_saved:,.2f}."

        preview = ", ".join([f"₱{amt:,.2f}" for amt in saved_amounts[:5]])
        extra_count = max(saved_count - 5, 0)
        extra_text = f", and {extra_count} more" if extra_count else ""

        return (
            f"{saved_count} donations saved successfully. "
            f"Total amount saved: ₱{total_saved:,.2f}. "
            f"Saved amounts: {preview}{extra_text}."
        )

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        staged_data = request.session.get(self.SESSION_KEY, [])

        if not staged_data:
            messages.warning(request, "No pending donations found. Please enter data first.")
            return self._entry_redirect(request, clear=True)

        staged_data = [self._normalize_staged_item(item) for item in staged_data]
        request.session[self.SESSION_KEY] = staged_data
        request.session.modified = True

        total_amount = sum(
            Decimal(str(item.get("amount") or "0.00"))
            for item in staged_data
        )

        return render(
            request,
            self.template_name,
            {
                "donations": staged_data,
                "total_amount": total_amount,
            }
        )

    def post(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        staged_data = request.session.get(self.SESSION_KEY, [])
        staged_data = [self._normalize_staged_item(item) for item in staged_data]

        if "cancel" in request.POST:
            self._cleanup_temp_files(staged_data)
            messages.info(request, "Entry cancelled. No data was saved.")
            return self._entry_redirect(request, clear=True)

        if "confirm" in request.POST:
            if not staged_data:
                messages.error(request, "Session expired. Please re-enter data.")
                return self._entry_redirect(request, clear=True)

            saved_amounts = []
            user_church = request.user.church
            temp_ids_to_delete = []

            try:
                user_model = Donations._meta.get_field("user").remote_field.model

                with transaction.atomic():
                    for item in staged_data:
                        category_id = item.get("donations_type_id")
                        if not category_id:
                            raise ValidationError("Donation category is missing in one staged item.")

                        category = DonationCategory.objects.get(
                            id=category_id,
                            church=user_church,
                        )

                        donation_date = parse_date(str(item.get("donations_date") or ""))
                        if not donation_date:
                            raise ValidationError(
                                f"Invalid donation date: {item.get('donations_date')}"
                            )

                        amount = Decimal(str(item.get("amount") or "0.00"))
                        donor_type = item.get("donor_type") or Donations.DonorType.ANONYMOUS

                        linked_user = None
                        donor_value = None

                        if donor_type == Donations.DonorType.MEMBER:
                            user_id = item.get("user_id")
                            if not user_id:
                                raise ValidationError("A member-type donation is missing the selected user.")

                            linked_user = user_model.objects.filter(
                                id=user_id,
                                church=user_church,
                                user_type__in=["Member", "Pastor", "Treasurer"],
                                is_active=True,
                            ).first()

                            if not linked_user:
                                raise ValidationError(
                                    "One selected member/pastor/treasurer is invalid or no longer belongs to your church."
                                )

                            donor_value = self._member_display_name(linked_user)

                        elif donor_type == Donations.DonorType.NON_MEMBER:
                            donor_value = str(item.get("donor") or "").strip()
                            if not donor_value:
                                raise ValidationError("A non-member donation is missing the donor name.")

                        elif donor_type == Donations.DonorType.ANONYMOUS:
                            linked_user = None
                            donor_value = None

                        else:
                            raise ValidationError("Invalid donor type found in one staged donation.")

                        donation = Donations(
                            amount=amount,
                            donations_type=category,
                            donor_type=donor_type,
                            donor=donor_value,
                            donations_date=donation_date,
                            other_donations_type=str(item.get("other_donations_type") or "").strip(),
                            user=linked_user,
                            church=user_church,
                            created_by=request.user,
                        )

                        ocr_text = (item.get("ocr_text") or "").strip()
                        if ocr_text:
                            donation.ocr_text = ocr_text
                            donation.ocr_extracted_at = timezone.now()

                        temp_file_id = item.get("temp_file_id")
                        if temp_file_id:
                            temp_file = TemporaryDonationFile.objects.filter(id=temp_file_id).first()
                            if temp_file and temp_file.file:
                                filename = os.path.basename(temp_file.file.name)
                                temp_file.file.open("rb")
                                donation.file.save(filename, File(temp_file.file), save=False)
                                temp_file.file.close()
                                temp_ids_to_delete.append(temp_file.id)

                        temp_receipt_image_id = item.get("temp_receipt_image_id")
                        if temp_receipt_image_id:
                            temp_img = TemporaryDonationFile.objects.filter(id=temp_receipt_image_id).first()
                            if temp_img and temp_img.file:
                                filename = os.path.basename(temp_img.file.name)
                                temp_img.file.open("rb")
                                donation.receipt_image.save(filename, File(temp_img.file), save=False)
                                temp_img.file.close()
                                temp_ids_to_delete.append(temp_img.id)

                        donation.save()
                        saved_amounts.append(amount)

                if temp_ids_to_delete:
                    TemporaryDonationFile.objects.filter(id__in=temp_ids_to_delete).delete()

                messages.success(request, self._build_success_message(saved_amounts))
                return self._entry_redirect(request, clear=True)

            except DonationCategory.DoesNotExist:
                messages.error(request, "One staged donation category no longer exists for your church.")
                return self._entry_redirect(request, clear=True)

            except ValidationError as e:
                messages.error(request, f"Validation Error: {e}")
                return redirect("confirm_donations")

            except Exception as e:
                messages.error(request, f"Database Error: {str(e)}")
                return redirect("confirm_donations")

        return self._entry_redirect(request, clear=False)


# ==========================================
# 4. Donation History & Search
# ==========================================
class DonationListView(FinanceRoleRequiredMixin, View):
    PERIOD_ALL = "all"
    PERIOD_WEEKLY = "weekly"
    PERIOD_MONTHLY = "monthly"
    PERIOD_YEARLY = "yearly"

    VALID_PERIODS = {
        PERIOD_ALL,
        PERIOD_WEEKLY,
        PERIOD_MONTHLY,
        PERIOD_YEARLY,
    }

    def _safe_decimal(self, value):
        try:
            return Decimal(str(value or "0"))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal("0.00")

    def _safe_date(self, value):
        if not value:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        parsed = parse_date(str(value))
        if parsed:
            return parsed

        try:
            return datetime.fromisoformat(str(value)).date()
        except Exception:
            return None

    def _get_row_date(self, row):
        possible_keys = ["donations_date", "date", "created_at"]
        for key in possible_keys:
            if key in row and row.get(key):
                return self._safe_date(row.get(key))
        return None

    def _get_row_amount(self, row):
        possible_keys = ["amount", "total_amount"]
        for key in possible_keys:
            if key in row:
                return self._safe_decimal(row.get(key))
        return Decimal("0.00")

    def _get_donor_name(self, row):
        donor_type = str(row.get("donor_type") or "").strip().lower()

        possible_name_keys = [
            "donor_display",
            "donor_name",
            "donor",
            "user_name",
            "member_name",
            "full_name",
        ]

        for key in possible_name_keys:
            value = str(row.get(key) or "").strip()
            if value:
                return value

        if donor_type == "anonymous":
            return "Anonymous"

        return "Unknown Donor"

    def _get_donation_type(self, row):
        possible_type_keys = [
            "category_name",
            "donations_type_name",
            "other_donations_type",
            "donation_type",
            "type_name",
        ]
        for key in possible_type_keys:
            value = str(row.get(key) or "").strip()
            if value:
                return value
        return "Unspecified"

    def _normalize_row(self, row):
        donation_date = self._get_row_date(row)
        amount = self._get_row_amount(row)
        donor_name = self._get_donor_name(row)
        donation_type = self._get_donation_type(row)

        return {
            "id": row.get("id"),
            "donations_date": donation_date,
            "amount": amount,
            "donor_display": donor_name,
            "donor_type": row.get("donor_type") or "",
            "category_name": donation_type,
        }

    def _get_period_bounds(self, period):
        today = timezone.localdate()

        if period == self.PERIOD_WEEKLY:
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            label = f"Weekly ({start_date} to {end_date})"
            return start_date, end_date, label

        if period == self.PERIOD_MONTHLY:
            start_date = today.replace(day=1)
            if start_date.month == 12:
                next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                next_month = start_date.replace(month=start_date.month + 1, day=1)

            end_date = next_month - timedelta(days=1)
            label = f"Monthly ({start_date.strftime('%B %Y')})"
            return start_date, end_date, label

        if period == self.PERIOD_YEARLY:
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
            label = f"Yearly ({today.year})"
            return start_date, end_date, label

        return None, None, "All Donations"

    def _filter_donations_by_period(self, donations, period):
        start_date, end_date, label = self._get_period_bounds(period)

        if not start_date or not end_date:
            return donations, label, None, None

        filtered = []
        for row in donations:
            donation_date = self._get_row_date(row)
            if donation_date and start_date <= donation_date <= end_date:
                filtered.append(row)

        return filtered, label, start_date, end_date

    def _filter_donations_by_search(self, donations, query):
        query = (query or "").strip().lower()
        if not query:
            return donations

        filtered = []
        for row in donations:
            donor_name = self._get_donor_name(row).lower()
            donation_type = self._get_donation_type(row).lower()
            if query in donor_name or query in donation_type:
                filtered.append(row)
        return filtered

    def _build_donor_totals(self, donations):
        donor_map = defaultdict(lambda: {
            "donor_name": "",
            "total_amount": Decimal("0.00"),
            "donation_count": 0,
            "last_donation_date": None,
        })

        for row in donations:
            donor_name = self._get_donor_name(row)
            amount = self._get_row_amount(row)
            donation_date = self._get_row_date(row)

            donor_map[donor_name]["donor_name"] = donor_name
            donor_map[donor_name]["total_amount"] += amount
            donor_map[donor_name]["donation_count"] += 1

            current_last = donor_map[donor_name]["last_donation_date"]
            if donation_date and (current_last is None or donation_date > current_last):
                donor_map[donor_name]["last_donation_date"] = donation_date

        donor_totals = list(donor_map.values())
        donor_totals.sort(key=lambda x: (-x["total_amount"], x["donor_name"].lower()))
        return donor_totals

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        church = request.user.church
        church_id = church.id
        period = (request.GET.get("period") or self.PERIOD_MONTHLY).lower()
        search_query = (request.GET.get("q") or "").strip()

        if period not in self.VALID_PERIODS:
            period = self.PERIOD_MONTHLY

        donations = []

        try:
            with connection.cursor() as cursor:
                cursor.callproc("GetDonationsByChurch", [church_id])

                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description] if cursor.description else []
                raw_donations = [dict(zip(columns, row)) for row in rows]

                try:
                    while cursor.nextset():
                        try:
                            cursor.fetchall()
                        except Exception:
                            pass
                except Exception:
                    pass

            donations = [self._normalize_row(row) for row in raw_donations]

        except Exception as e:
            logger.exception("DonationListView failed for church_id=%s", church_id)
            messages.error(request, f"Failed to load donations: {e}")
            donations = []

        filtered_donations, period_label, start_date, end_date = self._filter_donations_by_period(
            donations, period
        )
        filtered_donations = self._filter_donations_by_search(filtered_donations, search_query)

        donor_totals = self._build_donor_totals(filtered_donations)

        grand_total = sum((item["total_amount"] for item in donor_totals), Decimal("0.00"))
        total_count = sum(item["donation_count"] for item in donor_totals)

        context = {
            "donations": donations,
            "filtered_donations": filtered_donations,
            "donor_totals": donor_totals,
            "summary_period": period,
            "summary_period_label": period_label,
            "summary_start_date": start_date,
            "summary_end_date": end_date,
            "grand_total": grand_total,
            "grand_donation_count": total_count,
            "church": church,
            "treasurer": request.user,
            "search_query": search_query,
        }
        return render(request, "donation_history.html", context)

# ==========================================
# 2. RECEIPT VIEW (The Missing Piece)
# ==========================================
class DonationReceiptView(LoginRequiredMixin, View):
    def get(self, request, pk):
        # Security: Ensure user can only see their own church's receipts
        donation = get_object_or_404(Donations, pk=pk, church=request.user.church)

        context = {
            'donation': donation,
            'church': request.user.church,
            'treasurer': request.user,
        }
        # This renders the specific receipt for printing
        return render(request, 'donation_receipt.html', context)

# ==========================================
# OTHER INCOME MANAGEMENT (Class Based)
# ==========================================

def _is_other_income_fraud_detection_enabled(user) -> bool:
    church = getattr(user, "church", None)
    if not church:
        return False

    try:
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=church)
        return bool(settings_obj.enable_other_income_fraud_detection)
    except Exception as e:
        print("OTHER INCOME FRAUD TOGGLE READ ERROR:", repr(e))
        return False


def _verify_financial_receipt(
    uploaded_file,
    *,
    amount,
    txn_date,
    category_label,
    description,
    church,
    transaction_type,
):
    """
    Shared fraud detection for finance-related uploads.
    OCR still handles PDF separately.
    Fraud detection checks JPG/PNG only.
    """
    if not uploaded_file:
        return True, "No proof uploaded; fraud detection skipped."

    ext = (os.path.splitext(getattr(uploaded_file, "name", ""))[1] or "").lower()

    if ext == ".pdf":
        return True, "PDF proof detected. Fraud detection skipped; OCR remains available."

    if ext not in (".jpg", ".jpeg", ".png"):
        return True, f"Unsupported proof type '{ext}'. Fraud detection skipped."

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        is_valid, reason = verify_receipt_with_openai(
            image_file=uploaded_file,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": txn_date.isoformat() if txn_date else "",
                "context": {
                    "transaction_type": transaction_type,
                    "category_name": category_label or "",
                    "description": description or "",
                    "church_id": getattr(church, "id", None),
                }
            }
        )
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    return is_valid, reason


@require_POST
def scan_other_income_api(request):
    try:
        from openai import OpenAI
    except ImportError:
        return JsonResponse(
            {"error": "Server Error: Missing 'openai' library."},
            status=500
        )

    file = request.FILES.get("receipt_image")
    if not file:
        return JsonResponse({"error": "No file uploaded."}, status=400)

    api_key = getattr(settings, "OPENAI_API_KEY", None)
    if not api_key:
        return JsonResponse(
            {"error": "Server Error: OpenAI API Key is missing."},
            status=500
        )

    try:
        import fitz  # PyMuPDF

        filename = (file.name or "").lower()
        file.seek(0)
        file_bytes = file.read()
        mime_type = "image/jpeg"

        if filename.endswith(".pdf"):
            with fitz.open(stream=file_bytes, filetype="pdf") as doc:
                if doc.page_count < 1:
                    return JsonResponse({"error": "PDF is empty."}, status=400)

                pix = doc.load_page(0).get_pixmap(dpi=150)
                file_bytes = pix.tobytes("png")
                mime_type = "image/png"

        client = OpenAI(api_key=api_key)
        encoded_image = base64.b64encode(file_bytes).decode("utf-8")

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Return raw JSON with keys: "
                        "amount(number), date(YYYY-MM-DD), description(string), "
                        "category(string). If missing, return null values. No extra text."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Scan this receipt / proof of income."},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{encoded_image}"
                            },
                        },
                    ],
                },
            ],
            response_format={"type": "json_object"},
        )

        content = response.choices[0].message.content
        return JsonResponse(json.loads(content))

    except Exception as e:
        return JsonResponse({"error": f"Scan failed: {str(e)}"}, status=500)


class ManageIncomeCategoriesView(IsChurchAdmin, View):
    template_name = "manage_income_categories.html"

    def get(self, request):
        form = OtherIncomeCategoryForm()
        categories = OtherIncomeCategory.objects.filter(
            church=request.user.church
        ).order_by("name")
        return render(request, self.template_name, {
            "form": form,
            "categories": categories
        })

    def post(self, request):
        form = OtherIncomeCategoryForm(request.POST)

        if form.is_valid():
            name = form.cleaned_data["name"]

            if OtherIncomeCategory.objects.filter(
                church=request.user.church,
                name__iexact=name
            ).exists():
                messages.error(request, f"The category '{name}' already exists.")
            else:
                category = form.save(commit=False)
                category.church = request.user.church
                category.save()
                messages.success(request, f"Category '{name}' added successfully.")
                return redirect("manage_income_categories")

        categories = OtherIncomeCategory.objects.filter(
            church=request.user.church
        ).order_by("name")

        return render(request, self.template_name, {
            "form": form,
            "categories": categories
        })


@method_decorator(never_cache, name="dispatch")
class AddOtherIncomeView(LoginRequiredMixin, View):
    template_name = "add_other_income.html"
    unified_template_name = "unified_income_entry.html"

    SESSION_KEY = "staged_other_income"
    SOURCE_SESSION_KEY = "staged_other_income_source"
    UNIFIED_PANE = "pane-other-income"

    OtherIncomeFormSet = formset_factory(OtherIncomeForm, extra=1)

    def _cleanup_temp_files(self, staged_items):
        for item in staged_items or []:
            temp_id = item.get("temp_file_id")
            if temp_id:
                TemporaryOtherIncomeFile.objects.filter(id=temp_id).delete()

    def _render_form(self, request, formset):
        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return render(
                request,
                self.unified_template_name,
                build_unified_income_context(
                    request,
                    other_income_formset=formset,
                    active_pane=self.UNIFIED_PANE,
                ),
            )
        return render(request, self.template_name, {"formset": formset})

    def _abort_with_cleanup(self, request, formset, new_items, message=None):
        self._cleanup_temp_files(new_items)
        if message:
            messages.error(request, message)
        return self._render_form(request, formset)

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if request.GET.get("new_batch") == "1":
            staged = request.session.get(self.SESSION_KEY, [])
            self._cleanup_temp_files(staged)

            request.session.pop(self.SESSION_KEY, None)
            request.session.pop(self.SOURCE_SESSION_KEY, None)
            request.session.modified = True
            messages.info(request, "Started a new batch.")

        formset = self.OtherIncomeFormSet(
            prefix="income",
            form_kwargs={"user": request.user}
        )
        return render(request, self.template_name, {"formset": formset})

    def post(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)

        formset = self.OtherIncomeFormSet(
            request.POST,
            request.FILES,
            prefix="income",
            form_kwargs={"user": request.user},
        )

        if not formset.is_valid():
            messages.error(request, "Please check the form for errors.")
            return self._render_form(request, formset)

        existing_staged_list = request.session.get(self.SESSION_KEY, [])
        new_items = []
        fraud_detection_enabled = _is_other_income_fraud_detection_enabled(request.user)

        for form in formset:
            if not getattr(form, "cleaned_data", None):
                continue

            if form.cleaned_data.get("DELETE", False):
                continue

            data = form.cleaned_data

            if not any(data.get(k) for k in ["amount", "income_type", "date", "description", "file", "ocr_text"]):
                continue

            income = form.save(commit=False)
            income.church = request.user.church
            income.created_by = request.user

            try:
                income.clean()
            except ValidationError as e:
                form.add_error(None, e)
                return self._abort_with_cleanup(request, formset, new_items)

            category_obj = data.get("income_type")
            uploaded = data.get("file")

            if fraud_detection_enabled and not uploaded:
                form.add_error("file", "Proof file is required while other income fraud detection is enabled.")
                messages.error(
                    request,
                    "Please upload a proof file for every other income entry while fraud detection is enabled."
                )
                return self._abort_with_cleanup(request, formset, new_items)

            fraud_checked = False
            if not uploaded:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            else:
                fraud_reason = ""

            if fraud_detection_enabled and uploaded:
                is_valid, reason = _verify_financial_receipt(
                    uploaded_file=uploaded,
                    amount=data.get("amount"),
                    txn_date=data.get("date"),
                    category_label=getattr(category_obj, "name", "") if category_obj else "",
                    description=data.get("description"),
                    church=request.user.church,
                    transaction_type="other_income",
                )

                fraud_checked = True
                fraud_reason = reason or ""

                if not is_valid:
                    form.add_error("file", f"Fraud detection failed: {reason}")
                    messages.error(request, f"Potential Fraud Detected in one other income row: {reason}")
                    return self._abort_with_cleanup(request, formset, new_items)

            temp_file_id = None
            file_url = ""

            if uploaded:
                try:
                    try:
                        uploaded.seek(0)
                    except Exception:
                        pass

                    temp_file = TemporaryOtherIncomeFile.objects.create(file=uploaded)
                    temp_file_id = int(temp_file.id)
                    file_url = str(temp_file.file.url)
                except Exception as e:
                    return self._abort_with_cleanup(
                        request,
                        formset,
                        new_items,
                        message=f"File upload failed: {e}"
                    )

            staged_item = {
                "income_type_id": int(category_obj.id) if category_obj else None,
                "income_type_name": str(category_obj.name) if category_obj else "Uncategorized",
                "is_restricted": bool(getattr(category_obj, "is_restricted", False)) if category_obj else False,
                "amount": str(income.amount),
                "date": str(income.date),
                "description": str(income.description or ""),

                "has_file": bool(temp_file_id),
                "temp_file_id": temp_file_id,
                "file_url": file_url,

                "ocr_text": str(data.get("ocr_text", "") or ""),

                "fraud_checked": fraud_checked,
                "fraud_reason": fraud_reason,
                "fraud_detection_enabled": fraud_detection_enabled,
            }

            new_items.append(staged_item)

        combined_staged_list = existing_staged_list + new_items

        if not combined_staged_list:
            messages.warning(request, "No valid data found. Please fill at least one row.")
            return self._render_form(request, formset)

        try:
            json.dumps(combined_staged_list, cls=DjangoJSONEncoder)
        except TypeError as e:
            self._cleanup_temp_files(new_items)
            messages.error(request, f"Data Error: {e}")
            return self._render_form(request, formset)

        request.session[self.SESSION_KEY] = combined_staged_list
        request.session[self.SOURCE_SESSION_KEY] = request.POST.get("entry_source", "")
        request.session.modified = True

        if is_from_unified_income_page(request):
            set_unified_income_active_pane(request, self.UNIFIED_PANE)

        return redirect("review_other_income")


@method_decorator(never_cache, name="dispatch")
class ReviewOtherIncomeView(LoginRequiredMixin, View):
    template_name = "review_other_income.html"
    SESSION_KEY = "staged_other_income"
    SOURCE_SESSION_KEY = "staged_other_income_source"
    UNIFIED_PANE = "pane-other-income"

    def _entry_redirect(self, request, clear=False):
        source = request.session.get(self.SOURCE_SESSION_KEY, "")

        if clear:
            request.session.pop(self.SESSION_KEY, None)
            request.session.pop(self.SOURCE_SESSION_KEY, None)
            request.session.modified = True

        if source == "unified_income_entry":
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return redirect_unified_income_entry(self.UNIFIED_PANE)

        return redirect("add_other_income")

    def _dec(self, value, default="0.00"):
        try:
            return Decimal(str(value if value not in (None, "") else default)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal(default).quantize(Decimal("0.01"))

    def _normalize_staged_item(self, item):
        """
        Ensures older and newer staged rows always have the same keys.
        """
        p = dict(item)

        has_file = bool(
            p.get("has_file")
            or p.get("temp_file_id")
            or p.get("file_url")
        )

        fraud_detection_enabled = bool(p.get("fraud_detection_enabled", False))
        fraud_checked = bool(p.get("fraud_checked", False))
        fraud_reason = str(p.get("fraud_reason") or "").strip()

        if not fraud_reason:
            if not has_file:
                fraud_reason = "No proof uploaded; fraud detection skipped."
            elif not fraud_detection_enabled:
                fraud_reason = "Fraud detection is OFF."
            elif fraud_checked:
                fraud_reason = "Fraud check completed."
            else:
                fraud_reason = "Fraud check not available for this staged item."

        p["income_type_id"] = p.get("income_type_id")
        p["income_type_name"] = str(p.get("income_type_name") or "Uncategorized")
        p["description"] = str(p.get("description") or "")
        p["date"] = str(p.get("date") or "")
        p["amount"] = str(self._dec(p.get("amount") or "0.00"))
        p["ocr_text"] = str(p.get("ocr_text") or "")

        p["has_file"] = has_file
        p["temp_file_id"] = p.get("temp_file_id") or None
        p["file_url"] = str(p.get("file_url") or "")

        p["fraud_detection_enabled"] = fraud_detection_enabled
        p["fraud_checked"] = fraud_checked
        p["fraud_reason"] = fraud_reason

        return p

    def _normalize_items(self, items):
        return [self._normalize_staged_item(item) for item in (items or [])]

    def _save_session_items(self, request, items):
        request.session[self.SESSION_KEY] = items
        request.session.modified = True

    def _cleanup_temp_file(self, temp_id):
        if temp_id:
            TemporaryOtherIncomeFile.objects.filter(id=temp_id).delete()

    def _cleanup_temp_files(self, items):
        for item in items or []:
            self._cleanup_temp_file(item.get("temp_file_id"))

    def _build_success_message(self, saved_amounts):
        saved_amounts = [self._dec(x) for x in saved_amounts]
        saved_count = len(saved_amounts)
        total_saved = sum(saved_amounts, Decimal("0.00")).quantize(Decimal("0.01"))

        if saved_count <= 0:
            return "No other income entries were saved."

        if saved_count == 1:
            return f"Other income saved successfully. Amount saved: ₱{total_saved:,.2f}."

        preview = ", ".join([f"₱{amt:,.2f}" for amt in saved_amounts[:5]])
        extra_count = max(saved_count - 5, 0)
        extra_text = f", and {extra_count} more" if extra_count else ""

        return (
            f"{saved_count} other income entries saved successfully. "
            f"Total amount saved: ₱{total_saved:,.2f}. "
            f"Saved amounts: {preview}{extra_text}."
        )

    def get(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        items = request.session.get(self.SESSION_KEY, [])
        if not items:
            messages.warning(request, "No transaction found. Please enter details first.")
            return self._entry_redirect(request, clear=True)

        items = self._normalize_items(items)
        self._save_session_items(request, items)

        total = sum((self._dec(i.get("amount")) for i in items), Decimal("0.00")).quantize(Decimal("0.01"))

        return render(request, self.template_name, {
            "items": items,
            "total": total,
            "item_count": len(items),
        })

    def post(self, request):
        if not hasattr(request.user, "church") or not request.user.church:
            messages.error(request, "Access Denied: You are not linked to a church.")
            return redirect("home")

        items = request.session.get(self.SESSION_KEY, [])
        items = self._normalize_items(items)
        self._save_session_items(request, items)

        if not items:
            messages.warning(request, "No transaction found. Please enter details first.")
            return self._entry_redirect(request, clear=True)

        if "cancel" in request.POST:
            self._cleanup_temp_files(items)
            messages.info(request, "Batch cancelled.")
            return self._entry_redirect(request, clear=True)

        if "add_another" in request.POST:
            set_unified_income_active_pane(request, self.UNIFIED_PANE)
            return self._entry_redirect(request, clear=False)

        if "remove" in request.POST:
            try:
                idx = int(request.POST.get("remove"))
                if 0 <= idx < len(items):
                    removed = items.pop(idx)
                    self._cleanup_temp_file(removed.get("temp_file_id"))
                    self._save_session_items(request, items)
                    messages.info(request, "Entry removed.")
                else:
                    messages.error(request, "Invalid item selected.")
            except Exception:
                messages.error(request, "Invalid item selected.")
            return redirect("review_other_income")

        if "confirm" in request.POST:
            if not items:
                messages.warning(request, "No entries to save.")
                return self._entry_redirect(request, clear=True)

            temp_ids_to_delete = []
            saved_amounts = []

            try:
                with transaction.atomic():
                    apps.get_model("Church", "Church").objects.select_for_update().get(
                        id=request.user.church.id
                    )

                    for idx, item in enumerate(items, start=1):
                        income_type_id = item.get("income_type_id")
                        if not income_type_id:
                            raise ValueError(f"Entry #{idx} is missing an income type.")

                        amount = self._dec(item.get("amount"))
                        income_date = parse_date(str(item.get("date") or ""))

                        if amount <= 0:
                            raise ValueError(f"Entry #{idx} has an invalid amount.")

                        if not income_date:
                            raise ValueError(f"Entry #{idx} has an invalid date.")

                        if (
                            item.get("fraud_detection_enabled")
                            and item.get("has_file")
                            and not item.get("fraud_checked")
                        ):
                            raise ValueError(
                                f"Entry #{idx} has not completed fraud detection. Please re-upload and review again."
                            )

                        category = get_object_or_404(
                            OtherIncomeCategory,
                            pk=income_type_id,
                            church=request.user.church
                        )

                        oi = OtherIncome(
                            church=request.user.church,
                            income_type=category,
                            amount=amount,
                            date=income_date,
                            description=item.get("description", "") or "",
                            created_by=request.user,
                            ocr_text=item.get("ocr_text", "") or "",
                        )

                        if oi.ocr_text:
                            oi.ocr_processed_at = timezone.now()

                        temp_id = item.get("temp_file_id")
                        if temp_id:
                            temp_obj = TemporaryOtherIncomeFile.objects.filter(id=temp_id).first()
                            if temp_obj and temp_obj.file:
                                filename = os.path.basename(temp_obj.file.name)

                                temp_obj.file.open("rb")
                                oi.file.save(filename, File(temp_obj.file), save=False)
                                temp_obj.file.close()

                                temp_ids_to_delete.append(temp_id)

                        oi.full_clean()
                        oi.save()
                        saved_amounts.append(amount)

                if temp_ids_to_delete:
                    TemporaryOtherIncomeFile.objects.filter(id__in=temp_ids_to_delete).delete()

                messages.success(request, self._build_success_message(saved_amounts))
                return self._entry_redirect(request, clear=True)

            except Exception as e:
                messages.error(request, f"Error saving data: {str(e)}")
                return redirect("review_other_income")

        return redirect("review_other_income")


class OtherIncomeListView(LoginRequiredMixin, ListView):
    model = OtherIncome
    template_name = 'other_income_list.html'
    context_object_name = 'incomes'
    paginate_by = 20

    def get_queryset(self):
        return OtherIncome.objects.filter(
            church=self.request.user.church
        ).select_related('income_type', 'created_by').order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        context['total_income'] = qs.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        return context





class LandingView(LoginRequiredMixin, TemplateView):
    template_name = 'landing_Page.html'


class TreasurerLoginView( LoginView):
    """
    Class-based view for Treasurer login.
    """
    form_class = TreasurerLoginForm
    template_name = 'treasurer_login.html'  # Create this template
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy('home')



class MemberLoginView(LoginView):
    form_class = MemberLoginForm
    template_name = 'member_login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        """
        Redirect Members to their dashboard upon successful login.
        """
        user = self.request.user
        if user.user_type == 'Member':
            return reverse_lazy('home')  # Replace with your member dashboard URL
        return reverse_lazy('home')

class AdminLoginView(LoginView):
    form_class = AdminLoginForm
    template_name = 'admin_login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        """
        Redirect Members to their dashboard upon successful login.
        """
        user = self.request.user
        if user.user_type == 'Admin':
            return reverse_lazy('home')  # Replace with your member dashboard URL
        return reverse_lazy('home')


class FinanceOverview(LoginRequiredMixin, View):
    template_name = 'finance_overview.html'

    def get(self, request, *args, **kwargs):
        # 1. Fetch data grouped by date
        models = self.get_model_form_mapping()
        data_by_date = self.aggregate_data_by_date(models, request.user.church)

        # 2. Sort the dates in reverse chronological order
        sorted_dates = sorted(data_by_date.keys(), reverse=True)

        # 3. Calculate Financial Totals
        cash_on_hand = self.calculate_cash_on_hand(request.user.church)

        # A. Restricted Funds (Total Net Balance of Designated Categories)
        restricted_funds = self.get_restricted_net_balance(request.user.church.id)

        # B. Unrestricted Balance (Available for General Use)
        unrestricted_balance = cash_on_hand - restricted_funds

        # C. Unrestricted Income Stats (Breakdown of income sources)
        unrestricted_stats = self.get_unrestricted_income_stats(request.user.church.id)

        # 4. Prepare the context
        context = {
            'data_by_date': {date: data_by_date[date] for date in sorted_dates},
            'sorted_dates': sorted_dates,

            # --- NEW DASHBOARD METRICS ---
            'cash_on_hand': cash_on_hand,
            'restricted_funds': restricted_funds,
            'unrestricted_balance': unrestricted_balance,
            'unrestricted_stats': unrestricted_stats,
            # -----------------------------

            'modal_message': None,
            'modal_status': None,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        item_id = request.POST.get('item_id')
        action = request.POST.get('action', 'edit')
        category = request.POST.get('category')

        # Retrieve the appropriate model and form
        model, form_class = self.get_model_and_form(category)
        if not model:
            return self.render_with_message(request, "Invalid category", "error")

        # Fetch the item (Strictly filter by the user's church for security)
        item = get_object_or_404(model, id=item_id, church=request.user.church)

        # =========================================================
        # SECURITY CHECK: ACCOUNTING LOCK & FUTURE DATE
        # =========================================================
        item_date = self.get_entry_date(item)

        if item_date:
            # 1. Check Past Lock (Accounting Settings)
            if hasattr(request.user.church, 'accounting_settings'):
                lock_date = request.user.church.accounting_settings.lock_date
                if lock_date and item_date <= lock_date:
                    return self.render_with_message(
                        request,
                        f"ACCESS DENIED: The books are closed for {lock_date}. You cannot {action} this record.",
                        "error"
                    )

            # 2. Check Future Lock (Week After)
            today = timezone.now().date()
            future_limit = today + timedelta(days=7)
            if item_date > future_limit:
                return self.render_with_message(
                    request,
                    f"ACCESS DENIED: Cannot {action} records dated past {future_limit}.",
                    "error"
                )
        # =========================================================

        # ACTION: DELETE
        if action == 'delete':
            item.delete()
            return self.render_with_message(request, "Item deleted successfully", "success")

        # ACTION: EDIT
        # Special handling: OtherIncomeForm requires 'user' kwarg
        if category == 'Other Income':
            form = form_class(request.POST, request.FILES, instance=item, user=request.user)
        else:
            form = form_class(request.POST, request.FILES, instance=item)

        if form.is_valid():
            try:
                # form.save() triggers the model's clean() method (Lock Date check)
                form.instance.edited_by = request.user
                form.save()
                return self.render_with_message(request, "Item updated successfully", "success")
            except Exception as e:
                # Catch validation errors from the model
                return self.render_with_message(request, f"Error: {str(e)}", "error")
        else:
            error_message = ", ".join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()])
            return self.render_with_message(request, f"Form validation failed: {error_message}", "error")

    def render_with_message(self, request, message, status):
        """Helper to re-render the page with an error/success modal and REFRESHED TOTALS."""
        models = self.get_model_form_mapping()
        data_by_date = self.aggregate_data_by_date(models, request.user.church)
        sorted_dates = sorted(data_by_date.keys(), reverse=True)

        # --- RECALCULATE TOTALS (So they don't disappear after an edit) ---
        cash_on_hand = self.calculate_cash_on_hand(request.user.church)
        restricted_funds = self.get_restricted_net_balance(request.user.church.id)
        unrestricted_balance = cash_on_hand - restricted_funds
        unrestricted_stats = self.get_unrestricted_income_stats(request.user.church.id)
        # ------------------------------------------------------------------

        context = {
            'data_by_date': {date: data_by_date[date] for date in sorted_dates},
            'sorted_dates': sorted_dates,

            # --- NEW DASHBOARD METRICS ---
            'cash_on_hand': cash_on_hand,
            'restricted_funds': restricted_funds,
            'unrestricted_balance': unrestricted_balance,
            'unrestricted_stats': unrestricted_stats,
            # -----------------------------

            'modal_message': message,
            'modal_status': status,
        }
        return render(request, self.template_name, context)

    # =========================================================
    # NEW HELPERS: STORED PROCEDURES INTERFACE
    # =========================================================

    def get_restricted_net_balance(self, church_id):
        """Calls Stored Procedure: Finance_RestrictedFundsSum"""
        if not church_id: return 0
        try:
            with connection.cursor() as cursor:
                cursor.callproc('Finance_RestrictedFundsSum', [church_id])
                result = cursor.fetchone()
            # Result is a tuple: (Decimal('100.00'),)
            return result[0] if result and result[0] else 0
        except Exception as e:
            print(f"Error fetching restricted funds: {e}")
            return 0

    def get_unrestricted_income_stats(self, church_id):
        """Calls Stored Procedure: Finance_UnrestrictedSums"""
        if not church_id: return {}
        try:
            with connection.cursor() as cursor:
                cursor.callproc('Finance_UnrestrictedSums', [church_id])
                row = cursor.fetchone()

                # The SP returns: Tithes, Offerings, Donations, OtherIncome, GrandTotal
                if row:
                    return {
                        'tithes': row[0] or 0,
                        'offerings': row[1] or 0,
                        'donations': row[2] or 0,
                        'other_income': row[3] or 0,
                        'total_income': row[4] or 0
                    }
        except Exception as e:
            print(f"Error fetching unrestricted sums: {e}")

        return {}

    # =========================================================
    # EXISTING HELPERS
    # =========================================================

    def get_model_and_form(self, category):
        for model, form, cat in self.get_model_form_mapping():
            if cat.lower() == category.lower():
                return model, form
        return None, None

    @staticmethod
    def get_entry_date(item):
        """Helper to safely get the date field from different models."""
        for attr in ['date', 'expense_date', 'donations_date']:
            if hasattr(item, attr):
                return getattr(item, attr)
        return None

    def calculate_cash_on_hand(self, church):
        """
        Calculate total available cash: (Tithes + Offerings + Donations + Other Income) - Expenses
        """
        if not church: return 0

        totals = {
            'tithes': Tithe.objects.filter(church=church).aggregate(Sum('amount'))['amount__sum'] or 0,
            'offerings': Offering.objects.filter(church=church).aggregate(Sum('amount'))['amount__sum'] or 0,
            'donations': Donations.objects.filter(church=church).aggregate(Sum('amount'))['amount__sum'] or 0,
            'other_income': OtherIncome.objects.filter(church=church).aggregate(Sum('amount'))['amount__sum'] or 0,
            'expenses': Expense.objects.filter(church=church).aggregate(Sum('amount'))['amount__sum'] or 0,
        }
        return totals['tithes'] + totals['offerings'] + totals['donations'] + totals['other_income'] - totals[
            'expenses']

    def aggregate_data_by_date(self, models, church):
        """
        Aggregate data, strictly filtered by the current user's CHURCH.
        Groups all transactions by Date.
        """
        data_by_date = defaultdict(lambda: {
            'entries': [],
            'totals': {'tithes': 0, 'offerings': 0, 'expenses': 0, 'donations': 0, 'other_income': 0, 'accumulated': 0},
            'edited_by': None
        })

        if not church: return data_by_date

        for Model, Form, category in models:
            # 1. Fetch records for this church
            records = Model.objects.filter(church=church).select_related('edited_by')

            for item in records:
                entry_date = self.get_entry_date(item)
                amount = getattr(item, 'amount', 0)

                # 2. Instantiate Form (Handle 'user' arg for OtherIncomeForm)
                if category == 'Other Income':
                    # We use self.request.user because 'aggregate_data_by_date' is called from get/post
                    form_instance = Form(instance=item, user=self.request.user)
                else:
                    form_instance = Form(instance=item)

                entry = {
                    'data': item,
                    'form': form_instance,
                    'category': category,
                    'date': entry_date,
                    'edited_by': item.edited_by,
                }

                # 3. Add to grouped data
                data_by_date[entry_date]['entries'].append(entry)

                # FIX: Convert category name to underscore key (e.g., "Other Income" -> "other_income")
                key_name = category.lower().replace(' ', '_')

                # Update the specific total if the key exists
                if key_name in data_by_date[entry_date]['totals']:
                    data_by_date[entry_date]['totals'][key_name] += amount

                if item.edited_by:
                    data_by_date[entry_date]['edited_by'] = item.edited_by

        # 4. Calculate Accumulated Totals per Day
        for date, data in data_by_date.items():
            totals = data['totals']
            totals['accumulated'] = (
                    totals['tithes'] +
                    totals['offerings'] +
                    totals['donations'] +
                    totals['other_income'] -  # Use the underscore key
                    totals['expenses']
            )
        return data_by_date

    @staticmethod
    def get_model_form_mapping():
        """
        Defines the list of models to include in the overview.
        Structure: (ModelClass, FormClass, 'Display Name')
        """
        return [
            (Tithe, TitheForm, 'Tithes'),
            (Offering, OfferingForm, 'Offerings'),
            (Expense, ExpenseForm, 'Expenses'),
            (Donations, DonationsForm, 'Donations'),
            (OtherIncome, OtherIncomeForm, 'Other Income'),
        ]

class CustomLogoutView(LoginRequiredMixin, LogoutView):
    next_page = reverse_lazy('landing')


class CustomUserDetailView(LoginRequiredMixin, DetailView):
    model = CustomUser
    template_name = 'user_detail.html'
    context_object_name = 'user_obj'

    def get_object(self, queryset=None):
        # Always show the currently logged-in user
        return CustomUser.objects.select_related(
            'church', 'denomination'
        ).get(pk=self.request.user.pk)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Handle profile picture upload
        if request.POST.get('image_upload') == 'true':
            uploaded_file = request.FILES.get('profile_picture')

            if uploaded_file:
                self.object.profile_picture = uploaded_file
                self.object.save(update_fields=['profile_picture'])
                messages.success(request, 'Profile picture updated successfully.')
            else:
                messages.warning(request, 'No file selected.')

            return redirect(request.path_info)

        messages.warning(request, 'Invalid request.')
        return redirect(request.path_info)

class CustomUserUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomUser
    form_class = CustomUserForm
    template_name = 'user_edit.html'
    success_url = reverse_lazy('user_detail')
    context_object_name = 'user_obj'

    def get_object(self, queryset=None):
        return CustomUser.objects.select_related('church', 'denomination').get(pk=self.request.user.pk)

    def form_valid(self, form):
        # Wrap save in try/except to catch model validation errors that
        # can occur when the model's clean() enforces related fields
        # (e.g., ChurchAdmin must have a church). Without this guard,
        # a ValidationError raised by model.full_clean() will bubble up
        # and cause a crash (500) when a user edits only a subset of
        # editable fields.
        from django.core.exceptions import ValidationError as ModelValidationError

        try:
            # Save safely: use commit=False to avoid accidentally clearing
            # related fields that are not part of the form (e.g., church, user_type).
            existing = self.get_object()
            obj = form.save(commit=False)

            # Preserve fields that are not present in the form to avoid
            # unintentionally setting them to NULL/default.
            preserve_fields = [
                'church', 'denomination', 'user_type', 'status',
                'year_term', 'organization', 'is_active', 'disabled_by',
            ]
            for f in preserve_fields:
                if f not in form.fields:
                    try:
                        setattr(obj, f, getattr(existing, f))
                    except Exception:
                        # If attribute missing on existing, skip
                        pass

            # Now save the instance (this will call full_clean inside model.save)
            obj.save()
            # Ensure self.object is set for the UpdateView lifecycle
            self.object = obj
            messages.success(self.request, 'Profile details updated successfully.')
            return redirect(self.get_success_url())
        except ModelValidationError as e:
            # Attach model validation errors to the form and show them to the user
            # so they can correct missing/invalid related fields instead of crashing.
            if hasattr(e, 'message_dict'):
                for field, errs in e.message_dict.items():
                    for err in errs:
                        form.add_error(field, err)
            else:
                form.add_error(None, str(e))

            messages.error(self.request, 'Please correct the errors below.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)



class ChurchAccountListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = CustomUser
    template_name = "church_account_list.html"
    context_object_name = "accounts"

    def test_func(self):
        return self.request.user.user_type in [
            CustomUser.UserType.ADMIN,
            CustomUser.UserType.CHURCH_ADMIN,
        ]

    def get_queryset(self):
        church_id = self.request.user.church_id

        # Always scope this page to the logged-in user's church
        if not church_id:
            return CustomUser.objects.none()

        return (
            CustomUser.objects.filter(
                church_id=church_id,
                user_type__in=[
                    CustomUser.UserType.TREASURER,
                    CustomUser.UserType.PASTOR,
                    CustomUser.UserType.MEMBER,
                ],
            )
            .select_related("church", "denomination")
            .order_by("user_type", "first_name", "last_name", "username")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        for user in context["accounts"]:
            user.computed_inactive_days = user.inactive_days

        context["current_church_id"] = self.request.user.church_id
        return context


class ToggleChurchAccountStatusView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.user_type in [
            CustomUser.UserType.ADMIN,
            CustomUser.UserType.CHURCH_ADMIN,
        ]

    def get_queryset(self):
        church_id = self.request.user.church_id

        if not church_id:
            return CustomUser.objects.none()

        return CustomUser.objects.filter(
            church_id=church_id,
            user_type__in=[
                CustomUser.UserType.TREASURER,
                CustomUser.UserType.PASTOR,
                CustomUser.UserType.MEMBER,
            ],
        )

    def post(self, request, pk):
        user = get_object_or_404(self.get_queryset(), pk=pk)
        action = request.POST.get("action")

        if action == "disable":
            user.disable_account(by_user=request.user)
            messages.success(request, f"{user.username} has been disabled.")
        elif action == "enable":
            user.enable_account(by_user=request.user)
            messages.success(request, f"{user.username} has been enabled.")
        else:
            messages.error(request, "Invalid action.")

        return redirect("church-account-list")



class PastorEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    template_name = 'pastor_edit.html'
    fields = [
        'username', 'email', 'first_name', 'middle_name', 'last_name',
        'birthdate', 'province', 'municipality_or_city', 'barangay', 'purok',
        'user_type', 'year_term', 'organization',
    ]
    success_url = reverse_lazy('church-account-list')

    def test_func(self):
        return self.request.user.user_type in [
            CustomUser.UserType.ADMIN,
            CustomUser.UserType.CHURCH_ADMIN,
        ]

    def get_queryset(self):
        qs = CustomUser.objects.filter(user_type=CustomUser.UserType.PASTOR)
        if self.request.user.user_type == CustomUser.UserType.CHURCH_ADMIN:
            qs = qs.filter(church_id=self.request.user.church_id)
        return qs


class MemberEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    template_name = 'member_edit.html'
    fields = [
        'username', 'email', 'first_name', 'middle_name', 'last_name',
        'birthdate', 'province', 'municipality_or_city', 'barangay', 'purok',
        'user_type', 'year_term', 'organization',
    ]
    success_url = reverse_lazy('church-account-list')

    def test_func(self):
        return self.request.user.user_type in [
            CustomUser.UserType.ADMIN,
            CustomUser.UserType.CHURCH_ADMIN,
        ]

    def get_queryset(self):
        qs = CustomUser.objects.filter(user_type=CustomUser.UserType.MEMBER)
        if self.request.user.user_type == CustomUser.UserType.CHURCH_ADMIN:
            qs = qs.filter(church_id=self.request.user.church_id)
        return qs


class TreasurerEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    template_name = 'treasurer_edit.html'
    fields = [
        'username', 'email', 'first_name', 'middle_name', 'last_name',
        'birthdate', 'province', 'municipality_or_city', 'barangay', 'purok',
        'user_type', 'year_term', 'organization',
    ]
    success_url = reverse_lazy('church-account-list')

    def test_func(self):
        return self.request.user.user_type in [
            CustomUser.UserType.ADMIN,
            CustomUser.UserType.CHURCH_ADMIN,
        ]

    def get_queryset(self):
        qs = CustomUser.objects.filter(user_type=CustomUser.UserType.TREASURER)
        if self.request.user.user_type == CustomUser.UserType.CHURCH_ADMIN:
            qs = qs.filter(church_id=self.request.user.church_id)
        return qs


class FinancialOverviewView(TemplateView):
    template_name = "financial_overview.html"

    # =========================================================
    #  PERMISSION HELPER
    # =========================================================
    def _can_manage_financial_actions(self, request):
        return getattr(request.user, "user_type", None) in {"ChurchAdmin", "Admin"}

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")

        if not getattr(request.user, "church_id", None):
            messages.error(request, "No church is assigned to your account.")
            return redirect("home")

        return super().dispatch(request, *args, **kwargs)

    # =========================================================
    #  DB / CURSOR HELPERS
    # =========================================================
    def _drain_cursor(self, cursor):
        try:
            cursor.fetchall()
        except Exception:
            pass

        try:
            while cursor.nextset():
                try:
                    cursor.fetchall()
                except Exception:
                    pass
        except Exception:
            pass

    def _runtime_db_check(self):
        required_tables = [
            "register_tithe",
            "register_offering",
            "register_donations",
            "register_otherincome",
            "register_expense",
            "register_bankaccount",
            "register_cashbankmovement",
            "church_church",
        ]

        settings_db = connection.settings_dict
        engine = settings_db.get("ENGINE")
        host = settings_db.get("HOST")
        port = settings_db.get("PORT")
        name = settings_db.get("NAME")
        user = settings_db.get("USER")

        with connection.cursor() as cursor:
            cursor.execute("SELECT DATABASE()")
            row = cursor.fetchone()
            current_db = row[0] if row else None

            cursor.execute("SHOW TABLES")
            all_tables = [r[0] for r in cursor.fetchall()]

            missing_tables = [t for t in required_tables if t not in all_tables]

            cursor.execute("SHOW TABLES LIKE 'django_migrations'")
            has_migrations_table = bool(cursor.fetchone())

            migration_count = 0
            if has_migrations_table:
                cursor.execute("SELECT COUNT(*) FROM django_migrations")
                migration_count = cursor.fetchone()[0] or 0

        if missing_tables:
            raise Exception(
                "Runtime DB check failed. "
                f"ENGINE={engine}, HOST={host}, PORT={port}, USER={user}, "
                f"SETTINGS_DB={name}, CURRENT_DB={current_db}, "
                f"DJANGO_MIGRATIONS_COUNT={migration_count}, "
                f"MISSING_TABLES={', '.join(missing_tables)}"
            )

        return current_db

    # =========================================================
    #  STORED PROCEDURE HELPERS (STRICT)
    # =========================================================
    def sp_one(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            row = cursor.fetchone()
            self._drain_cursor(cursor)
        return row

    def sp_all(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            rows = cursor.fetchall()
            self._drain_cursor(cursor)
        return rows or []

    # =========================================================
    #  NORMALIZE TRANSACTION TYPE
    # =========================================================
    def _normalize_transaction_type(self, transaction_type):
        tx = (transaction_type or "").strip()

        aliases = {
            "Tithe": "Tithe",
            "TITHE": "Tithe",
            "Offering": "Offering",
            "OFFERING": "Offering",
            "Donations": "Donations",
            "Donation": "Donations",
            "DONATION": "Donations",
            "DONATIONS": "Donations",
            "OtherIncome": "OtherIncome",
            "Other Income": "OtherIncome",
            "OTHER_INCOME": "OtherIncome",
            "Expense": "Expense",
            "EXPENSE": "Expense",
            "Transfer": "Transfer",
            "TRANSFER": "Transfer",
            "ReleasedBudget": "ReleasedBudget",
        }
        return aliases.get(tx, tx)

    # =========================================================
    #  HELPER: Detect system-generated budget transactions
    # =========================================================
    def _is_budget_generated_transaction(self, transaction_type, transaction_id, church_id) -> bool:
        tx_type = self._normalize_transaction_type(transaction_type)

        if tx_type in {"Expense", "Transfer"}:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT
                        COALESCE(ec.is_system, 0) AS is_system,
                        COALESCE(ec.is_transfer, 0) AS is_transfer,
                        LOWER(TRIM(COALESCE(ec.name, ''))) AS cat_name,
                        LOWER(TRIM(COALESCE(e.description, ''))) AS exp_desc
                    FROM register_expense e
                    LEFT JOIN register_expensecategory ec
                           ON ec.id = e.category_id
                    WHERE e.id = %s
                      AND e.church_id = %s
                    LIMIT 1
                    """,
                    [transaction_id, church_id],
                )
                row = cursor.fetchone()

            if not row:
                return False

            is_system = int(row[0] or 0)
            is_transfer = int(row[1] or 0)
            cat_name = (row[2] or "").strip()
            exp_desc = (row[3] or "").strip()

            if is_system == 1:
                return True

            if cat_name in {
                "budget release",
                "budget release - cash",
                "budget release - bank",
                "budget return",
            }:
                return True

            if exp_desc.startswith("budget return deposit"):
                return True

            if exp_desc.startswith("budget return to bank"):
                return True

            if exp_desc.startswith("budget release - bank"):
                return True

            if exp_desc.startswith("budget release - cash"):
                return True

            if is_transfer == 1 and ("budget return" in exp_desc or "budget return" in cat_name):
                return True

            return False

        if tx_type == "OtherIncome":
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT
                        LOWER(TRIM(COALESCE(oic.name, ''))) AS income_cat_name,
                        LOWER(TRIM(COALESCE(oi.description, ''))) AS income_desc,
                        rb.id AS linked_release_id
                    FROM register_otherincome oi
                    LEFT JOIN register_otherincomecategory oic
                           ON oic.id = oi.income_type_id
                    LEFT JOIN register_releasedbudget rb
                           ON rb.budget_return_income_id = oi.id
                          AND rb.church_id = %s
                    WHERE oi.id = %s
                      AND oi.church_id = %s
                    LIMIT 1
                    """,
                    [church_id, transaction_id, church_id],
                )
                row = cursor.fetchone()

            if not row:
                return False

            income_cat_name = (row[0] or "").strip()
            income_desc = (row[1] or "").strip()
            linked_release_id = row[2]

            if linked_release_id:
                return True

            if income_cat_name == "budget return":
                return True

            if income_desc.startswith("budget return from"):
                return True

            return False

        return False

    # =========================================================
    #  HELPER: CASH/BANK MOVEMENT SYNC
    # =========================================================
    def _source_type_map(self, transaction_type):
        transaction_type = self._normalize_transaction_type(transaction_type)

        return {
            "Tithe": "TITHE",
            "Offering": "OFFERING",
            "Donations": "DONATION",
            "OtherIncome": "OTHER_INCOME",
            "Expense": "EXPENSE",
            "Transfer": "EXPENSE",
        }.get(transaction_type)

    def _default_direction_for_type(self, transaction_type):
        transaction_type = self._normalize_transaction_type(transaction_type)

        return {
            "Tithe": "CASH_IN",
            "Offering": "CASH_IN",
            "Donations": "CASH_IN",
            "OtherIncome": "CASH_IN",
            "Expense": "CASH_OUT",
            "Transfer": "CASH_OUT",
        }.get(transaction_type, "CASH_IN")

    def _get_file_name(self, obj, attr_name):
        f = getattr(obj, attr_name, None)
        if not f:
            return None
        return getattr(f, "name", f)

    def _source_tx_fields(self, transaction_type, obj):
        transaction_type = self._normalize_transaction_type(transaction_type)

        if transaction_type == "Tithe":
            return {
                "date": obj.date,
                "memo": getattr(obj, "description", "") or "",
                "proof_file": self._get_file_name(obj, "file"),
                "reference_no": None,
            }

        if transaction_type == "Offering":
            return {
                "date": obj.date,
                "memo": getattr(obj, "description", "") or "",
                "proof_file": self._get_file_name(obj, "proof_document"),
                "reference_no": None,
            }

        if transaction_type == "Donations":
            return {
                "date": obj.donations_date,
                "memo": (obj.other_donations_type or obj.donor or ""),
                "proof_file": self._get_file_name(obj, "file"),
                "reference_no": getattr(obj, "gcash_reference_number", None),
            }

        if transaction_type == "OtherIncome":
            return {
                "date": obj.date,
                "memo": obj.description or "",
                "proof_file": self._get_file_name(obj, "file"),
                "reference_no": None,
            }

        if transaction_type in {"Expense", "Transfer"}:
            return {
                "date": obj.expense_date,
                "memo": obj.description or "",
                "proof_file": self._get_file_name(obj, "file"),
                "reference_no": None,
            }

        return {
            "date": None,
            "memo": "",
            "proof_file": None,
            "reference_no": None,
        }

    def _recalculate_bank_balance(self, church_id):
        with connection.cursor() as cursor:
            cursor.callproc("sp_recalculate_bank_balance", [church_id])
            self._drain_cursor(cursor)

    def _sync_cashbankmovement(self, transaction_type, obj, movement_data=None):
        CashBankMovement = apps.get_model("Register", "CashBankMovement")
        transaction_type = self._normalize_transaction_type(transaction_type)
        source_type = self._source_type_map(transaction_type)

        if not source_type:
            return

        movement_data = movement_data or {}
        source_vals = self._source_tx_fields(transaction_type, obj)

        movement = CashBankMovement.objects.select_for_update().filter(
            church_id=obj.church_id,
            source_type=source_type,
            source_id=obj.id,
        ).first()

        direction_override = movement_data.get("direction")
        memo_override = movement_data.get("memo")
        reference_no_override = movement_data.get("reference_no")
        status_override = movement_data.get("status")

        final_direction = (
            direction_override
            if direction_override not in (None, "")
            else (movement.direction if movement else self._default_direction_for_type(transaction_type))
        )
        final_memo = memo_override if memo_override is not None else source_vals["memo"]
        final_reference_no = (
            reference_no_override if reference_no_override is not None else source_vals["reference_no"]
        )
        final_status = (
            status_override
            if status_override not in (None, "")
            else (movement.status if movement else "CONFIRMED")
        )

        if movement:
            movement.date = source_vals["date"]
            movement.amount = obj.amount
            movement.direction = final_direction
            movement.memo = final_memo
            movement.reference_no = final_reference_no
            movement.status = final_status

            if source_vals["proof_file"] is not None:
                movement.proof_file = source_vals["proof_file"]

            movement.save(
                update_fields=[
                    "date",
                    "amount",
                    "direction",
                    "memo",
                    "reference_no",
                    "status",
                    "proof_file",
                ]
            )
        else:
            CashBankMovement.objects.create(
                date=source_vals["date"],
                amount=obj.amount,
                direction=final_direction,
                source_type=source_type,
                source_id=obj.id,
                memo=final_memo,
                reference_no=final_reference_no,
                proof_file=source_vals["proof_file"],
                church_id=obj.church_id,
                created_by_id=getattr(obj, "created_by_id", None),
                status=final_status,
            )

        self._recalculate_bank_balance(obj.church_id)

    def _delete_cashbankmovement(self, transaction_type, transaction_id, church_id):
        CashBankMovement = apps.get_model("Register", "CashBankMovement")
        source_type = self._source_type_map(transaction_type)

        if not source_type:
            return

        CashBankMovement.objects.filter(
            church_id=church_id,
            source_type=source_type,
            source_id=transaction_id,
        ).delete()

        self._recalculate_bank_balance(church_id)

    # =========================================================
    #  HELPER: resolve type_id/category_id
    # =========================================================
    def _resolve_type_id_for_transaction(self, transaction_type, transaction_id, church_id):
        transaction_type = self._normalize_transaction_type(transaction_type)

        table_map = {
            "Donations": ("register_donations", "donations_type_id"),
            "OtherIncome": ("register_otherincome", "income_type_id"),
            "Expense": ("register_expense", "category_id"),
            "Transfer": ("register_expense", "category_id"),
        }

        if transaction_type not in table_map or not church_id:
            return None

        table_name, col_name = table_map[transaction_type]

        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT {col_name}
                FROM {table_name}
                WHERE id = %s AND church_id = %s
                LIMIT 1
                """,
                [transaction_id, church_id],
            )
            row = cursor.fetchone()

        return row[0] if row else None

    # =========================================================
    #  MAIN CONTEXT
    # =========================================================
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        church_id = self.request.user.church_id

        context["can_manage_financial_actions"] = self._can_manage_financial_actions(self.request)

        filter_date = self.request.GET.get("filter_date")
        monthly_filter_date = self.request.GET.get("monthly_filter_date")

        try:
            current_db = self._runtime_db_check()
            context["runtime_db_name"] = current_db

            # =========================================================
            #  A) ALWAYS-PRESENT TOP CARDS (SYSTEM TOTALS)
            # =========================================================
            res_phys = self.sp_one("Calculate_CashOnHand", [church_id])
            physical_cash = Decimal(res_phys[0] or 0) if res_phys and res_phys[0] is not None else Decimal("0.00")

            fund_row = self.sp_one("Finance_FundBalancesTotal", [church_id])

            if fund_row:
                unres_balance = Decimal(fund_row[2] or 0)
                res_balance = Decimal(fund_row[5] or 0)
                total_fund_balance = Decimal(fund_row[6] or 0)

                fund_stats = {
                    "unrestricted_income": Decimal(fund_row[0] or 0),
                    "unrestricted_expenses": Decimal(fund_row[1] or 0),
                    "unrestricted_balance": unres_balance,
                    "restricted_income": Decimal(fund_row[3] or 0),
                    "restricted_expenses": Decimal(fund_row[4] or 0),
                    "restricted_balance": res_balance,
                    "total_fund_balance": total_fund_balance,
                }
            else:
                unres_balance = Decimal("0.00")
                res_balance = Decimal("0.00")
                total_fund_balance = Decimal("0.00")
                fund_stats = {
                    "unrestricted_income": Decimal("0.00"),
                    "unrestricted_expenses": Decimal("0.00"),
                    "unrestricted_balance": Decimal("0.00"),
                    "restricted_income": Decimal("0.00"),
                    "restricted_expenses": Decimal("0.00"),
                    "restricted_balance": Decimal("0.00"),
                    "total_fund_balance": Decimal("0.00"),
                }

            bank_balance = Decimal("0.00")
            bank_name = "No Bank Account"
            bank_proof_url = None

            BankAccount = apps.get_model("Register", "BankAccount")
            bank = BankAccount.objects.filter(church_id=church_id).first()

            if bank:
                bank_balance = bank.current_balance or Decimal("0.00")
                bank_name = bank.bank_name or "Bank"
                if getattr(bank, "verification_image", None):
                    try:
                        bank_proof_url = bank.verification_image.url
                    except Exception:
                        bank_proof_url = None

            total_cash_position = Decimal(physical_cash) + Decimal(bank_balance)

            # Keep unrestricted SP only for unrestricted breakdown
            row_unrestricted = self.sp_one("Finance_UnrestrictedNet", [church_id])
            grand_total_unrestricted = Decimal("0.00")

            if row_unrestricted:
                gross_income = Decimal(row_unrestricted[5] or 0) if len(row_unrestricted) > 5 else Decimal("0.00")
                unres_expenses = Decimal(row_unrestricted[6] or 0) if len(row_unrestricted) > 6 else Decimal("0.00")
                net_unrestricted = Decimal(row_unrestricted[7] or 0) if len(row_unrestricted) > 7 else Decimal("0.00")

                grand_total_unrestricted = gross_income

                unrestricted_stats = {
                    "tithes": Decimal(row_unrestricted[0] or 0),
                    "offerings": Decimal(row_unrestricted[1] or 0),
                    "donations": Decimal(row_unrestricted[2] or 0),
                    "other_income": Decimal(row_unrestricted[3] or 0),
                    "budget_returns": Decimal(row_unrestricted[4] or 0),
                    "gross_income": gross_income,
                    "total_expenses": unres_expenses,
                    "net_unrestricted": net_unrestricted,
                    "total_income": gross_income,
                }
            else:
                unrestricted_stats = {
                    "tithes": Decimal("0.00"),
                    "offerings": Decimal("0.00"),
                    "donations": Decimal("0.00"),
                    "other_income": Decimal("0.00"),
                    "budget_returns": Decimal("0.00"),
                    "gross_income": Decimal("0.00"),
                    "total_income": Decimal("0.00"),
                    "total_expenses": Decimal("0.00"),
                    "net_unrestricted": Decimal("0.00"),
                }

            cash_on_hand = physical_cash
            restricted_funds = res_balance
            unrestricted_balance = unres_balance

            # =========================================================
            #  CORRECT BUSINESS RULE:
            #  Church Income = Restricted Funds + Unrestricted Balance
            #  Equivalent to total_fund_balance
            # =========================================================
            church_income = restricted_funds + unrestricted_balance

            context["physical_cash"] = physical_cash
            context["cash_on_hand"] = cash_on_hand
            context["restricted_funds"] = restricted_funds
            context["unrestricted_balance"] = unrestricted_balance
            context["total_fund_balance"] = total_fund_balance
            context["fund_stats"] = fund_stats
            context["unrestricted_stats"] = unrestricted_stats
            context["bank_balance"] = bank_balance
            context["bank_name"] = bank_name
            context["bank_proof_url"] = bank_proof_url
            context["total_funds"] = total_cash_position
            context["church_income"] = church_income
            context["grand_total_unrestricted"] = grand_total_unrestricted

            # =========================================================
            #  B) FILTERED VIEWS
            # =========================================================
            filtered_summary = []
            filtered_transactions = {}
            combined_data = []

            if monthly_filter_date:
                monthly_start_date = f"{monthly_filter_date}-01"

                filtered_summary_result = self.sp_all("Filter_Monthly_Summary", [monthly_start_date, church_id])
                filtered_summary = [
                    {
                        "total_tithes": row[2],
                        "total_offering": row[3],
                        "total_donations": row[4],
                        "total_other_income": row[5] if len(row) > 6 else 0,
                        "total_expenses": row[6] if len(row) > 6 else row[5],
                        "total_accumulated": row[7] if len(row) > 6 else row[6],
                    }
                    for row in filtered_summary_result
                ]

                monthly_results = self.sp_all("Finance_MonthlyTransactions", [monthly_start_date, church_id])
                filtered_transactions = self.parse_transaction_rows(monthly_results, church_id=church_id)

                context["filtered_summary"] = filtered_summary
                context["filtered_transactions"] = filtered_transactions
                context["monthly_filter_date"] = monthly_filter_date

            elif filter_date:
                filtered_summary_result = self.sp_all("Filtered_Financial_Summary", [filter_date, church_id])
                filtered_summary = [
                    {
                        "transaction_date": row[0],
                        "total_tithes": row[1],
                        "total_offering": row[2],
                        "total_donations": row[3],
                        "total_other_income": row[4],
                        "total_expenses": row[5],
                        "total_accumulated": row[6],
                    }
                    for row in filtered_summary_result
                ]

                filtered_results = self.sp_all("Finance_FilterByDate", [filter_date, church_id])
                filtered_transactions = self.parse_transaction_rows(filtered_results, church_id=church_id)

                context["filtered_summary"] = filtered_summary
                context["filtered_transactions"] = filtered_transactions
                context["filter_date"] = filter_date

            else:
                daily_summary_results = self.sp_all("Daily_Financial_Summary", [church_id])
                daily_summaries = [
                    {
                        "transaction_date": row[0],
                        "total_tithes": row[1],
                        "total_offering": row[2],
                        "total_donations": row[3],
                        "total_other_income": row[4],
                        "total_expenses": row[5],
                        "total_accumulated": row[6],
                    }
                    for row in daily_summary_results
                ]

                finance_sorted_results = self.sp_all("Finance_SortedbyDate", [church_id])
                transactions = defaultdict(list)

                parsed_rows = self.parse_transaction_rows(finance_sorted_results, church_id=church_id)
                for tx in parsed_rows:
                    transactions[tx["transaction_date"]].append(tx)

                for summary in daily_summaries:
                    date = summary["transaction_date"]
                    combined_data.append(
                        {
                            "date": date,
                            "summary": summary,
                            "transactions": transactions.get(date, []),
                        }
                    )

                context["combined_data"] = combined_data

        except Exception as e:
            logger.exception("FinancialOverviewView crashed for church_id=%s: %s", church_id, e)
            messages.error(self.request, f"Financial overview failed to load: {e}")

            context.setdefault("runtime_db_name", None)
            context.setdefault("physical_cash", Decimal("0.00"))
            context.setdefault("cash_on_hand", Decimal("0.00"))
            context.setdefault("restricted_funds", Decimal("0.00"))
            context.setdefault("unrestricted_balance", Decimal("0.00"))
            context.setdefault("total_fund_balance", Decimal("0.00"))
            context.setdefault("fund_stats", {
                "unrestricted_income": Decimal("0.00"),
                "unrestricted_expenses": Decimal("0.00"),
                "unrestricted_balance": Decimal("0.00"),
                "restricted_income": Decimal("0.00"),
                "restricted_expenses": Decimal("0.00"),
                "restricted_balance": Decimal("0.00"),
                "total_fund_balance": Decimal("0.00"),
            })
            context.setdefault("unrestricted_stats", {
                "tithes": Decimal("0.00"),
                "offerings": Decimal("0.00"),
                "donations": Decimal("0.00"),
                "other_income": Decimal("0.00"),
                "budget_returns": Decimal("0.00"),
                "gross_income": Decimal("0.00"),
                "total_income": Decimal("0.00"),
                "total_expenses": Decimal("0.00"),
                "net_unrestricted": Decimal("0.00"),
            })
            context.setdefault("bank_balance", Decimal("0.00"))
            context.setdefault("bank_name", "No Bank Account")
            context.setdefault("bank_proof_url", None)
            context.setdefault("total_funds", Decimal("0.00"))
            context.setdefault("church_income", Decimal("0.00"))
            context.setdefault("grand_total_unrestricted", Decimal("0.00"))
            context.setdefault("filtered_summary", [])
            context.setdefault("filtered_transactions", {})
            context.setdefault("combined_data", [])

        return context

    # =========================================================
    #  HELPER: Parse Rows
    # =========================================================
    def parse_transaction_rows(self, rows, church_id=None):
        data = []

        for row in rows:
            tx_type = self._normalize_transaction_type(row[0])
            transaction_id = row[1]

            is_budget_locked = (
                self._is_budget_generated_transaction(tx_type, transaction_id, church_id)
                if church_id else False
            )

            if len(row) >= 21:
                type_id = (
                    self._resolve_type_id_for_transaction(tx_type, transaction_id, church_id)
                    if church_id else None
                )

                parsed = {
                    "transaction_type": tx_type,
                    "transaction_id": transaction_id,
                    "transaction_date": row[2],
                    "typeof": row[3],
                    "others": row[4],
                    "amount": row[5],
                    "type_id": type_id,
                    "created_by_id": row[6],
                    "created_by": row[7],
                    "edited_by_id": row[8],
                    "edited_by": row[9],
                    "donor": row[10],
                    "vendor": row[11],
                    "file": row[12],
                    "movement_id": row[13],
                    "movement_direction": row[14],
                    "movement_status": row[15],
                    "movement_memo": row[16],
                    "movement_reference_no": row[17],
                    "movement_proof_file": row[18],
                    "movement_date": row[19],
                    "movement_amount": row[20],
                    "is_budget_locked": is_budget_locked,
                }
            else:
                parsed = {
                    "transaction_type": tx_type,
                    "transaction_id": transaction_id,
                    "transaction_date": row[2] if len(row) > 2 else None,
                    "typeof": row[3] if len(row) > 3 else None,
                    "others": row[4] if len(row) > 4 else None,
                    "amount": row[5] if len(row) > 5 else None,
                    "type_id": row[6] if len(row) > 6 else None,
                    "created_by": row[7] if len(row) > 7 else None,
                    "edited_by": row[9] if len(row) > 9 else None,
                    "donor": row[10] if len(row) > 10 else None,
                    "vendor": None,
                    "file": row[11] if len(row) > 11 else None,
                    "movement_id": row[12] if len(row) > 12 else None,
                    "movement_direction": row[13] if len(row) > 13 else None,
                    "movement_status": row[14] if len(row) > 14 else None,
                    "movement_memo": row[15] if len(row) > 15 else None,
                    "movement_reference_no": row[16] if len(row) > 16 else None,
                    "movement_proof_file": row[17] if len(row) > 17 else None,
                    "movement_date": row[18] if len(row) > 18 else None,
                    "movement_amount": row[19] if len(row) > 19 else None,
                    "is_budget_locked": is_budget_locked,
                }

            data.append(parsed)

        return data

    # =========================================================
    #  HELPER: LOCK VALIDATOR
    # =========================================================
    def check_lock_status(self, request, transaction_id, transaction_type, church_id):
        transaction_type = self._normalize_transaction_type(transaction_type)

        table_map = {
            "Tithe": ("register_tithe", "date"),
            "Offering": ("register_offering", "date"),
            "Donations": ("register_donations", "donations_date"),
            "Expense": ("register_expense", "expense_date"),
            "Transfer": ("register_expense", "expense_date"),
            "OtherIncome": ("register_otherincome", "date"),
        }

        if transaction_type not in table_map:
            return False, f"Invalid transaction type: {transaction_type}"

        table_name, date_col = table_map[transaction_type]

        transaction_date = None
        with connection.cursor() as cursor:
            query = f"SELECT {date_col} FROM {table_name} WHERE id = %s AND church_id = %s"
            cursor.execute(query, [transaction_id, church_id])
            result = cursor.fetchone()
            if result:
                transaction_date = result[0]

        if not transaction_date:
            return False, "Transaction not found."

        try:
            settings = apps.get_model("Register", "AccountingSettings").objects.get(church_id=church_id)
            if settings.lock_date and transaction_date <= settings.lock_date:
                return False, f"ACCESS DENIED: The books are closed for {settings.lock_date}."
        except Exception:
            pass

        today = timezone.now().date()
        future_limit = today + timedelta(days=7)
        if transaction_date > future_limit:
            return False, f"ACCESS DENIED: Cannot modify records dated past {future_limit}."

        return True, None

    def post(self, request, *args, **kwargs):
        if not self._can_manage_financial_actions(request):
            messages.error(request, "Only ChurchAdmin or Admin can perform financial actions.")
            return redirect("financial_overview")

        if "edit" in request.POST:
            return self.edit_transaction(request)
        elif "delete" in request.POST:
            return self.delete_transaction(request)
        elif "download_excel" in request.POST:
            return self.download_financial_report()

        return redirect("financial_overview")

    # =========================================================
    #  EDIT TRANSACTION
    # =========================================================
    def edit_transaction(self, request):
        if not self._can_manage_financial_actions(request):
            messages.error(request, "Only ChurchAdmin or Admin can edit transactions.")
            return redirect("financial_overview")

        transaction_id = request.POST.get("transaction_id")
        transaction_type = self._normalize_transaction_type(request.POST.get("transaction_type"))
        church_id = request.user.church_id

        is_allowed, error_msg = self.check_lock_status(request, transaction_id, transaction_type, church_id)
        if not is_allowed:
            messages.error(request, error_msg)
            return redirect("financial_overview")

        if self._is_budget_generated_transaction(transaction_type, transaction_id, church_id):
            messages.error(
                request,
                "You cannot edit this system-generated budget entry here. "
                "Use Release Budget or Budget Expenses (Liquidate)."
            )
            return redirect("financial_overview")

        amount = request.POST.get("amount")
        typeof = request.POST.get("typeof")
        others = request.POST.get("others")
        edited_by = request.user

        movement_data = {
            "direction": request.POST.get("movement_direction"),
            "memo": request.POST.get("movement_memo"),
            "reference_no": request.POST.get("movement_reference_no"),
        }

        Tithe = apps.get_model("Register", "Tithe")
        Offering = apps.get_model("Register", "Offering")
        Donations = apps.get_model("Register", "Donations")
        OtherIncome = apps.get_model("Register", "OtherIncome")
        Expense = apps.get_model("Register", "Expense")

        try:
            with transaction.atomic():
                obj = None

                if transaction_type == "Tithe":
                    obj = Tithe.objects.select_for_update().get(id=transaction_id, church_id=church_id)
                    obj.amount = amount
                    obj.edited_by = edited_by
                    obj.save(update_fields=["amount", "edited_by"])

                elif transaction_type == "Offering":
                    obj = Offering.objects.select_for_update().get(id=transaction_id, church_id=church_id)
                    obj.amount = amount
                    obj.edited_by = edited_by
                    obj.save(update_fields=["amount", "edited_by"])

                elif transaction_type == "Donations":
                    obj = Donations.objects.select_for_update().get(id=transaction_id, church_id=church_id)
                    obj.amount = amount
                    obj.donations_type_id = typeof
                    obj.other_donations_type = others
                    obj.edited_by = edited_by
                    obj.save(update_fields=["amount", "donations_type_id", "other_donations_type", "edited_by"])

                elif transaction_type == "OtherIncome":
                    obj = OtherIncome.objects.select_for_update().get(id=transaction_id, church_id=church_id)
                    obj.amount = amount
                    obj.income_type_id = typeof
                    obj.description = others
                    obj.edited_by = edited_by
                    obj.save(update_fields=["amount", "income_type_id", "description", "edited_by"])

                elif transaction_type in {"Expense", "Transfer"}:
                    obj = Expense.objects.select_for_update().get(id=transaction_id, church_id=church_id)
                    obj.amount = amount
                    obj.category_id = typeof
                    obj.description = others
                    obj.edited_by = edited_by
                    obj.save(update_fields=["amount", "category_id", "description", "edited_by"])

                else:
                    raise ValueError(f"Unsupported transaction type: {transaction_type}")

                if obj:
                    self._sync_cashbankmovement(transaction_type, obj, movement_data=movement_data)

            messages.success(request, "Transaction and linked cash/bank movement updated successfully.")
        except Exception as e:
            messages.error(request, f"Update failed: {e}")

        return redirect("financial_overview")

    # =========================================================
    #  DELETE TRANSACTION
    # =========================================================
    def delete_transaction(self, request):
        if not self._can_manage_financial_actions(request):
            messages.error(request, "Only ChurchAdmin or Admin can delete transactions.")
            return redirect("financial_overview")

        transaction_id = request.POST.get("transaction_id")
        transaction_type = self._normalize_transaction_type(request.POST.get("transaction_type"))
        church_id = request.user.church_id

        is_allowed, error_msg = self.check_lock_status(request, transaction_id, transaction_type, church_id)
        if not is_allowed:
            messages.error(request, error_msg)
            return redirect("financial_overview")

        if self._is_budget_generated_transaction(transaction_type, transaction_id, church_id):
            messages.error(
                request,
                "You cannot delete this system-generated budget entry here."
            )
            return redirect("financial_overview")

        Tithe = apps.get_model("Register", "Tithe")
        Offering = apps.get_model("Register", "Offering")
        Donations = apps.get_model("Register", "Donations")
        OtherIncome = apps.get_model("Register", "OtherIncome")
        Expense = apps.get_model("Register", "Expense")

        try:
            with transaction.atomic():
                deleted_count = 0

                if transaction_type == "Tithe":
                    deleted_count, _ = Tithe.objects.filter(
                        id=transaction_id,
                        church_id=church_id
                    ).delete()

                elif transaction_type == "Offering":
                    deleted_count, _ = Offering.objects.filter(
                        id=transaction_id,
                        church_id=church_id
                    ).delete()

                elif transaction_type == "Donations":
                    deleted_count, _ = Donations.objects.filter(
                        id=transaction_id,
                        church_id=church_id
                    ).delete()

                elif transaction_type == "OtherIncome":
                    deleted_count, _ = OtherIncome.objects.filter(
                        id=transaction_id,
                        church_id=church_id
                    ).delete()

                elif transaction_type in {"Expense", "Transfer"}:
                    deleted_count, _ = Expense.objects.filter(
                        id=transaction_id,
                        church_id=church_id
                    ).delete()

                else:
                    raise ValueError(f"Unsupported transaction type: {transaction_type}")

                if deleted_count == 0:
                    raise ValueError(
                        f"Source record was not deleted. "
                        f"transaction_type={transaction_type}, transaction_id={transaction_id}, church_id={church_id}"
                    )

                self._delete_cashbankmovement(transaction_type, transaction_id, church_id)

            messages.success(request, "Transaction and linked cash/bank movement deleted successfully.")
        except Exception as e:
            messages.error(request, f"Delete failed: {e}")

        return redirect("financial_overview")

    def download_financial_report(self):
        if not self._can_manage_financial_actions(self.request):
            messages.error(self.request, "Only ChurchAdmin or Admin can download the financial report.")
            return redirect("financial_overview")

        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception as e:
            messages.error(self.request, f"Excel export is unavailable: {e}")
            return redirect("financial_overview")

        try:
            church_id = self.request.user.church_id
            daily_summary_results = self.sp_all("Daily_Financial_Summary", [church_id])
            finance_sorted_results = self.sp_all("Finance_SortedbyDate", [church_id])
        except Exception as e:
            messages.error(self.request, f"Error generating report: {e}")
            return redirect("financial_overview")

        workbook = openpyxl.Workbook()

        daily_sheet = workbook.active
        daily_sheet.title = "Daily Financial Summary"

        daily_headers = [
            "Transaction Date",
            "Total Tithes",
            "Total Offering",
            "Total Donations",
            "Total Other Income",
            "Total Expenses",
            "Total Accumulated",
        ]
        for col_num, header in enumerate(daily_headers, 1):
            cell = daily_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        for row_num, row_data in enumerate(daily_summary_results, 2):
            for col_num, value in enumerate(row_data, 1):
                daily_sheet.cell(row=row_num, column=col_num, value=value)

        sorted_sheet = workbook.create_sheet(title="Finance Sorted by Date")
        sorted_headers = [
            "Transaction Type",
            "Transaction ID",
            "Transaction Date",
            "Type",
            "Others",
            "Amount",
            "Created By User ID",
            "Created By",
            "Edited By User ID",
            "Edited By",
            "Donor",
            "Vendor",
            "File",
            "Movement ID",
            "Movement Direction",
            "Movement Status",
            "Movement Memo",
            "Movement Reference No",
            "Movement Proof File",
            "Movement Date",
            "Movement Amount",
        ]

        for col_num, header in enumerate(sorted_headers, 1):
            cell = sorted_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        for row_num, row_data in enumerate(finance_sorted_results, 2):
            export_row = [
                row_data[0] if len(row_data) > 0 else None,
                row_data[1] if len(row_data) > 1 else None,
                row_data[2] if len(row_data) > 2 else None,
                row_data[3] if len(row_data) > 3 else None,
                row_data[4] if len(row_data) > 4 else None,
                row_data[5] if len(row_data) > 5 else None,
                row_data[6] if len(row_data) > 6 else None,
                row_data[7] if len(row_data) > 7 else None,
                row_data[8] if len(row_data) > 8 else None,
                row_data[9] if len(row_data) > 9 else None,
                row_data[10] if len(row_data) > 10 else None,
                row_data[11] if len(row_data) > 11 else None,
                row_data[12] if len(row_data) > 12 else None,
                row_data[13] if len(row_data) > 13 else None,
                row_data[14] if len(row_data) > 14 else None,
                row_data[15] if len(row_data) > 15 else None,
                row_data[16] if len(row_data) > 16 else None,
                row_data[17] if len(row_data) > 17 else None,
                row_data[18] if len(row_data) > 18 else None,
                row_data[19] if len(row_data) > 19 else None,
                row_data[20] if len(row_data) > 20 else None,
            ]

            for col_num, value in enumerate(export_row, 1):
                sorted_sheet.cell(row=row_num, column=col_num, value=value)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="financial_report.xlsx"'
        workbook.save(response)
        return response

class WeeklyReportView(LoginRequiredMixin, TemplateView):
    template_name = "weekly_report.html"

    # =========================================================
    #  STORED PROCEDURE HELPERS
    # =========================================================
    def sp_one(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            row = cursor.fetchone()

            try:
                cursor.fetchall()
            except Exception:
                pass

            while cursor.nextset():
                try:
                    cursor.fetchall()
                except Exception:
                    pass

        return row

    def sp_all(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            rows = cursor.fetchall()

            try:
                cursor.fetchall()
            except Exception:
                pass

            while cursor.nextset():
                try:
                    cursor.fetchall()
                except Exception:
                    pass

        return rows

    # =========================================================
    #  PERIOD LABEL FORMATTERS
    # =========================================================
    def _format_period_label(self, report_type, label):
        raw = str(label or "").strip()
        if not raw:
            return raw

        if report_type == "weekly":
            try:
                year_str, week_str = raw.split("-")
                year = int(year_str)
                week = int(week_str)

                start_date = date.fromisocalendar(year, week, 1)  # Monday
                end_date = date.fromisocalendar(year, week, 7)    # Sunday

                if start_date.month == end_date.month and start_date.year == end_date.year:
                    return f"{start_date.strftime('%b %d')}–{end_date.strftime('%d, %Y')}"
                elif start_date.year == end_date.year:
                    return f"{start_date.strftime('%b %d')}–{end_date.strftime('%b %d, %Y')}"
                else:
                    return f"{start_date.strftime('%b %d, %Y')}–{end_date.strftime('%b %d, %Y')}"
            except Exception:
                return raw

        if report_type == "monthly":
            try:
                year_str, month_str = raw.split("-")
                year = int(year_str)
                month = int(month_str)
                return date(year, month, 1).strftime("%B %Y")
            except Exception:
                return raw

        if report_type == "yearly":
            return raw

        return raw

    def _register_period(self, sort_key, label, available_periods, seen_sort_keys, label_to_sortkey, report_type):
        sort_key = str(sort_key or "")
        raw_label = str(label or "")

        if not sort_key or not raw_label:
            return

        display_label = self._format_period_label(report_type, raw_label)

        if sort_key not in seen_sort_keys:
            available_periods.append(display_label)
            seen_sort_keys.add(sort_key)

        label_to_sortkey[display_label] = sort_key

    # =========================================================
    #  MEMBER LOCK VISIBILITY HELPERS
    # =========================================================
    def _get_period_end_date(self, report_type, raw_label):
        raw = str(raw_label or "").strip()
        if not raw:
            return None

        try:
            if report_type == "weekly":
                year_str, week_str = raw.split("-")
                year = int(year_str)
                week = int(week_str)
                return date.fromisocalendar(year, week, 7)  # Sunday

            if report_type == "monthly":
                year_str, month_str = raw.split("-")
                year = int(year_str)
                month = int(month_str)
                last_day = monthrange(year, month)[1]
                return date(year, month, last_day)

            if report_type == "yearly":
                year = int(raw)
                return date(year, 12, 31)

        except Exception:
            return None

        return None

    def _can_view_period(self, user, report_type, raw_label, lock_date):
        """
        Members can only view a period if the entire period is already locked.
        Admin/staff/other roles are not restricted here.
        """
        if getattr(user, "user_type", "") != "Member":
            return True

        if not lock_date:
            return False

        period_end = self._get_period_end_date(report_type, raw_label)
        if not period_end:
            return False

        return period_end <= lock_date

    # =========================================================
    #  NORMALIZE REPORT TX TYPE
    # =========================================================
    def _normalize_report_tx_type(self, txn_type):
        tx = (txn_type or "").strip()
        key = tx.upper()

        aliases = {
            "TITHE": "Tithe",
            "OFFERING": "Offering",
            "UNRES_DONATION": "Unres_Donation",
            "UNRES_OTHER": "Unres_Other",
            "RES_DONATION": "Res_Donation",
            "RES_OTHER": "Res_Other",
            "GEN_EXPENSE": "Gen_Expense",
            "RES_EXPENSE": "Res_Expense",

            # Monthly / Yearly SP explicit non-fund-flow rows
            "TRANSFER_DEPOSIT": "Transfer_Deposit",
            "TRANSFER_WITHDRAW": "Transfer_Withdraw",
            "SYSTEM_BUDGETRELEASE": "System_BudgetRelease",
            "SYSTEM_BUDGETRETURN": "System_BudgetReturn",
        }
        return aliases.get(key, tx)

    # =========================================================
    #  CATEGORY DETECTORS
    # =========================================================
    def _is_budget_return_label(self, category_name: str) -> bool:
        c = (category_name or "").strip().lower()
        if not c:
            return False

        return (
            c == "budget return"
            or c.startswith("budget return:")
            or c.startswith("budget return -")
            or c.startswith("cash:budget return:")
            or c.startswith("bank:budget return:")
            or c.startswith("unposted:budget return:")
            or ":budget return:" in c
        )

    def _is_visible_budget_release_label(self, category_name: str) -> bool:
        c = (category_name or "").strip().lower()
        if not c:
            return False

        return (
            c in {
                "budget release - cash",
                "budget release - bank",
                "budget release (unposted)",
            }
            or c.startswith("cash:budget release:")
            or c.startswith("bank:budget release:")
            or c.startswith("unposted:budget release:")
            or ":budget release:" in c
        )

    def _coerce_report_tx_type(self, txn_type: str, category_name: str) -> str:
        """
        Budget Return belongs to General Income (Unrestricted).
        Visible Budget Release stays as restricted expense.
        """
        tx = self._normalize_report_tx_type(txn_type)

        if self._is_budget_return_label(category_name):
            return "Unres_Other"

        if self._is_visible_budget_release_label(category_name):
            return "Res_Expense"

        return tx

    # =========================================================
    #  DEFENSIVE DETECTORS
    # =========================================================
    def _looks_like_transfer(self, category_name: str) -> bool:
        c = (category_name or "").strip().lower()

        return (
            c.startswith("bank depos")
            or "bank deposit" in c
            or c.startswith("bank withdr")
            or "bank withdraw" in c
            or "bank withdrawal" in c
            or "transfer to bank" in c
            or "transfer from bank" in c
            or "cash to bank" in c
            or "bank to cash" in c
        )

    def _looks_like_internal_budget_expense(self, category_name: str) -> bool:
        """
        Exclude only true internal/system budget rows.

        KEEP as visible rows:
        - Budget Release (Unposted)
        - Budget Release - Cash
        - Budget Release - Bank
        - Budget Return rows
        """
        c = (category_name or "").strip().lower()

        if self._is_visible_budget_release_label(category_name):
            return False

        if self._is_budget_return_label(category_name):
            return False

        return c in {
            "budget release",
            "system budget release",
            "system budget return",
        }

    def _should_skip_detailed_row(self, txn_type: str, category_name: str) -> bool:
        tx = self._coerce_report_tx_type(txn_type, category_name)

        # Budget return is visible unrestricted income, never skip it here
        if self._is_budget_return_label(category_name):
            return False

        if tx in {
            "Transfer_Deposit",
            "Transfer_Withdraw",
            "System_BudgetRelease",
            "System_BudgetReturn",
        }:
            return True

        if tx in {"Gen_Expense", "Res_Expense"}:
            if self._looks_like_transfer(category_name):
                return True
            if self._looks_like_internal_budget_expense(category_name):
                return True

        return False

    # =========================================================
    #  MERGE SAME-NAME ROWS
    # =========================================================
    def _append_or_merge_amount(self, bucket, name, amount):
        clean_name = str(name or "").strip()

        for item in bucket:
            if item["name"] == clean_name:
                item["amount"] += amount
                return

        bucket.append(
            {
                "name": clean_name,
                "amount": amount,
            }
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        church_id = self.request.user.church_id

        # =========================================================
        #  0) Fetch accounting lock
        # =========================================================
        settings_obj = AccountingSettings.objects.filter(church_id=church_id).first()
        lock_date = settings_obj.lock_date if settings_obj else None

        # =========================================================
        #  1) Determine report type & filter
        # =========================================================
        report_type = (self.request.GET.get("report_type") or "weekly").strip().lower()
        date_filter = (self.request.GET.get("date_filter") or "").strip()

        if report_type == "monthly":
            detailed_sp = "Finance_MonthlyDetailedReport"
            context["report_title"] = "Monthly Financial Report"
        elif report_type == "yearly":
            detailed_sp = "Finance_YearlyDetailedReport"
            context["report_title"] = "Yearly Financial Report"
        else:
            report_type = "weekly"
            detailed_sp = "Finance_WeeklyDetailedReport"
            context["report_title"] = "Weekly Financial Report"

        balances_sp = "Finance_PeriodBalancesSummary"
        restricted_breakdown_sp = "Finance_RestrictedFunds_ByPeriod"
        bank_cashflow_sp = "Finance_BankCashFlow_ByPeriod"
        restricted_net_balance_sp = "Finance_RestrictedDonationsNetBalance"

        # =========================================================
        #  2) Fetch data
        # =========================================================
        raw_data = self.sp_all(detailed_sp, [church_id])
        balance_rows = self.sp_all(balances_sp, [church_id, report_type])
        restricted_rows = self.sp_all(restricted_breakdown_sp, [church_id, report_type])
        bank_rows = self.sp_all(bank_cashflow_sp, [church_id, report_type])
        restricted_net_rows = self.sp_all(restricted_net_balance_sp, [church_id])

        # =========================================================
        #  3A) Restricted fund map (per period)
        # =========================================================
        restricted_map = defaultdict(list)
        for r in restricted_rows:
            sk = str(r[0] or "")
            restricted_map[sk].append(
                {
                    "category": str(r[1] or ""),
                    "income": Decimal(str(r[2] or 0)),
                    "expenses": Decimal(str(r[3] or 0)),
                    "balance": Decimal(str(r[4] or 0)),
                }
            )

        # =========================================================
        #  3B) Balance summary map
        # =========================================================
        balance_by_sortkey = {}
        balance_by_label = {}

        for r in balance_rows:
            sort_key = str(r[0] or "")
            raw_label = str(r[1] or "")
            display_label = self._format_period_label(report_type, raw_label)

            payload = {
                "unrestricted_income": Decimal(str(r[2] or 0)),
                "unrestricted_expenses": Decimal(str(r[3] or 0)),
                "unrestricted_balance": Decimal(str(r[4] or 0)),
                "total_restricted_income": Decimal(str(r[5] or 0)),
                "total_restricted_expenses": Decimal(str(r[6] or 0)),
                "total_restricted_balance": Decimal(str(r[7] or 0)),
                "overall_income": Decimal(str(r[8] or 0)),
                "overall_expenses": Decimal(str(r[9] or 0)),
                "overall_balance": Decimal(str(r[10] or 0)),
                "restricted_funds": [],
            }

            if sort_key:
                balance_by_sortkey[sort_key] = payload
            if display_label:
                balance_by_label[display_label] = payload

        # =========================================================
        #  3C) Bank cashflow map
        # =========================================================
        bank_by_sortkey = {}
        bank_by_label = {}

        for r in bank_rows:
            sk = str(r[0] or "")
            raw_label = str(r[1] or "")
            display_label = self._format_period_label(report_type, raw_label)

            deposits = Decimal(str(r[2] or 0))
            withdrawals = Decimal(str(r[3] or 0))

            bank_payload = {
                "deposits": deposits,
                "withdrawals": withdrawals,
                "net_to_physical_cash": withdrawals - deposits,
                "net_to_bank": deposits - withdrawals,
            }

            if sk:
                bank_by_sortkey[sk] = bank_payload
            if display_label:
                bank_by_label[display_label] = bank_payload

        # =========================================================
        #  3D) Aggregate detailed report
        # =========================================================
        report = defaultdict(
            lambda: {
                "tithes": Decimal("0.00"),
                "offering": Decimal("0.00"),

                "unres_donations": [],
                "total_unres_donations": Decimal("0.00"),

                "unres_other": [],
                "total_unres_other": Decimal("0.00"),

                "res_donations": [],
                "total_res_donations": Decimal("0.00"),

                "res_other": [],
                "total_res_other": Decimal("0.00"),

                "gen_expenses": [],
                "total_gen_expenses": Decimal("0.00"),

                "res_expenses": [],
                "total_res_expenses": Decimal("0.00"),
            }
        )

        available_periods = []
        seen_sort_keys = set()
        label_to_sortkey = {}
        sortkey_to_rawlabel = {}

        for row in raw_data:
            sort_key = str(row[0] or "") if len(row) > 0 else ""
            raw_label = str(row[1] or "") if len(row) > 1 else ""

            if sort_key and raw_label:
                sortkey_to_rawlabel[sort_key] = raw_label

            self._register_period(
                row[0] if len(row) > 0 else None,
                row[1] if len(row) > 1 else None,
                available_periods,
                seen_sort_keys,
                label_to_sortkey,
                report_type,
            )

        for row in balance_rows:
            sort_key = str(row[0] or "") if len(row) > 0 else ""
            raw_label = str(row[1] or "") if len(row) > 1 else ""

            if sort_key and raw_label:
                sortkey_to_rawlabel[sort_key] = raw_label

            self._register_period(
                row[0] if len(row) > 0 else None,
                row[1] if len(row) > 1 else None,
                available_periods,
                seen_sort_keys,
                label_to_sortkey,
                report_type,
            )

        for row in bank_rows:
            sort_key = str(row[0] or "") if len(row) > 0 else ""
            raw_label = str(row[1] or "") if len(row) > 1 else ""

            if sort_key and raw_label:
                sortkey_to_rawlabel[sort_key] = raw_label

            self._register_period(
                row[0] if len(row) > 0 else None,
                row[1] if len(row) > 1 else None,
                available_periods,
                seen_sort_keys,
                label_to_sortkey,
                report_type,
            )

        # =========================================================
        #  3D-1) Restrict visible periods for Members based on lock date
        # =========================================================
        if getattr(self.request.user, "user_type", "") == "Member":
            filtered_periods = []
            filtered_label_to_sortkey = {}

            for period_label in available_periods:
                sort_key = label_to_sortkey.get(period_label)
                raw_label = sortkey_to_rawlabel.get(sort_key, "")

                if self._can_view_period(self.request.user, report_type, raw_label, lock_date):
                    filtered_periods.append(period_label)
                    filtered_label_to_sortkey[period_label] = sort_key

            available_periods = filtered_periods
            label_to_sortkey = filtered_label_to_sortkey

            if date_filter and date_filter not in available_periods:
                messages.warning(
                    self.request,
                    "That report period is not yet available because accounting for that period is not locked yet."
                )
                date_filter = ""

        # =========================================================
        #  3D-2) Aggregate the detailed rows
        # =========================================================
        for row in raw_data:
            sort_key = str(row[0] or "")
            raw_period_label = str(row[1] or "")
            display_period_label = self._format_period_label(report_type, raw_period_label)

            raw_txn_type = row[2] if len(row) > 2 else ""
            category = str(row[3] or "") if len(row) > 3 else ""
            amount = Decimal(str(row[4] or 0)) if len(row) > 4 else Decimal("0.00")

            txn_type = self._coerce_report_tx_type(raw_txn_type, category)

            if date_filter and date_filter != display_period_label:
                continue

            if self._should_skip_detailed_row(txn_type, category):
                continue

            d = report[sort_key]

            if txn_type == "Tithe":
                d["tithes"] += amount

            elif txn_type == "Offering":
                d["offering"] += amount

            elif txn_type == "Unres_Donation":
                self._append_or_merge_amount(d["unres_donations"], category, amount)
                d["total_unres_donations"] += amount

            elif txn_type == "Unres_Other":
                self._append_or_merge_amount(d["unres_other"], category, amount)
                d["total_unres_other"] += amount

            elif txn_type == "Res_Donation":
                self._append_or_merge_amount(d["res_donations"], category, amount)
                d["total_res_donations"] += amount

            elif txn_type == "Res_Other":
                self._append_or_merge_amount(d["res_other"], category, amount)
                d["total_res_other"] += amount

            elif txn_type == "Gen_Expense":
                self._append_or_merge_amount(d["gen_expenses"], category, amount)
                d["total_gen_expenses"] += amount

            elif txn_type == "Res_Expense":
                self._append_or_merge_amount(d["res_expenses"], category, amount)
                d["total_res_expenses"] += amount

        # =========================================================
        #  3E) Current restricted income update snapshot
        # =========================================================
        restricted_income_updates = []
        total_restricted_collected = Decimal("0.00")
        total_restricted_spent = Decimal("0.00")
        total_restricted_balance = Decimal("0.00")

        for r in restricted_net_rows:
            category_name = str(r[0] or "")
            total_collected = Decimal(str(r[1] or 0))
            total_spent = Decimal(str(r[2] or 0))
            current_balance = Decimal(str(r[3] or 0))

            restricted_income_updates.append(
                {
                    "category": category_name,
                    "total_collected": total_collected,
                    "total_spent": total_spent,
                    "current_balance": current_balance,
                }
            )

            total_restricted_collected += total_collected
            total_restricted_spent += total_spent
            total_restricted_balance += current_balance

        # =========================================================
        #  3F) Build restricted fund history per category (for modal)
        # =========================================================
        restricted_history_by_category = defaultdict(list)

        for r in restricted_rows:
            sort_key = str(r[0] or "")
            category_name = str(r[1] or "")
            income = Decimal(str(r[2] or 0))
            expenses = Decimal(str(r[3] or 0))
            balance = Decimal(str(r[4] or 0))

            raw_label = sortkey_to_rawlabel.get(sort_key, "")
            display_label = self._format_period_label(report_type, raw_label or sort_key)

            if getattr(self.request.user, "user_type", "") == "Member":
                if raw_label and not self._can_view_period(self.request.user, report_type, raw_label, lock_date):
                    continue

            if date_filter:
                target_sort_key = label_to_sortkey.get(date_filter)
                if target_sort_key and sort_key != target_sort_key:
                    continue

            restricted_history_by_category[category_name].append(
                {
                    "sort_key": sort_key,
                    "raw_label": raw_label,
                    "period_label": display_label,
                    "income": income,
                    "expenses": expenses,
                    "balance": balance,
                }
            )

        for category_name, history_rows in restricted_history_by_category.items():
            history_rows.sort(
                key=lambda x: self._get_period_end_date(report_type, x["raw_label"]) or date.min,
                reverse=True,
            )

        for item in restricted_income_updates:
            item["history"] = restricted_history_by_category.get(item["category"], [])
            item["history_count"] = len(item["history"])

        # =========================================================
        #  4) Merge / Final report
        # =========================================================
        ZERO_BAL = {
            "unrestricted_income": Decimal("0.00"),
            "unrestricted_expenses": Decimal("0.00"),
            "unrestricted_balance": Decimal("0.00"),
            "total_restricted_income": Decimal("0.00"),
            "total_restricted_expenses": Decimal("0.00"),
            "total_restricted_balance": Decimal("0.00"),
            "overall_income": Decimal("0.00"),
            "overall_expenses": Decimal("0.00"),
            "overall_balance": Decimal("0.00"),
            "restricted_funds": [],
        }

        ZERO_BANK = {
            "deposits": Decimal("0.00"),
            "withdrawals": Decimal("0.00"),
            "net_to_physical_cash": Decimal("0.00"),
            "net_to_bank": Decimal("0.00"),
        }

        final_report = []

        for period_label in available_periods:
            if date_filter and date_filter != period_label:
                continue

            sort_key = label_to_sortkey.get(period_label)
            if not sort_key:
                continue

            d = report.get(sort_key) or report[sort_key]

            total_general_income = (
                d["tithes"]
                + d["offering"]
                + d["total_unres_donations"]
                + d["total_unres_other"]
            )

            total_restricted_income = (
                d["total_res_donations"]
                + d["total_res_other"]
            )

            total_period_income = total_general_income + total_restricted_income

            total_real_expenses = (
                d["total_gen_expenses"]
                + d["total_res_expenses"]
            )

            net_fund_flow = total_period_income - total_real_expenses

            bank_payload = (
                bank_by_sortkey.get(sort_key)
                or bank_by_label.get(period_label)
                or ZERO_BANK
            )

            deposits = Decimal(str(bank_payload["deposits"] or 0))
            withdrawals = Decimal(str(bank_payload["withdrawals"] or 0))
            net_to_physical_cash = Decimal(str(bank_payload["net_to_physical_cash"] or 0))
            net_to_bank = Decimal(str(bank_payload["net_to_bank"] or 0))

            period_physical_cash_change = net_fund_flow + net_to_physical_cash

            base_balances = (
                balance_by_sortkey.get(sort_key)
                or balance_by_label.get(period_label)
                or copy.deepcopy(ZERO_BAL)
            )

            true_balances = copy.deepcopy(base_balances)
            true_balances["restricted_funds"] = restricted_map.get(sort_key, [])

            final_report.append(
                {
                    "week": period_label,
                    "data": d,
                    "bank": {
                        "deposits": deposits,
                        "withdrawals": withdrawals,
                        "net_to_physical_cash": net_to_physical_cash,
                        "net_to_bank": net_to_bank,
                    },
                    "summary": {
                        "total_general_income": total_general_income,
                        "total_restricted_income": total_restricted_income,
                        "total_weekly_income": total_period_income,
                        "total_expenses": total_real_expenses,
                        "total_real_expenses": total_real_expenses,
                        "net_fund_flow": net_fund_flow,
                        "weekly_cash_on_hand": net_fund_flow,
                        "weekly_cash_on_hand_change": net_fund_flow,
                        "total_bank_deposits": deposits,
                        "total_bank_withdrawals": withdrawals,
                        "net_transfer_to_bank": net_to_bank,
                        "net_transfer_to_physical_cash": net_to_physical_cash,
                        "weekly_physical_cash_change": period_physical_cash_change,
                    },
                    "balances": true_balances,
                }
            )

        context["weekly_report"] = final_report
        context["available_periods"] = available_periods
        context["current_report_type"] = report_type
        context["current_filter"] = date_filter
        context["accounting_lock_date"] = lock_date

        context["restricted_income_updates"] = restricted_income_updates
        context["restricted_income_summary"] = {
            "total_collected": total_restricted_collected,
            "total_spent": total_restricted_spent,
            "total_balance": total_restricted_balance,
        }

        return context

class MonthlyFinancialSummaryView(View):
    def get(self, request, *args, **kwargs):
        try:
            # Execute the stored procedure
            with connection.cursor() as cursor:
                cursor.callproc('Monthly_Financial_Summary')
                result = cursor.fetchall()

            # Define column names matching the SELECT statement in the stored procedure
            columns = [
                'Year',
                'Month',
                'TotalTithes',
                'TotalOffering',
                'TotalDonations',
                'TotalExpenses',
                'TotalAccumulated'
            ]

            # Convert the result into a list of dictionaries and map numeric months to month names
            data = [
                {
                    **dict(zip(columns, row)),
                    'Month': calendar.month_name[row[1]]  # Convert numeric month to name
                }
                for row in result
            ]

            # Render the template with data
            return render(request, 'monthly_summary.html', {
                'data': data,
                'year': datetime.now().year
            })

        except Exception as e:
            # Handle any errors and return an appropriate response
            return JsonResponse({'error': str(e)}, status=500)



class FinanceChartView(View):
    template_name = "finance_chart.html"  # Template for rendering both charts

    def get(self, request, *args, **kwargs):
        # Data for daily financial chart
        daily_chart_data = self.get_daily_finance_data()

        # Data for monthly financial chart
        monthly_chart_data = self.get_monthly_finance_data()

        # Pass both datasets to the template
        context = {
            "daily_chart_data": daily_chart_data,
            "monthly_chart_data": monthly_chart_data,
        }
        return render(request, self.template_name, context)

    def get_daily_finance_data(self):
        # Fetch data for daily finance chart
        with connection.cursor() as cursor:
            cursor.callproc("Finance_SortedbyDate")
            results = cursor.fetchall()

        chart_data = {
            "labels": [],  # Dates
            "datasets": {
                "Expense": [],
                "Offering": [],
                "Donations": [],
                "Tithe": [],
            },
        }

        data_dict = {}
        for row in results:
            transaction_type = row[0]
            transaction_date = str(row[2])  # Convert date to string
            amount = float(row[5])  # Ensure numeric value

            if transaction_date not in data_dict:
                data_dict[transaction_date] = {
                    "Expense": 0,
                    "Offering": 0,
                    "Donations": 0,
                    "Tithe": 0,
                }

            data_dict[transaction_date][transaction_type] += amount

        for date in sorted(data_dict.keys(), reverse=True):
            chart_data["labels"].append(date)
            for transaction_type in chart_data["datasets"].keys():
                chart_data["datasets"][transaction_type].append(
                    data_dict[date][transaction_type]
                )

        return chart_data

    def get_monthly_finance_data(self):
        # Fetch data for monthly finance chart
        with connection.cursor() as cursor:
            cursor.callproc("Monthly_Financial_Summary")
            results = cursor.fetchall()

        chart_data = {
            "labels": [],  # Months (e.g., "2024-12")
            "datasets": {
                "TotalTithes": [],
                "TotalOffering": [],
                "TotalDonations": [],
                "TotalExpenses": [],
                "TotalAccumulated": [],
            },
        }

        for row in results:
            year = row[0]
            month = row[1]
            chart_data["labels"].append(f"{year}-{month:02d}")  # Format as YYYY-MM

            chart_data["datasets"]["TotalTithes"].append(float(row[2]))
            chart_data["datasets"]["TotalOffering"].append(float(row[3]))
            chart_data["datasets"]["TotalDonations"].append(float(row[4]))
            chart_data["datasets"]["TotalExpenses"].append(float(row[5]))
            chart_data["datasets"]["TotalAccumulated"].append(float(row[6]))

        return chart_data

class ChangePasswordView(FormView):
    template_name = 'change_password.html'
    form_class = PasswordChangeForm
    success_url = reverse_lazy('change_password_done')  # Not used in this case but kept for fallback

    def get_form_kwargs(self):
        kwargs = super(ChangePasswordView, self).get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        user = form.save()
        update_session_auth_hash(self.request, user)  # Important to keep the user logged in
        return JsonResponse({'status': 'success', 'message': 'Your password was successfully updated!'})

    def form_invalid(self, form):
        return JsonResponse({'status': 'error', 'message': 'Please correct the errors below.', 'errors': form.errors})


class RoleAwareLoginView(LoginView):
    template_name = "login.html"
    form_class = RoleAwareAuthenticationForm
    redirect_authenticated_user = True

    def get_success_url(self):
        user = self.request.user
        role = getattr(user, "user_type", None)

        url = self.get_redirect_url()
        if url:
            return url

        if role == "Member":
            return reverse_lazy("weekly_financial_report")

        if role == "DenominationAdmin":
            return reverse_lazy("denomination_dashboard")  # change to your actual denomination page name

        if role in ["Admin", "Treasurer", "ChurchAdmin", "Pastor"]:
            return reverse_lazy("home")

        return reverse_lazy("home")





_CURRENCY_PATTERN = re.compile(r"[^\d\.\-]")


def _normalize_amount(raw):
    if raw is None:
        return None
    return _CURRENCY_PATTERN.sub("", str(raw)).strip()


class AdminRequiredMixin(LoginRequiredMixin):
    """
    Allow only logged-in users whose user_type is Admin or DenominationAdmin.
    """
    login_url = reverse_lazy("login")
    allowed_roles = ["Admin", "DenominationAdmin","ChurchAdmin"]

    def dispatch(self, request, *args, **kwargs):
        user = request.user

        if not user.is_authenticated:
            return redirect(self.login_url)

        if getattr(user, "user_type", None) not in self.allowed_roles:
            return redirect("home")

        return super().dispatch(request, *args, **kwargs)


# -----------------
# Ministries CRUD
# -----------------
def dictfetchall(cursor):
    """Return all rows from a cursor as a dict"""
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]


class MinistryManageView(LoginRequiredMixin, View):
    template_name = "ministries_manage.html"
    login_url = "login"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(self.login_url)

        if request.user.user_type not in ["Admin", "ChurchAdmin", "Treasurer","Pastor"]:
            messages.error(request, "You do not have permission to access ministry management.")
            return redirect("home")

        if not getattr(request.user, "church_id", None):
            messages.error(request, "No church is assigned to your account.")
            return redirect("home")

        return super().dispatch(request, *args, **kwargs)

    def get_ministries_list(self, church_id):
        return (
            Ministry.objects
            .filter(church_id=church_id)
            .select_related("assigned_requester")
            .order_by("name")
        )

    def get(self, request):
        form = MinistryForm(request=request)
        ministries = self.get_ministries_list(request.user.church_id)
        return render(request, self.template_name, {
            "form": form,
            "ministries": ministries,
        })

    def post(self, request):
        form = MinistryForm(request.POST, request=request)

        if form.is_valid():
            ministry = form.save(commit=False)
            ministry.church = request.user.church
            ministry.created_by = request.user
            ministry.save()

            messages.success(request, "Ministry added successfully.")
            return redirect("ministries_manage")

        ministries = self.get_ministries_list(request.user.church_id)
        return render(request, self.template_name, {
            "form": form,
            "ministries": ministries,
        })

class MinistryEditView(AdminRequiredMixin, View):
    # REUSE the same template
    template_name = "ministries_manage.html"

    def get_ministries_list(self, church_id):
        """Helper to fetch ministries using ORM (Same as Manage View)"""
        return Ministry.objects.filter(church_id=church_id) \
            .select_related('assigned_requester') \
            .order_by('name')

    def get(self, request, pk):
        # 1. Fetch the specific ministry to edit
        ministry = get_object_or_404(Ministry, pk=pk, church=request.user.church)

        # 2. Pre-fill the form (Pass request for user filtering)
        form = MinistryForm(instance=ministry, request=request)

        # 3. Fetch the list (so the table on the right still shows up!)
        ministries = self.get_ministries_list(request.user.church.id)

        context = {
            "form": form,
            "ministries": ministries,
            "edit_mode": True,  # Flag to tell template we are editing
            "ministry_id": pk  # Pass ID for cancel link
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        ministry = get_object_or_404(Ministry, pk=pk, church=request.user.church)

        # Validate input (Pass request and instance)
        form = MinistryForm(request.POST, instance=ministry, request=request)

        if form.is_valid():
            # SWITCH TO ORM: This handles the new 'assigned_requester' field automatically
            m = form.save(commit=False)
            m.edited_by = request.user
            m.save()

            messages.success(request, "Ministry updated successfully.")
            return redirect("ministries_manage")

        # If error, re-render everything
        ministries = self.get_ministries_list(request.user.church.id)
        return render(request, self.template_name, {
            "form": form,
            "ministries": ministries,
            "edit_mode": True,
            "ministry_id": pk
        })


class MinistryToggleActiveView(AdminRequiredMixin, View):
    def post(self, request, pk):
        # USE DJANGO ORM: Update Status
        ministry = get_object_or_404(Ministry, pk=pk, church=request.user.church)

        ministry.is_active = not ministry.is_active
        ministry.edited_by = request.user
        ministry.save()

        status = 'Active' if ministry.is_active else 'Inactive'
        messages.success(request, f"{ministry.name} set to {status}.")
        return redirect("ministries_manage")


class MinistryDeleteView(AdminRequiredMixin, View):
    def post(self, request, pk):
        # USE STORED PROCEDURE: Safe Delete (Checks budget history)
        try:
            with connection.cursor() as cursor:
                cursor.callproc('sp_DeleteMinistry', [
                    pk,
                    request.user.church.id
                ])
            messages.success(request, "Ministry deleted successfully.")

        except Exception as e:
            # Handle the "Cannot delete" error from SQL if budget exists
            error_msg = str(e)
            if 'Cannot delete' in error_msg:
                messages.error(request,
                               "⚠️ Cannot delete this Ministry because it has financial history. Please Deactivate it instead.")
            else:
                messages.error(request, f"Error: {error_msg}")

        return redirect("ministries_manage")


# -----------------
# Budgets Helper
# -----------------
def _normalize_amount(val):
    """Helper to clean currency strings"""
    if val:
        return str(val).replace(',', '')
    return val


# ==========================================
# 1. BUDGET MANAGEMENT (Plans)
# ==========================================
def _normalize_amount(value):
    if not value: return None
    if isinstance(value, (int, float, Decimal)): return value
    return str(value).replace(',', '').strip()


class BudgetManageView(AdminRequiredMixin, View):
    template_name = "budgets_manage.html"

    # =========================
    # Helpers
    # =========================
    def _safe_int(self, v, default=None):
        try:
            return int(v)
        except Exception:
            return default

    def _clean_decimal(self, v) -> Decimal:
        raw = str(v).replace(",", "").replace("₱", "").strip()
        if raw == "" or raw.lower() == "nan":
            raise InvalidOperation("empty")
        return Decimal(raw)

    def _valid_month(self, m: int) -> bool:
        return m in range(0, 13)  # 0..12

    def _get_filters(self, request):
        now = datetime.now()
        cy = now.year
        cm = now.month

        y = self._safe_int(request.GET.get("year"), cy) or cy
        m = self._safe_int(request.GET.get("month"), cm) or cm
        if m not in range(1, 13):
            m = cm
        return cy, cm, y, m

    # =========================================================
    # FIX: Month Dropdown Injector
    # =========================================================
    def _inject_month_choices(self, form):
        month_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
        form.fields['month'] = forms.TypedChoiceField(
            choices=month_choices,
            coerce=int,
            required=False,
            widget=forms.Select(attrs={'class': 'form-select'})
        )
        return form

    def _render_page(self, request, *, selected_year, selected_month, form=None):
        budgets = (
            MinistryBudget.objects
            .filter(
                church=request.user.church,
                year=selected_year,
                month__in=[selected_month, 0]  # selected month + yearly pools
            )
            .select_related("ministry", "balance_account", "ministry__assigned_requester")
            .prefetch_related("released_budgets", "released_budgets__expenses")
            .order_by("ministry__name")
        )

        month_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
        year_choices = range(selected_year - 2, selected_year + 3)

        if form is None:
            form = BudgetForm(request=request, initial={"year": selected_year, "month": selected_month})
            # FIX: Inject choices so the blank form displays the dropdown properly
            form = self._inject_month_choices(form)

        return render(request, self.template_name, {
            "form": form,
            "budgets": budgets,
            "selected_year": selected_year,
            "selected_month": selected_month,
            "current_month_name": calendar.month_name[selected_month],
            "month_choices": month_choices,
            "year_choices": year_choices,
            "edit_mode": False,
            "budget_id": None,
        })

    # =========================
    # Option B: mutual exclusivity
    # =========================
    def _block_mixed_scope(self, *, church, ministry, year, incoming_month):
        """
        Rule:
          - if incoming_month == 0 (Yearly), block if any monthly exists (1..12)
          - if incoming_month in 1..12 (Monthly), block if yearly exists (0)
        """
        if incoming_month == 0:
            if MinistryBudget.objects.filter(
                    church=church, ministry=ministry, year=year, month__gte=1, month__lte=12
            ).exists():
                return True, "Cannot add Yearly Pool: monthly budgets already exist for this ministry and year."
            return False, None

        if MinistryBudget.objects.filter(
                church=church, ministry=ministry, year=year, month=0
        ).exists():
            return True, "Cannot add Monthly budget: Yearly Pool already exists for this ministry and year."
        return False, None

    # =========================
    # GET / POST
    # =========================
    def get(self, request):
        _, _, selected_year, selected_month = self._get_filters(request)
        return self._render_page(request, selected_year=selected_year, selected_month=selected_month)

    def post(self, request):
        if "excel_file" in request.FILES:
            return self.process_bulk_upload(request)
        return self.process_manual_entry(request)

    # =========================
    # BULK UPLOAD (CREATE ONLY)
    # =========================
    def process_bulk_upload(self, request):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect("budgets_manage")

        if not excel_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Error: Please upload a valid Excel (.xlsx) file.")
            return redirect("budgets_manage")

        _, _, selected_year, selected_month = self._get_filters(request)

        try:
            df = pd.read_excel(excel_file)

            # Normalize headers
            df.columns = df.columns.astype(str)
            clean_map = {c.strip().lower(): c for c in df.columns}

            required_keys = ["ministry name", "year", "month", "amount"]
            missing = [k for k in required_keys if k not in clean_map]
            if missing:
                messages.error(request, f"Missing columns: {', '.join(missing)}")
                return redirect(f"{reverse('budgets_manage')}?year={selected_year}&month={selected_month}")

            col_ministry = clean_map["ministry name"]
            col_year = clean_map["year"]
            col_month = clean_map["month"]
            col_amount = clean_map["amount"]

            created_ministries = 0
            created_count = 0
            warnings = []

            with transaction.atomic():
                for idx, row in df.iterrows():
                    row_num = idx + 2

                    ministry_name = str(row[col_ministry]).strip()
                    if not ministry_name or ministry_name.lower() == "nan":
                        continue

                    try:
                        year = self._safe_int(row[col_year], None)
                        month = self._safe_int(row[col_month], None)
                        if year is None or month is None:
                            raise ValueError("bad year/month")
                        if not self._valid_month(month):
                            warnings.append(f"Row {row_num}: Invalid month '{month}' (must be 0-12).")
                            continue

                        amount = self._clean_decimal(row[col_amount])
                    except Exception:
                        warnings.append(f"Row {row_num}: Invalid data format (year/month/amount).")
                        continue

                    ministry = (
                        Ministry.objects
                        .filter(church=request.user.church, name__iexact=ministry_name)
                        .first()
                    )
                    if not ministry:
                        ministry = Ministry.objects.create(
                            church=request.user.church,
                            name=ministry_name,
                            description="Auto-created via Bulk Upload",
                            created_by=request.user,
                            is_active=True,
                        )
                        created_ministries += 1

                    # Option B guard
                    blocked, reason = self._block_mixed_scope(
                        church=request.user.church,
                        ministry=ministry,
                        year=year,
                        incoming_month=month,
                    )
                    if blocked:
                        warnings.append(f"Row {row_num}: Skipped. {reason} ({ministry_name}, {year})")
                        continue

                    # CREATE-ONLY: warn if exists
                    if MinistryBudget.objects.filter(
                            church=request.user.church,
                            ministry=ministry,
                            year=year,
                            month=month
                    ).exists():
                        warnings.append(
                            f"Row {row_num}: Budget already exists for '{ministry_name}' ({year}, month={month}). Use Edit to modify."
                        )
                        continue

                    try:
                        MinistryBudget.objects.create(
                            church=request.user.church,
                            ministry=ministry,
                            year=year,
                            month=month,
                            amount_allocated=amount,
                            is_active=True,
                            created_by=request.user,
                        )
                        created_count += 1
                    except IntegrityError:
                        warnings.append(
                            f"Row {row_num}: Duplicate blocked for '{ministry_name}' ({year}, month={month})."
                        )

            msg = f"Bulk upload complete: {created_count} created."
            if created_ministries:
                msg += f" Created {created_ministries} new ministries."
            if warnings:
                messages.warning(request, msg + " Warnings: " + " | ".join(warnings[:6]))
            else:
                messages.success(request, msg)

        except Exception as e:
            messages.error(request, f"System Error reading file: {str(e)}")

        return redirect(f"{reverse('budgets_manage')}?year={selected_year}&month={selected_month}")

    # =========================
    # MANUAL ENTRY (CREATE ONLY)
    # =========================
    def process_manual_entry(self, request):
        now = datetime.now()
        current_year = now.year

        submitted_year = self._safe_int(request.POST.get("year"), current_year) or current_year
        submitted_month = self._safe_int(request.POST.get("month"), now.month) or now.month
        if submitted_month not in range(1, 13):
            submitted_month = now.month

        # Normalize amount (assuming _normalize_amount is imported at the top of the file)
        data = request.POST.copy()
        if data.get("amount_allocated") and "_normalize_amount" in globals():
            data["amount_allocated"] = _normalize_amount(data["amount_allocated"])

        form = BudgetForm(data, request=request)

        # FIX: Inject choices before validation so the form knows `month` is a valid choice
        form = self._inject_month_choices(form)

        if not form.is_valid():
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, f"{field}: {err}")
            return self._render_page(
                request,
                selected_year=submitted_year,
                selected_month=submitted_month,
                form=form,
            )

        clean = form.cleaned_data
        scope = clean.get("budget_scope")  # 'Yearly' or 'Monthly'
        repeat = bool(clean.get("repeat_for_year"))  # Monthly repeat
        ministry = clean.get("ministry")
        year = clean.get("year")
        month = clean.get("month")
        amount = clean.get("amount_allocated")
        is_active = clean.get("is_active", True)

        base_url = reverse("budgets_manage")

        try:
            with transaction.atomic():
                # --- Option B guard BEFORE any create ---
                # Determine target month for single create
                if scope == "Yearly":
                    target_month = 0
                else:
                    target_month = month  # 1..12

                # Repeat block is special (months 1..12)
                if scope == "Monthly" and repeat:
                    blocked, reason = self._block_mixed_scope(
                        church=request.user.church,
                        ministry=ministry,
                        year=year,
                        incoming_month=1,
                    )
                    if blocked:
                        messages.error(request, reason)
                        return redirect(f"{base_url}?year={year}&month={submitted_month}")

                    created_c = 0
                    skipped_c = 0

                    for m in range(1, 13):
                        if MinistryBudget.objects.filter(
                                church=request.user.church,
                                ministry=ministry,
                                year=year,
                                month=m
                        ).exists():
                            skipped_c += 1
                            continue

                        try:
                            MinistryBudget.objects.create(
                                church=request.user.church,
                                ministry=ministry,
                                year=year,
                                month=m,
                                amount_allocated=amount,
                                is_active=is_active,
                                created_by=request.user,
                            )
                            created_c += 1
                        except IntegrityError:
                            skipped_c += 1

                    if skipped_c:
                        messages.warning(
                            request,
                            f"Batch created {created_c}. Skipped {skipped_c} existing month(s). Use Edit to modify existing budgets."
                        )
                    else:
                        messages.success(request, f"Batch created Jan–Dec budgets for {ministry.name} ({year}).")
                    return redirect(f"{base_url}?year={year}&month={submitted_month}")

                # Single create validation
                if not self._valid_month(target_month):
                    messages.error(request, "Invalid month value.")
                    return redirect(f"{base_url}?year={year}&month={submitted_month}")

                blocked, reason = self._block_mixed_scope(
                    church=request.user.church,
                    ministry=ministry,
                    year=year,
                    incoming_month=target_month,
                )
                if blocked:
                    messages.error(request, reason)
                    return redirect(f"{base_url}?year={year}&month={submitted_month}")

                # CREATE-ONLY: warn if exists
                if MinistryBudget.objects.filter(
                        church=request.user.church,
                        ministry=ministry,
                        year=year,
                        month=target_month
                ).exists():
                    period = "Yearly Pool" if target_month == 0 else calendar.month_name[target_month]
                    messages.warning(
                        request,
                        f"Budget already exists for {ministry.name} ({period} {year}). Click the pencil icon to edit."
                    )
                    return redirect(f"{base_url}?year={year}&month={submitted_month}")

                MinistryBudget.objects.create(
                    church=request.user.church,
                    ministry=ministry,
                    year=year,
                    month=target_month,
                    amount_allocated=amount,
                    is_active=is_active,
                    created_by=request.user,
                )

                period = "Yearly Pool" if target_month == 0 else calendar.month_name[target_month]
                messages.success(request, f"Budget created for {ministry.name} ({period} {year}).")

        except IntegrityError:
            messages.warning(request, "Budget already exists. Use Edit to modify.")
        except Exception as e:
            messages.error(request, f"System Error: {str(e)}")

        return redirect(f"{base_url}?year={submitted_year}&month={submitted_month}")

class BudgetEditView(AdminRequiredMixin, View):
    template_name = "budgets_manage.html"

    def _safe_int(self, v, default=None):
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    def _valid_month(self, m: int) -> bool:
        return m in range(0, 13)

    def _get_budgets_list(self, request, year, month):
        return (
            MinistryBudget.objects
            .filter(
                church=request.user.church,
                year=year,
                month__in=[month, 0]
            )
            .select_related("ministry", "balance_account")
            .prefetch_related("released_budgets", "released_budgets__expenses")
            .order_by("-is_active", "-month", "ministry__name")
        )

    # =========================================================
    # HELPER 1: Global Expense Check
    # =========================================================
    def _has_expenses_for_year(self, request, budget):
        return MinistryBudget.objects.filter(
            church=request.user.church,
            ministry=budget.ministry,
            year=budget.year,
            released_budgets__expenses__isnull=False
        ).exists()

    # =========================================================
    # HELPER 2: Fix Blank Month Dropdown
    # =========================================================
    def _inject_month_choices(self, form):
        month_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
        form.fields['month'] = forms.TypedChoiceField(
            choices=month_choices,
            coerce=int,
            required=False,
            widget=forms.Select(attrs={'class': 'form-select'})
        )
        return form

    def _render_edit_page(self, request, *, budget, form, selected_year, selected_month, has_expenses=False):
        budgets = self._get_budgets_list(request, selected_year, selected_month)
        month_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
        year_choices = range(selected_year - 2, selected_year + 3)

        return render(request, self.template_name, {
            "form": form,
            "budgets": budgets,
            "selected_year": selected_year,
            "selected_month": selected_month,
            "current_month_name": calendar.month_name[selected_month],
            "month_choices": month_choices,
            "year_choices": year_choices,
            "edit_mode": True,
            "budget_id": budget.id,
            "has_expenses": has_expenses,
        })

    def get(self, request, pk):
        budget = get_object_or_404(MinistryBudget, pk=pk, church=request.user.church)

        has_expenses = self._has_expenses_for_year(request, budget)
        fallback_month = budget.month if budget.month > 0 else datetime.now().month

        form = BudgetForm(
            instance=budget,
            request=request,
            selected_year=budget.year,
            selected_month=fallback_month,
        )
        form = self._inject_month_choices(form)

        selected_year = budget.year
        selected_month = self._safe_int(request.GET.get("month"), fallback_month) or fallback_month

        if selected_month not in range(1, 13):
            selected_month = datetime.now().month

        return self._render_edit_page(
            request, budget=budget, form=form,
            selected_year=selected_year, selected_month=selected_month,
            has_expenses=has_expenses,
        )

    def post(self, request, pk):
        budget = get_object_or_404(MinistryBudget, pk=pk, church=request.user.church)

        # =========================================================
        # CRITICAL FIX: Capture the original state BEFORE the form
        # mutates the object in memory.
        # =========================================================
        original_scope = "Yearly" if budget.month == 0 else "Monthly"
        original_ministry = budget.ministry
        original_year = budget.year

        has_expenses = self._has_expenses_for_year(request, budget)

        data = request.POST.copy()

        # Assuming _normalize_amount is imported at the top of your views.py
        if data.get("amount_allocated") and "_normalize_amount" in globals():
            data["amount_allocated"] = _normalize_amount(data["amount_allocated"])

        if not data.get("budget_scope"):
            data["budget_scope"] = original_scope

        submitted_year = self._safe_int(data.get("year"), budget.year) or budget.year
        submitted_month = self._safe_int(
            data.get("month"),
            budget.month if budget.month in range(1, 13) else datetime.now().month
        )
        if submitted_month not in range(1, 13):
            submitted_month = budget.month if budget.month in range(1, 13) else datetime.now().month

        form = BudgetForm(
            data, instance=budget, request=request,
            selected_year=submitted_year, selected_month=submitted_month,
        )
        form = self._inject_month_choices(form)

        if not form.is_valid():
            for field, err_list in form.errors.items():
                for err in err_list:
                    prefix = "Form Error" if field == "__all__" else f"Error in {field.title().replace('_', ' ')}"
                    messages.error(request, f"{prefix}: {err}")

            return self._render_edit_page(
                request, budget=budget, form=form,
                selected_year=submitted_year, selected_month=submitted_month,
                has_expenses=has_expenses,
            )

        clean = form.cleaned_data
        scope = clean.get("budget_scope")
        repeat = bool(clean.get("repeat_for_year"))
        ministry = clean.get("ministry")
        year = clean.get("year")
        month = clean.get("month")
        amount = clean.get("amount_allocated")
        is_active = clean.get("is_active", True)

        # =========================================================
        # SMART EXPENSE LOCK
        # Compares against the ORIGINAL unmutated values
        # =========================================================
        if has_expenses:
            is_structure_changed = (
                scope != original_scope or
                ministry != original_ministry or
                year != original_year
            )
            if is_structure_changed:
                messages.error(request, "Structure Locked: Expenses exist for this ministry/year. You can only update the Amount and Status.")
                return redirect(request.path)

        base_url = reverse("budgets_manage")

        try:
            with transaction.atomic():

                # =========================================================
                # CASE 1: Convert to / Edit as YEARLY POOL
                # =========================================================
                if scope == "Yearly":
                    # Save first to establish it as the definitive yearly pool
                    budget.ministry = ministry
                    budget.year = year
                    budget.month = 0
                    budget.amount_allocated = amount
                    budget.is_active = is_active
                    budget.edited_by = request.user
                    budget.save()

                    # Wipe out any lingering monthly records for this year
                    MinistryBudget.objects.filter(
                        church=request.user.church, ministry=ministry, year=year,
                    ).exclude(pk=budget.pk).delete()

                    messages.success(request, f"Converted {budget.ministry.name} to a Yearly Pool for {budget.year}.")
                    return_month = submitted_month if submitted_month in range(1, 13) else datetime.now().month
                    return redirect(f"{base_url}?year={budget.year}&month={return_month}")

                # =========================================================
                # CASE 2: MONTHLY + APPLY TO ALL 12 MONTHS
                # =========================================================
                elif scope == "Monthly" and repeat:
                    if ministry != original_ministry or year != original_year:
                        messages.error(request, "When applying changes to all 12 months, ministry and year must remain the same.")
                        return redirect(request.path)

                    # Transform current record into January to prevent it from remaining as '0'
                    budget.ministry = ministry
                    budget.year = year
                    budget.month = 1
                    budget.amount_allocated = amount
                    budget.is_active = is_active
                    budget.edited_by = request.user
                    budget.save()

                    # Hard delete any leftover Yearly pools (month 0)
                    MinistryBudget.objects.filter(
                        church=request.user.church, ministry=ministry, year=year, month=0,
                    ).delete()

                    # Build the rest of the 12 months
                    for m in range(1, 13):
                        obj, created = MinistryBudget.objects.update_or_create(
                            church=request.user.church,
                            ministry=ministry,
                            year=year,
                            month=m,
                            defaults={
                                "amount_allocated": amount,
                                "is_active": is_active,
                                "edited_by": request.user,
                            },
                        )
                        if created and hasattr(obj, 'created_by') and not obj.created_by_id:
                            obj.created_by = request.user
                            obj.save(update_fields=["created_by"])

                    messages.success(request, f"Applied changes to Jan–Dec for {budget.ministry.name} ({budget.year}).")
                    return redirect(f"{base_url}?year={budget.year}&month=1")

                # =========================================================
                # CASE 3: MONTHLY SINGLE MONTH EDIT
                # =========================================================
                else:
                    target_month = month
                    if not target_month or target_month == 0 or not self._valid_month(target_month):
                        messages.error(request, "Please select a valid specific month.")
                        return redirect(request.path)

                    duplicate_exists = MinistryBudget.objects.filter(
                        church=request.user.church, ministry=ministry, year=year, month=target_month,
                    ).exclude(pk=budget.pk).exists()

                    if duplicate_exists:
                        messages.error(request, f"A budget for {calendar.month_name[target_month]} {year} already exists for {ministry.name}. Edit that record instead.")
                        return redirect(request.path)

                    # Update the record to the target month
                    budget.ministry = ministry
                    budget.year = year
                    budget.month = target_month
                    budget.amount_allocated = amount
                    budget.is_active = is_active
                    budget.edited_by = request.user
                    budget.save()

                    # Hard delete any leftover Yearly pools (month 0)
                    MinistryBudget.objects.filter(
                        church=request.user.church, ministry=ministry, year=year, month=0,
                    ).delete()

                    messages.success(request, f"Updated {budget.ministry.name} for {calendar.month_name[target_month]} {budget.year}.")
                    return redirect(f"{base_url}?year={budget.year}&month={target_month}")

        except IntegrityError:
            messages.error(request, "A duplicate budget was detected and blocked by the system.")
        except Exception as e:
            messages.error(request, f"System Error: {str(e)}")

        return redirect(request.path)

# ==========================================
# 2. REQUEST & APPROVAL FLOW
# ==========================================
class ApprovalPermissionMixin(LoginRequiredMixin):
    """
    Custom Mixin: Allows Admin, Treasurer, AND Pastor to approve/decline.
    """

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        # The list of roles allowed to Approve/Decline
        allowed_roles = ['Admin', 'ChurchAdmin', 'Treasurer', 'Pastor']

        if request.user.user_type not in allowed_roles:
            messages.error(request, "⛔ Access Denied: Only Pastors, Admins, and Treasurers can approve requests.")
            return redirect("budget_release")

        return super().dispatch(request, *args, **kwargs)


class BudgetReleaseView(LoginRequiredMixin, View):
    """
    Handles submission + tracking of budget requests across stages.

    Request validation source:
    - Uses Total Available Funds
    - Total Available Funds = Calculate_CashOnHand(church_id) + BankAccount.current_balance

    Declined visibility rule:
    - Privileged roles (Admin/ChurchAdmin/Treasurer/Pastor): see ONLY declines they made (declined_by=user)
    - Non-privileged users: see ONLY declines they requested (requested_by=user)
    """
    template_name = "budget_release.html"
    PRIVILEGED_ROLES = ['Admin', 'ChurchAdmin', 'Treasurer', 'Pastor']

    # =========================================================
    # STORED PROCEDURE HELPER
    # =========================================================
    def _sp_one(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            row = cursor.fetchone()

            try:
                cursor.fetchall()
            except Exception:
                pass

            while cursor.nextset():
                try:
                    cursor.fetchall()
                except Exception:
                    pass

        return row

    # =========================================================
    # TOTAL AVAILABLE FUNDS HELPER
    # Returns: (physical_cash, bank_balance, total_available_funds)
    # =========================================================
    def _get_total_available_funds(self, user):
        """
        Total Available Funds = physical cash + bank balance

        physical cash -> Calculate_CashOnHand SP
        bank balance  -> BankAccount.current_balance
        """
        church_id = getattr(user, "church_id", None)
        if not church_id:
            return Decimal("0.00"), Decimal("0.00"), Decimal("0.00")

        # 1) Physical cash from SP
        physical_cash = Decimal("0.00")
        try:
            row = self._sp_one("Calculate_CashOnHand", [church_id])
            if row and row[0] is not None:
                physical_cash = Decimal(str(row[0]))
        except Exception:
            physical_cash = Decimal("0.00")

        # 2) Bank balance from BankAccount
        bank_balance = Decimal("0.00")
        try:
            BankAccount = apps.get_model("Register", "BankAccount")
            bank = BankAccount.objects.filter(church_id=church_id).first()
            if bank and bank.current_balance is not None:
                bank_balance = Decimal(str(bank.current_balance))
        except Exception:
            bank_balance = Decimal("0.00")

        return physical_cash, bank_balance, physical_cash + bank_balance

    # =========================================================
    # QUERY HELPERS
    # =========================================================
    def _get_pending_requests(self, user):
        """Stage 1: Pending (Waiting for Approval)"""
        if not getattr(user, "church_id", None):
            return BudgetReleaseRequest.objects.none()

        qs = (
            BudgetReleaseRequest.objects
            .filter(church_id=user.church_id, status='Pending')
            .select_related("budget", "ministry", "requested_by")
            .order_by("-requested_at")
        )

        if user.user_type not in self.PRIVILEGED_ROLES:
            qs = qs.filter(requested_by=user)

        return qs

    def _get_approved_items(self, user):
        """Stage 2: Approved (Waiting for Cash Release)"""
        if not getattr(user, "church_id", None):
            return ApprovedReleaseRequest.objects.none()

        qs = (
            ApprovedReleaseRequest.objects
            .filter(church_id=user.church_id)
            .select_related(
                "budget",
                "ministry",
                "approved_request",
                "approved_request__requested_by",
                "approved_by",
            )
            .order_by("-approved_at")
        )

        if user.user_type not in self.PRIVILEGED_ROLES:
            qs = qs.filter(ministry__assigned_requester=user)

        return qs

    def _get_released_items(self, user):
        """Stage 3 & 4: Released (Liquidate & Complete)"""
        if not getattr(user, "church_id", None):
            return ReleasedBudget.objects.none()

        qs = (
            ReleasedBudget.objects
            .filter(church_id=user.church_id)
            .select_related("budget", "ministry", "approved_request", "approved_request__requested_by")
            .order_by("-date_released")
        )

        if user.user_type not in self.PRIVILEGED_ROLES:
            qs = qs.filter(ministry__assigned_requester=user)

        return qs

    def _get_declined_items(self, user):
        """
        Declined history:
        - Privileged roles: only items they declined
        - Non-privileged: only items they requested
        """
        if not getattr(user, "church_id", None):
            return DeclinedReleaseRequest.objects.none()

        qs = (
            DeclinedReleaseRequest.objects
            .filter(budget__church_id=user.church_id)
            .select_related("budget", "budget__ministry", "declined_by", "requested_by")
            .order_by("-declined_at")
        )

        if user.user_type in self.PRIVILEGED_ROLES:
            return qs.filter(declined_by=user)

        return qs.filter(requested_by=user)

    # =========================================================
    # COMMON CONTEXT
    # =========================================================
    def _build_context(self, request, form, physical_cash=None, bank_balance=None, total_available_funds=None):
        pending = self._get_pending_requests(request.user)
        approved = self._get_approved_items(request.user)
        released = self._get_released_items(request.user)
        declined = self._get_declined_items(request.user)

        # Compute funds if not already provided by caller
        if total_available_funds is None:
            physical_cash, bank_balance, total_available_funds = self._get_total_available_funds(request.user)

        return {
            "form": form,
            "pending": pending,
            "approved": approved,
            "to_liquidate": released.filter(is_liquidated=False),
            "completed": released.filter(is_liquidated=True)[:50],
            "declined": declined[:20],
            "physical_cash": physical_cash,
            "bank_balance": bank_balance,
            "total_available_funds": total_available_funds,
        }

    # =========================================================
    # GET
    # =========================================================
    def get(self, request):
        physical_cash, bank_balance, total_available_funds = self._get_total_available_funds(request.user)
        form = BudgetReleaseRequestForm(
            request=request,
            physical_cash=physical_cash,
            bank_balance=bank_balance,
            total_available_funds=total_available_funds,
        )
        return render(
            request,
            self.template_name,
            self._build_context(request, form, physical_cash, bank_balance, total_available_funds),
        )

    # =========================================================
    # POST
    # =========================================================
    def post(self, request):
        data = request.POST.copy()

        # Normalize amount
        if "amount" in data:
            data["amount"] = str(data["amount"]).replace(",", "").strip()

        physical_cash, bank_balance, total_available_funds = self._get_total_available_funds(request.user)

        form = BudgetReleaseRequestForm(
            data,
            request=request,
            physical_cash=physical_cash,
            bank_balance=bank_balance,
            total_available_funds=total_available_funds,
        )

        if form.is_valid():
            req = form.save(commit=False)
            req.requested_by = request.user

            if getattr(request.user, "church_id", None):
                req.church_id = request.user.church_id

            req.save()

            messages.success(
                request,
                f"Request submitted successfully. Available funds checked against ₱{total_available_funds:,.2f}."
            )
            return redirect("budget_release")

        return render(
            request,
            self.template_name,
            self._build_context(request, form, physical_cash, bank_balance, total_available_funds),
        )


def _sp_one_helper(sp_name, params):
    with connection.cursor() as cursor:
        cursor.callproc(sp_name, params)
        row = cursor.fetchone()
        try:
            cursor.fetchall()
        except Exception:
            pass
        while cursor.nextset():
            try:
                cursor.fetchall()
            except Exception:
                pass
    return row


def get_total_available_funds(church_id):
    """
    Returns: (physical_cash, bank_balance, total_available_funds)
    """
    if not church_id:
        return Decimal("0.00"), Decimal("0.00"), Decimal("0.00")

    physical_cash = Decimal("0.00")
    try:
        row = _sp_one_helper("Calculate_CashOnHand", [church_id])
        if row and row[0] is not None:
            physical_cash = Decimal(str(row[0]))
    except Exception:
        pass

    bank_balance = Decimal("0.00")
    try:
        BankAccount = apps.get_model("Register", "BankAccount")
        bank = BankAccount.objects.filter(church_id=church_id).first()
        if bank and bank.current_balance is not None:
            bank_balance = Decimal(str(bank.current_balance))
    except Exception:
        pass

    return physical_cash, bank_balance, physical_cash + bank_balance


class BudgetSummaryView(FinanceRoleRequiredMixin, View):
    template_name = "budget_summary.html"

    def get(self, request):
        # 1. Fetch ALL budgets for this church
        raw_budgets = (
            MinistryBudget.objects
            .filter(church=request.user.church)
            .select_related('ministry')
            .order_by('-year', 'ministry__name')
        )

        # 2. Grouping Logic (The "One Row" Magic)
        # Key: (Year, MinistryID) -> Value: Aggregate Data
        grouped_data = {}

        for b in raw_budgets:
            key = (b.year, b.ministry.id)

            if key not in grouped_data:
                grouped_data[key] = {
                    'year': b.year,
                    'ministry_name': b.ministry.name,
                    'total_allocated': Decimal(0),
                    'total_spent': Decimal(0),
                    'total_unliquidated': Decimal(0),
                    'type_label': 'Yearly Pool', # Default
                    'badge_class': 'bg-info-subtle text-info-emphasis',
                    'record_count': 0
                }

            # Accumulate Totals
            group = grouped_data[key]
            group['total_allocated'] += b.amount_allocated
            group['total_spent'] += b.total_spent
            group['total_unliquidated'] += b.cash_on_hand
            group['record_count'] += 1

            # Determine Label
            # If we find a month > 0, or if we have multiple records, it's a Monthly structure
            if b.month > 0 or group['record_count'] > 1:
                group['type_label'] = 'Monthly Aggregated'
                group['badge_class'] = 'bg-primary-subtle text-primary-emphasis'

        # 3. Calculate Final "Available" for the group
        summary_list = []
        for key, data in grouped_data.items():
            # Available = Allocated - Spent
            data['total_available'] = data['total_allocated'] - data['total_spent']
            summary_list.append(data)

        # 4. Sort (Year Descending, then Ministry Name)
        summary_list.sort(key=lambda x: (-x['year'], x['ministry_name']))

        return render(request, self.template_name, {
            "summary_list": summary_list
        })


class BudgetReleaseApproveView(ApprovalPermissionMixin, View):
    """
    GET:  Renders the review page with fund availability breakdown.
    POST: Processes the approval with all safety checks.

    Checks:
      1. Permissions (via ApprovalPermissionMixin)
      2. Ministry Wallet Balance
      3. Total Available Funds
         = Calculate_CashOnHand(church_id) + BankAccount.current_balance
    """
    template_name = "review_budget_request.html"

    def _normalize_amount(self, value):
        if value in (None, ""):
            return None

        if isinstance(value, str):
            value = value.replace(",", "").strip()

        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _get_request_or_redirect(self, request, pk, lock=False):
        """
        Fetch the budget request with related data.
        If lock=True, uses select_for_update() for atomic safety.
        Returns (req, None) on success or (None, redirect_response) on failure.
        """
        church_id = getattr(request.user, "church_id", None)
        if not church_id:
            messages.error(request, "No church is linked to your account.")
            return None, redirect("budget_release")

        qs = BudgetReleaseRequest.objects.select_related(
            "budget",
            "budget__ministry",
            "requested_by",
        )

        if lock:
            qs = qs.select_for_update()

        req = get_object_or_404(qs, pk=pk, church_id=church_id)

        if req.status != "Pending":
            messages.warning(request, f"This request is already marked as '{req.status}'.")
            return None, redirect("budget_release")

        return req, None

    # =========================================================
    # GET — Review Page
    # =========================================================
    def get(self, request, pk):
        req, error_response = self._get_request_or_redirect(request, pk)
        if error_response:
            return error_response

        church_id = request.user.church_id

        # Ministry wallet balance
        ministry_balance = Decimal("0.00")
        try:
            wallet = BudgetBalance.objects.get(budget=req.budget)
            ministry_balance = Decimal(str(wallet.current_amount or 0))
        except BudgetBalance.DoesNotExist:
            ministry_balance = Decimal("0.00")

        # Total available funds from shared service
        physical_cash, bank_balance, total_available_funds = get_total_available_funds(church_id)

        db_amount = Decimal(str(req.amount or 0))

        sufficient_funds = (
            db_amount <= ministry_balance
            and db_amount <= total_available_funds
        )

        return render(request, self.template_name, {
            "req": req,
            "ministry_balance": ministry_balance,
            "physical_cash": physical_cash,
            "bank_balance": bank_balance,
            "total_available_funds": total_available_funds,
            "sufficient_funds": sufficient_funds,
        })

    # =========================================================
    # POST — Process Approval
    # =========================================================
    def post(self, request, pk):
        church_id = getattr(request.user, "church_id", None)
        if not church_id:
            messages.error(request, "No church is linked to your account.")
            return redirect("budget_release")

        posted_amount = self._normalize_amount(request.POST.get("amount"))

        try:
            with transaction.atomic():
                # 1. Lock request row to prevent double approval
                req, error_response = self._get_request_or_redirect(request, pk, lock=True)
                if error_response:
                    return error_response

                budget = req.budget
                ministry = getattr(budget, "ministry", None)
                db_amount = Decimal(str(req.amount or 0))

                # 2. Anti-tamper check for posted amount
                if posted_amount is not None and posted_amount != db_amount:
                    messages.error(
                        request,
                        "Approval denied: posted amount does not match the saved request amount."
                    )
                    return redirect("budget_release")

                # 3. Prevent duplicate approval ticket
                if ApprovedReleaseRequest.objects.filter(approved_request=req).exists():
                    messages.warning(
                        request,
                        "This request already has an approval record."
                    )
                    return redirect("budget_release")

                # 4. Lock ministry wallet
                try:
                    wallet = BudgetBalance.objects.select_for_update().get(budget=budget)
                except BudgetBalance.DoesNotExist:
                    messages.error(
                        request,
                        "Critical Error: Budget Wallet (BudgetBalance) not found."
                    )
                    return redirect("budget_release")

                ministry_balance = Decimal(str(wallet.current_amount or 0))

                # 5. Total available funds from shared service
                physical_cash, bank_balance, total_available_funds = get_total_available_funds(church_id)

                # 6. Check ministry wallet
                if db_amount > ministry_balance:
                    messages.error(
                        request,
                        f"Insufficient Ministry Funds. "
                        f"Wallet: ₱{ministry_balance:,.2f}. Requested: ₱{db_amount:,.2f}."
                    )
                    return redirect("budget_release")

                # 7. Check total available funds
                if db_amount > total_available_funds:
                    messages.error(
                        request,
                        f"Insufficient Total Available Funds. "
                        f"Physical Cash: ₱{physical_cash:,.2f}. "
                        f"Bank Balance: ₱{bank_balance:,.2f}. "
                        f"Total Available: ₱{total_available_funds:,.2f}. "
                        f"Requested: ₱{db_amount:,.2f}."
                    )
                    return redirect("budget_release")

                # 8. Create approval ticket
                ApprovedReleaseRequest.objects.create(
                    church_id=church_id,
                    ministry=ministry,
                    budget=budget,
                    date_released=req.date_released,
                    amount=db_amount,
                    remarks=req.remarks,
                    approved_by=request.user,
                    approved_request=req,
                )

                # 9. Update original request status
                req.status = "Approved"
                req.approved_by = request.user
                req.approved_at = timezone.now()
                req.save(update_fields=["status", "approved_by", "approved_at"])

            role_title = getattr(request.user, "user_type", "User")
            messages.success(
                request,
                f"Request for ₱{db_amount:,.2f} approved by {role_title} {request.user.username}."
            )
            return redirect("budget_release")

        except Exception as e:
            messages.error(request, f"Approval failed: {e}")
            return redirect("budget_release")

class BudgetReleaseDeclineView(ApprovalPermissionMixin, View):
    """
    Declines a budget request.
    ✅ Records: original_request, requested_by, requested_at, declined_reason, declined_by
    ✅ Keeps original request (status updated) so history is preserved.
    """
    def post(self, request, pk):
        # 1) Fetch ONLY pending requests (prevents declining released/approved items)
        req = get_object_or_404(
            BudgetReleaseRequest,
            pk=pk,
            church=request.user.church,
            status="Pending"
        )

        # 2) Reason
        reason = (request.POST.get("declined_reason") or "").strip()
        if not reason:
            reason = "Declined by Leadership"

        with transaction.atomic():
            # 3) Create Declined record (copy requester + timestamps)
            DeclinedReleaseRequest.objects.create(
                original_request=req,          # ✅ link back
                requested_by=req.requested_by, # ✅ requester tracking
                requested_at=req.requested_at, # ✅ original request time

                budget=req.budget,
                date_released=req.date_released,
                amount=req.amount,
                remarks=req.remarks,

                declined_by=request.user,
                declined_reason=reason,
                # declined_at is auto_now_add=True in your model
            )

            # 4) Update status (use ONE label consistently across system)
            # Choose either 'Rejected' or 'Declined' but keep it consistent.
            req.status = "Rejected"
            req.save(update_fields=["status"])

        role = request.user.user_type
        messages.success(request, f"Request declined by {role} {request.user.username}.")
        return redirect("budget_release")

class ReleaseBudgetView(FinanceRoleRequiredMixin, View):
    template_name = "release_budget.html"

    def _get_bank_account(self, church):
        try:
            return church.bank_account
        except Exception:
            return BankAccount.objects.filter(church=church).first()

    def get(self, request):
        approved = (
            ApprovedReleaseRequest.objects
            .filter(church=request.user.church)
            .select_related("budget", "ministry", "approved_by")
            .order_by("date_released")
        )

        released = (
            ReleasedBudget.objects
            .filter(church=request.user.church)
            .select_related("budget", "ministry", "released_by")
            .order_by("-released_at")[:100]
        )

        bank_acct = self._get_bank_account(request.user.church)

        lock_modal = request.session.pop("lock_modal", None)

        return render(
            request,
            self.template_name,
            {
                "approved": approved,
                "released": released,
                "bank_acct": bank_acct,
                "show_lock_modal": bool(lock_modal and lock_modal.get("show")),
                "lock_message": (lock_modal or {}).get("message", ""),
            },
        )


class ReleaseBudgetPerformView(FinanceRoleRequiredMixin, View):
    """
    POST: convert one ApprovedReleaseRequest into a ReleasedBudget.

    Rules:
    - cash release: deduct treasury + unrestricted
    - bank release: deduct bank + unrestricted
    - if unrestricted available is less than requested amount, block release
    """

    CASH_RELEASE_CATEGORY = "Budget Release - Cash"
    BANK_RELEASE_CATEGORY = "Budget Release - Bank"

    def _back_url(self, request):
        return request.META.get("HTTP_REFERER") or reverse("release_budget")

    def _get_lock_date(self, church):
        if hasattr(church, "accounting_settings") and church.accounting_settings:
            return church.accounting_settings.lock_date
        return None

    def _posting_date(self, church):
        lock_date = self._get_lock_date(church)
        d = timezone.localdate()
        if lock_date and d < lock_date:
            d = lock_date
        return d

    def _extract_validation_message(self, e: ValidationError) -> str:
        if hasattr(e, "message_dict") and e.message_dict:
            for _, msgs in e.message_dict.items():
                if msgs:
                    return str(msgs[0])
        if getattr(e, "messages", None):
            return str(e.messages[0])
        return "PERIOD LOCKED: You cannot add or edit records in this closed period."

    def _redirect_back_with_lock_modal(self, request, message: str):
        request.session["lock_modal"] = {
            "show": True,
            "message": message,
        }
        messages.error(request, message)
        return redirect(self._back_url(request))

    def _get_cash_on_hand(self, church):
        cash_on_hand = Decimal("0.00")
        with connection.cursor() as cursor:
            cursor.execute("CALL Calculate_CashOnHand(%s)", [church.pk])
            result = cursor.fetchone()
            if result and result[0] is not None:
                cash_on_hand = Decimal(str(result[0]))
        return cash_on_hand

    def _get_unrestricted_available(self, church):
        unrestricted = Decimal("0.00")

        with connection.cursor() as cursor:
            cursor.execute("CALL Finance_UnrestrictedNet(%s)", [church.pk])

            row = cursor.fetchone()
            if row:
                columns = [col[0] for col in cursor.description] if cursor.description else []
                result_map = dict(zip(columns, row))

                unrestricted = Decimal(str(
                    result_map.get("NetGrandTotalUnrestricted") or "0.00"
                ))

            try:
                while cursor.nextset():
                    pass
            except Exception:
                pass

        return unrestricted if unrestricted > 0 else Decimal("0.00")

    def _get_bank_account(self, church):
        try:
            return church.bank_account
        except Exception:
            return BankAccount.objects.filter(church=church).first()

    def _create_release_expense(
        self, *,
        church, category_name, amount, budget, release, request_user, posting_date, released_date
    ):
        release_cat, _ = ExpenseCategory.objects.get_or_create(
            church=church,
            name=category_name,
            defaults={"description": ""},
        )

        expense_kwargs = dict(
            church=church,
            category=release_cat,
            expense_date=posting_date,
            amount=amount,
            description=(
                f"{category_name}: {budget.ministry.name} | "
                f"Release ID {release.id} | "
                f"Released on {released_date}"
            ),
            created_by=request_user,
        )

        if hasattr(Expense, "user"):
            expense_kwargs["user"] = request_user
        if hasattr(Expense, "status"):
            expense_kwargs["status"] = "Approved"

        return Expense.objects.create(**expense_kwargs)

    def post(self, request, pk):
        deduct_from = (request.POST.get("deduct_from") or "").strip().lower()

        if deduct_from not in {
            ReleasedBudget.DEDUCT_CASH,
            ReleasedBudget.DEDUCT_BANK,
        }:
            messages.error(request, "Please choose whether to deduct from Physical Cash or Bank.")
            return redirect(self._back_url(request))

        try:
            with transaction.atomic():
                appr = get_object_or_404(
                    ApprovedReleaseRequest.objects
                    .select_for_update()
                    .select_related("budget", "ministry", "approved_request"),
                    pk=pk,
                    church=request.user.church,
                )

                budget = appr.budget
                original_request = appr.approved_request

                try:
                    wallet = BudgetBalance.objects.select_for_update().get(budget=budget)
                except BudgetBalance.DoesNotExist:
                    messages.error(request, "Critical Error: Budget Wallet (BudgetBalance) not found.")
                    return redirect(self._back_url(request))

                wallet_amount = wallet.current_amount or Decimal("0.00")

                if appr.amount > wallet_amount:
                    messages.error(
                        request,
                        f"Insufficient Ministry Budget. "
                        f"Allocated Wallet: ₱{wallet_amount:,.2f}. "
                        f"Required: ₱{appr.amount:,.2f}."
                    )
                    return redirect(self._back_url(request))

                try:
                    unrestricted_available = self._get_unrestricted_available(request.user.church)
                except Exception as e:
                    messages.error(request, f"System Error reading Unrestricted Balance: {e}")
                    return redirect(self._back_url(request))

                if appr.amount > unrestricted_available:
                    messages.error(
                        request,
                        f"CANNOT CONTINUE: Insufficient Unrestricted Funds. "
                        f"Available Unrestricted: ₱{unrestricted_available:,.2f}. "
                        f"Required: ₱{appr.amount:,.2f}."
                    )
                    return redirect(self._back_url(request))

                posting_date = self._posting_date(request.user.church)
                actual_release_date = timezone.localdate()

                # -------------------------------------------------
                # CASH PATH
                # -------------------------------------------------
                if deduct_from == ReleasedBudget.DEDUCT_CASH:
                    try:
                        cash_on_hand = self._get_cash_on_hand(request.user.church)
                    except Exception as e:
                        messages.error(request, f"System Error reading Treasury Balance: {e}")
                        return redirect(self._back_url(request))

                    if appr.amount > cash_on_hand:
                        messages.error(
                            request,
                            f"CANNOT RELEASE: Insufficient Main Treasury Funds. "
                            f"Physical Cash Available: ₱{cash_on_hand:,.2f}. "
                            f"Required: ₱{appr.amount:,.2f}."
                        )
                        return redirect(self._back_url(request))

                    release = ReleasedBudget.objects.create(
                        church=request.user.church,
                        budget=budget,
                        ministry=appr.ministry,
                        date_released=actual_release_date,
                        amount=appr.amount,
                        remarks=appr.remarks,
                        deduct_from=ReleasedBudget.DEDUCT_CASH,
                        released_by=request.user,
                        approved_request=original_request,
                    )

                    try:
                        self._create_release_expense(
                            church=request.user.church,
                            category_name=self.CASH_RELEASE_CATEGORY,
                            amount=appr.amount,
                            budget=budget,
                            release=release,
                            request_user=request.user,
                            posting_date=posting_date,
                            released_date=actual_release_date,
                        )
                    except ValidationError as e:
                        transaction.set_rollback(True)
                        lock_msg = self._extract_validation_message(e)
                        return self._redirect_back_with_lock_modal(request, lock_msg)

                    source_label = "Main Treasury"

                # -------------------------------------------------
                # BANK PATH
                # -------------------------------------------------
                else:
                    bank_acct = self._get_bank_account(request.user.church)

                    if not bank_acct:
                        messages.warning(request, "No bank account is set. Please configure bank settings first.")
                        return redirect("bank_settings")

                    bank_acct = BankAccount.objects.select_for_update().get(pk=bank_acct.pk)

                    # Always refresh balance from ledger first
                    _call_recalc_bank_balance(request.user.church.id)
                    bank_acct.refresh_from_db(fields=["current_balance"])

                    bank_balance = bank_acct.current_balance or Decimal("0.00")

                    if appr.amount > bank_balance:
                        messages.error(
                            request,
                            f"CANNOT CONTINUE: Insufficient Bank Balance. "
                            f"Bank Balance: ₱{bank_balance:,.2f}. "
                            f"Required: ₱{appr.amount:,.2f}."
                        )
                        return redirect(self._back_url(request))

                    release = ReleasedBudget.objects.create(
                        church=request.user.church,
                        budget=budget,
                        ministry=appr.ministry,
                        date_released=actual_release_date,
                        amount=appr.amount,
                        remarks=appr.remarks,
                        deduct_from=ReleasedBudget.DEDUCT_BANK,
                        released_by=request.user,
                        approved_request=original_request,
                    )

                    try:
                        release_expense = self._create_release_expense(
                            church=request.user.church,
                            category_name=self.BANK_RELEASE_CATEGORY,
                            amount=appr.amount,
                            budget=budget,
                            release=release,
                            request_user=request.user,
                            posting_date=posting_date,
                            released_date=actual_release_date,
                        )
                    except ValidationError as e:
                        transaction.set_rollback(True)
                        lock_msg = self._extract_validation_message(e)
                        return self._redirect_back_with_lock_modal(request, lock_msg)

                    # Record the actual BANK movement in the ledger
                    mv = CashBankMovement(
                        church=request.user.church,
                        date=posting_date,
                        amount=appr.amount,
                        direction=CashBankMovement.Direction.BANK_PAID_EXPENSE,
                        source_type=CashBankMovement.SourceType.EXPENSE,
                        source_id=release_expense.id,
                        memo=(
                            f"Budget release deducted from bank | "
                            f"Release ID {release.id} | "
                            f"Ministry: {budget.ministry.name}"
                        ),
                        reference_no=f"REL-{release.id}",
                        created_by=request.user,
                        status=CashBankMovement.Status.CONFIRMED,
                    )
                    mv.full_clean()
                    mv.save()

                    # Recalculate bank balance AFTER movement save
                    _call_recalc_bank_balance(request.user.church.id)
                    bank_acct.refresh_from_db(fields=["current_balance"])

                    bank_field_names = {f.name for f in BankAccount._meta.fields}
                    update_fields = []

                    if "updated_by" in bank_field_names:
                        bank_acct.updated_by = request.user
                        update_fields.append("updated_by")

                    if "last_updated" in bank_field_names:
                        bank_acct.last_updated = timezone.now()
                        update_fields.append("last_updated")

                    if update_fields:
                        bank_acct.save(update_fields=update_fields)

                    source_label = "Bank"

                appr.delete()

        except Exception as e:
            messages.error(request, f"Release failed: {e}")
            return redirect(self._back_url(request))

        messages.success(
            request,
            f"₱{release.amount:,.2f} successfully released to "
            f"{budget.ministry.name} and deducted from {source_label}."
        )
        return redirect(self._back_url(request))

# =========================================================
# 3) DECLINE RELEASE
# =========================================================
class ReleaseBudgetDeclineView(FinanceRoleRequiredMixin, View):
    def post(self, request, pk):
        reason = (request.POST.get("declined_reason") or "").strip()

        if not reason:
            messages.error(request, "Action Failed: You must provide a reason for declining.")
            return redirect("release_budget")

        appr = get_object_or_404(
            ApprovedReleaseRequest.objects.select_related("approved_request", "budget", "budget__ministry"),
            pk=pk,
            church=request.user.church
        )

        # original request (the one submitted by the requester)
        orig = getattr(appr, "approved_request", None)

        with transaction.atomic():
            DeclinedReleaseRequest.objects.create(
                # ✅ Tracking fields (so requester can see it)
                original_request=orig,
                requested_by=getattr(orig, "requested_by", None),
                requested_at=getattr(orig, "requested_at", None),

                # ✅ Decline snapshot fields
                budget=appr.budget,
                date_released=appr.date_released,
                amount=appr.amount,
                remarks=appr.remarks,

                declined_by=request.user,
                declined_reason=reason,
                # declined_at auto_now_add=True
            )

            # Remove from Approved queue
            appr.delete()

            # Optional: also mark the original request status for consistency
            # (only if your original request has these fields)
            if orig and hasattr(orig, "status"):
                orig.status = "Rejected"
                orig.save(update_fields=["status"])

        messages.success(
            request,
            f"Release for {appr.budget.ministry.name} has been declined and archived."
        )
        return redirect("release_budget")




# =========================================================
# 4) REQUEST APPROVAL (VIEW ONLY)
# =========================================================
class BudgetRequestApprovalView(ApprovalPermissionMixin, View):
    """
    GET:  Renders the review page with fund availability breakdown.
    POST: Processes the approval with all safety checks.
    """
    template_name = "budget_request_approval.html"

    def _normalize_amount(self, value):
        if value in (None, ""):
            return None

        if isinstance(value, str):
            value = value.replace(",", "").strip()

        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _get_request_or_redirect(self, request, pk, lock=False):
        """
        Fetch the budget request with related data.
        If lock=True, uses select_for_update() for atomic safety.
        Returns (req, None) on success or (None, redirect_response) on failure.
        """
        church_id = getattr(request.user, "church_id", None)
        if not church_id:
            messages.error(request, "No church is linked to your account.")
            return None, redirect("budget_release")

        qs = BudgetReleaseRequest.objects.select_related(
            "budget",
            "budget__ministry",
            "requested_by",
        )

        if lock:
            qs = qs.select_for_update()

        req = get_object_or_404(qs, pk=pk, church_id=church_id)

        if req.status != "Pending":
            messages.warning(request, f"This request is already marked as '{req.status}'.")
            return None, redirect("budget_release")

        return req, None

    # =========================================================
    # GET — Review Page
    # =========================================================
    def get(self, request, pk):
        req, error_response = self._get_request_or_redirect(request, pk)
        if error_response:
            return error_response

        church_id = request.user.church_id

        # Ministry wallet balance
        ministry_balance = Decimal("0.00")
        try:
            wallet = BudgetBalance.objects.get(budget=req.budget)
            ministry_balance = Decimal(str(wallet.current_amount or 0))
        except BudgetBalance.DoesNotExist:
            ministry_balance = Decimal("0.00")

        # Total available funds from shared service
        physical_cash, bank_balance, total_available_funds = get_total_available_funds(church_id)

        db_amount = Decimal(str(req.amount or 0))

        # All checks must pass
        sufficient_funds = (
            db_amount <= ministry_balance
            and db_amount <= total_available_funds
        )

        return render(request, self.template_name, {
            "req": req,
            "ministry_balance": ministry_balance,
            "physical_cash": physical_cash,
            "bank_balance": bank_balance,
            "total_available_funds": total_available_funds,
            "sufficient_funds": sufficient_funds,
        })

    # =========================================================
    # POST — Process Approval
    # =========================================================
    def post(self, request, pk):
        church_id = getattr(request.user, "church_id", None)
        if not church_id:
            messages.error(request, "No church is linked to your account.")
            return redirect("budget_release")

        posted_amount = self._normalize_amount(request.POST.get("amount"))

        try:
            with transaction.atomic():
                # 1. Lock request row to prevent double approval
                req, error_response = self._get_request_or_redirect(request, pk, lock=True)
                if error_response:
                    return error_response

                budget = req.budget
                ministry = getattr(budget, "ministry", None)
                db_amount = Decimal(str(req.amount or 0))

                # 2. Anti-tamper check for posted amount
                if posted_amount is not None and posted_amount != db_amount:
                    messages.error(
                        request,
                        "Approval denied: posted amount does not match the saved request amount."
                    )
                    return redirect("budget_release")

                # 3. Prevent duplicate approval ticket
                if ApprovedReleaseRequest.objects.filter(approved_request=req).exists():
                    messages.warning(
                        request,
                        "This request already has an approval record."
                    )
                    return redirect("budget_release")

                # 4. Lock ministry wallet
                try:
                    wallet = BudgetBalance.objects.select_for_update().get(budget=budget)
                except BudgetBalance.DoesNotExist:
                    messages.error(
                        request,
                        "Critical Error: Budget Wallet (BudgetBalance) not found."
                    )
                    return redirect("budget_release")

                ministry_balance = Decimal(str(wallet.current_amount or 0))

                # 5. Total available funds from shared service
                physical_cash, bank_balance, total_available_funds = get_total_available_funds(church_id)

                # 6. Check ministry wallet
                if db_amount > ministry_balance:
                    messages.error(
                        request,
                        f"Insufficient Ministry Funds. "
                        f"Wallet: ₱{ministry_balance:,.2f}. Requested: ₱{db_amount:,.2f}."
                    )
                    return redirect("budget_release")

                # 7. Check total available funds
                if db_amount > total_available_funds:
                    messages.error(
                        request,
                        f"Insufficient Total Available Funds. "
                        f"Physical Cash: ₱{physical_cash:,.2f}. "
                        f"Bank Balance: ₱{bank_balance:,.2f}. "
                        f"Total Available: ₱{total_available_funds:,.2f}. "
                        f"Requested: ₱{db_amount:,.2f}."
                    )
                    return redirect("budget_release")

                # 8. Create approval ticket
                ApprovedReleaseRequest.objects.create(
                    church_id=church_id,
                    ministry=ministry,
                    budget=budget,
                    date_released=req.date_released,
                    amount=db_amount,
                    remarks=req.remarks,
                    approved_by=request.user,
                    approved_request=req,
                )

                # 9. Update original request status
                req.status = "Approved"
                req.approved_by = request.user
                req.approved_at = timezone.now()
                req.save(update_fields=["status", "approved_by", "approved_at"])

            role_title = getattr(request.user, "user_type", "User")
            messages.success(
                request,
                f"Request for ₱{db_amount:,.2f} approved by {role_title} {request.user.username}."
            )
            return redirect("budget_release")

        except Exception as e:
            messages.error(request, f"Approval failed: {e}")
            return redirect("budget_release")

# =========================================================
# 5) EXPENSES (LIQUIDATION & AUTO-RETURN) + LOCK CHECK FIX
# =========================================================
class BudgetExpenseView(FinanceRoleRequiredMixin, View):
    template_name = "budget_expense.html"

    def build_sidebar_context(self, request, selected_release=None):
        return {
            "pending_liquidations": (
                ReleasedBudget.objects
                .filter(church=request.user.church, is_liquidated=False)
                .select_related("budget", "budget__ministry")
                .order_by("-date_released")
            ),
            "liquidated_history": (
                ReleasedBudget.objects
                .filter(church=request.user.church, is_liquidated=True)
                .select_related("budget", "budget__ministry")
                .order_by("-released_at")[:50]
            ),
            "selected_release": selected_release,
            "bank_acct": BankAccount.objects.filter(church=request.user.church).first(),
        }

    def _render_page(self, request, selected_release, formset, existing_expenses=None):
        if existing_expenses is None:
            existing_expenses = (
                selected_release.expenses.all().order_by("-date_incurred")
                if selected_release
                else BudgetExpense.objects.none()
            )

        context = self.build_sidebar_context(request, selected_release)
        context.update({
            "formset": formset,
            "existing_expenses": existing_expenses,
        })
        return render(request, self.template_name, context)

    def _get_lock_date(self, church):
        if hasattr(church, "accounting_settings") and church.accounting_settings:
            return church.accounting_settings.lock_date
        return None

    def _posting_date(self, church):
        """
        IMPORTANT:
        OtherIncome.clean() blocks date <= lock_date.
        So the valid posting date must be the NEXT OPEN DATE.
        """
        lock_date = self._get_lock_date(church)
        today = timezone.localdate()

        if lock_date and today <= lock_date:
            return lock_date + timedelta(days=1)

        return today

    # -------------------------------------------------
    # Expense Fraud Detection (shared with regular Expense)
    # -------------------------------------------------
    def _expense_fraud_enabled(self, request) -> bool:
        return _is_expense_fraud_detection_enabled(request.user)

    def _proof_field_name(self, form):
        """
        BudgetExpenseForm uses receipt_proof.
        """
        return "receipt_proof" if "receipt_proof" in getattr(form, "fields", {}) else None

    def _existing_proof_name(self, instance, field_name):
        if not instance or not getattr(instance, "pk", None) or not field_name:
            return ""

        current_file = getattr(instance, field_name, None)
        return getattr(current_file, "name", "") if current_file else ""

    def _run_budget_expense_fraud_checks(self, request, formset):
        """
        Reuse the SAME expense fraud toggle + verifier for budget expenses.

        Rules:
        - If expense fraud detection is OFF -> do nothing
        - If ON:
            * new entries must have a receipt proof
            * edited entries may keep existing proof
            * only newly uploaded proofs are re-verified
        """
        if not self._expense_fraud_enabled(request):
            return True

        has_errors = False

        for form in formset.forms:
            cd = getattr(form, "cleaned_data", None) or {}
            if not cd:
                continue

            if cd.get("DELETE", False):
                continue

            field_name = self._proof_field_name(form)
            if not field_name:
                form.add_error(
                    None,
                    "Expense fraud detection is enabled, but no receipt proof field is configured for budget expenses."
                )
                has_errors = True
                continue

            instance = form.instance
            current_proof_name = self._existing_proof_name(instance, field_name)
            uploaded_file = cd.get("receipt_proof")

            amount = cd.get("amount")
            date_incurred = cd.get("date_incurred")
            description = cd.get("description") or ""

            # Skip totally empty extra rows
            if (
                not getattr(instance, "pk", None)
                and not any([amount, date_incurred, description, uploaded_file])
            ):
                continue

            # Require proof when ON
            if not uploaded_file and not current_proof_name:
                form.add_error(
                    "receipt_proof",
                    "Receipt proof is required while expense fraud detection is enabled."
                )
                has_errors = True
                continue

            # Verify only newly uploaded proof
            if uploaded_file:
                is_valid, reason = _verify_expense_receipt(
                    uploaded_file,
                    amount=amount,
                    expense_date=date_incurred,
                    category_obj=None,   # BudgetExpense has no category field in your form/model flow here
                    description=description,
                    church=request.user.church,
                )

                if not is_valid:
                    form.add_error("receipt_proof", f"Fraud detection failed: {reason}")
                    has_errors = True

        if has_errors:
            messages.error(
                request,
                "Please fix the flagged receipt proof issue(s). Expense fraud detection is enabled for budget expenses."
            )
            return False

        return True

    # -------------------------------------------------
    # Existing helpers
    # -------------------------------------------------
    def _inject_release(self, formset, release):
        formset.release_instance = release
        for f in formset.forms:
            f.release_instance = release

    def _get_bank_account(self, church):
        try:
            return church.bank_account
        except Exception:
            return BankAccount.objects.filter(church=church).first()

    def _touch_bank_account(self, bank_acct, request_user):
        if not bank_acct:
            return

        bank_field_names = {f.name for f in BankAccount._meta.fields}
        update_fields = []

        if "updated_by" in bank_field_names:
            bank_acct.updated_by = request_user
            update_fields.append("updated_by")

        if "last_updated" in bank_field_names:
            bank_acct.last_updated = timezone.now()
            update_fields.append("last_updated")

        if update_fields:
            bank_acct.save(update_fields=update_fields)

    def _bank_return_direction(self):
        try:
            return CashBankMovement.Direction.DIRECT_BANK_RECEIPT
        except AttributeError:
            raise ValidationError(
                "CashBankMovement.Direction.DIRECT_BANK_RECEIPT is not defined."
            )

    def _bank_other_income_source_type(self):
        try:
            return CashBankMovement.SourceType.OTHER_INCOME
        except AttributeError:
            raise ValidationError(
                "CashBankMovement.SourceType.OTHER_INCOME is not defined."
            )

    def _return_reference_no(self, release):
        return f"RET-{release.id}"

    def _resolve_return_to(self, request, release, require_explicit=False):
        raw = request.POST.get("return_to")

        allowed = {
            ReleasedBudget.DEDUCT_CASH,
            ReleasedBudget.DEDUCT_BANK,
        }

        if raw is None or str(raw).strip() == "":
            if require_explicit and not release.is_liquidated:
                raise ValidationError(
                    "Please choose where the unused budget will be returned "
                    "(Physical Cash or Bank)."
                )
            return release.return_to or ReleasedBudget.DEDUCT_CASH

        value = str(raw).strip().lower()
        if value not in allowed:
            raise ValidationError("Invalid return destination selected.")

        if value == ReleasedBudget.DEDUCT_BANK and not self._get_bank_account(release.church):
            raise ValidationError("No bank account is set. Please configure bank settings first.")

        return value

    def _recompute_liquidation(self, release, request_user, return_to, finalize=False):
        total_spent = release.expenses.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        amount_released = release.amount or Decimal("0.00")

        change = amount_released - total_spent
        if change < 0:
            change = Decimal("0.00")

        release.amount_returned = change
        release.liquidated_date = self._posting_date(release.church)
        release.return_to = return_to

        update_fields = ["amount_returned", "liquidated_date", "return_to"]

        if finalize and not release.is_liquidated:
            release.is_liquidated = True
            update_fields.append("is_liquidated")

        release.save(update_fields=update_fields)

        # budget_return_income is synced inside ReleasedBudget.save()
        self._sync_return_destination(release, request_user)

        return change

    def _sync_return_destination(self, release, request_user):
        self._sync_bank_return_movement(release, request_user)
        self._sync_cash_return_transaction(release, request_user)

    def _cleanup_legacy_bank_return_expense(self, release):
        """
        Transitional cleanup for old logic that stored bank returns as Expense.
        Safe even after field removal.
        """
        legacy_exp = getattr(release, "bank_return_expense", None)
        if legacy_exp:
            legacy_exp.delete()
            try:
                release.bank_return_expense = None
                release.save(update_fields=["bank_return_expense"])
            except Exception:
                pass

    def _sync_bank_return_movement(self, release, request_user):
        """
        Bank return rules:
        - returned amount is already recorded as OtherIncome via release.budget_return_income
        - this method only mirrors the destination in the bank ledger
        - NO Expense is created here
        """
        posting_date = self._posting_date(release.church)
        amount_to_bank = Decimal(release.amount_returned or 0)
        reference_no = self._return_reference_no(release)
        bank_source_type = self._bank_other_income_source_type()

        # remove legacy expense-based artifacts
        self._cleanup_legacy_bank_return_expense(release)

        # if not returning to bank, or nothing to return, or no linked OtherIncome -> remove bank movement
        if (
            release.return_to != ReleasedBudget.DEDUCT_BANK
            or amount_to_bank <= 0
            or not release.is_liquidated
            or not release.budget_return_income_id
        ):
            CashBankMovement.objects.filter(
                church=release.church,
                reference_no=reference_no,
            ).delete()

            _call_recalc_bank_balance(release.church.id)

            bank_acct = self._get_bank_account(release.church)
            if bank_acct:
                bank_acct.refresh_from_db(fields=["current_balance"])
                self._touch_bank_account(bank_acct, request_user)
            return

        bank_acct = self._get_bank_account(release.church)
        if not bank_acct:
            raise ValidationError("No bank account is set. Please configure bank settings first.")

        bank_acct = BankAccount.objects.select_for_update().get(pk=bank_acct.pk)

        # refresh relation after ReleasedBudget.save() synced OtherIncome
        release.refresh_from_db(fields=["budget_return_income"])
        if not release.budget_return_income_id:
            raise ValidationError("Budget return income entry was not created.")

        linked_income = release.budget_return_income
        bank_direction = self._bank_return_direction()

        income_field_names = {f.name for f in OtherIncome._meta.fields}
        income_update_fields = []

        if "created_by" in income_field_names and not linked_income.created_by_id:
            linked_income.created_by = request_user
            income_update_fields.append("created_by")

        if "edited_by" in income_field_names:
            linked_income.edited_by = request_user
            income_update_fields.append("edited_by")

        if income_update_fields:
            linked_income.save(update_fields=income_update_fields)

        movement_qs = CashBankMovement.objects.filter(
            church=release.church,
            reference_no=reference_no,
        ).order_by("-id")

        primary_mv = movement_qs.first()

        if primary_mv:
            movement_qs.exclude(pk=primary_mv.pk).delete()

            primary_mv.date = posting_date
            primary_mv.amount = amount_to_bank
            primary_mv.direction = bank_direction
            primary_mv.source_type = bank_source_type
            primary_mv.source_id = linked_income.id
            primary_mv.memo = f"Budget Return to Bank | Release ID {release.id}"
            primary_mv.reference_no = reference_no
            primary_mv.created_by = request_user
            primary_mv.status = CashBankMovement.Status.CONFIRMED
            primary_mv.full_clean()
            primary_mv.save()
        else:
            mv = CashBankMovement(
                church=release.church,
                date=posting_date,
                amount=amount_to_bank,
                direction=bank_direction,
                source_type=bank_source_type,
                source_id=linked_income.id,
                memo=f"Budget Return to Bank | Release ID {release.id}",
                reference_no=reference_no,
                created_by=request_user,
                status=CashBankMovement.Status.CONFIRMED,
            )
            mv.full_clean()
            mv.save()

        _call_recalc_bank_balance(release.church.id)
        bank_acct.refresh_from_db(fields=["current_balance"])
        self._touch_bank_account(bank_acct, request_user)

    def _sync_cash_return_transaction(self, release, request_user):
        posting_date = self._posting_date(release.church)
        amount_to_cash = Decimal(release.amount_returned or 0)

        old_txn = release.cash_return_txn

        if (
            release.return_to != ReleasedBudget.DEDUCT_CASH
            or amount_to_cash <= 0
            or not release.is_liquidated
        ):
            if old_txn:
                old_txn.delete()
                release.cash_return_txn = None
                release.save(update_fields=["cash_return_txn"])
            return

        txn_kwargs = dict(
            church=release.church,
            txn_date=posting_date,
            direction=CashTransaction.IN,
            amount=amount_to_cash,
            source_type=CashTransaction.SOURCE_BUDGET_RETURN,
            description=f"Budget Return to Physical Cash | Release ID {release.id}",
            created_by=request_user,
        )

        if old_txn:
            for k, v in txn_kwargs.items():
                setattr(old_txn, k, v)
            old_txn.save()
        else:
            linked_txn = CashTransaction.objects.create(**txn_kwargs)
            release.cash_return_txn = linked_txn
            release.save(update_fields=["cash_return_txn"])

    def get(self, request):
        release_id = request.GET.get("release_id")
        selected_release = None
        existing_expenses = BudgetExpense.objects.none()

        if release_id:
            selected_release = get_object_or_404(
                ReleasedBudget,
                id=release_id,
                church=request.user.church,
            )
            existing_expenses = selected_release.expenses.all().order_by("-date_incurred")

        formset = BudgetExpenseFormSet(queryset=existing_expenses)

        if selected_release:
            self._inject_release(formset, selected_release)

        return self._render_page(request, selected_release, formset, existing_expenses)

    def post(self, request):
        if request.POST.get("action") == "OCR":
            uploaded = request.FILES.get("receipt_proof") or request.FILES.get("receipt")
            if not uploaded:
                return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)
            try:
                result = analyze_receipt_with_openai(uploaded)
                if isinstance(result, dict) and "error" in result:
                    return JsonResponse({"success": False, "error": result["error"]}, status=500)
                return JsonResponse({"success": True, "data": result})
            except Exception as e:
                return JsonResponse({"success": False, "error": str(e)}, status=500)

        release_id = request.POST.get("release_id")
        if not release_id:
            messages.error(request, "No release selected.")
            return redirect("budget_expense")

        selected_release = get_object_or_404(
            ReleasedBudget,
            id=release_id,
            church=request.user.church,
        )

        qs = selected_release.expenses.all()
        formset = BudgetExpenseFormSet(request.POST, request.FILES, queryset=qs)
        self._inject_release(formset, selected_release)

        if not formset.is_valid():
            return self._render_page(
                request,
                selected_release,
                formset,
                qs.order_by("-date_incurred"),
            )

        lock_date = self._get_lock_date(request.user.church)
        if lock_date:
            for form in formset.forms:
                cd = getattr(form, "cleaned_data", None) or {}
                if not cd:
                    continue

                instance = form.instance
                marked_for_delete = cd.get("DELETE", False)
                new_date = cd.get("date_incurred")

                if instance and instance.pk:
                    original_date = instance.date_incurred

                    if original_date and original_date <= lock_date:
                        if form.has_changed() or marked_for_delete:
                            form.add_error(None, f"LOCKED: Cannot edit/delete entries dated on or before {lock_date}.")
                            continue

                    if new_date and new_date <= lock_date:
                        form.add_error("date_incurred", f"LOCKED: Date must be after {lock_date}.")
                else:
                    if marked_for_delete:
                        continue
                    if new_date and new_date <= lock_date:
                        form.add_error("date_incurred", f"LOCKED: Date must be after {lock_date}.")

            if any(f.errors for f in formset.forms):
                return self._render_page(
                    request,
                    selected_release,
                    formset,
                    qs.order_by("-date_incurred"),
                )

        # Reuse the SAME expense fraud detection for budget expenses
        if not self._run_budget_expense_fraud_checks(request, formset):
            return self._render_page(
                request,
                selected_release,
                formset,
                qs.order_by("-date_incurred"),
            )

        finalize_requested = request.POST.get("finalize") == "true"
        must_recompute_liquidation = finalize_requested or selected_release.is_liquidated

        try:
            with transaction.atomic():
                for f in formset.deleted_forms:
                    if f.instance and f.instance.pk:
                        f.instance.delete()

                objs = formset.save(commit=False)
                for exp in objs:
                    exp.release = selected_release
                    exp.church = request.user.church
                    exp.ministry = selected_release.ministry

                    if exp._state.adding:
                        exp.created_by = request.user

                    exp.save()

                if hasattr(formset, "save_m2m"):
                    formset.save_m2m()

                if must_recompute_liquidation:
                    return_to = self._resolve_return_to(
                        request,
                        selected_release,
                        require_explicit=finalize_requested and not selected_release.is_liquidated,
                    )

                    change = self._recompute_liquidation(
                        selected_release,
                        request.user,
                        return_to=return_to,
                        finalize=finalize_requested,
                    )

                    if finalize_requested:
                        messages.success(
                            request,
                            f"Liquidation Finalized! ₱{change:,.2f} returned to "
                            f"{'Bank' if return_to == ReleasedBudget.DEDUCT_BANK else 'Physical Cash'}."
                        )
                        return redirect("budget_expense")

        except ValidationError as ve:
            formset._non_form_errors = formset.error_class([str(ve)])
            return self._render_page(
                request,
                selected_release,
                formset,
                selected_release.expenses.all().order_by("-date_incurred"),
            )

        if selected_release.is_liquidated:
            messages.success(request, "Expenses updated successfully. Liquidation totals were recalculated.")
        else:
            messages.success(request, "Expenses saved/updated successfully.")

        return redirect(f"{request.path}?release_id={release_id}")


class DeleteExpenseView(BudgetExpenseView):
    def post(self, request, pk):
        from django.urls import reverse

        try:
            with transaction.atomic():
                expense = get_object_or_404(
                    BudgetExpense.objects.select_related("release"),
                    pk=pk,
                    church=request.user.church,
                )

                release = get_object_or_404(
                    ReleasedBudget.objects.select_for_update(),
                    pk=expense.release_id,
                    church=request.user.church,
                )
                release_id = release.id

                lock_date = self._get_lock_date(request.user.church)
                if lock_date and expense.date_incurred and expense.date_incurred <= lock_date:
                    messages.error(request, f"LOCKED: Cannot delete entries dated on or before {lock_date}.")
                    base_url = reverse("budget_expense")
                    return redirect(f"{base_url}?release_id={release_id}")

                expense.delete()

                if release.is_liquidated:
                    return_to = release.return_to or ReleasedBudget.DEDUCT_CASH

                    self._recompute_liquidation(
                        release,
                        request.user,
                        return_to=return_to,
                        finalize=False,
                    )

                    messages.success(
                        request,
                        "Receipt deleted. Liquidation totals were recalculated."
                    )
                else:
                    messages.success(request, "Receipt deleted.")

        except ValidationError as ve:
            messages.error(request, str(ve))
            release_id = getattr(getattr(locals().get("release", None), "id", None), "__str__", lambda: "")()
            base_url = reverse("budget_expense")
            if release_id:
                return redirect(f"{base_url}?release_id={release_id}")
            return redirect(base_url)

        base_url = reverse("budget_expense")
        return redirect(f"{base_url}?release_id={release_id}")

class FinancialReportView(TemplateView):
    """
    Reports page: weekly / monthly / yearly summaries + full Expenses list (paginated)
    Accepts GET params:
      - period: weekly|monthly|yearly  (default: monthly)
      - start: YYYY-MM-DD (optional)
      - end:   YYYY-MM-DD (optional)
      - epage: expense list page number (optional)
    POST with download_excel=1 exports the summary table to Excel (respects filters).
    """
    template_name = "financial_reports.html"

    def get(self, request, *args, **kwargs):
        context = self._build_context(request)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        if request.POST.get("download_excel"):
            period = (request.POST.get("period") or "monthly").lower()
            start = request.POST.get("start") or ""
            end = request.POST.get("end") or ""
            data = self._aggregate(period, start, end)
            return self._export_excel(period, data, start, end)
        return redirect("financial_reports")

    # ---------------- helpers ----------------
    def _build_context(self, request):
        period = (request.GET.get("period") or "monthly").lower()
        start = request.GET.get("start") or ""
        end = request.GET.get("end") or ""

        rows = self._aggregate(period, start, end)

        totals = {
            "tithes": sum(r["tithes"] for r in rows),
            "offerings": sum(r["offerings"] for r in rows),
            "donations": sum(r["donations"] for r in rows),
            "expenses": sum(r["expenses"] for r in rows),
            "accumulated": sum(r["accumulated"] for r in rows),
        }

        # All expenses (respect date range), with pagination
        expenses_qs = self._fetch_expenses(start, end)
        expense_total = expenses_qs.aggregate(total=Sum("amount"))["total"] or 0

        paginator = Paginator(expenses_qs, 25)  # show 25 rows per page
        page_num = request.GET.get("epage") or 1
        try:
            expense_page = paginator.page(page_num)
        except PageNotAnInteger:
            expense_page = paginator.page(1)
        except EmptyPage:
            expense_page = paginator.page(paginator.num_pages)

        return {
            "rows": rows,
            "totals": totals,
            "period": period,
            "start": start,
            "end": end,
            "expense_page": expense_page,
            "expense_total": expense_total,
        }

    def _fetch_expenses(self, start, end):
        start_date = parse_date(start) if start else None
        end_date = parse_date(end) if end else None

        qs = Expense.objects.select_related("created_by", "edited_by").order_by("-expense_date")
        if start_date:
            qs = qs.filter(expense_date__gte=start_date)
        if end_date:
            qs = qs.filter(expense_date__lte=end_date)
        return qs

    def _aggregate(self, period, start, end):
        """
        Returns list of dicts:
        [
          {
            "label": "2025-W44" | "2025-11" | "2025",
            "date_key": datetime/date,
            "tithes": Decimal,
            "offerings": Decimal,
            "donations": Decimal,
            "expenses": Decimal,
            "accumulated": Decimal,
          },
          ...
        ]
        """
        trunc = {
            "weekly": TruncWeek,
            "monthly": TruncMonth,
            "yearly": TruncYear,
        }.get(period, TruncMonth)

        start_date = parse_date(start) if start else None
        end_date = parse_date(end) if end else None

        def _range(qs, field):
            if start_date:
                qs = qs.filter(**{f"{field}__gte": start_date})
            if end_date:
                qs = qs.filter(**{f"{field}__lte": end_date})
            return qs

        # group each model by chosen period (adjust field names if your models differ)
        t_qs = _range(Tithe.objects.all(), "date").annotate(p=trunc("date")).values("p").annotate(total=Sum("amount"))
        o_qs = _range(Offering.objects.all(), "date").annotate(p=trunc("date")).values("p").annotate(total=Sum("amount"))
        d_qs = _range(Donations.objects.all(), "donations_date").annotate(p=trunc("donations_date")).values("p").annotate(total=Sum("amount"))
        e_qs = _range(Expense.objects.all(), "expense_date").annotate(p=trunc("expense_date")).values("p").annotate(total=Sum("amount"))

        # index by the period key date
        bucket = {}
        def put(qs, key):
            for row in qs:
                p = row["p"]
                bucket.setdefault(p, {"tithes": 0, "offerings": 0, "donations": 0, "expenses": 0})
                bucket[p][key] = row["total"] or 0

        put(t_qs, "tithes")
        put(o_qs, "offerings")
        put(d_qs, "donations")
        put(e_qs, "expenses")

        rows = []
        for p, vals in bucket.items():
            t = vals["tithes"] or 0
            o = vals["offerings"] or 0
            d = vals["donations"] or 0
            ex = vals["expenses"] or 0
            acc = (t + o + d) - ex

            if period == "weekly":
                # p is the week's Monday date; get ISO year/week for label
                iso_year, iso_week, _ = p.isocalendar()
                label = f"{iso_year}-W{iso_week:02d}"
            elif period == "yearly":
                label = f"{p.year}"
            else:
                label = f"{p.year}-{p.month:02d}"

            rows.append({
                "label": label,
                "date_key": p,
                "tithes": t,
                "offerings": o,
                "donations": d,
                "expenses": ex,
                "accumulated": acc,
            })

        rows.sort(key=lambda r: r["date_key"], reverse=True)
        return rows

    def _export_excel(self, period, data, start, end):
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{period.capitalize()} Report"

        hdr = [
            {"weekly": "Week", "monthly": "Month", "yearly": "Year"}[period],
            "Total Tithes", "Total Offering", "Total Donations", "Total Expenses", "Accumulated"
        ]
        for c, h in enumerate(hdr, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True)

        for r, row in enumerate(data, start=2):
            ws.cell(row=r, column=1, value=row["label"])
            ws.cell(row=r, column=2, value=float(row["tithes"]))
            ws.cell(row=r, column=3, value=float(row["offerings"]))
            ws.cell(row=r, column=4, value=float(row["donations"]))
            ws.cell(row=r, column=4, value=float(row["ApprovedReference"]))
            ws.cell(row=r, column=5, value=float(row["expenses"]))
            ws.cell(row=r, column=6, value=float(row["accumulated"]))

        # Totals row
        last = len(data) + 2
        ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, 7):
            col_letter = get_column_letter(col)
            ws.cell(row=last, column=col, value=f"=SUM({col_letter}2:{col_letter}{last-1})").font = Font(bold=True)

        fname = f"financial_{period}_{start or 'all'}_{end or 'all'}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp



class BudgetRequest(View):
    """
    Queue release requests for approval.
    """
    template_name = "budget_request.html"

    def _form(self, *args, **kwargs):
        f = BudgetReleaseRequestForm(*args, **kwargs)
        try: f.fields["date_released"].widget.input_type = "date"
        except Exception: pass
        return f

    def get(self, request):
        form = self._form()
        pending = BudgetReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:100]
        approved = ApprovedReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:100]
        declined = DeclinedReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:50]
        return render(request, self.template_name, {
            "form": form, "pending": pending, "approved": approved, "declined": declined
        })

    def post(self, request):
        data = request.POST.copy()
        if "amount" in data:
            data["amount"] = _normalize_amount(data["amount"])
        form = self._form(data)
        if form.is_valid():
            req = form.save(commit=False)
            req.requested_by = request.user
            req.save()
            messages.success(request, "Release request submitted for approval.")
            return redirect("budget_request")
        pending = BudgetReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:100]
        approved = ApprovedReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:100]
        declined = DeclinedReleaseRequest.objects.select_related("budget", "budget__ministry").all()[:50]
        return render(request, self.template_name, {
            "form": form, "pending": pending, "approved": approved, "declined": declined
        })

class PrescriptiveBudgetOptimizerView(LoginRequiredMixin, TemplateView):
    template_name = "prescriptive_budget_optimizer.html"

class PendingReferenceCreateView(LoginRequiredMixin, View):
    template_name = 'submit_reference.html'

    def get(self, request):
        form = PendingReferenceForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = PendingReferenceForm(request.POST, request.FILES)
        if form.is_valid():
            pending = form.save(commit=False)
            pending.submitted_by = request.user
            pending.save()
            messages.success(request, "Reference submitted for approval.")
            return redirect('reference_submit')
        return render(request, self.template_name, {'form': form})

class PendingReferenceListView(LoginRequiredMixin, View):
    template_name = 'approve_reference.html'

    def get(self, request):
        pending_items = PendingReference.objects.order_by('-submitted_at')
        return render(request, self.template_name, {'pending_items': pending_items})


class ApproveReferenceView(LoginRequiredMixin, View):
    def post(self, request, pk):
        pending = get_object_or_404(PendingReference, pk=pk)

        # Create ApprovedReference from PendingReference
        approved = ApprovedReference.objects.create(
            reference_number=pending.reference_number,
            date=pending.date,
            amount=pending.amount,
            name=pending.name,
            file=pending.file,  # same file path
            submitted_by=pending.submitted_by,
            submitted_at=pending.submitted_at,
            approved_by=request.user,
            approved_at=timezone.now(),
        )

        # now delete from pending DB
        pending.delete()

        messages.success(request, f"Reference {approved.reference_number} approved.")
        return redirect('reference_approve_list')

class ReferenceHistoryView(View):
    template_name = 'reference_history.html'

    def get(self, request):
        pending_items = PendingReference.objects.order_by('-submitted_at')
        approved_items = ApprovedReference.objects.order_by('-approved_at')
        context = {
            'pending_items': pending_items,
            'approved_items': approved_items,
        }
        return render(request, self.template_name, context)


class SendThankYouLetterView(LoginRequiredMixin, View):
    template_name = 'send_thank_you_letter.html'

    def get(self, request):
        form = ThankYouLetterForm()
        approved_references = ApprovedReference.objects.filter(submitted_by__isnull=False)
        return render(request, self.template_name, {
            'form': form,
            'approved_references': approved_references
        })

    def post(self, request):
        form = ThankYouLetterForm(request.POST)
        if not form.is_valid():
            messages.error(request, "There was an error sending the thank you letters.")
            return render(request, self.template_name, {
                'form': form,
                'approved_references': ApprovedReference.objects.filter(submitted_by__isnull=False),
            })

        selected_references = request.POST.getlist('approved_references')
        base_message = form.cleaned_data['message']

        for ref_id in selected_references:
            approved_ref = get_object_or_404(ApprovedReference, id=ref_id)
            recipient_user = approved_ref.submitted_by
            recipient_email = getattr(recipient_user, 'email', '') or ''

            # Build email body using ApprovedReference fields
            ctx = {
                'recipient': recipient_user,
                'sender': request.user,
                'message': base_message,
                'ref': approved_ref,   # has: reference_number, date, amount, name, file
            }

            try:
                html_body = render_to_string('emails/thank_you_letter.html', ctx)
                text_body = strip_tags(html_body)
            except TemplateDoesNotExist:
                # Plain-text fallback
                text_body = (
                    f"Dear {recipient_user.first_name or recipient_user.username},\n\n"
                    f"{base_message}\n\n"
                    f"Reference Details:\n"
                    f"  Reference No.: {approved_ref.reference_number}\n"
                    f"  Name: {approved_ref.name}\n"
                    f"  Amount: {approved_ref.amount}\n"
                    f"  Date: {approved_ref.date}\n\n"
                    f"With gratitude,\n{request.user.get_full_name() or request.user.username}"
                )
                html_body = None

            # Save the letter (matches your model exactly — no snapshot fields)
            ThankYouLetter.objects.create(
                reference=approved_ref,
                recipient=recipient_user,
                sender=request.user,
                message=base_message,
            )

            # Send the email (HTML optional)
            if recipient_email:
                send_mail(
                    subject='Thank You for Your Reference Submission',
                    message=text_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[recipient_email],
                    fail_silently=False,
                    html_message=html_body,
                )

            messages.success(request, f"Thank-you letter prepared for {recipient_user.username}.")

        return redirect('reference_history')


class ThankYouLetterView(LoginRequiredMixin, View):
    template_name = 'thank_you_letter_list.html'

    def get(self, request):
        # Get all the thank-you letters where the logged-in user is the recipient
        thank_you_letters = ThankYouLetter.objects.filter(recipient=request.user).order_by('-sent_at')

        return render(request, self.template_name, {'thank_you_letters': thank_you_letters})

class ReferenceHistoryViewMember(LoginRequiredMixin, View):
    template_name = 'reference_history_member.html'

    def get(self, request):
        user = request.user

        date_from = request.GET.get('from')
        date_to   = request.GET.get('to')

        pending_qs  = PendingReference.objects.filter(submitted_by=user)
        approved_qs = ApprovedReference.objects.filter(submitted_by=user)

        if date_from:
            approved_qs = approved_qs.filter(date__gte=date_from)
            pending_qs  = pending_qs.filter(submitted_at__date__gte=date_from)
        if date_to:
            approved_qs = approved_qs.filter(date__lte=date_to)
            pending_qs  = pending_qs.filter(submitted_at__date__lte=date_to)

        pending_items  = pending_qs.order_by('-submitted_at')
        approved_items = approved_qs.order_by('-approved_at')

        approved_total = approved_items.aggregate(total=Sum('amount'))['total'] or 0
        # If PendingReference has no amount field, set pending_total = 0 or remove it in the template
        pending_total  = pending_items.aggregate(total=Sum('amount'))['total'] or 0

        context = {
            'pending_items': pending_items,
            'approved_items': approved_items,
            'approved_total': approved_total,
            'pending_total': pending_total,
            'date_from': date_from or '',
            'date_to': date_to or '',
            'today': timezone.now().date(),
        }
        return render(request, self.template_name, context)


def build_priority_features(model, feats_5):
    """
    Aligns feature vector length with the trained model.

    Your intended inference features (5):
      [pool_used_monthly, mode_is_historical, ui_is_monthly, min_req, max_cap]

    If the trained model expects 7, we pad with 2 zeros (or truncate if smaller).
    """
    expected = getattr(model, "n_features_in_", None)
    if expected is None:
        # If model doesn't expose this, assume your 5-feature design.
        return feats_5

    feats = list(feats_5)
    if len(feats) < expected:
        feats.extend([0.0] * (expected - len(feats)))
    elif len(feats) > expected:
        feats = feats[:expected]
    return feats

def _latest_budget_monthly_map(church_id: int, year: int, month: int):
    """
    BudgetManage-aligned snapshot:
    - includes budgets for selected month + yearly pool (month=0)
    - yearly pool converted to monthly by /12
    Returns (monthly_map, total_need_monthly)
    """
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    rows = (
        MinistryBudget.objects
        .filter(church_id=church_id, year=year, month__in=[month, 0], is_active=True)
        .values("ministry_id", "month")
        .annotate(total=Coalesce(Sum("amount_allocated"), Decimal("0.00")))
    )

    monthly_map = {}
    for r in rows:
        mid = r["ministry_id"]
        amt = Decimal(str(r["total"] or "0.00"))
        if int(r["month"]) == 0:
            amt = amt / Decimal("12")  # yearly -> monthly equivalent
        monthly_map[mid] = monthly_map.get(mid, Decimal("0.00")) + amt

    total_need = sum(monthly_map.values(), Decimal("0.00"))
    return monthly_map, total_need


def build_manual_inputs_from_budgets(church_id: int, year: int, month: int):
    """
    Manual source = BudgetManageView budgets.
    Returns:
      monthly_pool (Decimal)
      ministry_rows: list[dict] with {id, name, min_req_m, max_cap_m, base_budget_m}
      meta: dict
    """
    Ministry = apps.get_model("Register", "Ministry")
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    # Budgets available for that year + selected month and yearly pool (0)
    budget_rows = list(
        MinistryBudget.objects.filter(
            church_id=church_id,
            year=year,
            month__in=[month, 0],
            is_active=True
        ).select_related("ministry")
    )

    # Pick "effective" monthly-equivalent budget per ministry
    # Prefer monthly (month==selected) over yearly (month==0)
    chosen = {}  # ministry_id -> monthly_equiv_budget
    for b in budget_rows:
        mid = b.ministry_id
        amt = Decimal(str(b.amount_allocated or "0.00"))
        if b.month == month:
            chosen[mid] = amt                      # already monthly
        elif b.month == 0 and mid not in chosen:
            chosen[mid] = (amt / Decimal("12"))    # yearly -> monthly

    ministries = list(
        Ministry.objects.filter(church_id=church_id, is_active=True).order_by("name")
    )

    ministry_rows = []
    for m in ministries:
        base_budget_m = chosen.get(m.id, Decimal("0.00"))

        db_min_m = Decimal(str(getattr(m, "min_monthly_budget", 0) or 0))
        db_cap_m = Decimal(str(getattr(m, "max_monthly_cap", 0) or 0))  # 0=no cap

        # Manual min requirement comes from budget manage view (base budget), but never below DB min.
        min_req_m = max(base_budget_m, db_min_m)

        # Manual cap: if DB cap exists use it, else allow some headroom above base budget
        if db_cap_m > 0:
            max_cap_m = db_cap_m
        else:
            max_cap_m = (base_budget_m * Decimal("1.25")) if base_budget_m > 0 else Decimal("0.00")

        if max_cap_m > 0 and max_cap_m < min_req_m:
            max_cap_m = min_req_m

        ministry_rows.append({
            "id": m.id,
            "name": m.name,
            "base_budget_m": base_budget_m,
            "min_req_m": min_req_m,
            "max_cap_m": max_cap_m,
        })

    # Budget-based “need” pool (monthly)
    budget_need_m = sum((r["base_budget_m"] for r in ministry_rows), Decimal("0.00"))

    meta = {
        "year": year,
        "month": month,
        "budget_need_monthly": budget_need_m
    }
    return budget_need_m, ministry_rows, meta

@require_POST
@login_required
def api_get_ministry_stats_year(request):
    if not hasattr(request.user, "church") or not request.user.church:
        return JsonResponse({"error": "No church assigned"}, status=403)

    try:
        data = json.loads(request.body or "{}")
        year = int(data.get("year"))
    except Exception:
        return JsonResponse({"error": "Invalid year"}, status=400)

    church_id = request.user.church.id

    # IMPORTANT: ReleasedBudget must be defined here
    ReleasedBudget = apps.get_model("Register", "ReleasedBudget")

    # 1) Determine which ministries belong to that YEAR:
    #    - ministries that had a budget in that year OR had a release in that year
    budget_ministry_ids = set(
        MinistryBudget.objects.filter(church_id=church_id, year=year)
        .values_list("ministry_id", flat=True)
    )

    release_ministry_ids = set(
        ReleasedBudget.objects.filter(church_id=church_id, date_released__year=year)
        .exclude(ministry_id__isnull=True)
        .values_list("ministry_id", flat=True)
    )

    ministry_ids = budget_ministry_ids | release_ministry_ids

    # If absolutely none, fallback to current active ministries
    if not ministry_ids:
        ministry_qs = Ministry.objects.filter(church_id=church_id, is_active=True).order_by("name")
    else:
        ministry_qs = Ministry.objects.filter(church_id=church_id, id__in=ministry_ids).order_by("name")

    # 2) Get stats for selected year (allocated/spent/suggested_score etc.)
    stats_map = analyze_ministry_performance(church_id, year=year)

    items = []
    for m in ministry_qs:
        s = stats_map.get(m.id, {})
        items.append({
            "ministry_id": m.id,
            "ministry_name": m.name,
            "spent": float(s.get("spent", 0.0)),
            "min_req": float(s.get("allocated", 0.0)),          # allocated already includes fallback
            "suggested_score": int(s.get("suggested_score", 5)),
            "default_priority": int(getattr(m, "priority_score", 5) or 5),
        })

    return JsonResponse({"year": year, "items": items})


def sp_unrestricted_income_for_year(church_id: int, year: int) -> Decimal:
    """
    Calls Finance_UnrestrictedIncomeByYear(church_id) and returns the
    GrandTotalUnrestricted for the requested year.
    """
    with connection.cursor() as cursor:
        cursor.callproc("Finance_UnrestrictedIncomeByYear", [church_id])
        rows = cursor.fetchall() or []

    # row shape:
    # (Year, TotalTithes, TotalOfferings, TotalUnrestrictedDonations,
    #  TotalUnrestrictedOtherIncome, GrandTotalUnrestricted)
    for r in rows:
        try:
            if int(r[0]) == int(year):
                return Decimal(str(r[5] or "0.00"))
        except Exception:
            continue

    return Decimal("0.00")


@csrf_exempt
def api_get_historical_stats(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    try:
        data = json.loads(request.body or "{}")

        if not hasattr(request.user, "church") or not request.user.church:
            return JsonResponse({"error": "No church assigned"}, status=403)

        church_id = request.user.church.id

        # ---- validate year ----
        try:
            year = int(data.get("year"))
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid year"}, status=400)

        timeframe = (data.get("timeframe") or "yearly").lower()  # 'yearly' or 'monthly'
        if timeframe not in ("yearly", "monthly"):
            timeframe = "yearly"

        # =========================
        # 1) Total Income via SP (UNRESTRICTED, NO expenses, BY YEAR)
        # =========================
        annual_income = sp_unrestricted_income_for_year(church_id, year)  # Decimal
        total_income_raw = float(annual_income)

        # ==========================================================
        # 2) Ministry Spending Reference (ReleasedBudget net)
        # ==========================================================
        ReleasedBudget = apps.get_model("Register", "ReleasedBudget")

        released_total = ReleasedBudget.objects.filter(
            church_id=church_id,
            date_released__year=year
        ).aggregate(
            t=Coalesce(Sum("amount"), Decimal("0.00"))
        )["t"]

        returned_total = ReleasedBudget.objects.filter(
            church_id=church_id,
            is_liquidated=True,
            liquidated_date__isnull=False,
            liquidated_date__year=year
        ).aggregate(
            t=Coalesce(Sum("amount_returned"), Decimal("0.00"))
        )["t"]

        net_spent = released_total - returned_total
        if net_spent < 0:
            net_spent = Decimal("0.00")

        expense_total_raw = float(net_spent)

        # =========================
        # 3) Funds to optimize
        # =========================
        funds_to_optimize = (total_income_raw / 12.0) if timeframe == "monthly" else total_income_raw

        return JsonResponse({
            "raw_income": total_income_raw,          # annual unrestricted income for that year
            "funds_to_optimize": funds_to_optimize,  # monthly avg if monthly
            "expenses_reference": expense_total_raw,
            "released_total": float(released_total),
            "returned_total": float(returned_total),
            "year": year,
            "timeframe": timeframe
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def _manual_budget_effective_map(church_id: int, year: int, month: int):
    """
    Manual source aligned to BudgetManageView budgets:
      - If monthly exists for (year, month) -> use it (monthly units)
      - Else if yearly pool exists for (year, month=0) -> use yearly/12 (monthly units)

    Returns:
      chosen_monthly: {ministry_id: Decimal(monthly_equiv)}
      budget_need_monthly: Decimal(sum monthly_equiv)
      budget_need_yearly: Decimal(sum monthly_equiv * 12)
    """
    MinistryBudget = apps.get_model("Register", "MinistryBudget")
    qs = (
        MinistryBudget.objects
        .filter(church_id=church_id, year=year, month__in=[month, 0], is_active=True)
        .select_related("ministry")
    )

    chosen = {}  # ministry_id -> monthly equiv
    for b in qs:
        amt = _to_decimal(getattr(b, "amount_allocated", None), "0.00")
        if b.month == month:
            chosen[b.ministry_id] = amt
        elif b.month == 0 and b.ministry_id not in chosen:
            chosen[b.ministry_id] = (amt / Decimal("12"))

    need_m = sum(chosen.values(), Decimal("0.00"))
    need_y = (need_m * Decimal("12"))
    return chosen, need_m, need_y


def latest_budget_monthly_map(church_id: int):
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    # get latest year per ministry
    latest_year = (
        MinistryBudget.objects
        .filter(church_id=church_id, is_active=True)
        .values("ministry_id")
        .annotate(y=Max("year"))
    )
    latest_year_map = {r["ministry_id"]: r["y"] for r in latest_year}

    # now for each ministry, find latest month in that year
    chosen = {}
    for mid, y in latest_year_map.items():
        qs = MinistryBudget.objects.filter(
            church_id=church_id, ministry_id=mid, year=y, is_active=True
        ).order_by("-month")  # month 12..1..0 (0 last, so we adjust below)

        # Prefer monthly (1..12) if any exist in latest year; else take yearly (0)
        monthly = qs.exclude(month=0).first()
        if monthly:
            chosen[mid] = Decimal(str(monthly.amount_allocated or "0.00"))
            continue

        yearly = qs.filter(month=0).first()
        if yearly:
            chosen[mid] = (Decimal(str(yearly.amount_allocated or "0.00")) / Decimal("12"))

    return chosen


def _to_decimal(v, default="0.00") -> Decimal:
    try:
        if v is None or v == "":
            return Decimal(default)
        return Decimal(str(v))
    except Exception:
        return Decimal(default)


def _get_restricted_balance_now(church_id: int) -> Decimal:
    try:
        with connection.cursor() as cursor:
            cursor.callproc("Finance_RestrictedBalanceNow", [church_id])
            row = cursor.fetchone()
        # row[3] = RestrictedBalanceOutstanding
        return _to_decimal(row[3] if row and len(row) > 3 else "0.00")
    except Exception:
        return Decimal("0.00")


def _get_current_fund_balances(church_id: int):
    total_cash = _get_total_cash(church_id)  # existing helper using Calculate_CashOnHand
    restricted_balance = _get_restricted_balance_now(church_id)

    unrestricted_cash = total_cash - restricted_balance
    if unrestricted_cash < 0:
        unrestricted_cash = Decimal("0.00")

    return {
        "total_cash": total_cash,
        "restricted_balance_outstanding": restricted_balance,
        "unrestricted_cash_available": unrestricted_cash,
    }


def get_projected_expenses_unrestricted(church_id):
    Expense = apps.get_model("Register", "Expense")

    today = date.today()
    start_date = today - timedelta(days=90)

    fixed_keywords = [
        "Utilities", "Electric Bill", "Water Bill", "Internet",
        "Salaries", "Wages", "Honorarium", "Pastor's Honorarium",
        "Rent", "Lease", "Fixed Costs"
    ]

    recent_bills = (
        Expense.objects.filter(
            church_id=church_id,
            expense_date__gte=start_date,
            category__name__in=fixed_keywords
        )
        .filter(Q(category__is_restricted=False) | Q(category__is_restricted__isnull=True))
        .exclude(status__in=["Declined", "Rejected"])
        .aggregate(t=Coalesce(Sum("amount"), Decimal("0.00")))["t"]
    )

    projected = _to_decimal(recent_bills) / Decimal("3")
    return projected.quantize(Decimal("0.01"))

from ai_analytics.forecasting import (
    calculate_financial_health,
    get_current_fund_balances,
    get_projected_expenses_unrestricted,
    analyze_ministry_performance,
)

@method_decorator(ensure_csrf_cookie, name="dispatch")
class PrescriptiveBudgetOptimizersView(LoginRequiredMixin, TemplateView):
    template_name = "prescriptive_budgets_optimizer.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # -----------------------------
        # 0) Security / Church check
        # -----------------------------
        if not hasattr(self.request.user, "church") or not self.request.user.church:
            today = date.today()
            context.update({
                "health": {
                    "status": "No Church",
                    "ratio": 0,
                    "color": "secondary",
                    "advice": "No church is assigned to your account."
                },
                "ministries": Ministry.objects.none(),

                # fund tiles
                "total_cash": 0.00,
                "restricted_funds_held": 0.00,
                "unrestricted_cash_available": 0.00,
                "projected_bills": 0.00,
                "safe_to_spend": 0.00,

                # controls
                "manual_year": today.year,
                "selected_year": today.year,
                "year_choices": [today.year],
                "default_hist_year": today.year,
                "timeframe": "yearly",
                "manual_month": today.month,
                "month_choices": [(i, calendar.month_name[i]) for i in range(1, 13)],

                # optimizer display
                "recommended_pool": 0.00,
                "recommended_pool_label": "0.00",
            })
            return context

        church_id = self.request.user.church.id
        today = date.today()

        # -----------------------------
        # 1) Manual year fixed = current year
        # -----------------------------
        manual_year = today.year
        selected_year = manual_year

        # -----------------------------
        # 2) User choice: yearly/monthly
        # -----------------------------
        timeframe = (self.request.GET.get("timeframe") or "yearly").lower()
        if timeframe not in ("yearly", "monthly"):
            timeframe = "yearly"

        manual_month = _safe_int(self.request.GET.get("month"), default=None) or today.month
        if manual_month < 1 or manual_month > 12:
            manual_month = today.month

        context["manual_year"] = manual_year
        context["selected_year"] = selected_year
        context["timeframe"] = timeframe
        context["manual_month"] = manual_month
        context["month_choices"] = [(i, calendar.month_name[i]) for i in range(1, 13)]

        # History dropdown years (last 6 years incl current)
        context["year_choices"] = [manual_year - i for i in range(0, 6)]
        context["default_hist_year"] = (
            manual_year - 1 if (manual_year - 1) in context["year_choices"] else manual_year
        )

        # -----------------------------
        # 3) Financial Health (SP-based, selected year)
        # -----------------------------
        context["health"] = calculate_financial_health(church_id, year=selected_year)

        # -----------------------------
        # 4) Current fund balances (restricted-aware)
        #    Option A:
        #    Total Cash on Hand
        #    Restricted Funds Held
        #    Unrestricted Cash Available
        #    Projected Unrestricted Bills
        #    Unrestricted Safe to Spend
        # -----------------------------
        funds = get_current_fund_balances(church_id)

        total_cash = _to_decimal(funds.get("total_cash"), "0.00")
        restricted_held = _to_decimal(funds.get("restricted_balance_outstanding"), "0.00")
        unrestricted_cash = _to_decimal(funds.get("unrestricted_cash_available"), "0.00")

        projected_bills = _to_decimal(get_projected_expenses_unrestricted(church_id), "0.00")

        safe_to_spend = unrestricted_cash - projected_bills
        if safe_to_spend < 0:
            safe_to_spend = Decimal("0.00")

        context["total_cash"] = float(total_cash)
        context["restricted_funds_held"] = float(restricted_held)
        context["unrestricted_cash_available"] = float(unrestricted_cash)
        context["projected_bills"] = float(projected_bills)
        context["safe_to_spend"] = float(safe_to_spend)

        # -----------------------------
        # 5) Manual budget snapshot (BudgetManageView-aligned)
        # -----------------------------
        budget_map_m, budget_need_m, budget_need_y = _manual_budget_effective_map(
            church_id=church_id,
            year=selected_year,
            month=manual_month
        )

        # IMPORTANT:
        # safe_to_spend is a CURRENT live cash amount, so do NOT annualize it.
        # Keep it as the real available cap regardless of UI timeframe.
        safe_ui = safe_to_spend

        # Budget need display follows timeframe
        need_ui = budget_need_m if timeframe == "monthly" else budget_need_y

        # Recommended pool = current unrestricted safe cash capped by displayed need
        recommended_pool_ui = min(safe_ui, need_ui) if need_ui > 0 else safe_ui
        if recommended_pool_ui < 0:
            recommended_pool_ui = Decimal("0.00")

        context["recommended_pool"] = float(recommended_pool_ui.quantize(Decimal("0.01")))
        context["recommended_pool_label"] = f"{context['recommended_pool']:.2f}"

        # -----------------------------
        # 6) Ministry stats for display
        # -----------------------------
        performance_stats = analyze_ministry_performance(church_id, year=selected_year) or {}

        ministries = (
            Ministry.objects
            .filter(church_id=church_id, is_active=True)
            .order_by("name")
        )

        # For display: if user selects yearly, show yearly equivalents; else monthly.
        display_multiplier = Decimal("1") if timeframe == "monthly" else Decimal("12")

        for m in ministries:
            stats = performance_stats.get(m.id, {
                "utilization": 0,
                "share_percentage": 0,
                "suggested_score": 5,
                "spent": 0.0,
                "allocated": 0.0
            })

            # base budget monthly-equivalent (BudgetManage aligned)
            base_budget_m = budget_map_m.get(m.id, Decimal("0.00"))

            # DB min monthly fallback
            db_min_m = _to_decimal(getattr(m, "min_monthly_budget", 0), "0.00")

            # Manual min requirement in MONTHLY units
            min_req_m = max(base_budget_m, db_min_m)

            # Display min requirement in the user-selected timeframe
            min_req_display = (min_req_m * display_multiplier).quantize(Decimal("0.01"))

            m.stats_utilization = stats.get("utilization", 0)
            m.stats_share = stats.get("share_percentage", 0)
            m.stats_suggested = stats.get("suggested_score", 5)
            m.stats_spent = stats.get("spent", 0.0)
            m.stats_min_req = float(min_req_display)
            m.stats_min_req_unit = timeframe

        context["ministries"] = ministries
        return context

from decimal import Decimal
from django.apps import apps
from django.db.models import Sum
from django.db.models.functions import Coalesce

def _all_years_budget_monthly_map(church_id: int):
    """
    Returns:
      budget_monthly_map: dict[ministry_id] = monthly budget baseline
      total_need_monthly: sum of monthly baselines

    Rules:
      - month = 0   => yearly pool, convert to monthly by /12
      - month = 1..12 => monthly rows, use average monthly amount for that year
      - across all years, average the per-year monthly baselines
    """
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    rows = (
        MinistryBudget.objects
        .filter(church_id=church_id, is_active=True, ministry_id__isnull=False)
        .values("ministry_id", "year", "month")
        .annotate(total_amt=Coalesce(Sum("amount_allocated"), Decimal("0.00")))
        .order_by("ministry_id", "year", "month")
    )

    per_ministry_year = {}

    for r in rows:
        mid = r["ministry_id"]
        year = r["year"]
        month = int(r["month"] or 0)
        amt = _to_decimal(r["total_amt"], "0.00")

        if amt <= 0:
            continue

        key = (mid, year)
        entry = per_ministry_year.setdefault(key, {
            "yearly_pool": None,
            "monthly_rows": []
        })

        if month == 0:
            entry["yearly_pool"] = amt
        else:
            entry["monthly_rows"].append(amt)

    per_ministry_values = {}

    for (mid, year), entry in per_ministry_year.items():
        if entry["yearly_pool"] is not None:
            monthly_val = (entry["yearly_pool"] / Decimal("12")).quantize(Decimal("0.01"))
        elif entry["monthly_rows"]:
            monthly_val = (
                sum(entry["monthly_rows"], Decimal("0.00")) /
                Decimal(str(len(entry["monthly_rows"])))
            ).quantize(Decimal("0.01"))
        else:
            continue

        per_ministry_values.setdefault(mid, []).append(monthly_val)

    budget_monthly_map = {}
    total_need_monthly = Decimal("0.00")

    for mid, vals in per_ministry_values.items():
        avg = (
            sum(vals, Decimal("0.00")) /
            Decimal(str(len(vals)))
        ).quantize(Decimal("0.01"))

        budget_monthly_map[mid] = avg
        total_need_monthly += avg

    return budget_monthly_map, total_need_monthly.quantize(Decimal("0.01"))


def _parse_release_ministry(desc: str) -> str:
    if not desc:
        return ""
    if not desc.lower().startswith("budget release:"):
        return ""
    rest = desc.split(":", 1)[1].strip()
    return rest.split("|", 1)[0].strip()


def _parse_return_ministry(desc: str) -> str:
    if not desc:
        return ""
    prefix = "budget return from"
    if not desc.lower().startswith(prefix):
        return ""
    rest = desc[len(prefix):].strip()
    return rest.split("|", 1)[0].strip()

def _get_total_cash(church_id: int) -> Decimal:
    try:
        with connection.cursor() as cursor:
            cursor.callproc("Calculate_CashOnHand", [church_id])
            row = cursor.fetchone()
        return _to_decimal(row[0] if row and row[0] is not None else "0.00")
    except Exception:
        return Decimal("0.00")

def _manual_budget_snapshot(church_id: int, year: int, month: int):
    """
    Manual data source = BudgetManageView logic.
    - Pull budgets for (year, month in [selected_month, 0]).
    - For each ministry: prefer monthly budget; else yearly pool / 12.
    Returns:
      budget_monthly_map: {ministry_id: Decimal(monthly_equiv)}
      budget_need_monthly: Decimal(sum of monthly equivs)
    """
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    qs = (
        MinistryBudget.objects
        .filter(church_id=church_id, year=year, month__in=[month, 0], is_active=True)
        .select_related("ministry")
    )

    chosen = {}  # ministry_id -> monthly_equiv (Decimal)
    for b in qs:
        amt = _to_decimal(getattr(b, "amount_allocated", None), "0.00")
        if b.month == month:
            chosen[b.ministry_id] = amt
        elif b.month == 0 and b.ministry_id not in chosen:
            chosen[b.ministry_id] = (amt / Decimal("12"))

    need = sum(chosen.values(), Decimal("0.00"))
    return chosen, need




def _to_decimal(v, default="0.00") -> Decimal:
    try:
        if v is None or v == "":
            return Decimal(default)
        return Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)

def _safe_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(v)
    except Exception:
        return default


def _latest_budget_monthly_map(church_id: int):
    MinistryBudget = apps.get_model("Register", "MinistryBudget")

    # Pull all budgets for church; only what we need
    rows = (
        MinistryBudget.objects
        .filter(church_id=church_id, is_active=True)
        .values("ministry_id", "year", "month", "amount_allocated")
        .order_by("ministry_id", "-year", "-month")  # latest year first; month 12..0
    )

    latest = {}  # ministry_id -> (year, month, amount_decimal)
    for r in rows:
        mid = r["ministry_id"]
        if mid in latest:
            continue  # keep first (latest by ordering)

        amt = r["amount_allocated"]
        try:
            amt = Decimal(str(amt or "0.00"))
        except Exception:
            amt = Decimal("0.00")

        latest[mid] = (int(r["year"] or 0), int(r["month"] or 0), amt)

    latest_monthly = {}
    for mid, (y, mo, amt) in latest.items():
        if mo == 0:
            # yearly pool -> monthly equivalent
            m_amt = (amt / Decimal("12")) if amt > 0 else Decimal("0.00")
        else:
            # monthly budget
            m_amt = amt
        latest_monthly[mid] = m_amt

    budget_need_monthly = sum(latest_monthly.values(), Decimal("0.00"))
    return latest_monthly, budget_need_monthly


def _all_years_spent_monthly_map(church_id):
    """
    Returns:
      spent_monthly_map: dict[ministry_id:int] -> Decimal(avg monthly net spent)
      total_months_observed: int (months with any release activity across all ministries)
    """
    ReleasedBudget = apps.get_model("Register", "ReleasedBudget")

    # 1) Aggregate released per ministry across all time
    released_rows = (
        ReleasedBudget.objects
        .filter(church_id=church_id, ministry_id__isnull=False)
        .values("ministry_id")
        .annotate(released=Coalesce(Sum("amount"), Decimal("0.00")))
    )
    released_by_id = {r["ministry_id"]: (r["released"] or Decimal("0.00")) for r in released_rows}

    # 2) Aggregate returned per ministry across all time (only liquidated with date)
    returned_rows = (
        ReleasedBudget.objects
        .filter(
            church_id=church_id,
            ministry_id__isnull=False,
            is_liquidated=True,
            liquidated_date__isnull=False,
        )
        .values("ministry_id")
        .annotate(returned=Coalesce(Sum("amount_returned"), Decimal("0.00")))
    )
    returned_by_id = {r["ministry_id"]: (r["returned"] or Decimal("0.00")) for r in returned_rows}

    # 3) Estimate months observed (for normalization)
    months_qs = (
        ReleasedBudget.objects
        .filter(church_id=church_id, ministry_id__isnull=False)
        .annotate(m=TruncMonth("date_released"))
        .values("m")
        .distinct()
    )
    total_months_observed = months_qs.count() or 0

    # Safety: avoid divide-by-zero; if no months, monthly is 0
    denom = Decimal(str(total_months_observed)) if total_months_observed > 0 else Decimal("0")

    spent_monthly = {}
    for mid, released in released_by_id.items():
        returned = returned_by_id.get(mid, Decimal("0.00"))
        net = released - returned
        if net < 0:
            net = Decimal("0.00")

        if denom > 0:
            spent_monthly[mid] = (net / denom)
        else:
            spent_monthly[mid] = Decimal("0.00")

    return spent_monthly, total_months_observed



def _net_spent_ytd_map(church_id, year, ministry_ids=None, as_of_date=None):
    ReleasedBudget = apps.get_model("Register", "ReleasedBudget")

    rel_qs = ReleasedBudget.objects.filter(
        church_id=church_id,
        ministry_id__isnull=False,
        date_released__year=year,
    )

    ret_qs = ReleasedBudget.objects.filter(
        church_id=church_id,
        ministry_id__isnull=False,
        is_liquidated=True,
        liquidated_date__isnull=False,
        liquidated_date__year=year,
    )

    if ministry_ids:
        rel_qs = rel_qs.filter(ministry_id__in=ministry_ids)
        ret_qs = ret_qs.filter(ministry_id__in=ministry_ids)

    if as_of_date is not None:
        rel_qs = rel_qs.filter(date_released__lte=as_of_date)
        ret_qs = ret_qs.filter(liquidated_date__lte=as_of_date)

    released_by_id = {
        r["ministry_id"]: _to_decimal(r["released"], "0.00")
        for r in rel_qs.values("ministry_id").annotate(
            released=Coalesce(Sum("amount"), Decimal("0.00"))
        )
    }

    returned_by_id = {
        r["ministry_id"]: _to_decimal(r["returned"], "0.00")
        for r in ret_qs.values("ministry_id").annotate(
            returned=Coalesce(Sum("amount_returned"), Decimal("0.00"))
        )
    }

    out = {}
    all_ids = set(released_by_id.keys()) | set(returned_by_id.keys())
    for mid in all_ids:
        net = released_by_id.get(mid, Decimal("0.00")) - returned_by_id.get(mid, Decimal("0.00"))
        if net < 0:
            net = Decimal("0.00")
        out[mid] = net
    return out


@csrf_exempt
@require_POST
@login_required
def api_run_optimizer(request):
    try:
        # 0) Parse JSON safely
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        # 1) Church security
        if not hasattr(request.user, "church") or not request.user.church:
            return JsonResponse(
                {"error": "Security Violation: No church assigned to this user."},
                status=403
            )

        church_id = request.user.church.id

        ReleasedBudget = apps.get_model("Register", "ReleasedBudget")
        Ministry = apps.get_model("Register", "Ministry")
        MinistryBudget = apps.get_model("Register", "MinistryBudget")

        # --- PARAMETERS ---
        raw_use_history = data.get("use_history", False)
        if isinstance(raw_use_history, bool):
            use_history = raw_use_history
        else:
            use_history = str(raw_use_history).strip().lower() in ("1", "true", "yes", "on")

        timeframe = (data.get("timeframe") or "yearly").lower()
        if timeframe not in ("yearly", "monthly"):
            timeframe = "yearly"

        historical_year = _safe_int(data.get("historical_year"), default=None)
        if use_history and not historical_year:
            return JsonResponse(
                {"error": "historical_year is required when use_history is true."},
                status=400
            )

        try:
            shortage_priority_boost = float(data.get("shortage_priority_boost", 1.0))
        except Exception:
            shortage_priority_boost = 1.0
        if shortage_priority_boost < 0:
            shortage_priority_boost = 0.0

        budget_map = {}
        spent_map = {}
        months_observed = 0

        # =========================================================
        # 2) POOL SOURCE (SOLVER WORKS IN MONTHLY UNITS)
        # =========================================================
        manual_funds_meta = {}

        if use_history:
            income_basis = "historical_unrestricted_sp"

            annual_income = sp_unrestricted_income_for_year(
                church_id, historical_year
            ) or Decimal("0.00")

            if annual_income <= 0:
                return JsonResponse(
                    {"error": f"No unrestricted income found for year {historical_year}."},
                    status=400
                )

            monthly_pool = annual_income / Decimal("12")
            pool_used_ui = monthly_pool if timeframe == "monthly" else (monthly_pool * Decimal("12"))

        else:
            income_basis = "current_unrestricted_cash_minus_projected_unrestricted_bills"

            budget_map, budget_need_monthly = _all_years_budget_monthly_map(church_id)
            spent_map, months_observed = _all_years_spent_monthly_map(church_id)

            funds = get_current_fund_balances(church_id)

            total_cash = _to_decimal(funds.get("total_cash"), "0.00")
            restricted_held = _to_decimal(funds.get("restricted_balance_outstanding"), "0.00")
            unrestricted_cash = _to_decimal(funds.get("unrestricted_cash_available"), "0.00")

            projected_bills = _to_decimal(get_projected_expenses_unrestricted(church_id), "0.00")

            safe_to_spend = unrestricted_cash - projected_bills
            if safe_to_spend < 0:
                safe_to_spend = Decimal("0.00")

            safe_monthly = safe_to_spend if timeframe == "monthly" else (safe_to_spend / Decimal("12"))

            spend_need_monthly = sum(
                (_to_decimal(v, "0.00") for v in spent_map.values()),
                Decimal("0.00")
            )

            if months_observed and months_observed >= 3:
                baseline_need_monthly = (
                    (budget_need_monthly * Decimal("0.60")) +
                    (spend_need_monthly * Decimal("0.40"))
                )
            else:
                baseline_need_monthly = budget_need_monthly

            if baseline_need_monthly > 0:
                monthly_pool = min(safe_monthly, baseline_need_monthly)
            else:
                monthly_pool = safe_monthly

            pool_used_ui = monthly_pool if timeframe == "monthly" else (monthly_pool * Decimal("12"))

            manual_funds_meta = {
                "total_cash": float(total_cash.quantize(Decimal("0.01"))),
                "restricted_funds_held": float(restricted_held.quantize(Decimal("0.01"))),
                "unrestricted_cash_available": float(unrestricted_cash.quantize(Decimal("0.01"))),
                "projected_unrestricted_bills": float(projected_bills.quantize(Decimal("0.01"))),
                "safe_to_spend": float(safe_to_spend.quantize(Decimal("0.01"))),
                "budget_need_monthly": float(budget_need_monthly.quantize(Decimal("0.01"))),
                "spend_need_monthly": float(spend_need_monthly.quantize(Decimal("0.01"))),
                "baseline_need_monthly": float(baseline_need_monthly.quantize(Decimal("0.01"))),
            }

        if monthly_pool <= 0:
            return JsonResponse({"error": "Funds to optimize must be greater than 0."}, status=400)

        # =========================================================
        # 3) MINISTRIES
        # =========================================================
        if use_history:
            budget_ministry_ids = set(
                MinistryBudget.objects.filter(
                    church_id=church_id,
                    year=historical_year
                ).values_list("ministry_id", flat=True)
            )

            release_ministry_ids = set(
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    date_released__year=historical_year
                )
                .exclude(ministry_id__isnull=True)
                .values_list("ministry_id", flat=True)
            )

            ministry_ids = budget_ministry_ids | release_ministry_ids

            if not ministry_ids:
                return JsonResponse(
                    {
                        "error": (
                            f"No ministries found for year {historical_year}. "
                            f"Add a budget or a release for that year first."
                        )
                    },
                    status=400
                )

            ministries = list(
                Ministry.objects.filter(
                    church_id=church_id,
                    is_active=True,
                    id__in=ministry_ids
                ).order_by("name")
            )
        else:
            ministries = list(
                Ministry.objects.filter(
                    church_id=church_id,
                    is_active=True
                ).order_by("name")
            )

        if not ministries:
            return JsonResponse({"error": "No eligible ministries found."}, status=400)

        # =========================================================
        # 3A) PRIORITIES from UI (manual overrides)
        # =========================================================
        raw_priorities = data.get("priorities") or {}
        user_priorities = (
            {str(k): v for k, v in raw_priorities.items()}
            if isinstance(raw_priorities, dict) else {}
        )

        # =========================================================
        # 3B) HISTORY NET SPENT (only if history)
        # =========================================================
        historical_net_spent_annual = {}
        if use_history:
            rel_rows = (
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    ministry_id__isnull=False,
                    date_released__year=historical_year,
                    ministry_id__in=[m.id for m in ministries],
                )
                .values("ministry_id")
                .annotate(released=Coalesce(Sum("amount"), Decimal("0.00")))
            )
            released_by_id = {
                r["ministry_id"]: _to_decimal(r["released"], "0.00")
                for r in rel_rows
            }

            ret_rows = (
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    ministry_id__isnull=False,
                    is_liquidated=True,
                    liquidated_date__isnull=False,
                    liquidated_date__year=historical_year,
                    ministry_id__in=[m.id for m in ministries],
                )
                .values("ministry_id")
                .annotate(returned=Coalesce(Sum("amount_returned"), Decimal("0.00")))
            )
            returned_by_id = {
                r["ministry_id"]: _to_decimal(r["returned"], "0.00")
                for r in ret_rows
            }

            for m in ministries:
                net = released_by_id.get(m.id, Decimal("0.00")) - returned_by_id.get(m.id, Decimal("0.00"))
                if net < 0:
                    net = Decimal("0.00")
                historical_net_spent_annual[m.id] = net

        # Current YTD map for manual mode
        current_ytd_spent_map = {}
        if not use_history:
            today = timezone.now().date()
            current_ytd_spent_map = _net_spent_ytd_map(
                church_id=church_id,
                year=today.year,
                ministry_ids=[m.id for m in ministries],
                as_of_date=today,
            )

        # =========================================================
        # 3D) BUILD ministry_data (MONTHLY UNITS)
        # =========================================================
        ministry_data = []
        manual_override_count = 0
        ai_predicted_count = 0
        db_fallback_count = 0

        for m in ministries:
            db_min_monthly = _to_decimal(getattr(m, "min_monthly_budget", 0), "0.00")
            db_max_monthly = _to_decimal(getattr(m, "max_monthly_cap", 0), "0.00")  # 0 => no cap

            if use_history:
                annual_spent = historical_net_spent_annual.get(m.id, Decimal("0.00"))
                hist_avg_monthly = (annual_spent / Decimal("12")) if annual_spent > 0 else Decimal("0.00")

                min_req_m = max(db_min_monthly, hist_avg_monthly)

                hist_cap_m = (hist_avg_monthly * Decimal("1.20")) if hist_avg_monthly > 0 else Decimal("0.00")
                max_cap_m = max(db_max_monthly, hist_cap_m)

                spent_ytd_val = float(annual_spent)
            else:
                base_budget_m = _to_decimal(budget_map.get(m.id, 0), "0.00")
                spent_avg_m = _to_decimal(spent_map.get(m.id, 0), "0.00")

                min_req_m = max(base_budget_m, spent_avg_m, db_min_monthly)

                if db_max_monthly > 0:
                    max_cap_m = db_max_monthly
                else:
                    cap_from_budget = (base_budget_m * Decimal("1.25")) if base_budget_m > 0 else Decimal("0.00")
                    cap_from_spend = (spent_avg_m * Decimal("1.20")) if spent_avg_m > 0 else Decimal("0.00")
                    max_cap_m = max(cap_from_budget, cap_from_spend, Decimal("0.00"))

                spent_ytd_val = float(_to_decimal(current_ytd_spent_map.get(m.id, 0), "0.00"))

            if max_cap_m > 0 and max_cap_m < min_req_m:
                max_cap_m = min_req_m

            mid_str = str(m.id)
            manual_override = user_priorities.get(mid_str)

            if manual_override is not None:
                score = _safe_int(manual_override, default=5) or 5
                priority_source = "manual_override"
                manual_override_count += 1
            else:
                feats_6 = [
                    float(monthly_pool),
                    1.0 if use_history else 0.0,
                    1.0 if timeframe == "monthly" else 0.0,
                    float(min_req_m),
                    float(max_cap_m),
                    float(spent_ytd_val),
                ]

                ai_score = predict_priority(feats_6, church_id=church_id)

                if ai_score is not None:
                    score = int(ai_score)
                    priority_source = "ai_model"
                    ai_predicted_count += 1
                else:
                    score = int(getattr(m, "priority_score", 5) or 5)
                    priority_source = "db_default"
                    db_fallback_count += 1

            ministry_data.append({
                "id": m.id,
                "name": m.name,
                "min_req": float(min_req_m),
                "max_cap": float(max_cap_m),
                "spent_ytd": float(spent_ytd_val),
                "priority_score": int(score),
                "priority_source": priority_source,
            })

        # Dedup safety
        ministry_data = list({str(md["id"]): md for md in ministry_data}.values())

        # =========================================================
        # 4) OPTIMIZE (MONTHLY solver)
        # =========================================================
        opt_result = optimize_budget_distribution_ai(
            float(monthly_pool),
            ministry_data,
            reserve_ratio=0.00,
            w_priority=1.0,
            w_fairness=0.15,
            w_stability=0.00,
            shortage_priority_boost=shortage_priority_boost,
        )

        if isinstance(opt_result, dict) and "allocations" in opt_result:
            allocations_monthly = opt_result.get("allocations", {})
            optimizer_source = opt_result.get("optimizer_source", "unknown")
            optimizer_status = opt_result.get("optimizer_status", "unknown")
        else:
            allocations_monthly = opt_result or {}
            optimizer_source = "unknown_legacy_shape"
            optimizer_status = "unknown"

        # =========================================================
        # 5) FORMAT OUTPUT for UI timeframe
        # =========================================================
        multiplier = 1 if timeframe == "monthly" else 12

        allocations_out = []
        for md in ministry_data:
            mid = str(md["id"])
            amt_m = _to_decimal(allocations_monthly.get(mid, 0.0), "0.00")
            amt_display = (amt_m * Decimal(multiplier)).quantize(Decimal("0.01"))

            allocations_out.append({
                "ministry_id": mid,
                "ministry_name": md["name"],
                "allocated_amount": float(amt_display),
                "priority_score": int(md["priority_score"]),
                "priority_source": md.get("priority_source", "unknown"),
                "spent_ytd": float(_to_decimal(md.get("spent_ytd", 0.0), "0.00")),
                "min_req_monthly": float(_to_decimal(md["min_req"], "0.00")),
                "max_cap_monthly": float(_to_decimal(md["max_cap"], "0.00")),
            })

        result = {
            "allocations": allocations_out,
            "meta": {
                "mode": "Historical" if use_history else "Manual",
                "reference_year": historical_year if use_history else None,
                "ui_timeframe": timeframe,
                "optimizer_unit": "monthly",
                "income_basis": income_basis,
                "pool_used_ui": float(pool_used_ui.quantize(Decimal("0.01"))),
                "pool_used_monthly": float(monthly_pool.quantize(Decimal("0.01"))),
                "pool_equivalent_yearly": float((monthly_pool * Decimal("12")).quantize(Decimal("0.01"))),
                "optimizer_source": optimizer_source,
                "optimizer_status": optimizer_status,
                "shortage_priority_boost": shortage_priority_boost,
                "manual_override_count": manual_override_count,
                "ai_predicted_count": ai_predicted_count,
                "db_fallback_count": db_fallback_count,
                **manual_funds_meta,
            }
        }

        # =========================================================
        # 6) LOG RUN
        # =========================================================
        try:
            priorities_used = {str(mm["id"]): int(mm["priority_score"]) for mm in ministry_data}
            ministry_constraints = {
                str(mm["id"]): {
                    "min_req": float(mm["min_req"]),
                    "max_cap": float(mm["max_cap"]),
                    "spent_ytd": float(mm.get("spent_ytd", 0.0)),
                    "priority_score": int(mm["priority_score"]),
                    "priority_source": mm.get("priority_source", "unknown"),
                }
                for mm in ministry_data
            }
            allocations_result = {a["ministry_id"]: float(a["allocated_amount"]) for a in allocations_out}

            OptimizerRunLog.objects.create(
                church_id=church_id,
                mode="Historical" if use_history else "Manual",
                reference_year=historical_year if use_history else None,
                timeframe=timeframe,
                pool_used_monthly=monthly_pool.quantize(Decimal("0.01")),
                priorities_used=priorities_used,
                ministry_constraints=ministry_constraints,
                allocations_result=allocations_result,
                created_by=request.user,
            )
        except Exception as log_err:
            print("OptimizerRunLog failed:", log_err)

        return JsonResponse(result)

    except Exception as e:
        print("--- OPTIMIZER ERROR ---")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)

DEPOSIT_CATEGORY_NAME = "Bank Deposit"
WITHDRAW_CATEGORY_NAME = "Bank Withdraw"

class BankSettingsView(FinanceRoleRequiredMixin, View):
    template_name = "bank_settings.html"

    def get(self, request):
        # 1. Check if account already exists to pre-fill the form
        try:
            bank = BankAccount.objects.get(church=request.user.church)
            form = BankInfoForm(instance=bank)
        except BankAccount.DoesNotExist:
            bank = None
            form = BankInfoForm()

        return render(request, self.template_name, {'form': form, 'bank': bank})

    def post(self, request):
        try:
            bank = BankAccount.objects.get(church=request.user.church)
        except BankAccount.DoesNotExist:
            bank = None

        form = BankInfoForm(request.POST, request.FILES, instance=bank)

        if form.is_valid():
            # 2. Get data for AI Verification
            bank_name = form.cleaned_data['bank_name']
            acct_num = form.cleaned_data['account_number']
            acct_name = form.cleaned_data['account_name']
            image = form.cleaned_data['verification_image']

            # 3. AI CHECK: IDENTITY
            # We ask AI: "Does this document match these account details?"
            is_valid, reason = verify_receipt_with_openai(
                image_file=image,
                check_type="IDENTITY",
                expected_data={
                    'bank_name': bank_name,
                    'account_number': acct_num,
                    'account_name': acct_name
                }
            )

            if not is_valid:
                messages.error(request, f"Verification Failed: {reason}")
                return render(request, self.template_name, {'form': form, 'bank': bank})

            # 4. Save Logic
            obj = form.save(commit=False)
            if not bank:
                # If creating new, link church and set initial balance to 0 (user sets it in Dashboard)
                obj.church = request.user.church
                obj.current_balance = 0.00

            obj.updated_by = request.user
            obj.save()

            if not bank:
                messages.success(request, "Bank Identity Verified! Now please set your current balance.")
                return redirect('manage_bank')  # Redirect new users to Dashboard
            else:
                messages.success(request, "Bank Information Updated successfully.")
                return redirect('financial_overview')

        return render(request, self.template_name, {'form': form, 'bank': bank})


# --- VIEW 2: BANK DASHBOARD (Balance & History) ---
# Purpose: Update the Balance amount and view logs.
# Register/views.py (Partial Update)

# --- VIEW 2: BANK DASHBOARD (Balance & History) ---
def _call_recalc_bank_balance(church_id):
    """
    Calls your bank balance stored procedure.
    Assumes you already created:
        sp_recalculate_bank_balance(church_id)
    """
    with connection.cursor() as cursor:
        cursor.callproc("sp_recalculate_bank_balance", [church_id])

        # Important for some MySQL setups after stored procedures
        try:
            while cursor.nextset():
                pass
        except Exception:
            pass


def _save_uploaded_file_to_field(instance, field_name, uploaded_file):
    """
    Safely saves the same uploaded file into a FileField/ImageField.
    """
    if not uploaded_file or not hasattr(instance, field_name):
        return

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    filename = os.path.basename(getattr(uploaded_file, "name", "")) or "proof"
    getattr(instance, field_name).save(filename, uploaded_file, save=False)


class ManageBankView(FinanceRoleRequiredMixin, View):
    template_name = "manage_bank.html"

    def get_bank(self, request):
        return BankAccount.objects.filter(church=request.user.church).first()

    def get_deposit_history(self, request):
        qs = Expense.objects.filter(church=request.user.church)

        deposit_qs = qs.filter(category__name=DEPOSIT_CATEGORY_NAME)

        if not deposit_qs.exists():
            deposit_qs = qs.filter(description__icontains=DEPOSIT_CATEGORY_NAME)

        return deposit_qs.order_by("-expense_date")

    def get_context(self, request, bank, form=None):
        deposit_history = self.get_deposit_history(request)

        if form is None:
            form = BankBalanceForm(
                instance=bank,
                initial={
                    "current_balance": bank.current_balance,
                },
            )

        return {
            "form": form,
            "bank": bank,
            "deposit_history": deposit_history,
        }

    def get(self, request):
        bank = self.get_bank(request)
        if not bank:
            messages.warning(request, "Please set up your Bank Information first.")
            return redirect("bank_settings")

        return render(request, self.template_name, self.get_context(request, bank))

    def post(self, request):
        bank = self.get_bank(request)
        if not bank:
            messages.warning(request, "Please set up your Bank Information first.")
            return redirect("bank_settings")

        # You may still use the form for validation,
        # but DO NOT use form.save() for current_balance anymore.
        form = BankBalanceForm(request.POST, request.FILES, instance=bank)

        if not form.is_valid():
            return render(request, self.template_name, self.get_context(request, bank, form=form))

        verified_balance = form.cleaned_data["current_balance"]
        image = form.cleaned_data.get("verification_image")

        if verified_balance is None or verified_balance < Decimal("0.00"):
            messages.error(request, "Verified bank balance cannot be negative.")
            return render(request, self.template_name, self.get_context(request, bank, form=form))

        if not image:
            messages.error(request, "Please upload a verification image for the balance update.")
            return render(request, self.template_name, self.get_context(request, bank, form=form))

        # 1) AI verification first
        is_valid, reason = verify_receipt_with_openai(
            image_file=image,
            check_type="TRANSACTION",
            expected_data={
                "amount": verified_balance,
                "date": "Recent Balance",
                "context": {
                    "transaction_type": "bank_balance_reconciliation",
                    "bank_name": getattr(bank, "bank_name", ""),
                    "account_name": getattr(bank, "account_name", ""),
                    "account_number_last4": (bank.account_number[-4:] if bank.account_number else ""),
                },
            },
        )

        if not is_valid:
            messages.error(request, f"Balance Update Rejected: {reason}")
            return render(request, self.template_name, self.get_context(request, bank, form=form))

        try:
            with transaction.atomic():
                # 2) Lock bank row and refresh computed balance first
                bank = BankAccount.objects.select_for_update().get(pk=bank.pk)

                _call_recalc_bank_balance(bank.church_id)
                bank.refresh_from_db(fields=["current_balance"])

                system_balance = bank.current_balance or Decimal("0.00")
                difference = verified_balance - system_balance

                # 3) If there is a difference, post ONE adjustment movement
                if difference != Decimal("0.00"):
                    movement_kwargs = dict(
                        church=bank.church,
                        date=timezone.localdate(),
                        amount=abs(difference),
                        direction=(
                            CashBankMovement.Direction.DIRECT_BANK_RECEIPT
                            if difference > 0
                            else CashBankMovement.Direction.BANK_TO_CASH
                        ),
                        # Make sure ADJUSTMENT exists in your SourceType choices
                        source_type=getattr(CashBankMovement.SourceType, "ADJUSTMENT", "ADJUSTMENT"),
                        source_id=None,
                        memo=(
                            f"Manual bank reconciliation adjustment. "
                            f"System balance: {system_balance}. "
                            f"Verified balance: {verified_balance}."
                        ),
                        reference_no="BANK-RECON",
                        created_by=request.user,
                        status=CashBankMovement.Status.CONFIRMED,
                    )

                    movement = CashBankMovement(**movement_kwargs)

                    movement_field_names = {f.name for f in CashBankMovement._meta.fields}
                    if "proof_file" in movement_field_names:
                        _save_uploaded_file_to_field(movement, "proof_file", image)

                    movement.full_clean()
                    movement.save()

                    # 4) Recalculate again after the adjustment movement
                    _call_recalc_bank_balance(bank.church_id)
                    bank.refresh_from_db(fields=["current_balance"])

                # 5) Save only proof/audit metadata to BankAccount
                bank_field_names = {f.name for f in BankAccount._meta.fields}

                if "verification_image" in bank_field_names:
                    _save_uploaded_file_to_field(bank, "verification_image", image)

                if "updated_by" in bank_field_names:
                    bank.updated_by = request.user

                if "last_updated" in bank_field_names:
                    bank.last_updated = timezone.now()

                update_fields = []
                if "verification_image" in bank_field_names:
                    update_fields.append("verification_image")
                if "updated_by" in bank_field_names:
                    update_fields.append("updated_by")
                if "last_updated" in bank_field_names:
                    update_fields.append("last_updated")

                if update_fields:
                    bank.save(update_fields=update_fields)

            if difference == Decimal("0.00"):
                messages.success(request, "Bank balance proof verified. No adjustment was needed.")
            else:
                messages.success(
                    request,
                    f"Bank balance verified and reconciled successfully. "
                    f"Adjustment applied: ₱{abs(difference):,.2f}."
                )

            return redirect("financial_overview")

        except Exception as e:
            messages.error(request, f"Balance reconciliation failed: {e}")
            return render(request, self.template_name, self.get_context(request, bank, form=form))


def _call_recalc_bank_balance(church_id):
    """
    Rebuild BankAccount.current_balance from CashBankMovement ledger.
    Assumes you already created:
        sp_recalculate_bank_balance(church_id)
    """
    with connection.cursor() as cursor:
        cursor.callproc("sp_recalculate_bank_balance", [church_id])

        # Some MySQL setups need this after stored procedures
        try:
            while cursor.nextset():
                pass
        except Exception:
            pass


def _save_uploaded_file_to_field(instance, field_name, uploaded_file):
    """
    Safely saves the uploaded file into a FileField/ImageField.
    """
    if not uploaded_file or not hasattr(instance, field_name):
        return

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    filename = os.path.basename(getattr(uploaded_file, "name", "")) or "proof"
    getattr(instance, field_name).save(filename, uploaded_file, save=False)


def get_or_create_expense_category(church, name, description=None):
    CategoryModel = Expense._meta.get_field("category").related_model
    defaults = {}
    if description is not None:
        defaults["description"] = description

    obj, _ = CategoryModel.objects.get_or_create(
        church=church,
        name=name,
        defaults=defaults
    )
    return obj


def _attach_proof(expense_kwargs, image):
    field_names = {f.name for f in Expense._meta.fields}
    if "receipt_image" in field_names:
        expense_kwargs["receipt_image"] = image
    elif "file" in field_names:
        expense_kwargs["file"] = image
    else:
        raise ValueError("Expense model has no receipt_image or file field to store proof.")
    return expense_kwargs


def get_cash_on_hand(church_id):
    with connection.cursor() as cursor:
        cursor.callproc("Calculate_CashOnHand", [church_id])
        row = cursor.fetchone()

        # drain remaining result sets to avoid "commands out of sync"
        while cursor.nextset():
            pass

    return Decimal(row[0] or 0)


class BankDepositView(FinanceRoleRequiredMixin, View):
    template_name = "bank_deposit.html"

    def get(self, request):
        bank_acct = BankAccount.objects.filter(church=request.user.church).first()
        if not bank_acct:
            messages.warning(request, "Please configure Bank Settings before depositing.")
            return redirect("bank_settings")

        form = BankDepositForm()
        cash_on_hand = get_cash_on_hand(request.user.church.id)

        return render(
            request,
            self.template_name,
            {
                "form": form,
                "bank": bank_acct,
                "cash_on_hand": cash_on_hand,
            },
        )

    def post(self, request):
        bank_acct = BankAccount.objects.filter(church=request.user.church).first()
        if not bank_acct:
            messages.warning(request, "Please configure Bank Settings before depositing.")
            return redirect("bank_settings")

        form = BankDepositForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "cash_on_hand": get_cash_on_hand(request.user.church.id),
                },
            )

        amount = form.cleaned_data["amount"]
        date = form.cleaned_data["deposit_date"]
        image = form.cleaned_data.get("proof_image")
        notes = (form.cleaned_data.get("notes") or "").strip()

        if not image:
            messages.error(request, "Please upload proof image for deposit.")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "cash_on_hand": get_cash_on_hand(request.user.church.id),
                },
            )

        # CHECK PHYSICAL CASH FIRST
        cash_on_hand = get_cash_on_hand(request.user.church.id)
        if amount > cash_on_hand:
            messages.error(
                request,
                f"Insufficient physical cash. Cash on hand is only ₱{cash_on_hand:,.2f}. "
                f"You cannot deposit ₱{amount:,.2f}."
            )
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "cash_on_hand": cash_on_hand,
                },
            )

        # AI CHECK
        is_valid, reason = verify_receipt_with_openai(
            image_file=image,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": str(date),
                "context": {
                    "bank_name": getattr(bank_acct, "bank_name", ""),
                    "account_name": getattr(bank_acct, "account_name", ""),
                    "account_number_last4": (
                        bank_acct.account_number[-4:] if bank_acct.account_number else ""
                    ),
                    "transaction_type": "deposit",
                },
            },
        )

        if not is_valid:
            messages.error(request, f"Potential Fraud Detected: {reason}")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "cash_on_hand": cash_on_hand,
                },
            )

        try:
            with transaction.atomic():
                bank_acct = BankAccount.objects.select_for_update().get(pk=bank_acct.pk)

                # RE-CHECK INSIDE TRANSACTION
                cash_on_hand = get_cash_on_hand(request.user.church.id)
                if amount > cash_on_hand:
                    messages.error(
                        request,
                        f"Insufficient physical cash. Cash on hand is only ₱{cash_on_hand:,.2f}. "
                        f"You cannot deposit ₱{amount:,.2f}."
                    )
                    return render(
                        request,
                        self.template_name,
                        {
                            "form": form,
                            "bank": bank_acct,
                            "cash_on_hand": cash_on_hand,
                        },
                    )

                deposit_category = get_or_create_expense_category(
                    request.user.church,
                    DEPOSIT_CATEGORY_NAME,
                    description="Deposits made into the bank account",
                )

                description = "Bank Deposit"
                if notes:
                    description = f"Bank Deposit - {notes}"

                expense_kwargs = dict(
                    church=request.user.church,
                    category=deposit_category,
                    amount=amount,
                    expense_date=date,
                    created_by=request.user,
                    description=description,
                    status="Approved",
                )
                expense_kwargs = _attach_proof(expense_kwargs, image)
                expense = Expense.objects.create(**expense_kwargs)

                movement = CashBankMovement(
                    church=request.user.church,
                    date=date,
                    amount=amount,
                    direction=CashBankMovement.Direction.CASH_TO_BANK,
                    source_type=CashBankMovement.SourceType.EXPENSE,
                    source_id=expense.id,
                    memo=description,
                    reference_no=f"DEP-EXP-{expense.id}",
                    created_by=request.user,
                    status=CashBankMovement.Status.CONFIRMED,
                )

                movement_field_names = {f.name for f in CashBankMovement._meta.fields}
                if "proof_file" in movement_field_names:
                    _save_uploaded_file_to_field(movement, "proof_file", image)

                movement.full_clean()
                movement.save()

                _call_recalc_bank_balance(request.user.church.id)
                bank_acct.refresh_from_db(fields=["current_balance"])

                updated_cash_on_hand = get_cash_on_hand(request.user.church.id)

            messages.success(
                request,
                f"Deposit verified and recorded successfully. "
                f"Updated bank balance: ₱{(bank_acct.current_balance or Decimal('0.00')):,.2f}. "
                f"Remaining physical cash: ₱{updated_cash_on_hand:,.2f}."
            )
            return redirect("financial_overview")

        except Exception as e:
            messages.error(request, f"Database Error: {str(e)}")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "cash_on_hand": get_cash_on_hand(request.user.church.id),
                },
            )


class BankWithdrawView(FinanceRoleRequiredMixin, View):
    template_name = "bank_withdraw.html"

    def get(self, request):
        bank_acct = BankAccount.objects.filter(church=request.user.church).first()
        if not bank_acct:
            messages.warning(request, "Please configure Bank Settings before withdrawing.")
            return redirect("bank_settings")

        _call_recalc_bank_balance(request.user.church.id)
        bank_acct.refresh_from_db(fields=["current_balance"])

        form = BankWithdrawForm()
        return render(
            request,
            self.template_name,
            {
                "form": form,
                "bank": bank_acct,
                "current_balance": bank_acct.current_balance or Decimal("0.00"),
            },
        )

    def post(self, request):
        bank_acct = BankAccount.objects.filter(church=request.user.church).first()
        if not bank_acct:
            messages.warning(request, "Please configure Bank Settings before withdrawing.")
            return redirect("bank_settings")

        _call_recalc_bank_balance(request.user.church.id)
        bank_acct.refresh_from_db(fields=["current_balance"])
        current_balance = bank_acct.current_balance or Decimal("0.00")

        form = BankWithdrawForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "current_balance": current_balance,
                },
            )

        amount = form.cleaned_data["amount"]
        date = form.cleaned_data["withdraw_date"]
        image = form.cleaned_data.get("proof_image")
        notes = (form.cleaned_data.get("notes") or "").strip()

        if not image:
            messages.error(request, "Please upload proof image for withdrawal.")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "current_balance": current_balance,
                },
            )

        # CHECK BANK BALANCE FIRST
        if amount > current_balance:
            messages.error(
                request,
                f"Insufficient bank balance. Available bank balance is only ₱{current_balance:,.2f}. "
                f"You cannot withdraw ₱{amount:,.2f}."
            )
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "current_balance": current_balance,
                },
            )

        # AI CHECK
        is_valid, reason = verify_receipt_with_openai(
            image_file=image,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": str(date),
                "context": {
                    "bank_name": getattr(bank_acct, "bank_name", ""),
                    "account_name": getattr(bank_acct, "account_name", ""),
                    "account_number_last4": (
                        bank_acct.account_number[-4:] if bank_acct.account_number else ""
                    ),
                    "transaction_type": "withdraw",
                }
            }
        )

        if not is_valid:
            messages.error(request, f"Potential Fraud Detected: {reason}")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "current_balance": current_balance,
                },
            )

        try:
            with transaction.atomic():
                bank_acct = BankAccount.objects.select_for_update().get(pk=bank_acct.pk)

                # RECHECK INSIDE TRANSACTION
                _call_recalc_bank_balance(request.user.church.id)
                bank_acct.refresh_from_db(fields=["current_balance"])

                current_balance = bank_acct.current_balance or Decimal("0.00")
                if amount > current_balance:
                    messages.error(
                        request,
                        f"Insufficient bank balance. Available bank balance is only ₱{current_balance:,.2f}. "
                        f"You cannot withdraw ₱{amount:,.2f}."
                    )
                    return render(
                        request,
                        self.template_name,
                        {
                            "form": form,
                            "bank": bank_acct,
                            "current_balance": current_balance,
                        },
                    )

                withdraw_category = get_or_create_expense_category(
                    request.user.church,
                    WITHDRAW_CATEGORY_NAME,
                    description="Withdrawals taken from the bank account"
                )

                description = "Bank Withdraw"
                if notes:
                    description = f"Bank Withdraw - {notes}"

                expense_kwargs = dict(
                    church=request.user.church,
                    category=withdraw_category,
                    amount=amount,
                    expense_date=date,
                    created_by=request.user,
                    description=description,
                    status="Approved",
                )
                expense_kwargs = _attach_proof(expense_kwargs, image)
                expense = Expense.objects.create(**expense_kwargs)

                movement = CashBankMovement(
                    church=request.user.church,
                    date=date,
                    amount=amount,
                    direction=CashBankMovement.Direction.BANK_TO_CASH,
                    source_type=CashBankMovement.SourceType.EXPENSE,
                    source_id=expense.id,
                    memo=description,
                    reference_no=f"WDR-EXP-{expense.id}",
                    created_by=request.user,
                    status=CashBankMovement.Status.CONFIRMED,
                )

                movement_field_names = {f.name for f in CashBankMovement._meta.fields}
                if "proof_file" in movement_field_names:
                    _save_uploaded_file_to_field(movement, "proof_file", image)

                movement.full_clean()
                movement.save()

                _call_recalc_bank_balance(request.user.church.id)
                bank_acct.refresh_from_db(fields=["current_balance"])

            messages.success(
                request,
                f"Withdrawal verified and recorded successfully. "
                f"Updated bank balance: ₱{(bank_acct.current_balance or Decimal('0.00')):,.2f}."
            )
            return redirect("financial_overview")

        except Exception as e:
            messages.error(request, f"Database Error: {str(e)}")
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "bank": bank_acct,
                    "current_balance": bank_acct.current_balance or Decimal("0.00"),
                },
            )

# ==========================================
#  FORGOT PASSWORD (DIRECT EMAIL METHOD)
# ==========================================
User = get_user_model()
class ForgotPasswordForm(forms.Form):
    email = forms.EmailField(
        label="Enter your registered email",
        required=True,
        widget=forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'name@example.com'})
    )


class ForgotPasswordView(View):
    template_name = 'forgot_password.html'

    def get(self, request):
        return render(request, self.template_name, {'form': ForgotPasswordForm()})

    def post(self, request):
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']

            try:
                user = User.objects.get(email=email)

                # Generate 6-digit code
                otp = str(random.randint(100000, 999999))

                # Store in Session (Temporary Memory)
                request.session['reset_email'] = email
                request.session['reset_otp'] = otp

                # Send Email
                send_mail(
                    'Password Reset Code',
                    f'Hello {user.username},\n\nYour verification code is: {otp}\n\nUse this code to set a new password.',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False
                )
                messages.success(request, "Code sent! Please check your email.")
                return redirect('reset_password_confirm')  # Go to next step

            except User.DoesNotExist:
                # Security: Don't reveal user doesn't exist
                messages.success(request, "Code sent! Please check your email.")
                return redirect('reset_password_confirm')

        return render(request, self.template_name, {'form': form})


# 2. VERIFY CODE & SET PASSWORD VIEW
class ResetPasswordConfirmView(View):
    template_name = 'reset_password_confirm.html'

    def get(self, request):
        # Security check: Must have requested a code first
        if 'reset_email' not in request.session:
            messages.error(request, "Session expired. Please start over.")
            return redirect('forgot_password')

        return render(request, self.template_name, {'form': ResetPasswordConfirmForm()})

    def post(self, request):
        form = ResetPasswordConfirmForm(request.POST)

        if form.is_valid():
            entered_code = form.cleaned_data['code']
            new_password = form.cleaned_data['new_password']

            # Get data from session
            session_code = request.session.get('reset_otp')
            email = request.session.get('reset_email')

            if entered_code == session_code and email:
                try:
                    user = User.objects.get(email=email)
                    user.set_password(new_password)
                    user.save()

                    # Clear session security data
                    del request.session['reset_otp']
                    del request.session['reset_email']

                    messages.success(request, "Password changed successfully! You can now login.")
                    return redirect('login')

                except User.DoesNotExist:
                    messages.error(request, "User not found.")
            else:
                messages.error(request, "Invalid Verification Code.")

        return render(request, self.template_name, {'form': form})


class AccountingLockView(IsChurchAdmin, View):
    template_name = "accounting_lock.html"

    # --- SECURITY: Church Admin Only ---
    def test_func(self):
        user = self.request.user
        # Check 1: User has a church
        if not hasattr(user, 'church') or not user.church:
            return False

        # Check 2: User is an Admin
        # This checks if they are marked as 'Admin' in your custom role field OR are a superuser/staff
        return user.is_staff or getattr(user, 'user_type', '') == 'ChurchAdmin'

    def handle_no_permission(self):
        messages.error(self.request, "Access Denied: Only Church Administrators can manage accounting periods.")
        # FIXED: Changed 'dashboard' to 'home' to match your URL patterns
        return redirect('home')

    def get(self, request):
        # Fetch existing settings or create blank ones for this church
        settings_obj, created = AccountingSettings.objects.get_or_create(church=request.user.church)

        form = AccountingLockForm(instance=settings_obj)

        return render(request, self.template_name, {
            "form": form,
            "current_lock": settings_obj.lock_date
        })

    def post(self, request):
        settings_obj, _ = AccountingSettings.objects.get_or_create(church=request.user.church)

        form = AccountingLockForm(request.POST, instance=settings_obj)

        if form.is_valid():
            form.save()

            date_val = form.cleaned_data.get('lock_date')
            if date_val:
                messages.success(request, f"System Locked. Transactions on or before {date_val} are now read-only.")
            else:
                messages.warning(request, "Lock removed. All past dates are now open for editing.")

            return redirect('accounting_lock')  # Refresh page

        return render(request, self.template_name, {"form": form})


class AnalyticsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "analytics_dashboard.html"

    # =========================================================
    # STORED PROCEDURE HELPER
    # =========================================================
    def sp_one(self, sp_name, params):
        with connection.cursor() as cursor:
            cursor.callproc(sp_name, params)
            row = cursor.fetchone()

            try:
                cursor.fetchall()
            except Exception:
                pass

            while cursor.nextset():
                try:
                    cursor.fetchall()
                except Exception:
                    pass

        return row

    # =========================================================
    # HELPERS
    # =========================================================
    def _safe_int(self, value, default):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def _sum_amount(self, model, church, date_field, start_date, end_date):
        return (
            model.objects.filter(
                church=church,
                **{f"{date_field}__range": (start_date, end_date)}
            ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        )

    def _sum_amount_datetime(self, model, church, datetime_field, start_date, end_date):
        return (
            model.objects.filter(
                church=church,
                **{f"{datetime_field}__date__range": (start_date, end_date)}
            ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        )

    def _count_users(self, UserModel, church, user_type=None, status=None):
        qs = UserModel.objects.filter(church=church)

        if user_type:
            qs = qs.filter(user_type=user_type)

        if status:
            qs = qs.filter(status=status)

        return qs.count()

    def _get_bank_balance(self, church):
        bank = BankAccount.objects.filter(church=church).first()
        if bank and bank.current_balance is not None:
            return Decimal(bank.current_balance)
        return Decimal("0.00")

    def _expense_queryset(self, church, start_date=None, end_date=None):
        """
        Exclude bank movement and budget release categories from expense analytics.
        """
        qs = Expense.objects.filter(church=church)

        if start_date and end_date:
            qs = qs.filter(expense_date__range=(start_date, end_date))

        qs = qs.exclude(
            Q(category__name__iexact="Bank Withdraw") |
            Q(category__name__iexact="Bank Deposit") |
            Q(category__name__iexact="Budget Release - Cash") |
            Q(category__name__iexact="Budget Release - Bank")
        )

        return qs

    def _get_available_years(self, church, current_year):
        years = set()

        sources = [
            (Expense, "expense_date"),
            (Tithe, "date"),
            (Offering, "date"),
            (Donations, "donations_date"),
            (OtherIncome, "date"),
            (BudgetReleaseRequest, "requested_at"),
            (ReleasedBudget, "date_released"),
            (ReleasedBudget, "liquidated_date"),
        ]

        for model, field_name in sources:
            result = (
                model.objects.filter(church=church)
                .annotate(year=ExtractYear(field_name))
                .values_list("year", flat=True)
                .distinct()
            )
            years.update(y for y in result if y is not None)

        years = sorted(years, reverse=True)
        if current_year not in years:
            years.insert(0, current_year)

        return years

    def _build_week_options(self, selected_year):
        max_week = date(selected_year, 12, 28).isocalendar().week
        options = []

        for week_num in range(1, max_week + 1):
            start_date = date.fromisocalendar(selected_year, week_num, 1)
            end_date = date.fromisocalendar(selected_year, week_num, 7)

            label = f"Week {week_num} — {start_date.strftime('%b %d')} to {end_date.strftime('%b %d, %Y')}"
            options.append({
                "value": week_num,
                "label": label,
                "start_date": start_date,
                "end_date": end_date,
            })

        return options

    def _get_period_range(self, selected_year, period, selected_month=None, selected_week=None):
        today = timezone.localdate()

        if period == "monthly":
            selected_month = self._safe_int(selected_month, today.month)
            if selected_month < 1 or selected_month > 12:
                selected_month = today.month

            start_date = date(selected_year, selected_month, 1)
            end_date = date(selected_year, selected_month, monthrange(selected_year, selected_month)[1])

        elif period == "weekly":
            current_week = today.isocalendar().week
            selected_week = self._safe_int(selected_week, current_week)

            max_week = date(selected_year, 12, 28).isocalendar().week
            if selected_week < 1 or selected_week > max_week:
                selected_week = current_week if selected_year == today.year and current_week <= max_week else 1

            try:
                start_date = date.fromisocalendar(selected_year, selected_week, 1)
                end_date = date.fromisocalendar(selected_year, selected_week, 7)
            except ValueError:
                selected_week = 1
                start_date = date.fromisocalendar(selected_year, selected_week, 1)
                end_date = date.fromisocalendar(selected_year, selected_week, 7)

        else:
            period = "yearly"
            selected_month = None
            selected_week = None
            start_date = date(selected_year, 1, 1)
            end_date = date(selected_year, 12, 31)

        return period, start_date, end_date, selected_month, selected_week

    def _build_chart_data(self, church, period, selected_year, selected_month, start_date, end_date):
        if period == "yearly":
            labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            income_data = [0.0] * 12
            expense_data = [0.0] * 12

            expense_qs = (
                self._expense_queryset(church, start_date, end_date)
                .annotate(month=TruncMonth("expense_date"))
                .values("month")
                .annotate(total=Sum("amount"))
                .order_by("month")
            )
            for entry in expense_qs:
                idx = entry["month"].month - 1
                expense_data[idx] = float(entry["total"] or 0)

            def add_income(model, date_field):
                qs = (
                    model.objects.filter(
                        church=church,
                        **{f"{date_field}__range": (start_date, end_date)}
                    )
                    .annotate(month=TruncMonth(date_field))
                    .values("month")
                    .annotate(total=Sum("amount"))
                    .order_by("month")
                )
                for entry in qs:
                    idx = entry["month"].month - 1
                    income_data[idx] += float(entry["total"] or 0)

        elif period == "monthly":
            days_in_month = monthrange(selected_year, selected_month)[1]
            labels = [str(i) for i in range(1, days_in_month + 1)]
            income_data = [0.0] * days_in_month
            expense_data = [0.0] * days_in_month

            expense_qs = (
                self._expense_queryset(church, start_date, end_date)
                .values("expense_date")
                .annotate(total=Sum("amount"))
                .order_by("expense_date")
            )
            for entry in expense_qs:
                idx = entry["expense_date"].day - 1
                expense_data[idx] = float(entry["total"] or 0)

            def add_income(model, date_field):
                qs = (
                    model.objects.filter(
                        church=church,
                        **{f"{date_field}__range": (start_date, end_date)}
                    )
                    .values(date_field)
                    .annotate(total=Sum("amount"))
                    .order_by(date_field)
                )
                for entry in qs:
                    current_date = entry[date_field]
                    idx = current_date.day - 1
                    income_data[idx] += float(entry["total"] or 0)

        else:  # weekly
            labels = [(start_date + timedelta(days=i)).strftime("%a") for i in range(7)]
            income_data = [0.0] * 7
            expense_data = [0.0] * 7

            expense_qs = (
                self._expense_queryset(church, start_date, end_date)
                .values("expense_date")
                .annotate(total=Sum("amount"))
                .order_by("expense_date")
            )
            for entry in expense_qs:
                idx = (entry["expense_date"] - start_date).days
                if 0 <= idx < 7:
                    expense_data[idx] = float(entry["total"] or 0)

            def add_income(model, date_field):
                qs = (
                    model.objects.filter(
                        church=church,
                        **{f"{date_field}__range": (start_date, end_date)}
                    )
                    .values(date_field)
                    .annotate(total=Sum("amount"))
                    .order_by(date_field)
                )
                for entry in qs:
                    current_date = entry[date_field]
                    idx = (current_date - start_date).days
                    if 0 <= idx < 7:
                        income_data[idx] += float(entry["total"] or 0)

        add_income(Tithe, "date")
        add_income(Offering, "date")
        add_income(Donations, "donations_date")
        add_income(OtherIncome, "date")

        return labels, income_data, expense_data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        church = self.request.user.church
        today = timezone.localdate()
        User = get_user_model()

        # =========================================================
        # 1. HANDLE FILTER SELECTION
        # =========================================================
        available_years = self._get_available_years(church, today.year)

        selected_year = self._safe_int(self.request.GET.get("year"), today.year)
        period = (self.request.GET.get("period") or "yearly").lower()
        selected_month = self.request.GET.get("month")
        selected_week = self.request.GET.get("week")

        period, start_date, end_date, selected_month, selected_week = self._get_period_range(
            selected_year=selected_year,
            period=period,
            selected_month=selected_month,
            selected_week=selected_week,
        )

        # =========================================================
        # 2. KPI TOTALS (FOR SELECTED RANGE)
        # =========================================================
        t_total = self._sum_amount(Tithe, church, "date", start_date, end_date)
        o_total = self._sum_amount(Offering, church, "date", start_date, end_date)
        d_total = self._sum_amount(Donations, church, "donations_date", start_date, end_date)
        oi_total = self._sum_amount(OtherIncome, church, "date", start_date, end_date)

        period_income = t_total + o_total + d_total + oi_total

        period_expense = (
            self._expense_queryset(church, start_date, end_date)
            .aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        )

        # current cash on hand, not filtered by range
        cash_on_hand = Decimal("0.00")
        result = self.sp_one("Calculate_CashOnHand", [church.id])
        if result and result[0] is not None:
            cash_on_hand = Decimal(result[0] or 0)

        # current bank balance, not filtered by range
        bank_balance = self._get_bank_balance(church)

        # =========================================================
        # 2B. BUDGET KPIs (FOR SELECTED RANGE)
        # =========================================================
        total_budget_request = self._sum_amount_datetime(
            BudgetReleaseRequest, church, "requested_at", start_date, end_date
        )

        total_budget_unliquidated = (
            ReleasedBudget.objects.filter(
                church=church,
                is_liquidated=False,
                date_released__range=(start_date, end_date),
            ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        )

        total_budget_liquidated = (
            ReleasedBudget.objects.filter(
                church=church,
                is_liquidated=True,
                liquidated_date__range=(start_date, end_date),
            ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
        )

        total_budget_request_count = BudgetReleaseRequest.objects.filter(
            church=church,
            requested_at__date__range=(start_date, end_date),
        ).count()

        total_budget_unliquidated_count = ReleasedBudget.objects.filter(
            church=church,
            is_liquidated=False,
            date_released__range=(start_date, end_date),
        ).count()

        total_budget_liquidated_count = ReleasedBudget.objects.filter(
            church=church,
            is_liquidated=True,
            liquidated_date__range=(start_date, end_date),
        ).count()

        # =========================================================
        # 3. USER KPIs (FILTERED BY CURRENT USER'S CHURCH)
        # =========================================================
        total_users = self._count_users(User, church)
        total_active_users = self._count_users(User, church, status=User.Status.ACTIVE)
        total_inactive_users = self._count_users(User, church, status=User.Status.INACTIVE)

        total_pastors = self._count_users(User, church, user_type=User.UserType.PASTOR)
        total_active_pastors = self._count_users(
            User, church,
            user_type=User.UserType.PASTOR,
            status=User.Status.ACTIVE
        )
        total_inactive_pastors = self._count_users(
            User, church,
            user_type=User.UserType.PASTOR,
            status=User.Status.INACTIVE
        )

        total_treasurers = self._count_users(User, church, user_type=User.UserType.TREASURER)
        total_active_treasurers = self._count_users(
            User, church,
            user_type=User.UserType.TREASURER,
            status=User.Status.ACTIVE
        )
        total_inactive_treasurers = self._count_users(
            User, church,
            user_type=User.UserType.TREASURER,
            status=User.Status.INACTIVE
        )

        total_members = self._count_users(User, church, user_type=User.UserType.MEMBER)
        total_active_members = self._count_users(
            User, church,
            user_type=User.UserType.MEMBER,
            status=User.Status.ACTIVE
        )
        total_inactive_members = self._count_users(
            User, church,
            user_type=User.UserType.MEMBER,
            status=User.Status.INACTIVE
        )

        # =========================================================
        # 4. CHART DATA
        # =========================================================
        chart_labels, income_data, expense_data = self._build_chart_data(
            church=church,
            period=period,
            selected_year=selected_year,
            selected_month=selected_month,
            start_date=start_date,
            end_date=end_date,
        )

        # =========================================================
        # 5. PIE CHARTS (FOR SELECTED RANGE)
        # =========================================================
        expense_breakdown = (
            self._expense_queryset(church, start_date, end_date)
            .values("category__name")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )
        pie_labels = [x["category__name"] or "Uncategorized" for x in expense_breakdown]
        pie_data = [float(x["total"] or 0) for x in expense_breakdown]

        fundtype_breakdown = (
            self._expense_queryset(church, start_date, end_date)
            .annotate(
                fund_type=Case(
                    When(
                        category__description__isnull=False,
                        category__description__gt="",
                        then=Value("Restricted")
                    ),
                    default=Value("Unrestricted"),
                    output_field=CharField()
                )
            )
            .values("fund_type")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )
        fund_pie_labels = [x["fund_type"] for x in fundtype_breakdown]
        fund_pie_data = [float(x["total"] or 0) for x in fundtype_breakdown]

        # =========================================================
        # 6. FILTER OPTIONS
        # =========================================================
        month_options = [
            (1, "January"), (2, "February"), (3, "March"), (4, "April"),
            (5, "May"), (6, "June"), (7, "July"), (8, "August"),
            (9, "September"), (10, "October"), (11, "November"), (12, "December"),
        ]

        week_options = self._build_week_options(selected_year)

        if period == "yearly":
            filter_label = f"Year {selected_year}"
        elif period == "monthly":
            filter_label = f"{month_options[selected_month - 1][1]} {selected_year}"
        else:
            filter_label = f"Week {selected_week} — {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"

        # =========================================================
        # 7. CONTEXT
        # =========================================================
        context.update({
            "period": period,
            "selected_year": selected_year,
            "selected_month": selected_month,
            "selected_week": selected_week,
            "available_years": available_years,
            "month_options": month_options,
            "week_options": week_options,
            "filter_label": filter_label,
            "range_start": start_date,
            "range_end": end_date,

            "kpi_cash": cash_on_hand,
            "kpi_bank_balance": bank_balance,
            "kpi_income": period_income,
            "kpi_expense": period_expense,
            "kpi_net": period_income - period_expense,

            "kpi_budget_total_request": total_budget_request,
            "kpi_budget_unliquidated": total_budget_unliquidated,
            "kpi_budget_liquidated": total_budget_liquidated,

            "kpi_budget_total_request_count": total_budget_request_count,
            "kpi_budget_unliquidated_count": total_budget_unliquidated_count,
            "kpi_budget_liquidated_count": total_budget_liquidated_count,

            "kpi_total_users": total_users,
            "kpi_total_active_users": total_active_users,
            "kpi_total_inactive_users": total_inactive_users,

            "kpi_total_pastors": total_pastors,
            "kpi_total_active_pastors": total_active_pastors,
            "kpi_total_inactive_pastors": total_inactive_pastors,

            "kpi_total_treasurers": total_treasurers,
            "kpi_total_active_treasurers": total_active_treasurers,
            "kpi_total_inactive_treasurers": total_inactive_treasurers,

            "kpi_total_members": total_members,
            "kpi_total_active_members": total_active_members,
            "kpi_total_inactive_members": total_inactive_members,

            "chart_labels": json.dumps(chart_labels),
            "chart_income": json.dumps(income_data),
            "chart_expense": json.dumps(expense_data),

            "pie_labels": json.dumps(pie_labels),
            "pie_data": json.dumps(pie_data),

            "fund_pie_labels": json.dumps(fund_pie_labels),
            "fund_pie_data": json.dumps(fund_pie_data),
        })
        return context

def _delete_movement_and_source(movement: CashBankMovement):
    """
    Deletes CashBankMovement and (if present) deletes the linked source record.
    Safe: always church-scoped deletes to prevent cross-church deletion.
    """
    church = movement.church
    sid = movement.source_id
    st = movement.source_type

    # Delete the linked record only if the movement created/linked one
    if sid:
        if st == CashBankMovement.SourceType.TITHE:
            Tithe.objects.filter(id=sid, church=church).delete()

        elif st == CashBankMovement.SourceType.OFFERING:
            Offering.objects.filter(id=sid, church=church).delete()

        elif st == CashBankMovement.SourceType.DONATION:
            Donations.objects.filter(id=sid, church=church).delete()

        elif st == CashBankMovement.SourceType.OTHER_INCOME:
            OtherIncome.objects.filter(id=sid, church=church).delete()

        elif st == CashBankMovement.SourceType.EXPENSE:
            Expense.objects.filter(id=sid, church=church).delete()

    # Finally delete movement record
    movement.delete()


# =========================================================
#  1) CREATE/ENTRY VIEW (single page)
#  - Creates draft records immediately (like your tithe/offering flow)
#  - Stores movement_id in session for review
#  - Auto-clean abandoned drafts when user returns to entry page
# =========================================================
SESSION_KEY = "staged_bank_movement"


# -----------------------------
# Helpers
# -----------------------------
def _require_church(user):
    if not hasattr(user, "church") or not user.church:
        raise PermissionError("Security Violation: No church assigned to this user.")
    return user.church


def _get_bank_account(church):
    """
    If BankAccount is OneToOne -> church.bank_account may raise RelatedObjectDoesNotExist.
    This handles both OneToOne and fallback query.
    """
    try:
        return church.bank_account
    except ObjectDoesNotExist:
        return BankAccount.objects.filter(church=church).first()


def _lock_date_guard(church, txn_date):
    lock_date = getattr(getattr(church, "accounting_settings", None), "lock_date", None)
    if lock_date and txn_date and txn_date <= lock_date:
        raise ValueError(f"Transactions on or before {lock_date} are locked.")


def _parse_session_date(v):
    """
    Accepts: date, datetime, "YYYY-MM-DD", "YYYY-MM-DDTHH:MM:SS", etc.
    Returns: date or None
    """
    if v is None:
        return None

    if isinstance(v, date_cls) and not isinstance(v, datetime_cls):
        return v

    if isinstance(v, datetime_cls):
        return v.date()

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Most drafts: "YYYY-MM-DD"
        try:
            return date_cls.fromisoformat(s[:10])
        except Exception:
            # Sometimes full ISO datetime
            try:
                return datetime_cls.fromisoformat(s).date()
            except Exception:
                return None

    return None


def _attach_temp_file_to_field(instance, field_name: str, temp_obj):
    """
    Copy TemporaryBankMovementFile.file to instance.<field_name> (FileField/ImageField).
    Safe if temp_obj missing or field missing.
    """
    if not temp_obj or not getattr(temp_obj, "file", None):
        return
    if not hasattr(instance, field_name):
        return

    f = temp_obj.file
    filename = os.path.basename(getattr(f, "name", "")) or "proof"

    f.open("rb")
    try:
        try:
            f.seek(0)
        except Exception:
            pass
        getattr(instance, field_name).save(filename, f, save=False)
    finally:
        try:
            f.close()
        except Exception:
            pass


def _attach_temp_file_to_any_known_field(instance, temp_obj):
    """
    If your models use different proof field names, this tries common ones.
    (No crash if field doesn't exist.)
    """
    for field_name in ("proof_file", "file", "proof_document", "proof_image", "receipt"):
        _attach_temp_file_to_field(instance, field_name, temp_obj)


def _ai_verify_from_temp(*, temp_obj, bank_acct, amount: Decimal, txn_date, direction, source_type):
    """
    Runs verify_receipt_with_openai() using the staged temp file.
    ✅ Enforces JPG/PNG because your AI verifier expects images.
    Returns (is_valid: bool, reason: str)
    """
    if not temp_obj or not getattr(temp_obj, "file", None):
        return False, "Missing proof file."

    ext = (os.path.splitext(getattr(temp_obj.file, "name", ""))[1] or "").lower()
    if ext not in (".jpg", ".jpeg", ".png"):
        return False, "AI verification requires a JPG/PNG image proof (PDF not supported)."

    # AI-friendly transaction type label
    txn_type = {
        CashBankMovement.Direction.CASH_TO_BANK: "deposit",
        CashBankMovement.Direction.BANK_TO_CASH: "withdraw",
        CashBankMovement.Direction.BANK_TO_BANK: "bank_transfer",
        CashBankMovement.Direction.DIRECT_BANK_RECEIPT: "direct_bank_receipt",
        CashBankMovement.Direction.BANK_PAID_EXPENSE: "bank_paid_expense",
    }.get(direction, str(direction))

    # If BANK_TO_BANK is being used as "income transfer", give more context
    if direction == CashBankMovement.Direction.BANK_TO_BANK and source_type in (
        CashBankMovement.SourceType.TITHE,
        CashBankMovement.SourceType.OFFERING,
        CashBankMovement.SourceType.DONATION,
        CashBankMovement.SourceType.OTHER_INCOME,
    ):
        txn_type = "bank_transfer_in"

    f = temp_obj.file
    f.open("rb")
    try:
        try:
            f.seek(0)
        except Exception:
            pass

        return verify_receipt_with_openai(
            image_file=f,
            check_type="TRANSACTION",
            expected_data={
                "amount": amount,
                "date": str(txn_date),
                "context": {
                    "transaction_type": txn_type,
                    "bank_name": getattr(bank_acct, "bank_name", ""),
                    "account_name": getattr(bank_acct, "account_name", ""),
                    "account_number_last4": (
                        bank_acct.account_number[-4:] if getattr(bank_acct, "account_number", "") else ""
                    ),
                }
            }
        )
    finally:
        try:
            f.close()
        except Exception:
            pass


def _build_error_summary(form):
    """
    Used to avoid template crash from form[name] indexing.
    Returns: list of dicts {label, errors_text}
    """
    out = []
    for field_name, errs in form.errors.items():
        if field_name == "__all__":
            label = "General"
        else:
            label = form.fields.get(field_name).label if field_name in form.fields else field_name
        out.append({"label": label, "errors": errs})
    return out


# -----------------------------
# Bank Movement: Create Draft
# -----------------------------
def _movement_txn_type(direction):
    D = CashBankMovement.Direction
    if direction in (D.CASH_TO_BANK, D.DIRECT_BANK_RECEIPT):
        return "deposit"
    if direction == D.BANK_TO_CASH:
        return "withdrawal"
    if direction == D.BANK_TO_BANK:
        return "transfer"
    if direction == D.BANK_PAID_EXPENSE:
        return "expense"
    return "bank_movement"


def _get_live_bank_balance(church):
    bank_acct = _get_bank_account(church)
    if not bank_acct:
        raise ValidationError("Please configure Bank Settings first.")

    _call_recalc_bank_balance(church.id)
    bank_acct.refresh_from_db(fields=["current_balance"])
    return bank_acct, Decimal(str(bank_acct.current_balance or "0.00"))


def _get_restricted_balance_for_expense_category(church, expense_category):
    """
    Returns: (fund_name, balance) or (None, None)

    Uses the same restricted-fund setup already used in your expense flow:
    - expense_category.is_restricted
    - expense_category.restricted_source
    - expense_category.restricted_category_id
    - sp_get_restricted_balances(church.id)
    """
    if not expense_category or not getattr(expense_category, "is_restricted", False):
        return None, None

    restricted_source = getattr(expense_category, "restricted_source", None)
    restricted_category_id = getattr(expense_category, "restricted_category_id", None)

    if not restricted_source or not restricted_category_id:
        return None, None

    fund_name = None

    if restricted_source == "DONATION":
        fund_obj = DonationCategory.objects.filter(
            church=church,
            id=restricted_category_id,
            is_restricted=True,
        ).first()
        if fund_obj:
            fund_name = fund_obj.name

    elif restricted_source == "OTHER_INCOME":
        fund_obj = OtherIncomeCategory.objects.filter(
            church=church,
            id=restricted_category_id,
            is_restricted=True,
        ).first()
        if fund_obj:
            fund_name = fund_obj.name

    if not fund_name:
        return None, None

    restricted_balances_raw = sp_get_restricted_balances(church.id) or {}
    restricted_balances = {
        str(name).strip().lower(): Decimal(str(balance or "0.00"))
        for name, balance in restricted_balances_raw.items()
    }

    return fund_name, restricted_balances.get(fund_name.strip().lower(), Decimal("0.00"))


def _movement_source_label(source_type, direction):
    if (
        direction == CashBankMovement.Direction.BANK_PAID_EXPENSE
        and source_type == CashBankMovement.SourceType.EXPENSE
    ):
        return "Bank-paid Expense"

    labels = {}

    for attr_name, label in (
        ("TITHE", "Tithe"),
        ("OFFERING", "Offering"),
        ("DONATION", "Donation"),
        ("OTHER_INCOME", "Other Income"),
        ("EXPENSE", "Expense"),
        ("TRANSFER_ONLY", "Transfer Only"),
        ("TRANSFER", "Transfer"),
    ):
        enum_value = getattr(CashBankMovement.SourceType, attr_name, None)
        if enum_value is not None:
            labels[enum_value] = label

    if source_type in labels:
        return labels[source_type]

    raw_value = getattr(source_type, "value", source_type)
    raw_value = str(raw_value or "").strip()
    return raw_value.replace("_", " ").title() if raw_value else "Bank movement"


def _build_bank_movement_success_message(direction, source_type, amount, updated_balance):
    source_label = _movement_source_label(source_type, direction)
    peso_amount = f"₱{Decimal(str(amount or '0.00')):,.2f}"
    peso_balance = f"₱{Decimal(str(updated_balance or '0.00')):,.2f}"

    income_like = source_type in (
        CashBankMovement.SourceType.TITHE,
        CashBankMovement.SourceType.OFFERING,
        CashBankMovement.SourceType.DONATION,
        CashBankMovement.SourceType.OTHER_INCOME,
    )

    if direction in (
        CashBankMovement.Direction.DIRECT_BANK_RECEIPT,
        CashBankMovement.Direction.BANK_TO_BANK,
    ) and income_like:
        return (
            f"{source_label} successfully saved to bank. "
            f"Amount saved to bank: {peso_amount}. "
            f"Updated bank balance: {peso_balance}."
        )

    if direction == CashBankMovement.Direction.CASH_TO_BANK:
        return (
            f"Cash deposit successfully recorded. "
            f"Amount deposited to bank: {peso_amount}. "
            f"Updated bank balance: {peso_balance}."
        )

    if direction == CashBankMovement.Direction.BANK_TO_CASH:
        return (
            f"Cash withdrawal successfully recorded. "
            f"Amount withdrawn from bank: {peso_amount}. "
            f"Updated bank balance: {peso_balance}."
        )

    if direction == CashBankMovement.Direction.BANK_PAID_EXPENSE:
        return (
            f"{source_label} successfully recorded. "
            f"Amount charged to bank: {peso_amount}. "
            f"Updated bank balance: {peso_balance}."
        )

    return (
        f"{source_label} successfully recorded. "
        f"Amount: {peso_amount}. "
        f"Updated bank balance: {peso_balance}."
    )


@method_decorator(never_cache, name="dispatch")
class BankMovementCreateView(FinanceRoleRequiredMixin, View):
    template_name = "bank_movement_form.html"

    def _member_display_name(self, user_obj):
        if not user_obj:
            return ""
        if hasattr(user_obj, "get_full_name"):
            full_name = (user_obj.get_full_name() or "").strip()
            if full_name:
                return full_name
        return getattr(user_obj, "username", str(user_obj))

    def _validate_donation_party(self, form, cd, church):
        donor_type = cd.get("donor_type") or Donations.DonorType.ANONYMOUS
        donor_value = (cd.get("donor") or "").strip()
        donation_member = cd.get("donation_member")

        allowed_types = {"Member", "Pastor", "Treasurer"}

        if donor_type == Donations.DonorType.MEMBER:
            if not donation_member:
                form.add_error("donation_member", "Please select a member, pastor, or treasurer.")
                return donor_type, donor_value, donation_member, ""

            if getattr(donation_member, "church_id", None) != church.id:
                form.add_error("donation_member", "Selected user must belong to your church.")

            if getattr(donation_member, "user_type", None) not in allowed_types:
                form.add_error("donation_member", "Only Member, Pastor, or Treasurer can be selected.")

            donor_value = self._member_display_name(donation_member)
            return donor_type, donor_value, donation_member, donor_value

        if donor_type == Donations.DonorType.NON_MEMBER:
            if not donor_value:
                form.add_error("donor", "Please enter the non-member donor name.")
            return donor_type, donor_value, None, donor_value or "Non-member"

        if donor_type == Donations.DonorType.ANONYMOUS:
            return donor_type, "", None, "Anonymous"

        form.add_error("donor_type", "Invalid donor type selected.")
        return donor_type, donor_value, donation_member, ""

    def get(self, request):
        staged = request.session.get(SESSION_KEY)
        if staged and staged.get("temp_file_id"):
            TemporaryBankMovementFile.objects.filter(id=staged["temp_file_id"]).delete()

        request.session.pop(SESSION_KEY, None)
        request.session.modified = True

        form = BankMovementUnifiedForm(user=request.user)
        return render(request, self.template_name, {"form": form, "error_summary": []})

    def post(self, request):
        try:
            church = _require_church(request.user)
        except PermissionError as e:
            messages.error(request, str(e))
            return redirect("home")

        form = BankMovementUnifiedForm(request.POST, request.FILES, user=request.user)

        if not form.is_valid():
            messages.error(request, "Please correct the errors below.")
            return render(
                request,
                self.template_name,
                {"form": form, "error_summary": _build_error_summary(form)},
            )

        cd = form.cleaned_data

        try:
            _lock_date_guard(church, cd["date"])
        except ValueError as e:
            messages.error(request, str(e))
            return render(
                request,
                self.template_name,
                {"form": form, "error_summary": _build_error_summary(form)},
            )

        direction = cd.get("direction")
        source_type = cd.get("source_type")

        bank_affecting = {
            CashBankMovement.Direction.CASH_TO_BANK,
            CashBankMovement.Direction.BANK_TO_CASH,
            CashBankMovement.Direction.BANK_TO_BANK,
            CashBankMovement.Direction.DIRECT_BANK_RECEIPT,
            CashBankMovement.Direction.BANK_PAID_EXPENSE,
        }

        proof = cd.get("proof_file")
        if direction in bank_affecting and not proof:
            form.add_error("proof_file", "Proof file is required for this transaction.")
            messages.error(request, "Please correct the errors below.")
            return render(
                request,
                self.template_name,
                {"form": form, "error_summary": _build_error_summary(form)},
            )

        if direction in (
            CashBankMovement.Direction.DIRECT_BANK_RECEIPT,
            CashBankMovement.Direction.BANK_TO_BANK,
        ):
            if source_type == CashBankMovement.SourceType.DONATION and not cd.get("donations_type"):
                form.add_error("donations_type", "Donation category is required.")
            if source_type == CashBankMovement.SourceType.OTHER_INCOME and not cd.get("income_type"):
                form.add_error("income_type", "Other income category is required.")

        if (
            direction == CashBankMovement.Direction.BANK_TO_BANK
            and source_type == CashBankMovement.SourceType.EXPENSE
        ):
            form.add_error("source_type", "BANK_TO_BANK cannot be Expense. Use BANK_PAID_EXPENSE instead.")

        if (
            direction == CashBankMovement.Direction.BANK_PAID_EXPENSE
            and source_type == CashBankMovement.SourceType.EXPENSE
        ):
            expense_amount = Decimal(str(cd.get("amount") or "0.00"))
            expense_category = cd.get("expense_category")

            if not expense_category:
                form.add_error("expense_category", "Expense category is required for bank-paid expense.")
            else:
                try:
                    _bank_acct, current_bank_balance = _get_live_bank_balance(church)
                except ValidationError as e:
                    form.add_error(None, str(e))
                else:
                    if expense_amount > current_bank_balance:
                        form.add_error(
                            "amount",
                            f"Expense amount cannot exceed the current bank balance of ₱{current_bank_balance:,.2f}."
                        )

                if getattr(expense_category, "is_restricted", False):
                    fund_name, restricted_balance = _get_restricted_balance_for_expense_category(
                        church,
                        expense_category,
                    )

                    if restricted_balance is None:
                        form.add_error(
                            "expense_category",
                            "This restricted expense category is not properly linked to a restricted fund."
                        )
                    elif expense_amount > restricted_balance:
                        form.add_error(
                            "amount",
                            f"Expense amount cannot exceed the restricted fund balance for '{fund_name}' "
                            f"(Available: ₱{restricted_balance:,.2f})."
                        )

        donor_type = ""
        donor_value = ""
        donation_member = None
        donor_display = ""

        if source_type == CashBankMovement.SourceType.DONATION:
            donor_type, donor_value, donation_member, donor_display = self._validate_donation_party(
                form, cd, church
            )

        if form.errors:
            messages.error(request, "Please correct the errors below.")
            return render(
                request,
                self.template_name,
                {"form": form, "error_summary": _build_error_summary(form)},
            )

        ai_verified = True
        ai_reason = ""

        if direction in bank_affecting and proof:
            bank_acct = _get_bank_account(church)

            context = {
                "transaction_type": _movement_txn_type(direction),
                "direction": str(direction),
                "source_type": str(source_type),
                "reference_no": (cd.get("reference_no") or "").strip(),
                "memo": (cd.get("memo") or "").strip(),
                "bank_name": getattr(bank_acct, "bank_name", "") if bank_acct else "",
                "account_name": getattr(bank_acct, "account_name", "") if bank_acct else "",
                "account_number_last4": (
                    (bank_acct.account_number[-4:] if getattr(bank_acct, "account_number", None) else "")
                    if bank_acct else ""
                ),
            }

            if source_type == CashBankMovement.SourceType.DONATION:
                context.update(
                    {
                        "donor_type": donor_type,
                        "donor": donor_value,
                        "donor_display": donor_display,
                        "donations_type": str(cd.get("donations_type") or ""),
                        "donation_member": (
                            self._member_display_name(donation_member) if donation_member else ""
                        ),
                        "other_donations_type": (cd.get("other_donations_type") or "").strip(),
                    }
                )
            elif source_type == CashBankMovement.SourceType.OTHER_INCOME:
                context.update(
                    {
                        "income_type": str(cd.get("income_type") or ""),
                        "income_description": (cd.get("income_description") or "").strip(),
                    }
                )
            elif source_type == CashBankMovement.SourceType.EXPENSE:
                context.update(
                    {
                        "expense_category": str(cd.get("expense_category") or ""),
                        "expense_description": (cd.get("expense_description") or "").strip(),
                    }
                )

            try:
                ai_verified, ai_reason = verify_receipt_with_openai(
                    image_file=proof,
                    check_type="TRANSACTION",
                    expected_data={
                        "amount": cd["amount"],
                        "date": cd["date"].isoformat(),
                        "context": context,
                    },
                )
            except Exception as e:
                ai_verified = False
                ai_reason = f"AI verification error: {e}"

            if not ai_verified:
                form.add_error("proof_file", f"AI verification failed: {ai_reason}")
                messages.error(request, f"Potential Fraud Detected: {ai_reason}")
                return render(
                    request,
                    self.template_name,
                    {"form": form, "error_summary": _build_error_summary(form)},
                )

        temp_file_id = None
        file_url = ""
        if proof:
            tf = TemporaryBankMovementFile.objects.create(file=proof)
            temp_file_id = tf.id
            file_url = str(tf.file.url)

        staged = {
            "date": cd["date"].isoformat(),
            "amount": str(cd["amount"]),
            "direction": direction,
            "source_type": source_type,
            "memo": (cd.get("memo") or "").strip(),
            "reference_no": (cd.get("reference_no") or "").strip(),

            "donations_type_id": int(cd["donations_type"].id) if cd.get("donations_type") else None,
            "donor_type": donor_type if source_type == CashBankMovement.SourceType.DONATION else "",
            "donor": donor_value if source_type == CashBankMovement.SourceType.DONATION else "",
            "donor_display": donor_display if source_type == CashBankMovement.SourceType.DONATION else "",
            "other_donations_type": (cd.get("other_donations_type") or "").strip(),
            "donation_member_id": int(donation_member.id) if donation_member else None,
            "donation_member_name": (
                self._member_display_name(donation_member) if donation_member else ""
            ),

            "income_type_id": int(cd["income_type"].id) if cd.get("income_type") else None,
            "income_description": (cd.get("income_description") or "").strip(),

            "expense_category_id": int(cd["expense_category"].id) if cd.get("expense_category") else None,
            "expense_description": (cd.get("expense_description") or "").strip(),

            "temp_file_id": temp_file_id,
            "file_url": file_url,
            "ai_verified": bool(ai_verified),
            "ai_reason": ai_reason or "",
        }

        request.session[SESSION_KEY] = staged
        request.session.modified = True
        return redirect("review_bank_movement")


@method_decorator(never_cache, name="dispatch")
class BankMovementReviewView(FinanceRoleRequiredMixin, View):
    template_name = "bank_movement_review.html"

    def _member_display_name(self, user_obj):
        if not user_obj:
            return ""
        if hasattr(user_obj, "get_full_name"):
            full_name = (user_obj.get_full_name() or "").strip()
            if full_name:
                return full_name
        return getattr(user_obj, "username", str(user_obj))

    def _normalize_staged(self, staged):
        p = dict(staged or {})

        source_type = p.get("source_type")
        donor_type = p.get("donor_type")
        donor = str(p.get("donor") or "").strip()
        donor_display = str(p.get("donor_display") or "").strip()
        donation_member_id = p.get("donation_member_id")
        donation_member_name = str(p.get("donation_member_name") or "").strip()

        if source_type == CashBankMovement.SourceType.DONATION:
            if donor_type not in {
                Donations.DonorType.MEMBER,
                Donations.DonorType.NON_MEMBER,
                Donations.DonorType.ANONYMOUS,
            }:
                if donation_member_id:
                    donor_type = Donations.DonorType.MEMBER
                elif donor and donor.lower() != "anonymous":
                    donor_type = Donations.DonorType.NON_MEMBER
                else:
                    donor_type = Donations.DonorType.ANONYMOUS

            if donor_type == Donations.DonorType.MEMBER:
                donor = donor or donation_member_name or donor_display or "Selected Member"
                donor_display = donor

            elif donor_type == Donations.DonorType.NON_MEMBER:
                donor = donor or donor_display or ""
                donor_display = donor or "Non-member"
                donation_member_id = None
                donation_member_name = ""

            else:
                donor = ""
                donor_display = "Anonymous"
                donation_member_id = None
                donation_member_name = ""

            p["donor_type"] = donor_type
            p["donor"] = donor
            p["donor_display"] = donor_display
            p["donation_member_id"] = donation_member_id
            p["donation_member_name"] = donation_member_name

        return p

    def get(self, request):
        staged = request.session.get(SESSION_KEY)
        if not staged:
            messages.warning(request, "No draft found. Please add bank movement first.")
            return redirect("add_bank_movement")

        staged = self._normalize_staged(staged)
        request.session[SESSION_KEY] = staged
        request.session.modified = True

        return render(request, self.template_name, {"staged": staged})

    def post(self, request):
        staged = request.session.get(SESSION_KEY)
        if not staged:
            messages.warning(request, "Session expired. Please add bank movement again.")
            return redirect("add_bank_movement")

        staged = self._normalize_staged(staged)
        request.session[SESSION_KEY] = staged
        request.session.modified = True

        if "cancel" in request.POST:
            if staged.get("temp_file_id"):
                TemporaryBankMovementFile.objects.filter(id=staged["temp_file_id"]).delete()
            request.session.pop(SESSION_KEY, None)
            request.session.modified = True
            messages.info(request, "Draft cancelled.")
            return redirect("add_bank_movement")

        if "confirm" not in request.POST:
            return redirect("review_bank_movement")

        church = _require_church(request.user)

        txn_date_obj = _parse_session_date(staged.get("date"))
        if not txn_date_obj:
            if staged.get("temp_file_id"):
                TemporaryBankMovementFile.objects.filter(id=staged["temp_file_id"]).delete()
            request.session.pop(SESSION_KEY, None)
            request.session.modified = True
            messages.error(request, "Invalid date in session draft. Please create the draft again.")
            return redirect("add_bank_movement")

        try:
            txn_amount = Decimal(staged.get("amount") or "0")
        except Exception:
            messages.error(request, "Invalid amount in draft. Please create the draft again.")
            return redirect("add_bank_movement")

        direction = staged.get("direction")
        source_type = staged.get("source_type")

        try:
            _lock_date_guard(church, txn_date_obj)
        except Exception as e:
            messages.error(request, str(e))
            return redirect("review_bank_movement")

        bank_acct = _get_bank_account(church)
        if not bank_acct:
            messages.error(request, "Please configure Bank Settings first.")
            return redirect("bank_settings")

        temp = TemporaryBankMovementFile.objects.filter(id=staged.get("temp_file_id")).first()

        bank_affecting = {
            CashBankMovement.Direction.CASH_TO_BANK,
            CashBankMovement.Direction.BANK_TO_CASH,
            CashBankMovement.Direction.BANK_TO_BANK,
            CashBankMovement.Direction.DIRECT_BANK_RECEIPT,
            CashBankMovement.Direction.BANK_PAID_EXPENSE,
        }

        if direction in bank_affecting:
            is_valid, reason = _ai_verify_from_temp(
                temp_obj=temp,
                bank_acct=bank_acct,
                amount=txn_amount,
                txn_date=txn_date_obj,
                direction=direction,
                source_type=source_type,
            )
            if not is_valid:
                messages.error(request, f"Potential Fraud Detected: {reason}")
                return redirect("review_bank_movement")

        try:
            with transaction.atomic():
                bank_acct = BankAccount.objects.select_for_update().get(pk=bank_acct.pk)

                _call_recalc_bank_balance(church.id)
                bank_acct.refresh_from_db(fields=["current_balance"])

                created_source_id = None

                is_income_like = (
                    direction in (
                        CashBankMovement.Direction.DIRECT_BANK_RECEIPT,
                        CashBankMovement.Direction.BANK_TO_BANK,
                    )
                    and source_type in (
                        CashBankMovement.SourceType.TITHE,
                        CashBankMovement.SourceType.OFFERING,
                        CashBankMovement.SourceType.DONATION,
                        CashBankMovement.SourceType.OTHER_INCOME,
                    )
                )

                true_bank_outflow = direction in (
                    CashBankMovement.Direction.BANK_TO_CASH,
                    CashBankMovement.Direction.BANK_PAID_EXPENSE,
                )

                current_balance = Decimal(str(bank_acct.current_balance or "0.00"))
                if true_bank_outflow and txn_amount > current_balance:
                    raise ValidationError("Insufficient bank balance.")

                if is_income_like and source_type == CashBankMovement.SourceType.DONATION:
                    cat = get_object_or_404(
                        DonationCategory,
                        pk=staged.get("donations_type_id"),
                        church=church,
                    )

                    donor_type = staged.get("donor_type") or Donations.DonorType.ANONYMOUS
                    linked_user = None
                    donor_value = None

                    if donor_type == Donations.DonorType.MEMBER:
                        member_id = staged.get("donation_member_id")
                        if not member_id:
                            raise ValidationError("Member donor is missing the selected user.")

                        user_model = Donations._meta.get_field("user").remote_field.model
                        linked_user = user_model.objects.filter(
                            pk=member_id,
                            church=church,
                            is_active=True,
                            user_type__in=["Member", "Pastor", "Treasurer"],
                        ).first()

                        if not linked_user:
                            raise ValidationError(
                                "Selected member/pastor/treasurer is invalid or no longer belongs to your church."
                            )

                        donor_value = self._member_display_name(linked_user)

                    elif donor_type == Donations.DonorType.NON_MEMBER:
                        donor_value = str(staged.get("donor") or "").strip()
                        if not donor_value:
                            raise ValidationError("Non-member donor name is required.")

                    elif donor_type == Donations.DonorType.ANONYMOUS:
                        linked_user = None
                        donor_value = None

                    else:
                        raise ValidationError("Invalid donor type in staged donation.")

                    d = Donations(
                        church=church,
                        amount=txn_amount,
                        donations_type=cat,
                        donor_type=donor_type,
                        donor=donor_value,
                        donations_date=txn_date_obj,
                        other_donations_type=staged.get("other_donations_type", ""),
                        user=linked_user,
                        created_by=request.user,
                    )
                    _attach_temp_file_to_any_known_field(d, temp)
                    d.full_clean()
                    d.save()
                    created_source_id = d.id

                elif is_income_like and source_type == CashBankMovement.SourceType.OTHER_INCOME:
                    cat = get_object_or_404(
                        OtherIncomeCategory,
                        pk=staged.get("income_type_id"),
                        church=church,
                    )
                    oi = OtherIncome(
                        church=church,
                        income_type=cat,
                        amount=txn_amount,
                        date=txn_date_obj,
                        description=staged.get("income_description", ""),
                        created_by=request.user,
                    )
                    _attach_temp_file_to_any_known_field(oi, temp)
                    oi.full_clean()
                    oi.save()
                    created_source_id = oi.id

                elif is_income_like and source_type == CashBankMovement.SourceType.TITHE:
                    t = Tithe(
                        church=church,
                        amount=txn_amount,
                        date=txn_date_obj,
                        description=staged.get("memo") or "Bank Tithe",
                        created_by=request.user,
                        user_id=staged.get("donation_member_id"),
                    )
                    _attach_temp_file_to_any_known_field(t, temp)
                    t.full_clean()
                    t.save()
                    created_source_id = t.id

                elif is_income_like and source_type == CashBankMovement.SourceType.OFFERING:
                    o = Offering(
                        church=church,
                        amount=txn_amount,
                        date=txn_date_obj,
                        description=staged.get("memo") or "Bank Offering",
                        created_by=request.user,
                        user_id=staged.get("donation_member_id"),
                    )
                    _attach_temp_file_to_any_known_field(o, temp)
                    o.full_clean()
                    o.save()
                    created_source_id = o.id

                elif (
                    direction == CashBankMovement.Direction.BANK_PAID_EXPENSE
                    and source_type == CashBankMovement.SourceType.EXPENSE
                ):
                    cat = get_object_or_404(
                        ExpenseCategory,
                        pk=staged.get("expense_category_id"),
                        church=church,
                    )

                    if getattr(cat, "is_restricted", False):
                        fund_name, restricted_balance = _get_restricted_balance_for_expense_category(
                            church,
                            cat,
                        )

                        if restricted_balance is None:
                            raise ValidationError(
                                "This restricted expense category is not properly linked to a restricted fund."
                            )

                        if txn_amount > restricted_balance:
                            raise ValidationError(
                                f"Insufficient restricted fund balance for '{fund_name}'. "
                                f"Available: ₱{restricted_balance:,.2f}."
                            )

                    exp = Expense(
                        church=church,
                        amount=txn_amount,
                        expense_date=txn_date_obj,
                        description=staged.get("expense_description", ""),
                        category=cat,
                        created_by=request.user,
                        status="Pending",
                    )
                    _attach_temp_file_to_any_known_field(exp, temp)
                    exp.full_clean()
                    exp.save()
                    created_source_id = exp.id

                mv = CashBankMovement(
                    church=church,
                    date=txn_date_obj,
                    amount=txn_amount,
                    direction=direction,
                    source_type=source_type,
                    source_id=created_source_id,
                    memo=staged.get("memo", ""),
                    reference_no=staged.get("reference_no", ""),
                    created_by=request.user,
                    status=CashBankMovement.Status.CONFIRMED,
                )
                _attach_temp_file_to_field(mv, "proof_file", temp)
                mv.full_clean()
                mv.save()

                _call_recalc_bank_balance(church.id)
                bank_acct.refresh_from_db(fields=["current_balance"])

                bank_field_names = {f.name for f in BankAccount._meta.fields}
                update_fields = []

                if "updated_by" in bank_field_names:
                    bank_acct.updated_by = request.user
                    update_fields.append("updated_by")

                if "last_updated" in bank_field_names:
                    bank_acct.last_updated = timezone.now()
                    update_fields.append("last_updated")

                if update_fields:
                    bank_acct.save(update_fields=update_fields)

            updated_balance = bank_acct.current_balance or Decimal("0.00")
            success_message = _build_bank_movement_success_message(
                direction=direction,
                source_type=source_type,
                amount=txn_amount,
                updated_balance=updated_balance,
            )

            if staged.get("temp_file_id"):
                TemporaryBankMovementFile.objects.filter(id=staged["temp_file_id"]).delete()

            request.session.pop(SESSION_KEY, None)
            request.session.modified = True

            messages.success(request, success_message)
            return redirect("add_bank_movement")

        except Exception as e:
            messages.error(request, f"Save failed: {e}")
            return redirect("review_bank_movement")

class BankMovementHistoryView(FinanceRoleRequiredMixin, View):
    template_name = "bank_movement_history.html"

    def get(self, request):
        church = _require_church(request.user)
        qs = (
            CashBankMovement.objects
            .filter(church=church)
            .select_related("created_by")
            .order_by("-date", "-created_at")[:300]
        )
        return render(request, self.template_name, {"rows": qs})


UNIFIED_INCOME_PANE_SESSION_KEY = "unified_income_active_pane"

PANE_TITHE = "pane-tithe"
PANE_SINGLE_OFFERING = "pane-single"
PANE_LOOSE_OFFERING = "pane-loose"
PANE_BATCH_OFFERING = "pane-batch"
PANE_DONATION = "pane-donations"
PANE_OTHER_INCOME = "pane-income"


def set_unified_income_active_pane(request, pane_name):
    request.session[UNIFIED_INCOME_PANE_SESSION_KEY] = pane_name


def get_unified_income_active_pane(request, default=PANE_TITHE):
    return (
        request.GET.get("active_pane")
        or request.session.pop(UNIFIED_INCOME_PANE_SESSION_KEY, None)
        or request.POST.get("active_pane")
        or default
    )


def redirect_unified_income_entry(pane_name=None):
    url = reverse("unified_income_entry")
    if pane_name:
        return redirect(f"{url}?active_pane={pane_name}")
    return redirect(url)


def get_unified_income_redirect_target(request, fallback=PANE_TITHE):
    pane_name = (
        request.POST.get("active_pane")
        or request.GET.get("active_pane")
        or request.session.get(UNIFIED_INCOME_PANE_SESSION_KEY)
        or fallback
    )
    return pane_name


def preserve_unified_income_pane(request, fallback=PANE_TITHE):
    pane_name = get_unified_income_redirect_target(request, fallback=fallback)
    set_unified_income_active_pane(request, pane_name)
    return pane_name

def is_from_unified_income_page(request):
    return request.POST.get("entry_source") == "unified_income_entry"


def build_unified_income_context(request, **overrides):
    context = {
        "tithe_form": TitheForm(),
        "single_form": OfferingForm(prefix="single", user=request.user),
        "loose_form": LooseOfferingForm(
            initial={"date": timezone.now().date()},
            prefix="loose"
        ),
        "batch_formset": EnvelopeFormSet(
            prefix="envelope",
            form_kwargs={"user": request.user}
        ),
        "donation_formset": AddDonationsView.DonationFormSet(
            prefix="donation",
            form_kwargs={"user": request.user}
        ),
        "other_income_formset": AddOtherIncomeView.OtherIncomeFormSet(
            prefix="income",
            form_kwargs={"user": request.user}
        ),
        "active_pane": overrides.pop(
            "active_pane",
            get_unified_income_active_pane(request, default=PANE_TITHE)
        ),
    }
    context.update(overrides)
    return context

class UnifiedIncomeEntryPageView(FinanceRoleRequiredMixin, TemplateView):
    template_name = "unified_income_entry.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(build_unified_income_context(self.request))
        return context

class ChurchAdminLiquidationRequestsView(IsChurchAdmin, TemplateView):
    template_name = "church_admin_liquidation_requests.html"

    def post(self, request, *args, **kwargs):
        request_id = (request.POST.get("request_id") or "").strip()
        action = (request.POST.get("action") or "").strip().lower()
        remarks = (request.POST.get("remarks") or "").strip()

        if not request_id:
            messages.error(request, "Missing request ID.")
            return redirect("church_admin_liquidation_requests")

        access_request = get_object_or_404(
            DenominationLiquidationAccessRequest,
            id=request_id,
            church=request.user.church,
        )

        if action == "approve":
            access_request.approve(request.user, remarks)
            messages.success(
                request,
                f"Viewing access approved for {access_request.denomination_admin}."
            )

        elif action == "reject":
            access_request.reject(request.user, remarks)
            messages.success(
                request,
                f"Viewing access request rejected for {access_request.denomination_admin}."
            )

        elif action == "close":
            if access_request.status != DenominationLiquidationAccessRequest.STATUS_APPROVED:
                messages.warning(request, "Only approved access can be closed.")
                return redirect("church_admin_liquidation_requests")

            access_request.close_access(request.user, remarks)
            messages.success(
                request,
                f"Viewing access closed for {access_request.denomination_admin}."
            )

        else:
            messages.error(request, "Invalid action.")

        return redirect("church_admin_liquidation_requests")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        requests_qs = (
            DenominationLiquidationAccessRequest.objects
            .filter(church=self.request.user.church)
            .select_related("denomination_admin", "reviewed_by", "closed_by", "church")
            .order_by("-requested_at", "-id")
        )

        pending_count = requests_qs.filter(
            status=DenominationLiquidationAccessRequest.STATUS_PENDING
        ).count()
        approved_count = requests_qs.filter(
            status=DenominationLiquidationAccessRequest.STATUS_APPROVED
        ).count()
        rejected_count = requests_qs.filter(
            status=DenominationLiquidationAccessRequest.STATUS_REJECTED
        ).count()
        closed_count = requests_qs.filter(
            status=DenominationLiquidationAccessRequest.STATUS_CLOSED
        ).count()

        context["requests"] = requests_qs
        context["pending_count"] = pending_count
        context["approved_count"] = approved_count
        context["rejected_count"] = rejected_count
        context["closed_count"] = closed_count
        context["rejected_closed_count"] = rejected_count + closed_count

        return context


class MonthlyYearlyAnalysisView(LoginRequiredMixin, View):
    template_name = "monthly_yearly_analysis.html"
    allowed_roles = {"Admin", "ChurchAdmin", "Treasurer", "Pastor"}

    def _safe_int(self, value, default=None):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def _has_access(self, user):
        return getattr(user, "user_type", None) in self.allowed_roles

    def _get_selected_year(self, request):
        now = datetime.now()
        current_year = now.year

        selected_year = self._safe_int(request.GET.get("year"), current_year) or current_year

        if selected_year < current_year - 10:
            selected_year = current_year - 10
        elif selected_year > current_year + 10:
            selected_year = current_year + 10

        return selected_year

    def _get_selected_month(self, request):
        selected_month = self._safe_int(request.GET.get("month"), 0)
        if selected_month not in range(0, 13):
            selected_month = 0
        return selected_month

    def _empty_summary(self, year, month=0):
        zero = Decimal("0.00")

        if month in range(1, 13):
            monthly_series = [
                {
                    "month_no": month,
                    "month": calendar.month_name[month],
                    "amount": zero,
                }
            ]
        else:
            monthly_series = [
                {
                    "month_no": m,
                    "month": calendar.month_name[m],
                    "amount": zero,
                }
                for m in range(1, 13)
            ]

        return {
            "year": year,
            "selected_month": month if month in range(1, 13) else None,
            "ministry_count": 0,
            "allocated_total": zero,
            "released_total": zero,
            "returned_total": zero,
            "spent_total": zero,
            "unliquidated_total": zero,
            "available_total": zero,
            "unreleased_total": zero,
            "utilization_pct": zero,
            "burn_rate_monthly_avg": zero,
            "projected_year_end_spend": zero,
            "monthly_spent_series": monthly_series,
            "ministries": [],
        }

    def get(self, request, *args, **kwargs):
        if not self._has_access(request.user):
            messages.error(
                request,
                "Access denied. Only Admin, Church Admin, Treasurer, and Pastor can view this analysis."
            )
            return redirect("home")

        church = getattr(request.user, "church", None)
        if not church:
            messages.error(request, "No church is linked to your account.")
            return redirect("home")

        now = datetime.now()
        selected_year = self._get_selected_year(request)
        selected_month = self._get_selected_month(request)

        month_choices = [(0, "All Months")] + [
            (m, calendar.month_name[m]) for m in range(1, 13)
        ]

        current_scope_label = (
            f"{calendar.month_name[selected_month]} {selected_year}"
            if selected_month in range(1, 13)
            else f"All Months of {selected_year}"
        )

        try:
            summary = get_monthly_yearly_dashboard_summary(
                church,
                selected_year,
                selected_month if selected_month in range(1, 13) else None,
            )
            ministries = summary.get("ministries", [])

            zero = Decimal("0.00")

            top_spending_ministry = (
                max(ministries, key=lambda x: x.get("spent_total", zero))
                if ministries else None
            )

            highest_utilization_ministry = (
                max(ministries, key=lambda x: x.get("utilization_pct", zero))
                if ministries else None
            )

            highest_burn_rate_ministry = (
                max(ministries, key=lambda x: x.get("burn_rate_monthly_avg", zero))
                if ministries else None
            )

            at_risk_ministries = [
                row for row in ministries
                if (
                    row.get("allocated_total", zero) > zero and (
                        row.get("projected_year_end_spend", zero) > row.get("allocated_total", zero)
                        or row.get("unliquidated_total", zero) > zero
                    )
                )
            ]

            monthly_spent_series = summary.get("monthly_spent_series", [])
            peak_month = None
            if monthly_spent_series:
                temp_peak = max(monthly_spent_series, key=lambda x: x.get("amount", zero))
                if temp_peak.get("amount", zero) > zero:
                    peak_month = temp_peak

            ai_insights = None
            try:
                if ministries:
                    ai_insights = build_budget_insights(summary, current_scope_label)
            except Exception:
                ai_insights = None

            context = {
                "summary": summary,
                "ministries": ministries,
                "selected_year": selected_year,
                "selected_month": selected_month,
                "month_choices": month_choices,
                "year_choices": range(now.year - 2, now.year + 3),
                "current_scope_label": current_scope_label,

                "top_spending_ministry": top_spending_ministry,
                "highest_utilization_ministry": highest_utilization_ministry,
                "highest_burn_rate_ministry": highest_burn_rate_ministry,
                "at_risk_ministries": at_risk_ministries,
                "peak_month": peak_month,

                "ai_insights": ai_insights,
            }
            return render(request, self.template_name, context)

        except Exception as e:
            messages.error(request, f"Unable to load Monthly-to-Yearly Analysis: {str(e)}")

            context = {
                "summary": self._empty_summary(selected_year, selected_month),
                "ministries": [],
                "selected_year": selected_year,
                "selected_month": selected_month,
                "month_choices": month_choices,
                "year_choices": range(now.year - 2, now.year + 3),
                "current_scope_label": current_scope_label,
                "top_spending_ministry": None,
                "highest_utilization_ministry": None,
                "highest_burn_rate_ministry": None,
                "at_risk_ministries": [],
                "peak_month": None,
                "ai_insights": None,
            }
            return render(request, self.template_name, context)

def _get_openai_client():
    if not getattr(settings, "OPENAI_API_KEY", ""):
        raise ValueError("OPENAI_API_KEY is not configured.")
    return OpenAI(api_key=settings.OPENAI_API_KEY)


def _get_openai_budget_model():
    return getattr(settings, "OPENAI_BUDGET_MODEL", "gpt-5.4")


def _openai_budget_schema():
    return {
        "type": "object",
        "properties": {
            "summary": {"type": "string"},
            "reserve_amount_monthly": {"type": "number"},
            "allocations": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "ministry_id": {"type": "string"},
                        "allocated_amount_monthly": {"type": "number"},
                        "priority_score": {"type": "integer"},
                        "rationale": {"type": "string"},
                    },
                    "required": [
                        "ministry_id",
                        "allocated_amount_monthly",
                        "priority_score",
                        "rationale",
                    ],
                    "additionalProperties": False,
                },
            },
        },
        "required": ["summary", "reserve_amount_monthly", "allocations"],
        "additionalProperties": False,
    }


def _call_openai_budget_optimizer(*, church_id, monthly_pool, timeframe, use_history, historical_year, ministry_data):
    """
    OpenAI proposes the allocation plan in MONTHLY units.
    We still validate and clamp it on the server.
    """
    client = _get_openai_client()
    model_name = _get_openai_budget_model()

    payload = {
        "church_id": church_id,
        "mode": "Historical" if use_history else "Manual",
        "historical_year": historical_year,
        "timeframe_for_ui": timeframe,
        "optimizer_unit": "monthly",
        "monthly_pool": float(_to_decimal(monthly_pool, "0.00")),
        "ministries": [
            {
                "ministry_id": str(m["id"]),
                "ministry_name": m["name"],
                "min_req_monthly": float(_to_decimal(m["min_req"], "0.00")),
                "max_cap_monthly": float(_to_decimal(m["max_cap"], "0.00")),
                "priority_score": int(m["priority_score"]),
                "spent_ytd": float(_to_decimal(m.get("spent_ytd", 0.0), "0.00")),
                "priority_source": m.get("priority_source", "unknown"),
            }
            for m in ministry_data
        ],
        "rules": [
            "Return allocations in MONTHLY units only.",
            "Do not return negative numbers.",
            "Respect each ministry max_cap_monthly.",
            "Try to satisfy min_req_monthly when possible.",
            "If funds are insufficient, prioritize higher priority_score ministries.",
            "Use each ministry exactly once.",
            "Any unallocated remainder must go to reserve_amount_monthly.",
        ],
    }

    response = client.responses.create(
        model=model_name,
        store=False,
        input=[
            {
                "role": "system",
                "content": (
                    "You are a church budget optimization engine. "
                    "Return only structured allocation data that follows the schema."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(payload, ensure_ascii=False),
            },
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "budget_optimizer_plan",
                "strict": True,
                "schema": _openai_budget_schema(),
            }
        },
    )

    raw_text = response.output_text or "{}"
    data = json.loads(raw_text)

    return {
        "model_name": model_name,
        "summary": data.get("summary", ""),
        "reserve_amount_monthly": _to_decimal(data.get("reserve_amount_monthly", 0), "0.00"),
        "allocations": data.get("allocations", []),
    }


def _normalize_openai_budget_allocations(monthly_pool, ministry_data, ai_allocations):
    """
    Final server-side protection:
    - known ministry IDs only
    - no negatives
    - clamp to max_cap
    - total cannot exceed monthly_pool
    """
    monthly_pool = _to_decimal(monthly_pool, "0.00")
    ministry_map = {str(m["id"]): m for m in ministry_data}

    cleaned = {}
    for item in ai_allocations:
        mid = str(item.get("ministry_id"))
        if mid not in ministry_map:
            continue

        amt = _to_decimal(item.get("allocated_amount_monthly", 0), "0.00")
        if amt < 0:
            amt = Decimal("0.00")

        max_cap = _to_decimal(ministry_map[mid]["max_cap"], "0.00")
        if max_cap > 0 and amt > max_cap:
            amt = max_cap

        cleaned[mid] = amt

    # ensure every ministry exists
    for mid in ministry_map.keys():
        cleaned.setdefault(mid, Decimal("0.00"))

    total_alloc = sum(cleaned.values(), Decimal("0.00"))

    # scale down if model exceeded pool
    if total_alloc > monthly_pool and total_alloc > 0:
        ratio = monthly_pool / total_alloc
        for mid in cleaned:
            cleaned[mid] = (cleaned[mid] * ratio).quantize(Decimal("0.01"))

        # rounding guard
        total_alloc = sum(cleaned.values(), Decimal("0.00"))
        overflow = total_alloc - monthly_pool
        if overflow > 0:
            biggest = max(cleaned.keys(), key=lambda k: cleaned[k])
            cleaned[biggest] = max(Decimal("0.00"), cleaned[biggest] - overflow)

    reserve = monthly_pool - sum(cleaned.values(), Decimal("0.00"))
    if reserve < 0:
        reserve = Decimal("0.00")

    return cleaned, reserve.quantize(Decimal("0.01"))

def _get_openai_client():
    api_key = getattr(settings, "OPENAI_API_KEY", "")
    if not api_key:
        raise ValueError("OPENAI_API_KEY is not configured.")
    return OpenAI(api_key=api_key)


def _get_openai_budget_model():
    return getattr(settings, "OPENAI_BUDGET_MODEL", "gpt-5.4")


def _openai_optimizer_schema():
    return {
        "type": "object",
        "properties": {
            "summary": {"type": "string"},
            "reserve_amount_monthly": {"type": "number"},
            "allocations": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "ministry_id": {"type": "string"},
                        "allocated_amount_monthly": {"type": "number"},
                        "priority_score": {"type": "integer"},
                        "rationale": {"type": "string"},
                    },
                    "required": [
                        "ministry_id",
                        "allocated_amount_monthly",
                        "priority_score",
                        "rationale",
                    ],
                    "additionalProperties": False,
                },
            },
        },
        "required": ["summary", "reserve_amount_monthly", "allocations"],
        "additionalProperties": False,
    }


def _call_openai_optimizer_for_budget(
    *,
    church_id,
    monthly_pool,
    timeframe,
    use_history,
    historical_year,
    ministry_data,
):
    """
    Ask OpenAI for proposed MONTHLY allocations.
    We still validate and normalize server-side afterwards.
    """
    client = _get_openai_client()
    model_name = _get_openai_budget_model()

    payload = {
        "church_id": church_id,
        "mode": "Historical" if use_history else "Manual",
        "historical_year": historical_year,
        "timeframe_for_ui": timeframe,
        "optimizer_unit": "monthly",
        "monthly_pool": float(_to_decimal(monthly_pool, "0.00")),
        "ministries": [
            {
                "ministry_id": str(m["id"]),
                "ministry_name": m["name"],
                "min_req_monthly": float(_to_decimal(m["min_req"], "0.00")),
                "max_cap_monthly": float(_to_decimal(m["max_cap"], "0.00")),
                "priority_score": int(m["priority_score"]),
                "spent_ytd": float(_to_decimal(m.get("spent_ytd", 0), "0.00")),
                "priority_source": m.get("priority_source", "unknown"),
            }
            for m in ministry_data
        ],
        "rules": [
            "Return allocations in MONTHLY units only.",
            "Do not allocate negative values.",
            "Use every ministry exactly once.",
            "Do not exceed the provided monthly_pool total.",
            "Try to satisfy min_req_monthly where feasible.",
            "Respect max_cap_monthly if it is greater than zero.",
            "If funds are insufficient, higher priority_score ministries should receive relatively more.",
            "Put any leftover into reserve_amount_monthly.",
        ],
    }

    response = client.responses.create(
        model=model_name,
        store=False,
        input=[
            {
                "role": "system",
                "content": (
                    "You are a church budget optimization engine. "
                    "Return only valid structured JSON that matches the required schema."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(payload, ensure_ascii=False),
            },
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "budget_optimizer_plan",
                "strict": True,
                "schema": _openai_optimizer_schema(),
            }
        },
    )

    raw_text = response.output_text or "{}"
    parsed = json.loads(raw_text)

    return {
        "model_name": model_name,
        "summary": parsed.get("summary", ""),
        "reserve_amount_monthly": _to_decimal(parsed.get("reserve_amount_monthly", 0), "0.00"),
        "allocations": parsed.get("allocations", []),
    }


def _normalize_openai_allocations(monthly_pool, ministry_data, ai_allocations):
    """
    Clamp and repair OpenAI output so the final plan is safe:
    - known ministry IDs only
    - no negatives
    - do not exceed max caps
    - do not exceed total pool
    """
    monthly_pool = _to_decimal(monthly_pool, "0.00")
    ministry_map = {str(m["id"]): m for m in ministry_data}

    cleaned = {}
    rationale_map = {}

    for item in ai_allocations:
        mid = str(item.get("ministry_id"))
        if mid not in ministry_map:
            continue

        amt = _to_decimal(item.get("allocated_amount_monthly", 0), "0.00")
        if amt < 0:
            amt = Decimal("0.00")

        max_cap = _to_decimal(ministry_map[mid].get("max_cap", 0), "0.00")
        if max_cap > 0 and amt > max_cap:
            amt = max_cap

        cleaned[mid] = amt
        rationale_map[mid] = item.get("rationale", "")

    # ensure every ministry exists
    for mid in ministry_map.keys():
        cleaned.setdefault(mid, Decimal("0.00"))
        rationale_map.setdefault(mid, "")

    total_alloc = sum(cleaned.values(), Decimal("0.00"))

    # if AI exceeded total pool, scale down proportionally
    if total_alloc > monthly_pool and total_alloc > 0:
        ratio = monthly_pool / total_alloc
        for mid in cleaned:
            cleaned[mid] = (cleaned[mid] * ratio).quantize(Decimal("0.01"))

        total_alloc = sum(cleaned.values(), Decimal("0.00"))
        overflow = total_alloc - monthly_pool
        if overflow > 0:
            biggest = max(cleaned.keys(), key=lambda k: cleaned[k])
            cleaned[biggest] = max(Decimal("0.00"), cleaned[biggest] - overflow)

    reserve = monthly_pool - sum(cleaned.values(), Decimal("0.00"))
    if reserve < 0:
        reserve = Decimal("0.00")

    return cleaned, rationale_map, reserve.quantize(Decimal("0.01"))



def _clamp_decimal(val, low, high):
    if val < low:
        return low
    if val > high:
        return high
    return val

def _clamp_int(val, low=1, high=10):
    return max(low, min(high, int(val)))

def _current_year_monthly_spend_features(church_id, ministry_ids, year):
    """
    Returns per ministry:
      {
        ministry_id: {
          "monthly_totals": [Decimal x 12],
          "spent_ytd": Decimal,
          "avg_monthly": Decimal,
          "active_months": int,
          "recent_3m_avg": Decimal,
          "prev_3m_avg": Decimal,
          "trend_ratio": Decimal,
        }
      }

    Adjust the query below to your real spending source if needed.
    This version assumes BudgetExpense -> release -> budget -> ministry.
    """
    BudgetExpense = apps.get_model("Register", "BudgetExpense")

    rows = (
        BudgetExpense.objects
        .filter(
            release__budget__church_id=church_id,
            release__budget__ministry_id__in=ministry_ids,
            release__date_released__year=year,
        )
        .values("release__budget__ministry_id", "release__date_released__month")
        .annotate(total_amt=Coalesce(Sum("amount"), Decimal("0.00")))
        .order_by("release__budget__ministry_id", "release__date_released__month")
    )

    result = {}
    for mid in ministry_ids:
        result[mid] = {
            "monthly_totals": [Decimal("0.00")] * 12,
            "spent_ytd": Decimal("0.00"),
            "avg_monthly": Decimal("0.00"),
            "active_months": 0,
            "recent_3m_avg": Decimal("0.00"),
            "prev_3m_avg": Decimal("0.00"),
            "trend_ratio": Decimal("1.00"),
        }

    for r in rows:
        mid = r["release__budget__ministry_id"]
        month = int(r["release__date_released__month"])
        amt = _to_decimal(r["total_amt"], "0.00")
        if 1 <= month <= 12:
            result[mid]["monthly_totals"][month - 1] = amt

    for mid, item in result.items():
        monthly = item["monthly_totals"]
        spent_ytd = sum(monthly, Decimal("0.00"))
        active_months = sum(1 for x in monthly if x > 0)
        avg_monthly = (spent_ytd / Decimal("12")).quantize(Decimal("0.01")) if spent_ytd > 0 else Decimal("0.00")

        recent_3m = monthly[9:12]   # Oct-Dec
        prev_3m = monthly[6:9]      # Jul-Sep

        recent_3m_avg = (
            sum(recent_3m, Decimal("0.00")) / Decimal("3")
        ).quantize(Decimal("0.01"))
        prev_3m_avg = (
            sum(prev_3m, Decimal("0.00")) / Decimal("3")
        ).quantize(Decimal("0.01"))

        if prev_3m_avg > 0:
            trend_ratio = (recent_3m_avg / prev_3m_avg).quantize(Decimal("0.01"))
        elif recent_3m_avg > 0:
            trend_ratio = Decimal("1.50")
        else:
            trend_ratio = Decimal("1.00")

        item["spent_ytd"] = spent_ytd.quantize(Decimal("0.01"))
        item["avg_monthly"] = avg_monthly
        item["active_months"] = active_months
        item["recent_3m_avg"] = recent_3m_avg
        item["prev_3m_avg"] = prev_3m_avg
        item["trend_ratio"] = trend_ratio

    return result


def _compute_effective_priority(
    manual_priority,
    base_budget_monthly,
    spent_avg_monthly,
    spent_ytd,
    trend_ratio,
    active_months,
):
    """
    Blended score 1..10.

    Weights:
      50% manual priority
      25% utilization/spending intensity
      15% recent trend
      10% activity consistency
    """
    manual = Decimal(str(manual_priority or 5))

    budget_base = _to_decimal(base_budget_monthly, "0.00")
    spent_avg = _to_decimal(spent_avg_monthly, "0.00")
    spent_total = _to_decimal(spent_ytd, "0.00")
    trend = _to_decimal(trend_ratio, "1.00")

    if budget_base > 0:
        utilization_ratio = (spent_avg / budget_base).quantize(Decimal("0.01"))
    else:
        utilization_ratio = Decimal("0.00")

    utilization_ratio = _clamp_decimal(utilization_ratio, Decimal("0.00"), Decimal("2.00"))
    utilization_score = utilization_ratio * Decimal("5.00")   # 0..10 approximately

    trend_score = _clamp_decimal(trend * Decimal("5.00"), Decimal("1.00"), Decimal("10.00"))
    activity_score = Decimal(str(active_months)) / Decimal("12") * Decimal("10.00")

    weighted = (
        manual * Decimal("0.50") +
        utilization_score * Decimal("0.25") +
        trend_score * Decimal("0.15") +
        activity_score * Decimal("0.10")
    )

    return _clamp_int(int(weighted.quantize(Decimal("1"), rounding=ROUND_HALF_UP)), 1, 10)

@csrf_exempt
@require_POST
@login_required
def api_run_optimizer_openai(request):
    try:
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if not hasattr(request.user, "church") or not request.user.church:
            return JsonResponse(
                {"error": "Security Violation: No church assigned to this user."},
                status=403
            )

        church_id = request.user.church.id

        ReleasedBudget = apps.get_model("Register", "ReleasedBudget")
        Ministry = apps.get_model("Register", "Ministry")
        MinistryBudget = apps.get_model("Register", "MinistryBudget")

        # =========================
        # Configurable weights
        # =========================
        BUDGET_WEIGHT = Decimal("0.60")
        SPEND_WEIGHT = Decimal("0.40")

        raw_use_history = data.get("use_history", False)
        if isinstance(raw_use_history, bool):
            use_history = raw_use_history
        else:
            use_history = str(raw_use_history).strip().lower() in ("1", "true", "yes", "on")

        timeframe = (data.get("timeframe") or "yearly").lower()
        if timeframe not in ("yearly", "monthly"):
            timeframe = "yearly"

        historical_year = _safe_int(data.get("historical_year"), default=None)
        if use_history and not historical_year:
            return JsonResponse(
                {"error": "historical_year is required when use_history is true."},
                status=400
            )

        raw_dry_run = data.get("dry_run", False)
        if isinstance(raw_dry_run, bool):
            dry_run = raw_dry_run
        else:
            dry_run = str(raw_dry_run).strip().lower() in ("1", "true", "yes", "on")

        budget_map = {}
        spent_map = {}
        months_observed = 0
        manual_funds_meta = {}

        # Manual-mode working values
        total_cash = Decimal("0.00")
        restricted_held = Decimal("0.00")
        unrestricted_cash = Decimal("0.00")
        projected_bills = Decimal("0.00")
        reserve_target_monthly = Decimal("0.00")
        affordable_pool_monthly = Decimal("0.00")
        budget_need_monthly = Decimal("0.00")
        spend_need_monthly = Decimal("0.00")
        baseline_need_monthly = Decimal("0.00")
        need_total_monthly = Decimal("0.00")
        summed_ministry_need_monthly = Decimal("0.00")
        monthly_pool = Decimal("0.00")
        pool_used_ui = Decimal("0.00")

        # =========================================================
        # 1) BUILD MONTHLY POOL
        # =========================================================
        if use_history:
            income_basis = "historical_unrestricted_sp"

            annual_income = sp_unrestricted_income_for_year(
                church_id, historical_year
            ) or Decimal("0.00")

            if annual_income <= 0:
                return JsonResponse(
                    {"error": f"No unrestricted income found for year {historical_year}."},
                    status=400
                )

            monthly_pool = (annual_income / Decimal("12")).quantize(Decimal("0.01"))
            pool_used_ui = monthly_pool if timeframe == "monthly" else (monthly_pool * Decimal("12"))

        else:
            income_basis = "current_unrestricted_cash_minus_projected_unrestricted_bills_capped_by_need"

            # All-years historical need signals
            budget_map, budget_need_monthly = _all_years_budget_monthly_map(church_id)
            spent_map, months_observed = _all_years_spent_monthly_map(church_id)

            # Current financial capacity
            funds = get_current_fund_balances(church_id)

            total_cash = _to_decimal(funds.get("total_cash"), "0.00")
            restricted_held = _to_decimal(funds.get("restricted_balance_outstanding"), "0.00")
            unrestricted_cash = _to_decimal(funds.get("unrestricted_cash_available"), "0.00")

            projected_bills = _to_decimal(get_projected_expenses_unrestricted(church_id), "0.00")

            # Simple reserve rule
            reserve_target_monthly = projected_bills

            affordable_pool_monthly = unrestricted_cash - reserve_target_monthly
            if affordable_pool_monthly < 0:
                affordable_pool_monthly = Decimal("0.00")

            spend_need_monthly = sum(
                (_to_decimal(v, "0.00") for v in spent_map.values()),
                Decimal("0.00")
            ).quantize(Decimal("0.01"))

            # Overall baseline pool need signal
            if months_observed and months_observed >= 3:
                baseline_need_monthly = (
                    (budget_need_monthly * BUDGET_WEIGHT) +
                    (spend_need_monthly * SPEND_WEIGHT)
                ).quantize(Decimal("0.01"))
            else:
                baseline_need_monthly = _to_decimal(budget_need_monthly, "0.00").quantize(Decimal("0.01"))

            monthly_pool = (
                min(affordable_pool_monthly, baseline_need_monthly)
                if baseline_need_monthly > 0
                else affordable_pool_monthly
            ).quantize(Decimal("0.01"))

            pool_used_ui = monthly_pool if timeframe == "monthly" else (monthly_pool * Decimal("12"))

            manual_funds_meta = {
                "total_cash": float(total_cash.quantize(Decimal("0.01"))),
                "restricted_funds_held": float(restricted_held.quantize(Decimal("0.01"))),
                "unrestricted_cash_available": float(unrestricted_cash.quantize(Decimal("0.01"))),
                "projected_unrestricted_bills_monthly": float(projected_bills.quantize(Decimal("0.01"))),
                "reserve_target_monthly": float(reserve_target_monthly.quantize(Decimal("0.01"))),
                "affordable_pool_monthly": float(affordable_pool_monthly.quantize(Decimal("0.01"))),
                "budget_need_monthly": float(_to_decimal(budget_need_monthly, "0.00").quantize(Decimal("0.01"))),
                "spend_need_monthly": float(spend_need_monthly.quantize(Decimal("0.01"))),
                "baseline_need_monthly": float(baseline_need_monthly.quantize(Decimal("0.01"))),
                "initial_recommended_pool_monthly": float(monthly_pool.quantize(Decimal("0.01"))),
            }

        # =========================================================
        # 2) LOAD ELIGIBLE MINISTRIES
        # =========================================================
        if use_history:
            budget_ministry_ids = set(
                MinistryBudget.objects.filter(
                    church_id=church_id,
                    year=historical_year
                ).values_list("ministry_id", flat=True)
            )

            release_ministry_ids = set(
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    date_released__year=historical_year
                )
                .exclude(ministry_id__isnull=True)
                .values_list("ministry_id", flat=True)
            )

            ministry_ids = budget_ministry_ids | release_ministry_ids

            if not ministry_ids:
                return JsonResponse(
                    {
                        "error": (
                            f"No ministries found for year {historical_year}. "
                            f"Add a budget or a release for that year first."
                        )
                    },
                    status=400
                )

            ministries = list(
                Ministry.objects.filter(
                    church_id=church_id,
                    is_active=True,
                    id__in=ministry_ids
                ).order_by("name")
            )
        else:
            ministries = list(
                Ministry.objects.filter(
                    church_id=church_id,
                    is_active=True
                ).order_by("name")
            )

        if not ministries:
            return JsonResponse({"error": "No eligible ministries found."}, status=400)

        # =========================================================
        # 3) PRIORITIES + CONSTRAINTS
        # =========================================================
        raw_priorities = data.get("priorities") or {}
        user_priorities = (
            {str(k): v for k, v in raw_priorities.items()}
            if isinstance(raw_priorities, dict) else {}
        )

        historical_net_spent_annual = {}
        if use_history:
            rel_rows = (
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    ministry_id__isnull=False,
                    date_released__year=historical_year,
                    ministry_id__in=[m.id for m in ministries],
                )
                .values("ministry_id")
                .annotate(released=Coalesce(Sum("amount"), Decimal("0.00")))
            )
            released_by_id = {
                r["ministry_id"]: _to_decimal(r["released"], "0.00")
                for r in rel_rows
            }

            ret_rows = (
                ReleasedBudget.objects.filter(
                    church_id=church_id,
                    ministry_id__isnull=False,
                    is_liquidated=True,
                    liquidated_date__isnull=False,
                    liquidated_date__year=historical_year,
                    ministry_id__in=[m.id for m in ministries],
                )
                .values("ministry_id")
                .annotate(returned=Coalesce(Sum("amount_returned"), Decimal("0.00")))
            )
            returned_by_id = {
                r["ministry_id"]: _to_decimal(r["returned"], "0.00")
                for r in ret_rows
            }

            for m in ministries:
                net = released_by_id.get(m.id, Decimal("0.00")) - returned_by_id.get(m.id, Decimal("0.00"))
                if net < 0:
                    net = Decimal("0.00")
                historical_net_spent_annual[m.id] = net

        current_ytd_spent_map = {}
        if not use_history:
            today = timezone.now().date()
            current_ytd_spent_map = _net_spent_ytd_map(
                church_id=church_id,
                year=today.year,
                ministry_ids=[m.id for m in ministries],
                as_of_date=today,
            )

        ministry_data = []
        manual_override_count = 0

        for m in ministries:
            db_min_monthly = _to_decimal(getattr(m, "min_monthly_budget", 0), "0.00")
            db_max_monthly = _to_decimal(getattr(m, "max_monthly_cap", 0), "0.00")

            if use_history:
                annual_spent = historical_net_spent_annual.get(m.id, Decimal("0.00"))
                hist_avg_monthly = (annual_spent / Decimal("12")) if annual_spent > 0 else Decimal("0.00")

                # Historical mode stays spending-based
                min_req_m = max(db_min_monthly, hist_avg_monthly).quantize(Decimal("0.01"))

                hist_cap_m = (hist_avg_monthly * Decimal("1.20")) if hist_avg_monthly > 0 else Decimal("0.00")
                max_cap_m = max(db_max_monthly, hist_cap_m).quantize(Decimal("0.01"))

                spent_ytd_val = float(annual_spent.quantize(Decimal("0.01")))
                base_budget_m = Decimal("0.00")
                spent_avg_m = hist_avg_monthly

            else:
                base_budget_m = _to_decimal(budget_map.get(m.id, 0), "0.00")
                spent_avg_m = _to_decimal(spent_map.get(m.id, 0), "0.00")

                # =====================================================
                # UPDATED LOGIC:
                # Blend budget basis with actual spending pattern
                # =====================================================
                if base_budget_m > 0 and spent_avg_m > 0:
                    blended_need_m = (
                        (base_budget_m * BUDGET_WEIGHT) +
                        (spent_avg_m * SPEND_WEIGHT)
                    )
                    min_req_m = max(blended_need_m, db_min_monthly)

                elif base_budget_m > 0:
                    min_req_m = max(base_budget_m, db_min_monthly)

                else:
                    min_req_m = max(spent_avg_m, db_min_monthly)

                min_req_m = min_req_m.quantize(Decimal("0.01"))

                if db_max_monthly > 0:
                    max_cap_m = db_max_monthly
                else:
                    cap_from_budget = (base_budget_m * Decimal("1.25")) if base_budget_m > 0 else Decimal("0.00")
                    cap_from_spend = (spent_avg_m * Decimal("1.20")) if spent_avg_m > 0 else Decimal("0.00")
                    max_cap_m = max(cap_from_budget, cap_from_spend, Decimal("0.00"))

                max_cap_m = max_cap_m.quantize(Decimal("0.01"))
                spent_ytd_val = float(
                    _to_decimal(current_ytd_spent_map.get(m.id, 0), "0.00").quantize(Decimal("0.01"))
                )

            if max_cap_m > 0 and max_cap_m < min_req_m:
                max_cap_m = min_req_m

            mid_str = str(m.id)
            manual_override = user_priorities.get(mid_str)

            if manual_override is not None:
                score = _safe_int(manual_override, default=5) or 5
                priority_source = "manual_override"
                manual_override_count += 1
            else:
                score = int(getattr(m, "priority_score", 5) or 5)
                priority_source = "db_default"

            ministry_data.append({
                "id": m.id,
                "name": m.name,
                "min_req": float(min_req_m),
                "max_cap": float(max_cap_m),
                "spent_ytd": float(spent_ytd_val),
                "priority_score": int(score),
                "priority_source": priority_source,

                # Support/debug fields
                "base_budget_monthly": float(base_budget_m),
                "spent_avg_monthly": float(spent_avg_m),
                "db_min_monthly": float(db_min_monthly),
                "db_max_monthly": float(db_max_monthly),
            })

        # De-duplicate by ministry id while preserving the last computed value
        ministry_data = list({str(md["id"]): md for md in ministry_data}.values())

        # =========================================================
        # 3.5) FINAL MANUAL POOL CAP USING MINISTRY BASE REQUIREMENTS
        # =========================================================
        if not use_history:
            summed_ministry_need_monthly = sum(
                (_to_decimal(md.get("min_req", 0), "0.00") for md in ministry_data),
                Decimal("0.00")
            ).quantize(Decimal("0.01"))

            need_total_monthly = max(baseline_need_monthly, summed_ministry_need_monthly).quantize(Decimal("0.01"))

            monthly_pool = (
                min(affordable_pool_monthly, need_total_monthly)
                if need_total_monthly > 0
                else affordable_pool_monthly
            ).quantize(Decimal("0.01"))

            pool_used_ui = monthly_pool if timeframe == "monthly" else (monthly_pool * Decimal("12"))

            manual_funds_meta.update({
                "summed_ministry_need_monthly": float(summed_ministry_need_monthly),
                "need_total_monthly": float(need_total_monthly),
                "final_recommended_pool_monthly": float(monthly_pool.quantize(Decimal("0.01"))),
            })

        if monthly_pool <= 0:
            payload = {
                "error": "No safe funds available to optimize after reserve protection and ministry-need analysis."
            }
            if manual_funds_meta:
                payload["meta"] = manual_funds_meta
            return JsonResponse(payload, status=400)

        # =========================================================
        # 4) OPENAI OPTIMIZATION
        # =========================================================
        ai_result = _call_openai_optimizer_for_budget(
            church_id=church_id,
            monthly_pool=monthly_pool,
            timeframe=timeframe,
            use_history=use_history,
            historical_year=historical_year,
            ministry_data=ministry_data,
        )

        allocations_monthly, rationale_map, reserve_amount_monthly = _normalize_openai_allocations(
            monthly_pool=monthly_pool,
            ministry_data=ministry_data,
            ai_allocations=ai_result["allocations"],
        )

        # =========================================================
        # 5) FORMAT OUTPUT FOR UI TIMEFRAME
        # =========================================================
        multiplier = Decimal("1") if timeframe == "monthly" else Decimal("12")
        display_unit = "monthly" if timeframe == "monthly" else "yearly"

        projected_bills_ui = (projected_bills * multiplier).quantize(Decimal("0.01"))
        affordable_pool_ui = (affordable_pool_monthly * multiplier).quantize(Decimal("0.01"))
        budget_need_ui = (_to_decimal(budget_need_monthly, "0.00") * multiplier).quantize(Decimal("0.01"))
        spend_need_ui = (spend_need_monthly * multiplier).quantize(Decimal("0.01"))
        baseline_need_ui = (baseline_need_monthly * multiplier).quantize(Decimal("0.01"))
        summed_ministry_need_ui = (summed_ministry_need_monthly * multiplier).quantize(Decimal("0.01"))
        need_total_ui = (need_total_monthly * multiplier).quantize(Decimal("0.01"))

        allocations_out = []
        for md in ministry_data:
            mid = str(md["id"])

            amt_m = _to_decimal(allocations_monthly.get(mid, 0.0), "0.00")
            min_req_m = _to_decimal(md["min_req"], "0.00")
            max_cap_m = _to_decimal(md["max_cap"], "0.00")

            amt_display = (amt_m * multiplier).quantize(Decimal("0.01"))
            min_req_display = (min_req_m * multiplier).quantize(Decimal("0.01"))
            max_cap_display = (max_cap_m * multiplier).quantize(Decimal("0.01"))

            allocations_out.append({
                "ministry_id": mid,
                "ministry_name": md["name"],

                # Display-safe fields
                "allocated_amount": float(amt_display),
                "min_req_display": float(min_req_display),
                "max_cap_display": float(max_cap_display),
                "display_unit": display_unit,

                # Raw monthly debug fields
                "allocated_amount_monthly": float(amt_m.quantize(Decimal("0.01"))),
                "min_req_monthly": float(min_req_m),
                "max_cap_monthly": float(max_cap_m),

                "priority_score": int(md["priority_score"]),
                "priority_source": md.get("priority_source", "unknown"),

                "spent_ytd": float(_to_decimal(md.get("spent_ytd", 0.0), "0.00")),
                "spent_signal_display": float(_to_decimal(md.get("spent_ytd", 0.0), "0.00")),

                "rationale": rationale_map.get(mid, ""),
            })

        reserve_amount_monthly = _to_decimal(reserve_amount_monthly, "0.00")
        reserve_amount_ui = (reserve_amount_monthly * multiplier).quantize(Decimal("0.01"))

        result = {
            "allocations": allocations_out,
            "meta": {
                "mode": "Historical" if use_history else "Manual",
                "reference_year": historical_year if use_history else None,
                "ui_timeframe": timeframe,
                "display_unit": display_unit,

                "optimizer_unit": "monthly",
                "optimizer_internal_unit": "monthly",
                "projected_bills_unit": "monthly_estimate",
                "affordable_capacity_unit": "current_capacity",

                "income_basis": income_basis,

                "pool_used_ui": float(pool_used_ui.quantize(Decimal("0.01"))),
                "pool_used_monthly": float(monthly_pool.quantize(Decimal("0.01"))),
                "pool_equivalent_yearly": float((monthly_pool * Decimal("12")).quantize(Decimal("0.01"))),

                "reserve_amount_monthly": float(reserve_amount_monthly.quantize(Decimal("0.01"))),
                "reserve_amount_ui": float(reserve_amount_ui),

                "projected_unrestricted_bills_ui": float(projected_bills_ui),
                "affordable_pool_ui": float(affordable_pool_ui),
                "budget_need_ui": float(budget_need_ui),
                "spend_need_ui": float(spend_need_ui),
                "baseline_need_ui": float(baseline_need_ui),
                "summed_ministry_need_ui": float(summed_ministry_need_ui),
                "need_total_ui": float(need_total_ui),

                "optimizer_source": "openai_responses_api",
                "optimizer_status": "success",
                "manual_override_count": manual_override_count,
                "ai_model_name": ai_result.get("model_name"),
                "openai_summary": ai_result.get("summary"),

                **manual_funds_meta,
            }
        }

        # =========================================================
        # 6) OPTIONAL LOGGING
        # =========================================================
        if not dry_run:
            try:
                priorities_used = {str(mm["id"]): int(mm["priority_score"]) for mm in ministry_data}
                ministry_constraints = {
                    str(mm["id"]): {
                        "min_req": float(mm["min_req"]),
                        "max_cap": float(mm["max_cap"]),
                        "spent_ytd": float(mm.get("spent_ytd", 0.0)),
                        "priority_score": int(mm["priority_score"]),
                        "priority_source": mm.get("priority_source", "unknown"),
                        "base_budget_monthly": float(mm.get("base_budget_monthly", 0.0)),
                        "spent_avg_monthly": float(mm.get("spent_avg_monthly", 0.0)),
                        "db_min_monthly": float(mm.get("db_min_monthly", 0.0)),
                        "db_max_monthly": float(mm.get("db_max_monthly", 0.0)),
                    }
                    for mm in ministry_data
                }
                allocations_result = {
                    a["ministry_id"]: float(a["allocated_amount"])
                    for a in allocations_out
                }

                OptimizerRunLog.objects.create(
                    church_id=church_id,
                    mode="Historical" if use_history else "Manual",
                    reference_year=historical_year if use_history else None,
                    timeframe=timeframe,
                    pool_used_monthly=monthly_pool.quantize(Decimal("0.01")),
                    priorities_used=priorities_used,
                    ministry_constraints=ministry_constraints,
                    allocations_result=allocations_result,
                    created_by=request.user,
                )
            except Exception as log_err:
                print("OpenAI OptimizerRunLog failed:", log_err)

        return JsonResponse(result)

    except Exception as e:
        print("--- OPENAI OPTIMIZER ERROR ---")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)

@method_decorator(ensure_csrf_cookie, name="dispatch")
class OpenAIPrescriptiveBudgetOptimizersView(PrescriptiveBudgetOptimizersView):
    """
    OpenAI page version of the existing optimizer page.

    Keeps the SAME output structure by reusing the parent view's
    full context, then only adds a few extra template variables.
    """

    # NEW TEMPLATE FILE
    template_name = "prescriptive_budgets_optimizer_openai.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # same page data, just extra flags for the new template
        context["optimizer_engine"] = "openai"
        context["optimizer_post_url"] = reverse("api_run_optimizer_openai")
        context["historical_stats_url"] = reverse("api_get_historical_stats")
        context["ministry_year_stats_url"] = reverse("api_get_ministry_stats_year")

        # optional UI labels
        context["page_title"] = "AI Budget Optimizer"
        context["page_heading"] = "AI Budget Optimizer"
        context["page_subheading"] = "Define constraints or use historical data to generate the optimal allocation."
        context["system_badge"] = "System Online"

        return context

class DenominationAdminHomeView(IsDenominationAdmin, TemplateView):
    template_name = "denomination_admin_dashboard.html"

    def _handle_user_without_denomination(self, request):
        messages.error(request, "No denomination is assigned to your account.")
        return HttpResponseForbidden("No denomination is assigned to your account.")

    def _get_latest_liquidation_requests_map(self, churches):
        """
        Get only the latest liquidation access request per church
        for the current denomination admin.
        """
        RequestModel = apps.get_model("Church", "DenominationLiquidationAccessRequest")

        request_rows = (
            RequestModel.objects
            .filter(
                denomination_admin=self.request.user,
                church__in=churches,
            )
            .select_related("church")
            .order_by("church_id", "-requested_at", "-id")
        )

        latest_by_church = {}
        for req in request_rows:
            if req.church_id not in latest_by_church:
                latest_by_church[req.church_id] = req

        return latest_by_church

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        if not getattr(request.user, "denomination", None):
            return self._handle_user_without_denomination(request)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ChurchApplication = apps.get_model("Church", "ChurchApplication")
        Church = apps.get_model("Church", "Church")
        RequestModel = apps.get_model("Church", "DenominationLiquidationAccessRequest")

        denomination = getattr(self.request.user, "denomination", None)

        if not denomination:
            context.update({
                "denomination_name": "",
                "kpi_pending_applications": 0,
                "kpi_revoked_applications": 0,
                "kpi_active_member_churches": 0,
                "kpi_liquidation_requested": 0,
                "kpi_liquidation_approved": 0,
                "kpi_liquidation_rejected": 0,
                "kpi_liquidation_closed": 0,
                "recent_pending_applications": [],
                "recent_revoked_applications": [],
                "recent_liquidation_requests": [],
            })
            return context

        # =========================================================
        # 1. CHURCH APPLICATION KPIs
        # =========================================================
        pending_applications = ChurchApplication.objects.filter(
            denomination=denomination,
            status="Pending",
        ).count()

        revoked_applications = ChurchApplication.objects.filter(
            denomination=denomination,
            status="Revoked",
        ).count()

        # =========================================================
        # 2. ACTIVE MEMBER CHURCHES
        # =========================================================
        churches = Church.objects.filter(
            denomination=denomination
        ).order_by("name")

        active_member_churches = churches.count()

        # =========================================================
        # 3. LIQUIDATION ACCESS KPIs
        # Requested = Pending only
        # Approved = Approved only
        # based on latest request per church
        # =========================================================
        latest_request_by_church = self._get_latest_liquidation_requests_map(churches)

        liquidation_requested_count = 0
        liquidation_approved_count = 0
        liquidation_rejected_count = 0
        liquidation_closed_count = 0

        for req in latest_request_by_church.values():
            if req.status == RequestModel.STATUS_PENDING:
                liquidation_requested_count += 1
            elif req.status == RequestModel.STATUS_APPROVED:
                liquidation_approved_count += 1
            elif req.status == RequestModel.STATUS_REJECTED:
                liquidation_rejected_count += 1
            elif req.status == RequestModel.STATUS_CLOSED:
                liquidation_closed_count += 1

        # =========================================================
        # 4. RECENT DASHBOARD LISTS
        # =========================================================
        recent_pending_applications = (
            ChurchApplication.objects.filter(
                denomination=denomination,
                status="Pending",
            )
            .select_related("church", "decided_by", "denomination")
            .order_by("-applied_at")[:5]
        )

        recent_revoked_applications = (
            ChurchApplication.objects.filter(
                denomination=denomination,
                status="Revoked",
            )
            .select_related("church", "decided_by", "denomination")
            .order_by("-decided_at", "-applied_at")[:5]
        )

        recent_liquidation_requests = sorted(
            latest_request_by_church.values(),
            key=lambda x: (
                x.requested_at or timezone.now(),
                x.id or 0
            ),
            reverse=True
        )[:5]

        # =========================================================
        # 5. CONTEXT
        # =========================================================
        context.update({
            "denomination_name": getattr(denomination, "name", ""),
            "kpi_pending_applications": pending_applications,
            "kpi_revoked_applications": revoked_applications,
            "kpi_active_member_churches": active_member_churches,

            # Pending only
            "kpi_liquidation_requested": liquidation_requested_count,

            # Approved only
            "kpi_liquidation_approved": liquidation_approved_count,

            "kpi_liquidation_rejected": liquidation_rejected_count,
            "kpi_liquidation_closed": liquidation_closed_count,

            "recent_pending_applications": recent_pending_applications,
            "recent_revoked_applications": recent_revoked_applications,
            "recent_liquidation_requests": recent_liquidation_requests,
        })
        return context